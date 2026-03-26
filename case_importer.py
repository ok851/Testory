"""
用例导入导出管理器
支持 Excel 和 JSON 格式的用例导入导出
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from database import Database

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CaseImportExporter:
    """用例导入导出器"""
    
    def __init__(self, db: Database = None):
        self.db = db if db else Database()
        self.export_dir = "exports"
        
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
    
    # ==================== Excel 导出 ====================
    
    def export_cases_to_excel(self, project_id: int = None, case_ids: List[int] = None, 
                               filename: str = None) -> str:
        """导出用例到Excel文件
        
        Args:
            project_id: 项目ID，导出该项目下所有用例
            case_ids: 用例ID列表，指定导出哪些用例
            filename: 文件名（不含扩展名）
        
        Returns:
            导出文件的完整路径
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        # 获取用例数据
        cases = self._get_cases_for_export(project_id, case_ids)
        
        if not cases:
            raise ValueError("没有找到要导出的用例")
        
        # 创建工作簿
        wb = Workbook()
        
        # 用例Sheet
        ws_cases = wb.active
        ws_cases.title = "测试用例"
        self._write_cases_sheet(ws_cases, cases)
        
        # 步骤Sheet
        ws_steps = wb.create_sheet("测试步骤")
        self._write_steps_sheet(ws_steps, cases)
        
        # 保存文件
        if not filename:
            filename = f"cases_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.export_dir, f"{filename}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def _get_cases_for_export(self, project_id: int = None, 
                               case_ids: List[int] = None) -> List[Dict]:
        """获取要导出的用例"""
        cases = []
        
        if case_ids:
            for case_id in case_ids:
                case = self.db.get_test_case_v2(case_id)
                if case:
                    case['steps'] = self.db.get_case_steps(case_id)
                    cases.append(case)
        elif project_id:
            project_cases = self.db.get_project_cases(project_id)
            for case in project_cases:
                case['steps'] = self.db.get_case_steps(case['id'])
                cases.append(case)
        
        return cases
    
    def _write_cases_sheet(self, ws, cases: List[Dict]):
        """写入用例Sheet"""
        # 表头
        headers = ['用例ID', '用例名称', '测试URL', '描述', '前置条件', '预期结果', '项目ID', '创建时间']
        
        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=case.get('id', '')).border = thin_border
            ws.cell(row=row, column=2, value=case.get('name', '')).border = thin_border
            ws.cell(row=row, column=3, value=case.get('url', '')).border = thin_border
            ws.cell(row=row, column=4, value=case.get('description', '')).border = thin_border
            ws.cell(row=row, column=5, value=case.get('precondition', '')).border = thin_border
            ws.cell(row=row, column=6, value=case.get('expected_result', '')).border = thin_border
            ws.cell(row=row, column=7, value=case.get('project_id', '')).border = thin_border
            ws.cell(row=row, column=8, value=case.get('created_at', '')).border = thin_border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
    
    def _write_steps_sheet(self, ws, cases: List[Dict]):
        """写入步骤Sheet"""
        headers = ['用例ID', '步骤序号', '操作类型', '定位方式', '定位值', '输入值', 
                   '描述', '页面名称', 'URL', '对比方式', '进入iframe', 'iframe选择器']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        row = 2
        for case in cases:
            for step in case.get('steps', []):
                ws.cell(row=row, column=1, value=case.get('id', '')).border = thin_border
                ws.cell(row=row, column=2, value=step.get('step_order', '')).border = thin_border
                ws.cell(row=row, column=3, value=step.get('action', '')).border = thin_border
                ws.cell(row=row, column=4, value=step.get('selector_type', '')).border = thin_border
                ws.cell(row=row, column=5, value=step.get('selector_value', '')).border = thin_border
                ws.cell(row=row, column=6, value=step.get('input_value', '')).border = thin_border
                ws.cell(row=row, column=7, value=step.get('description', '')).border = thin_border
                ws.cell(row=row, column=8, value=step.get('page_name', '')).border = thin_border
                ws.cell(row=row, column=9, value=step.get('url', '')).border = thin_border
                ws.cell(row=row, column=10, value=step.get('compare_type', 'equals')).border = thin_border
                ws.cell(row=row, column=11, value='是' if step.get('enter_iframe') else '否').border = thin_border
                ws.cell(row=row, column=12, value=step.get('iframe_selector', '')).border = thin_border
                row += 1
        
        # 调整列宽
        widths = [10, 10, 15, 10, 40, 30, 30, 15, 40, 10, 10, 30]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def generate_excel_template(self, filename: str = None) -> str:
        """生成Excel导入模板"""
        if not EXCEL_AVAILABLE:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        wb = Workbook()
        
        # 用例Sheet
        ws_cases = wb.active
        ws_cases.title = "测试用例"
        
        headers = ['用例名称*', '测试URL', '描述', '前置条件', '预期结果', '项目ID*']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws_cases.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 示例数据
        ws_cases.cell(row=2, column=1, value="登录功能测试")
        ws_cases.cell(row=2, column=2, value="https://example.com/login")
        ws_cases.cell(row=2, column=3, value="测试用户登录功能")
        ws_cases.cell(row=2, column=4, value="用户已注册")
        ws_cases.cell(row=2, column=5, value="登录成功跳转首页")
        ws_cases.cell(row=2, column=6, value="1")
        
        # 步骤Sheet
        ws_steps = wb.create_sheet("测试步骤")
        
        step_headers = ['用例名称*', '步骤序号*', '操作类型*', '定位方式', '定位值', 
                        '输入值', '描述', '页面名称', 'URL', '对比方式']
        
        header_fill2 = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
        
        for col, header in enumerate(step_headers, 1):
            cell = ws_steps.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill2
        
        # 示例步骤
        example_steps = [
            ["登录功能测试", 1, "navigate", "", "", "", "打开登录页", "", "https://example.com/login", ""],
            ["登录功能测试", 2, "input", "css", "#username", "testuser", "输入用户名", "登录页", "", ""],
            ["登录功能测试", 3, "input", "css", "#password", "password123", "输入密码", "登录页", "", ""],
            ["登录功能测试", 4, "click", "css", "#login-btn", "", "点击登录", "登录页", "", ""],
            ["登录功能测试", 5, "extract_text", "css", ".welcome", "欢迎", "验证登录成功", "首页", "", "contains"],
        ]
        
        for row, step in enumerate(example_steps, 2):
            for col, value in enumerate(step, 1):
                ws_steps.cell(row=row, column=col, value=value)
        
        # 说明Sheet
        ws_help = wb.create_sheet("使用说明")
        instructions = [
            "=== 用例导入模板使用说明 ===",
            "",
            "【测试用例 Sheet】",
            "- 用例名称*: 必填，用例的名称",
            "- 测试URL: 用例起始URL",
            "- 描述: 用例描述",
            "- 前置条件: 执行用例前的条件",
            "- 预期结果: 用例执行的预期结果",
            "- 项目ID*: 必填，用例所属项目的ID",
            "",
            "【测试步骤 Sheet】",
            "- 用例名称*: 必填，与用例Sheet中的名称对应",
            "- 步骤序号*: 必填，执行顺序",
            "- 操作类型*: 必填，可选值：",
            "  - navigate: 导航到URL",
            "  - click: 点击元素",
            "  - input: 输入文本",
            "  - hover: 悬停",
            "  - double_click: 双击",
            "  - right_click: 右键点击",
            "  - wait: 等待",
            "  - scroll: 滚动",
            "  - extract_text: 提取文本并验证",
            "  - verify: 验证操作",
            "  - select: 下拉选择",
            "  - assert: 断言操作",
            "- 定位方式: css 或 xpath",
            "- 定位值: 元素选择器",
            "- 输入值: 输入的文本或预期值",
            "- 描述: 步骤描述",
            "- 对比方式: equals/not_equals/contains (用于extract_text)",
        ]
        
        for row, text in enumerate(instructions, 1):
            ws_help.cell(row=row, column=1, value=text)
        
        ws_help.column_dimensions['A'].width = 60
        
        if not filename:
            filename = "case_import_template"
        
        filepath = os.path.join(self.export_dir, f"{filename}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    # ==================== Excel 导入 ====================
    
    def import_cases_from_excel(self, filepath: str, project_id: int = None) -> Dict[str, Any]:
        """从Excel文件导入用例
        
        Args:
            filepath: Excel文件路径
            project_id: 默认项目ID（如果Excel中未指定）
        
        Returns:
            导入结果统计
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        
        results = {
            'success': True,
            'cases_created': 0,
            'steps_created': 0,
            'errors': [],
            'case_mapping': {}  # 用例名称 -> 用例ID 的映射
        }
        
        try:
            # 导入用例
            if '测试用例' in wb.sheetnames:
                ws_cases = wb['测试用例']
                self._import_cases_from_sheet(ws_cases, project_id, results)
            
            # 导入步骤
            if '测试步骤' in wb.sheetnames:
                ws_steps = wb['测试步骤']
                self._import_steps_from_sheet(ws_steps, results)
        finally:
            wb.close()
        
        return results
    
    def _import_cases_from_sheet(self, ws, default_project_id: int, results: Dict):
        """从Sheet导入用例"""
        headers = [cell.value for cell in ws[1]]
        
        # 找到各列的索引
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                h_lower = h.lower().replace('*', '').strip()
                if '名称' in h_lower or 'name' in h_lower:
                    col_map['name'] = i
                elif 'url' in h_lower:
                    col_map['url'] = i
                elif '描述' in h_lower or 'desc' in h_lower:
                    col_map['description'] = i
                elif '前置' in h_lower or 'precondition' in h_lower:
                    col_map['precondition'] = i
                elif '预期' in h_lower or 'expected' in h_lower:
                    col_map['expected_result'] = i
                elif '项目' in h_lower or 'project' in h_lower:
                    col_map['project_id'] = i
        
        for row_num in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_num]]
            
            # 跳过空行
            if not any(row_data):
                continue
            
            try:
                name = row_data[col_map.get('name', 0)] if col_map.get('name') is not None else None
                if not name:
                    results['errors'].append(f"行 {row_num}: 用例名称不能为空")
                    continue
                
                project_id = row_data[col_map.get('project_id', 5)] if col_map.get('project_id') is not None else default_project_id
                if not project_id:
                    results['errors'].append(f"行 {row_num}: 项目ID不能为空")
                    continue
                
                case_id = self.db.create_test_case_v2(
                    project_id=int(project_id),
                    name=str(name),
                    url=str(row_data[col_map.get('url', 1)] or '') if col_map.get('url') is not None else '',
                    description=str(row_data[col_map.get('description', 2)] or '') if col_map.get('description') is not None else '',
                    precondition=str(row_data[col_map.get('precondition', 3)] or '') if col_map.get('precondition') is not None else '',
                    expected_result=str(row_data[col_map.get('expected_result', 4)] or '') if col_map.get('expected_result') is not None else ''
                )
                
                results['cases_created'] += 1
                results['case_mapping'][str(name)] = case_id
                
            except Exception as e:
                results['errors'].append(f"行 {row_num}: {str(e)}")
    
    def _import_steps_from_sheet(self, ws, results: Dict):
        """从Sheet导入步骤"""
        headers = [cell.value for cell in ws[1]]
        
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                h_lower = h.lower().replace('*', '').strip()
                if '用例' in h_lower and '名称' in h_lower:
                    col_map['case_name'] = i
                elif '序号' in h_lower or 'order' in h_lower:
                    col_map['step_order'] = i
                elif '操作' in h_lower or 'action' in h_lower:
                    col_map['action'] = i
                elif '定位方式' in h_lower or 'selector_type' in h_lower:
                    col_map['selector_type'] = i
                elif '定位值' in h_lower or 'selector_value' in h_lower:
                    col_map['selector_value'] = i
                elif '输入' in h_lower or 'input' in h_lower:
                    col_map['input_value'] = i
                elif '描述' in h_lower or 'desc' in h_lower:
                    col_map['description'] = i
                elif '页面' in h_lower or 'page' in h_lower:
                    col_map['page_name'] = i
                elif 'url' in h_lower:
                    col_map['url'] = i
                elif '对比' in h_lower or 'compare' in h_lower:
                    col_map['compare_type'] = i
        
        for row_num in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_num]]
            
            if not any(row_data):
                continue
            
            try:
                case_name = row_data[col_map.get('case_name', 0)] if col_map.get('case_name') is not None else None
                if not case_name:
                    results['errors'].append(f"步骤行 {row_num}: 用例名称不能为空")
                    continue
                
                case_id = results['case_mapping'].get(str(case_name))
                if not case_id:
                    results['errors'].append(f"步骤行 {row_num}: 找不到用例 '{case_name}'")
                    continue
                
                action = row_data[col_map.get('action', 2)] if col_map.get('action') is not None else None
                if not action:
                    results['errors'].append(f"步骤行 {row_num}: 操作类型不能为空")
                    continue
                
                self.db.create_test_step(
                    case_id=case_id,
                    action=str(action),
                    selector_type=str(row_data[col_map.get('selector_type', 3)] or '') if col_map.get('selector_type') is not None else '',
                    selector_value=str(row_data[col_map.get('selector_value', 4)] or '') if col_map.get('selector_value') is not None else '',
                    input_value=str(row_data[col_map.get('input_value', 5)] or '') if col_map.get('input_value') is not None else '',
                    description=str(row_data[col_map.get('description', 6)] or '') if col_map.get('description') is not None else '',
                    step_order=int(row_data[col_map.get('step_order', 1)] or 0) if col_map.get('step_order') is not None else None,
                    page_name=str(row_data[col_map.get('page_name', 7)] or '') if col_map.get('page_name') is not None else '',
                    url=str(row_data[col_map.get('url', 8)] or '') if col_map.get('url') is not None else '',
                    compare_type=str(row_data[col_map.get('compare_type', 9)] or 'equals') if col_map.get('compare_type') is not None else 'equals'
                )
                
                results['steps_created'] += 1
                
            except Exception as e:
                results['errors'].append(f"步骤行 {row_num}: {str(e)}")
    
    # ==================== JSON 导出 ====================
    
    def export_cases_to_json(self, project_id: int = None, case_ids: List[int] = None,
                              filename: str = None) -> str:
        """导出用例到JSON文件"""
        cases = self._get_cases_for_export(project_id, case_ids)
        
        if not cases:
            raise ValueError("没有找到要导出的用例")
        
        # 构建导出数据
        export_data = {
            'version': '1.0',
            'export_time': datetime.now().isoformat(),
            'cases': []
        }
        
        for case in cases:
            case_data = {
                'name': case.get('name', ''),
                'url': case.get('url', ''),
                'description': case.get('description', ''),
                'precondition': case.get('precondition', ''),
                'expected_result': case.get('expected_result', ''),
                'steps': []
            }
            
            for step in case.get('steps', []):
                step_data = {
                    'step_order': step.get('step_order', 0),
                    'action': step.get('action', ''),
                    'selector_type': step.get('selector_type', ''),
                    'selector_value': step.get('selector_value', ''),
                    'input_value': step.get('input_value', ''),
                    'description': step.get('description', ''),
                    'page_name': step.get('page_name', ''),
                    'url': step.get('url', ''),
                    'compare_type': step.get('compare_type', 'equals'),
                    'enter_iframe': step.get('enter_iframe', False),
                    'iframe_selector': step.get('iframe_selector', '')
                }
                case_data['steps'].append(step_data)
            
            export_data['cases'].append(case_data)
        
        if not filename:
            filename = f"cases_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.export_dir, f"{filename}.json")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        return filepath
    
    # ==================== JSON 导入 ====================
    
    def import_cases_from_json(self, filepath: str = None, json_data: str = None,
                                project_id: int = None) -> Dict[str, Any]:
        """从JSON导入用例
        
        Args:
            filepath: JSON文件路径
            json_data: JSON字符串（二选一）
            project_id: 目标项目ID
        
        Returns:
            导入结果统计
        """
        if not project_id:
            raise ValueError("项目ID不能为空")
        
        if filepath:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
        elif json_data:
            data = json.loads(json_data)
        else:
            raise ValueError("必须提供文件路径或JSON数据")
        
        results = {
            'success': True,
            'cases_created': 0,
            'steps_created': 0,
            'errors': []
        }
        
        cases = data.get('cases', [])
        
        for i, case_data in enumerate(cases):
            try:
                name = case_data.get('name')
                if not name:
                    results['errors'].append(f"用例 {i+1}: 名称不能为空")
                    continue
                
                case_id = self.db.create_test_case_v2(
                    project_id=project_id,
                    name=name,
                    url=case_data.get('url', ''),
                    description=case_data.get('description', ''),
                    precondition=case_data.get('precondition', ''),
                    expected_result=case_data.get('expected_result', '')
                )
                
                results['cases_created'] += 1
                
                # 导入步骤
                for j, step_data in enumerate(case_data.get('steps', [])):
                    try:
                        action = step_data.get('action')
                        if not action:
                            results['errors'].append(f"用例 {name} 步骤 {j+1}: 操作类型不能为空")
                            continue
                        
                        self.db.create_test_step(
                            case_id=case_id,
                            action=action,
                            selector_type=step_data.get('selector_type', ''),
                            selector_value=step_data.get('selector_value', ''),
                            input_value=step_data.get('input_value', ''),
                            description=step_data.get('description', ''),
                            step_order=step_data.get('step_order'),
                            page_name=step_data.get('page_name', ''),
                            url=step_data.get('url', ''),
                            compare_type=step_data.get('compare_type', 'equals'),
                            enter_iframe=step_data.get('enter_iframe', False),
                            iframe_selector=step_data.get('iframe_selector', '')
                        )
                        
                        results['steps_created'] += 1
                        
                    except Exception as e:
                        results['errors'].append(f"用例 {name} 步骤 {j+1}: {str(e)}")
                
            except Exception as e:
                results['errors'].append(f"用例 {i+1}: {str(e)}")
        
        return results
    
    def generate_json_template(self) -> str:
        """生成JSON导入模板"""
        template = {
            "version": "1.0",
            "description": "用例导入模板",
            "cases": [
                {
                    "name": "登录功能测试",
                    "url": "https://example.com/login",
                    "description": "测试用户登录功能",
                    "precondition": "用户已注册",
                    "expected_result": "登录成功跳转首页",
                    "steps": [
                        {
                            "step_order": 1,
                            "action": "navigate",
                            "selector_type": "",
                            "selector_value": "",
                            "input_value": "",
                            "description": "打开登录页",
                            "page_name": "登录页",
                            "url": "https://example.com/login",
                            "compare_type": "equals"
                        },
                        {
                            "step_order": 2,
                            "action": "input",
                            "selector_type": "css",
                            "selector_value": "#username",
                            "input_value": "testuser",
                            "description": "输入用户名",
                            "page_name": "登录页",
                            "url": "",
                            "compare_type": "equals"
                        },
                        {
                            "step_order": 3,
                            "action": "click",
                            "selector_type": "css",
                            "selector_value": "#login-btn",
                            "input_value": "",
                            "description": "点击登录按钮",
                            "page_name": "登录页",
                            "url": "",
                            "compare_type": "equals"
                        },
                        {
                            "step_order": 4,
                            "action": "extract_text",
                            "selector_type": "css",
                            "selector_value": ".welcome-msg",
                            "input_value": "欢迎",
                            "description": "验证登录成功",
                            "page_name": "首页",
                            "url": "",
                            "compare_type": "contains"
                        }
                    ]
                }
            ]
        }
        
        return json.dumps(template, ensure_ascii=False, indent=2)


# 单例实例
case_importer = CaseImportExporter()
