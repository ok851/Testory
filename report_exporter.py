from test_report import TestReportGenerator
from typing import Dict, List, Any
from datetime import datetime
import os


class ReportExporter:
    """报告导出器"""
    
    def __init__(self, report_generator: TestReportGenerator = None):
        self.report_generator = report_generator if report_generator else TestReportGenerator()
        self.export_dir = "exports"
        
        # 确保导出目录存在
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
    
    def export_to_html(self, data: Dict[str, Any], filename: str = None) -> str:
        """导出报告为HTML格式
        
        Args:
            data: 报告数据
            filename: 文件名，不包含扩展名
        
        Returns:
            导出文件的完整路径
        """
        if not filename:
            filename = f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.export_dir, f"{filename}.html")
        
        html_content = self._generate_html_report(data)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return filepath
    
    def export_to_excel(self, data: Dict[str, Any], filename: str = None) -> str:
        """导出报告为Excel格式
        
        Args:
            data: 报告数据
            filename: 文件名，不包含扩展名
        
        Returns:
            导出文件的完整路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("请安装 openpyxl 库: pip install openpyxl")
        
        if not filename:
            filename = f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.export_dir, f"{filename}.xlsx")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试报告"
        
        # 定义样式 - 使用更美观的配色
        title_font = Font(name='微软雅黑', size=20, bold=True, color='1890FF')
        section_font = Font(name='微软雅黑', size=14, bold=True, color='333333')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        normal_font = Font(name='微软雅黑', size=10, color='333333')
        number_font = Font(name='微软雅黑', size=11, bold=True, color='1890FF')
        
        # 填充色
        primary_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
        success_fill = PatternFill(start_color='52C41A', end_color='52C41A', fill_type='solid')
        error_fill = PatternFill(start_color='FF4D4F', end_color='FF4D4F', fill_type='solid')
        header_fill = PatternFill(start_color='F0F2F5', end_color='F0F2F5', fill_type='solid')
        alt_row_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        card_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # 对齐方式
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 边框
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # 写入主标题
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = 'UI自动化测试报告'
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 35
        
        # 写入生成时间
        ws.merge_cells('A2:G2')
        time_cell = ws['A2']
        time_cell.value = f'生成时间: {datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}'
        time_cell.font = Font(name='微软雅黑', size=10, color='999999')
        time_cell.alignment = center_align
        ws.row_dimensions[2].height = 20
        
        current_row = 4
        overview = data.get('overview', {})
        
        # 统计概览 - 使用卡片式布局
        ws[f'A{current_row}'] = '统计概览'
        ws[f'A{current_row}'].font = section_font
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 1
        
        # 创建5个统计卡片
        card_row = current_row
        ws.row_dimensions[card_row].height = 60
        
        stats = [
            ('总执行次数', overview.get('total_runs', 0), '1890FF'),
            ('通过次数', overview.get('passed_runs', 0), '52C41A'),
            ('失败次数', overview.get('failed_runs', 0), 'FF4D4F'),
            ('通过率', f"{overview.get('pass_rate', 0):.1f}%", '1890FF'),
            ('平均耗时', f"{overview.get('avg_duration', 0):.1f}s", '1890FF')
        ]
        
        col_width = 16
        for i, (label, value, color) in enumerate(stats):
            col_start = i * 2 + 1
            col_end = col_start + 1
            
            # 合并单元格作为卡片
            ws.merge_cells(start_row=card_row, start_column=col_start, end_row=card_row, end_column=col_end)
            cell = ws.cell(row=card_row, column=col_start)
            cell.value = f'{label}\n{value}'
            cell.font = Font(name='微软雅黑', size=10, color=color, bold=True)
            cell.alignment = center_align
            cell.fill = card_fill
            cell.border = thin_border
        
        current_row += 2
        
        # 状态分布表格
        status_dist = data.get('status_distribution', [])
        if status_dist:
            ws[f'A{current_row}'] = '状态分布'
            ws[f'A{current_row}'].font = section_font
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1
            
            # 表头
            headers = ['状态', '次数', '占比']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            # 数据行
            total = overview.get('total_runs', 1)
            for i, item in enumerate(status_dist):
                status = item.get('status', '未知')
                count = item.get('count', 0)
                percentage = (count / total * 100) if total > 0 else 0
                status_label = '通过' if status == 'passed' else '失败' if status == 'failed' else status
                
                values = [status_label, count, f"{percentage:.2f}%"]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = val
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    if i % 2 == 1:
                        cell.fill = alt_row_fill
                current_row += 1
            current_row += 1
        
        # 耗时分布表格
        duration_dist_dict = data.get('duration_distribution', {})
        duration_dist = duration_dist_dict.get('distribution', []) if isinstance(duration_dist_dict, dict) else []
        if duration_dist:
            ws[f'A{current_row}'] = '耗时分布'
            ws[f'A{current_row}'].font = section_font
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1
            
            headers = ['耗时区间', '用例数', '占比']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            for i, item in enumerate(duration_dist):
                range_label = item.get('range', '未知')
                count = item.get('count', 0)
                percentage = (count / total * 100) if total > 0 else 0
                
                values = [range_label, count, f"{percentage:.2f}%"]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = val
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    if i % 2 == 1:
                        cell.fill = alt_row_fill
                current_row += 1
            current_row += 1
        
        # 用例统计表格
        case_stats = data.get('case_statistics', [])
        if case_stats:
            ws[f'A{current_row}'] = '用例统计'
            ws[f'A{current_row}'].font = section_font
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1
            
            headers = ['用例名称', '执行次数', '通过', '失败', '通过率', '平均耗时']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            for i, item in enumerate(case_stats):
                case_name = item.get('case_name', item.get('name', '未知'))
                values = [
                    case_name,
                    item.get('total_runs', 0),
                    item.get('passed_runs', 0),
                    item.get('failed_runs', 0),
                    f"{item.get('pass_rate', 0):.1f}%",
                    f"{item.get('avg_duration', 0):.1f}s"
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = val
                    cell.font = normal_font
                    cell.alignment = left_align if col == 1 else center_align
                    cell.border = thin_border
                    if i % 2 == 1:
                        cell.fill = alt_row_fill
                current_row += 1
            current_row += 1
        
        # 项目统计表格
        project_stats = data.get('project_statistics', [])
        if project_stats:
            ws[f'A{current_row}'] = '项目统计'
            ws[f'A{current_row}'].font = section_font
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1
            
            headers = ['项目名称', '用例数', '执行次数', '通过', '失败', '通过率']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            for i, item in enumerate(project_stats):
                project_name = item.get('project_name', item.get('name', '未知'))
                values = [
                    project_name,
                    item.get('total_cases', 0),
                    item.get('total_runs', 0),
                    item.get('passed_runs', 0),
                    item.get('failed_runs', 0),
                    f"{item.get('pass_rate', 0):.1f}%"
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = val
                    cell.font = normal_font
                    cell.alignment = left_align if col == 1 else center_align
                    cell.border = thin_border
                    if i % 2 == 1:
                        cell.fill = alt_row_fill
                current_row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 35  # 名称列
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # 冻结首行
        ws.freeze_panes = 'A5'
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def export_to_pdf(self, data: Dict[str, Any], filename: str = None) -> str:
        """导出报告为PDF格式
        
        Args:
            data: 报告数据
            filename: 文件名，不包含扩展名
        
        Returns:
            导出文件的完整路径
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.colors import HexColor, white, black
        
        if not filename:
            filename = "test_report_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        
        filepath = os.path.join(self.export_dir, filename + ".pdf")
        
        # 注册中文字体
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
        try:
            # 尝试使用Windows系统自带的微软雅黑字体
            font_paths = [
                "C:\\Windows\\Fonts\\msyh.ttc",
                "C:\\Windows\\Fonts\\msyh.ttf",
                "C:\\Windows\\Fonts\\simhei.ttf",
                "C:\\Windows\\Fonts\\simsun.ttc"
            ]
            for font_path in font_paths:
                if os.path.exists(font_path):
                    font_name = 'ChineseFont'
                    font_name_bold = 'ChineseFont'
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    break
        except Exception as e:
            print(f"注册中文字体失败: {e}")
        
        # 创建PDF画布
        c = canvas.Canvas(filepath, pagesize=A4)
        width, height = A4
        
        # 定义颜色
        primary_color = HexColor('#1890ff')
        success_color = HexColor('#52c41a')
        error_color = HexColor('#ff4d4f')
        header_bg = HexColor('#f0f2f5')
        border_color = HexColor('#d9d9d9')
        
        def draw_header():
            """绘制页眉"""
            c.setFillColor(primary_color)
            c.rect(0, height - 1.5*cm, width, 1.5*cm, fill=1, stroke=0)
            c.setFillColor(white)
            c.setFont(font_name_bold, 16)
            c.drawCentredString(width/2, height - 1*cm, 'UI自动化测试报告')
        
        def draw_footer(page_num):
            """绘制页脚"""
            c.setStrokeColor(border_color)
            c.line(2*cm, 1.5*cm, width - 2*cm, 1.5*cm)
            c.setFillColor(HexColor('#999999'))
            c.setFont(font_name, 9)
            c.drawString(2*cm, 1*cm, f'生成时间: {datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}')
            c.drawRightString(width - 2*cm, 1*cm, f'第 {page_num} 页')
        
        def draw_card(x, y, w, h, label, value, color):
            """绘制统计卡片"""
            # 卡片背景
            c.setFillColor(HexColor('#f8f9fa'))
            c.roundRect(x, y - h, w, h, 5, fill=1, stroke=0)
            # 顶部色条
            c.setFillColor(color)
            c.roundRect(x, y - 0.3*cm, w, 0.3*cm, 2, fill=1, stroke=0)
            # 标签
            c.setFillColor(HexColor('#666666'))
            c.setFont(font_name, 10)
            c.drawCentredString(x + w/2, y - 0.8*cm, label)
            # 数值
            c.setFillColor(color)
            c.setFont(font_name_bold, 20)
            c.drawCentredString(x + w/2, y - 1.5*cm, str(value))
        
        def draw_table_header(cols, y):
            """绘制表格表头"""
            c.setFillColor(header_bg)
            c.setStrokeColor(border_color)
            row_height = 0.6*cm
            c.rect(2*cm, y - row_height, width - 4*cm, row_height, fill=1, stroke=1)
            c.setFillColor(black)
            c.setFont(font_name_bold, 10)
            col_width = (width - 4*cm) / len(cols)
            for i, col in enumerate(cols):
                c.drawCentredString(2*cm + i*col_width + col_width/2, y - 0.4*cm, col)
            return y - row_height
        
        def draw_table_row(values, y, is_alt=False):
            """绘制表格行"""
            row_height = 0.5*cm
            if is_alt:
                c.setFillColor(HexColor('#fafafa'))
                c.rect(2*cm, y - row_height, width - 4*cm, row_height, fill=1, stroke=0)
            c.setStrokeColor(border_color)
            c.line(2*cm, y - row_height, width - 2*cm, y - row_height)
            c.setFillColor(black)
            c.setFont(font_name, 9)
            col_width = (width - 4*cm) / len(values)
            for i, val in enumerate(values):
                c.drawCentredString(2*cm + i*col_width + col_width/2, y - 0.35*cm, str(val))
            return y - row_height
        
        # 开始绘制第一页
        page_num = 1
        draw_header()
        
        y = height - 3*cm
        
        # 统计概览 - 卡片布局
        c.setFillColor(black)
        c.setFont(font_name_bold, 14)
        c.drawString(2*cm, y, '统计概览')
        y -= 0.5*cm
        
        overview = data.get('overview', {})
        card_width = (width - 4.8*cm) / 5
        card_height = 2*cm
        card_y = y - card_height
        
        draw_card(2*cm, card_y + card_height, card_width, card_height, 
                  '总执行次数', overview.get('total_runs', 0), primary_color)
        draw_card(2*cm + card_width + 0.2*cm, card_y + card_height, card_width, card_height,
                  '通过次数', overview.get('passed_runs', 0), success_color)
        draw_card(2*cm + 2*(card_width + 0.2*cm), card_y + card_height, card_width, card_height,
                  '失败次数', overview.get('failed_runs', 0), error_color)
        draw_card(2*cm + 3*(card_width + 0.2*cm), card_y + card_height, card_width, card_height,
                  '通过率', f"{overview.get('pass_rate', 0):.1f}%", primary_color)
        draw_card(2*cm + 4*(card_width + 0.2*cm), card_y + card_height, card_width, card_height,
                  '平均耗时', f"{overview.get('avg_duration', 0):.1f}s", primary_color)
        
        y = card_y - 0.8*cm
        
        # 检查是否需要新页
        def check_new_page(needed_height):
            nonlocal y, page_num
            if y - needed_height < 2.5*cm:
                draw_footer(page_num)
                c.showPage()
                page_num += 1
                draw_header()
                y = height - 3*cm
                return True
            return False
        
        # 状态分布表格
        check_new_page(3*cm)
        c.setFillColor(black)
        c.setFont(font_name_bold, 14)
        c.drawString(2*cm, y, '状态分布')
        y -= 0.6*cm
        
        status_dist = data.get('status_distribution', [])
        total = overview.get('total_runs', 1)
        y = draw_table_header(['状态', '次数', '占比'], y)
        for i, item in enumerate(status_dist):
            status = item.get('status', '未知')
            count = item.get('count', 0)
            percentage = (count / total * 100) if total > 0 else 0
            status_label = '通过' if status == 'passed' else '失败' if status == 'failed' else status
            y = draw_table_row([status_label, count, f"{percentage:.2f}%"], y, i % 2 == 1)
        y -= 0.5*cm
        
        # 耗时分布表格
        check_new_page(3*cm)
        c.setFillColor(black)
        c.setFont(font_name_bold, 14)
        c.drawString(2*cm, y, '耗时分布')
        y -= 0.6*cm
        
        duration_dist_dict = data.get('duration_distribution', {})
        duration_dist = duration_dist_dict.get('distribution', []) if isinstance(duration_dist_dict, dict) else []
        y = draw_table_header(['耗时区间', '用例数', '占比'], y)
        for i, item in enumerate(duration_dist):
            range_label = item.get('range', '未知')
            count = item.get('count', 0)
            percentage = (count / total * 100) if total > 0 else 0
            y = draw_table_row([range_label, count, f"{percentage:.2f}%"], y, i % 2 == 1)
        y -= 0.5*cm
        
        # 用例统计表格
        case_stats = data.get('case_statistics', [])
        if case_stats:
            check_new_page(3*cm)
            c.setFillColor(black)
            c.setFont(font_name_bold, 14)
            c.drawString(2*cm, y, '用例统计')
            y -= 0.6*cm
            
            y = draw_table_header(['用例名称', '执行次数', '通过', '失败', '通过率', '平均耗时'], y)
            for i, item in enumerate(case_stats):
                case_name = item.get('case_name', item.get('name', '未知'))
                if len(case_name) > 20:
                    case_name = case_name[:17] + '...'
                total_runs = item.get('total_runs', 0)
                passed = item.get('passed_runs', 0)
                failed = item.get('failed_runs', 0)
                pass_rate = item.get('pass_rate', 0)
                avg_dur = item.get('avg_duration', 0)
                y = draw_table_row([
                    case_name, total_runs, passed, failed,
                    f"{pass_rate:.1f}%", f"{avg_dur:.1f}s"
                ], y, i % 2 == 1)
            y -= 0.5*cm
        
        # 项目统计表格
        project_stats = data.get('project_statistics', [])
        if project_stats:
            check_new_page(3*cm)
            c.setFillColor(black)
            c.setFont(font_name_bold, 14)
            c.drawString(2*cm, y, '项目统计')
            y -= 0.6*cm
            
            y = draw_table_header(['项目名称', '用例数', '执行次数', '通过', '失败', '通过率'], y)
            for i, item in enumerate(project_stats):
                project_name = item.get('project_name', item.get('name', '未知'))
                if len(project_name) > 15:
                    project_name = project_name[:12] + '...'
                total_cases = item.get('total_cases', 0)
                total_runs = item.get('total_runs', 0)
                passed = item.get('passed_runs', 0)
                failed = item.get('failed_runs', 0)
                pass_rate = item.get('pass_rate', 0)
                y = draw_table_row([
                    project_name, total_cases, total_runs, passed, failed,
                    f"{pass_rate:.1f}%"
                ], y, i % 2 == 1)
        
        # 绘制页脚
        draw_footer(page_num)
        
        # 保存PDF
        c.save()
        return filepath
    
    def _generate_simplified_html_report(self, data: Dict[str, Any]) -> str:
        """生成简化版HTML报告内容，避免触发OCR
        
        Args:
            data: 报告数据
        
        Returns:
            简化版HTML字符串
        """
        overview = data.get('overview', {})
        status_dist = data.get('status_distribution', [])
        duration_dist_dict = data.get('duration_distribution', {})
        duration_dist = duration_dist_dict.get('distribution', []) if isinstance(duration_dist_dict, dict) else []
        case_stats = data.get('case_statistics', [])
        project_stats = data.get('project_statistics', [])
        
        # 使用更简单的HTML和CSS，确保xhtml2pdf能够正确渲染
        html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>测试报告</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            padding: 0;
        }
        .container {
            width: 100%;
            max-width: 1000px;
            margin: 0 auto;
        }
        h1 {
            color: #1890ff;
            text-align: center;
            margin-bottom: 30px;
        }
        .section {
            margin-bottom: 30px;
        }
        .section-title {
            font-size: 18px;
            font-weight: bold;
            color: #333;
            margin-bottom: 15px;
            padding-bottom: 5px;
            border-bottom: 2px solid #1890ff;
        }
        .overview-grid {
            display: table;
            width: 100%;
            margin-bottom: 30px;
        }
        .overview-row {
            display: table-row;
        }
        .overview-card {
            display: table-cell;
            padding: 15px;
            text-align: center;
            border: 1px solid #e8e8e8;
            margin: 5px;
        }
        .overview-card .label {
            font-size: 14px;
            color: #666;
            margin-bottom: 5px;
        }
        .overview-card .value {
            font-size: 24px;
            font-weight: bold;
            color: #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        th, td {
            padding: 10px;
            text-align: left;
            border: 1px solid #e8e8e8;
        }
        th {
            background-color: #f0f2f5;
            font-weight: bold;
            color: #333;
        }
        .footer {
            text-align: center;
            color: #999;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #e8e8e8;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>测试报告</h1>
        <p style="text-align: center; color: #999; margin-bottom: 30px;">
            生成时间: ''' + datetime.now().strftime('%Y年%m月%d日 %H:%M:%S') + '''
        </p>
        
        <div class="section">
            <div class="section-title">统计概览</div>
            <div class="overview-grid">
                <div class="overview-row">
                    <div class="overview-card">
                        <div class="label">总执行次数</div>
                        <div class="value">''' + str(overview.get('total_runs', 0)) + '''</div>
                    </div>
                    <div class="overview-card">
                        <div class="label">通过次数</div>
                        <div class="value">''' + str(overview.get('passed_runs', 0)) + '''</div>
                    </div>
                    <div class="overview-card">
                        <div class="label">失败次数</div>
                        <div class="value">''' + str(overview.get('failed_runs', 0)) + '''</div>
                    </div>
                    <div class="overview-card">
                        <div class="label">通过率</div>
                        <div class="value">''' + str(round(overview.get('pass_rate', 0), 2)) + '%' + '''</div>
                    </div>
                    <div class="overview-card">
                        <div class="label">平均执行时间</div>
                        <div class="value">''' + str(round(overview.get('avg_duration', 0), 2)) + '秒' + '''</div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title">状态分布</div>
            <table>
                <thead>
                    <tr>
                        <th>状态</th>
                        <th>次数</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + ''.join(["<tr><td>" + str(item.get('status', '未知')) + "</td><td>" + str(item.get('count', 0)) + "</td><td>" + str(round(item.get('percentage', 0), 2)) + "%</td></tr>" for item in status_dist]) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">耗时分布</div>
            <table>
                <thead>
                    <tr>
                        <th>耗时区间</th>
                        <th>用例数</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + ''.join(["<tr><td>" + str(item.get('range', '未知')) + "</td><td>" + str(item.get('count', 0)) + "</td><td>" + str(round(item.get('percentage', 0), 2)) + "%</td></tr>" for item in duration_dist]) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">用例统计</div>
            <table>
                <thead>
                    <tr>
                        <th>用例名称</th>
                        <th>总执行次数</th>
                        <th>通过次数</th>
                        <th>失败次数</th>
                        <th>通过率</th>
                        <th>平均耗时</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + ''.join(["<tr><td>" + str(item.get('name', '未知')) + "</td><td>" + str(item.get('total_runs', 0)) + "</td><td>" + str(item.get('passed_runs', 0)) + "</td><td>" + str(item.get('failed_runs', 0)) + "</td><td>" + str(round(item.get('pass_rate', 0), 2)) + "%</td><td>" + str(round(item.get('avg_duration', 0), 2)) + "秒</td></tr>" for item in case_stats]) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">项目统计</div>
            <table>
                <thead>
                    <tr>
                        <th>项目名称</th>
                        <th>用例数</th>
                        <th>总执行次数</th>
                        <th>通过次数</th>
                        <th>失败次数</th>
                        <th>通过率</th>
                        <th>平均耗时</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + ''.join(["<tr><td>" + str(item.get('project_name', '未知')) + "</td><td>" + str(item.get('total_cases', 0)) + "</td><td>" + str(item.get('total_runs', 0)) + "</td><td>" + str(item.get('passed_runs', 0)) + "</td><td>" + str(item.get('failed_runs', 0)) + "</td><td>" + str(round(item.get('pass_rate', 0), 2)) + "%</td><td>" + str(round(item.get('avg_duration', 0), 2)) + "秒</td></tr>" for item in project_stats]) + '''
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>本报告由UI自动化测试平台自动生成</p>
        </div>
    </div>
</body>
</html>
'''
        
        return html
    
    def _generate_html_report(self, data: Dict[str, Any]) -> str:
        """生成HTML报告内容
        
        Args:
            data: 报告数据
        
        Returns:
            HTML字符串
        """
        overview = data.get('overview', {})
        status_dist = data.get('status_distribution', [])
        duration_dist_dict = data.get('duration_distribution', {})
        duration_dist = duration_dist_dict.get('distribution', []) if isinstance(duration_dist_dict, dict) else []
        case_stats = data.get('case_statistics', [])
        project_stats = data.get('project_statistics', [])
        
        # 使用三重引号和转义来避免f-string解析问题
        html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>测试报告</title>
    <style>
        body {
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }
        h1 {
            color: #1890ff;
            text-align: center;
            margin-bottom: 30px;
        }
        .section {
            margin-bottom: 30px;
        }
        .section-title {
            font-size: 18px;
            font-weight: bold;
            color: #333;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 2px solid #1890ff;
        }
        .overview-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }
        .overview-card {
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .overview-card .label {
            color: #666;
            font-size: 14px;
            margin-bottom: 10px;
        }
        .overview-card .value {
            font-size: 28px;
            font-weight: bold;
            color: #1890ff;
        }
        .overview-card.passed .value {
            color: #52c41a;
        }
        .overview-card.failed .value {
            color: #ff4d4f;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #e8e8e8;
        }
        th {
            background-color: #fafafa;
            font-weight: bold;
            color: #333;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .status-badge {
            display: inline-block;
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: bold;
        }
        .status-passed {
            background-color: #f6ffed;
            color: #52c41a;
            border: 1px solid #b7eb8f;
        }
        .status-failed {
            background-color: #fff1f0;
            color: #ff4d4f;
            border: 1px solid #ffa39e;
        }
        .footer {
            text-align: center;
            color: #999;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #e8e8e8;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>测试报告</h1>
        <p style="text-align: center; color: #999; margin-bottom: 30px;">
            生成时间: ''' + datetime.now().strftime('%Y年%m月%d日 %H:%M:%S') + '''
        </p>
        
        <div class="section">
            <div class="section-title">统计概览</div>
            <div class="overview-grid">
                <div class="overview-card">
                    <div class="label">总执行次数</div>
                    <div class="value">''' + str(overview.get('total_runs', 0)) + '''</div>
                </div>
                <div class="overview-card passed">
                    <div class="label">通过次数</div>
                    <div class="value">''' + str(overview.get('passed_runs', 0)) + '''</div>
                </div>
                <div class="overview-card failed">
                    <div class="label">失败次数</div>
                    <div class="value">''' + str(overview.get('failed_runs', 0)) + '''</div>
                </div>
                <div class="overview-card">
                    <div class="label">通过率</div>
                    <div class="value">''' + str(overview.get('pass_rate', 0)) + '''%</div>
                </div>
                <div class="overview-card">
                    <div class="label">平均执行时间</div>
                    <div class="value">''' + str(overview.get('avg_duration', 0)) + '''秒</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title">状态分布</div>
            <table>
                <thead>
                    <tr>
                        <th>状态</th>
                        <th>数量</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + self._generate_status_rows(status_dist, overview.get('total_runs', 1)) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">耗时分布</div>
            <table>
                <thead>
                    <tr>
                        <th>耗时区间</th>
                        <th>数量</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + self._generate_duration_rows(duration_dist, overview.get('total_runs', 1)) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">用例统计</div>
            <table>
                <thead>
                    <tr>
                        <th>用例名称</th>
                        <th>总执行次数</th>
                        <th>通过次数</th>
                        <th>失败次数</th>
                        <th>通过率</th>
                        <th>平均耗时</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + self._generate_case_rows(case_stats) + '''
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">项目统计</div>
            <table>
                <thead>
                    <tr>
                        <th>项目名称</th>
                        <th>用例数量</th>
                        <th>总执行次数</th>
                        <th>通过次数</th>
                        <th>失败次数</th>
                        <th>通过率</th>
                        <th>平均耗时</th>
                    </tr>
                </thead>
                <tbody>
                    ''' + self._generate_project_rows(project_stats) + '''
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>本报告由UI自动化测试平台自动生成</p>
        </div>
    </div>
</body>
</html>'''
        
        return html
    
    def _generate_status_rows(self, status_dist: List[Dict], total: int) -> str:
        """生成状态表格行"""
        rows = []
        for item in status_dist:
            status = item.get('status', 'unknown')
            count = item.get('count', 0)
            percentage = (count / total * 100) if total > 0 else 0
            
            status_class = 'status-passed' if status == 'passed' else 'status-failed' if status == 'failed' else ''
            
            rows.append(f"""
                    <tr>
                        <td><span class="status-badge {status_class}">{status}</span></td>
                        <td>{count}</td>
                        <td>{percentage:.2f}%</td>
                    </tr>""")
        
        return '\n'.join(rows)
    
    def _generate_duration_rows(self, duration_dist: List[Dict], total: int) -> str:
        """生成耗时表格行"""
        rows = []
        for item in duration_dist:
            range_label = item.get('range', '')
            count = item.get('count', 0)
            percentage = (count / total * 100) if total > 0 else 0
            
            rows.append(f"""
                    <tr>
                        <td>{range_label}</td>
                        <td>{count}</td>
                        <td>{percentage:.2f}%</td>
                    </tr>""")
        
        return '\n'.join(rows)
    
    def _generate_case_rows(self, case_stats: List[Dict]) -> str:
        """生成用例表格行"""
        rows = []
        for case in case_stats:
            rows.append(f"""
                    <tr>
                        <td>{case.get('case_name', '')}</td>
                        <td>{case.get('total_runs', 0)}</td>
                        <td>{case.get('passed_runs', 0)}</td>
                        <td>{case.get('failed_runs', 0)}</td>
                        <td>{case.get('pass_rate', 0)}%</td>
                        <td>{case.get('avg_duration', 0)}秒</td>
                    </tr>""")
        
        return '\n'.join(rows)
    
    def _generate_project_rows(self, project_stats: List[Dict]) -> str:
        """生成项目表格行"""
        rows = []
        for project in project_stats:
            rows.append(f"""
                    <tr>
                        <td>{project.get('project_name', '')}</td>
                        <td>{project.get('total_cases', 0)}</td>
                        <td>{project.get('total_runs', 0)}</td>
                        <td>{project.get('passed_runs', 0)}</td>
                        <td>{project.get('failed_runs', 0)}</td>
                        <td>{project.get('pass_rate', 0)}%</td>
                        <td>{project.get('avg_duration', 0)}秒</td>
                    </tr>""")
        
        return '\n'.join(rows)
