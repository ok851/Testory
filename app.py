import os
import sys

try:
    from pathlib import Path

    from dotenv import load_dotenv

    _app_root = Path(__file__).resolve().parent

    def _pick_env_file() -> Path:
        for key in ("TESTORY_ENV_FILE",):
            raw = os.environ.get(key, "").strip()
            if raw:
                return Path(raw)
        uat_data = os.environ.get("UAT_DATA_DIR", "").strip()
        if uat_data:
            return Path(uat_data) / ".env"
        return _app_root / ".env"

    _env_path = _pick_env_file()
    try:
        from env_example_sync import sync_env_from_example

        # 桌面版 SKIP_ENV_EXAMPLE_SYNC=1；否则仅在可写目录合并 .env
        sync_root = Path(os.environ.get("UAT_DATA_DIR", "").strip() or _app_root)
        sync_env_from_example(sync_root)
    except Exception:
        pass
    if _env_path.is_file():
        load_dotenv(_env_path, encoding="utf-8-sig")
    elif (_app_root / ".env").is_file():
        load_dotenv(_app_root / ".env", encoding="utf-8-sig")
    else:
        load_dotenv(encoding="utf-8-sig")
except ImportError:
    pass

if sys.platform == "win32":
    # 桌面壳：侧车服务放到后台，避免阻塞 Flask 监听（启动页长时间停在「正在启动本地服务」）
    try:
        import threading as _boot_threading

        def _boot_desktop_side_services() -> None:
            try:
                from desktop_service_bootstrap import bootstrap_desktop_services

                bootstrap_desktop_services()
            except Exception:
                pass

        _lazy_boot = (
            os.environ.get("DESKTOP_LAZY_GATEWAY_BOOT", "").strip().lower() in ("1", "true", "yes", "on")
            or os.environ.get("UAT_DESKTOP_MODE", "").strip().lower() in ("1", "true", "yes")
        )
        if _lazy_boot:
            _boot_threading.Thread(
                target=_boot_desktop_side_services,
                daemon=True,
                name="desktop-side-boot",
            ).start()
        else:
            _boot_desktop_side_services()
    except Exception:
        pass

try:
    import threading as _mobile_boot_threading

    def _boot_mobile_side_services() -> None:
        try:
            from mobile_service_bootstrap import bootstrap_mobile_services

            bootstrap_mobile_services()
        except Exception:
            pass

    _lazy_mobile = (
        os.environ.get("DESKTOP_LAZY_GATEWAY_BOOT", "").strip().lower() in ("1", "true", "yes", "on")
        or os.environ.get("UAT_DESKTOP_MODE", "").strip().lower() in ("1", "true", "yes")
    )
    if _lazy_mobile:
        _mobile_boot_threading.Thread(
            target=_boot_mobile_side_services,
            daemon=True,
            name="mobile-side-boot",
        ).start()
    else:
        _boot_mobile_side_services()
except Exception:
    pass

try:
    from hermes_config import ensure_hermes_home

    ensure_hermes_home()
except Exception:
    pass

# Hermes Gateway：不在进程启动时自动拉起，由 AI 页「启动智能体」手动触发

from flask import Flask, render_template, request, jsonify, session, make_response, redirect, url_for, Response, stream_with_context
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import time
import shlex
import shutil
import secrets
import uuid
import json
import re
from typing import Any, Dict, List, Optional, Tuple
import functools
import threading
import io
import tempfile
import subprocess
from database import Database
from time_utils import beijing_now_iso, utc_now_sqlite_str as utc_sql_str
from api_spec_pipeline import run_api_spec_pipeline
from execution_context import ExecutionContext
from embedded_browser_client import embedded_gateway_config, embedded_gateway_enabled, embedded_gateway_json
from batch_input_parse import parse_batch_input_lines
from playwright_automation import (
    automation,
    normalize_playwright_browser_name,
    force_reset_execution_state,
    parse_platform_scroll_input_value,
    scroll_event_to_platform_input_value,
    _execution_lock,
    set_execution_in_progress,
    sync_analyze_page_content,
    sync_automation_session_usable,
    sync_browser_go_back,
    sync_browser_go_forward,
    sync_browser_keyboard_press,
    sync_browser_keyboard_type,
    sync_browser_mouse_click,
    sync_browser_mouse_wheel,
    sync_browser_reload,
    sync_click_element,
    sync_element_info_at_point,
    sync_get_interactive_page_snapshot,
    sync_get_accessibility_outline_text,
    sync_get_page_diagnostics,
    sync_get_viewport_size,
    sync_close_browser,
    sync_disable_element_selection,
    sync_double_click_element,
    sync_enable_element_selection,
    sync_enter_iframe,
    sync_execute_multiple_test_cases,
    sync_exit_iframe,
    sync_extract_element_data,
    sync_extract_element_json,
    sync_extract_element_text,
    sync_extract_json_from_selected_element,
    sync_fill_input,
    resolve_fill_step_text,
    step_description_implies_empty_input,
    sync_get_all_links,
    sync_get_current_url,
    sync_get_element_count,
    sync_get_page_data,
    sync_get_page_elements,
    sync_get_page_text,
    sync_get_page_title,
    sync_get_selected_element,
    sync_hover_element,
    sync_execute_script_steps,
    sync_run_api_case_for_batch,
    sync_navigate_to,
    sync_right_click_element,
    sync_scroll_by_delta,
    sync_scroll_page,
    sync_start_browser,
    sync_prepare_fresh_web_session,
    resolve_playwright_headless,
    sync_select_date,
    sync_select_option,
    sync_swipe_element,
    sync_take_screenshot,
    sync_take_screenshot_bytes,
    sync_verify_element,
    sync_wait_for_element_visible,
    sync_wait_for_page_stable,
    sync_wait_for_selector,
    sync_wait_for_timeout,
    worker,
    _url_assert_matches_pa,
)
from playwright_codegen_import import (
    enrich_steps_with_xpath_priority,
    parse_playwright_codegen_to_steps,
)
from selenium_ide_import import parse_selenium_ide_to_steps
from test_report import TestReportGenerator
from report_exporter import ReportExporter
from logger import uat_logger
from license_manager import license_manager, LicenseType
from deployment_hooks import (
    guard_billing_route,
    init_server_instance,
    patch_run_case_for_server,
    register_deployment_hooks,
)
from deployment_config import is_client_mode, should_delegate_execution_to_clients
from client_run_helpers import load_case_and_steps, sync_run_to_team_server
from cloud_llm_gateway import CloudLLMGateway
from ai_config_paths import ai_model_registry_path, ai_provider_catalog_path, load_ai_provider_catalog_dict
from ai_local_inference import local_ai_service
from mail_service import send_verify_code, verify_code
from ai_step_normalization import (
    ai_plan_steps_to_playwright_script_steps,
    apply_step_normalization_to_plan,
    dedupe_and_validate_ai_steps,
)
from ai_platform_audit import log_ai_plan_to_audit
import asyncio
import threading
import datetime
# `flask_migrate` 只用于数据库迁移；某些部署环境可能未安装，需保证应用可正常启动。
try:
    from flask_migrate import Migrate  # type: ignore
except ModuleNotFoundError:
    Migrate = None

# 数据驱动批量执行进度（内存态，按 run_id + 用户隔离；完成后首次拉取状态即清理）
_dataset_run_jobs: dict = {}
_dataset_run_lock = threading.Lock()
_case_run_jobs: dict = {}
_case_run_lock = threading.Lock()
_user_ui_run_locks: dict = {}
_ai_task_abort_events: Dict[str, threading.Event] = {}
_ai_task_abort_lock = threading.Lock()
_user_ui_run_locks_mu = threading.Lock()
_ai_model_cfg_lock = threading.Lock()
_login_fail_lock = threading.Lock()
_login_fail_timestamps: dict = {}

if sys.platform == "win32":
    # OpenCV 导入较慢：后台探测，避免拖慢 Flask 首响
    try:
        import threading as _desk_check_threading

        def _log_desktop_runtime() -> None:
            try:
                from desktop_runtime import (
                    desktop_runtime_available,
                    desktop_runtime_unavailable_reason,
                )

                if desktop_runtime_available():
                    uat_logger.info(
                        "桌面视觉自动化依赖就绪 (opencv+mss, 解释器=%s)",
                        sys.executable,
                    )
                else:
                    uat_logger.warning(
                        "桌面自动化不可用: %s",
                        desktop_runtime_unavailable_reason() or "未知原因",
                    )
            except Exception as _desk_boot_exc:
                uat_logger.debug("桌面依赖自检跳过: %s", _desk_boot_exc)

        _desk_check_threading.Thread(
            target=_log_desktop_runtime,
            daemon=True,
            name="desktop-runtime-check",
        ).start()
    except Exception:
        pass


def _is_production_env() -> bool:
    return (
        os.environ.get("FLASK_ENV", "").strip().lower() == "production"
        or os.environ.get("APP_ENV", "").strip().lower() == "production"
    )


def _login_client_ip() -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").strip()
    if xff:
        return xff.split(",")[0].strip()[:100] or (request.remote_addr or "")
    return request.remote_addr or ""


def _login_is_rate_limited(ip: str) -> tuple:
    max_n = int(os.environ.get("LOGIN_RATE_LIMIT_MAX", "40"))
    win = int(os.environ.get("LOGIN_RATE_LIMIT_WINDOW_SEC", "900"))
    if max_n <= 0:
        return False, 0
    now = time.time()
    with _login_fail_lock:
        lst = _login_fail_timestamps.get(ip, [])
        lst = [t for t in lst if now - t < win]
        if not lst:
            _login_fail_timestamps.pop(ip, None)
        else:
            _login_fail_timestamps[ip] = lst
        if len(lst) >= max_n:
            retry = int(win - (now - min(lst))) + 1
            return True, max(1, retry)
    return False, 0


def _login_record_failure(ip: str) -> None:
    win = int(os.environ.get("LOGIN_RATE_LIMIT_WINDOW_SEC", "900"))
    now = time.time()
    with _login_fail_lock:
        lst = _login_fail_timestamps.get(ip, [])
        lst = [t for t in lst if now - t < win]
        lst.append(now)
        _login_fail_timestamps[ip] = lst


def _login_clear_failures(ip: str) -> None:
    with _login_fail_lock:
        _login_fail_timestamps.pop(ip, None)


def _case_job_update(user_id: int, **kwargs):
    with _case_run_lock:
        if user_id in _case_run_jobs:
            _case_run_jobs[user_id].update(kwargs)


def _case_run_cancelled(user_id: int) -> bool:
    """当前用户是否已请求停止用例执行（与 /api/cases/current-run/stop 一致）。"""
    with _case_run_lock:
        job = _case_run_jobs.get(user_id)
        return bool(job and job.get('cancel_requested'))


def _get_user_ui_run_lock(user_id: int) -> threading.Lock:
    """同一用户 UI 用例运行串行化，避免多标签页/连点导致并发抢占浏览器。"""
    with _user_ui_run_locks_mu:
        lock = _user_ui_run_locks.get(user_id)
        if lock is None:
            lock = threading.Lock()
            _user_ui_run_locks[user_id] = lock
        return lock


def _record_run_history_rejected(db, case_id: int, reason: str, duration: float = 0.0):
    """锁冲突或未启动执行时写入历史，避免运行历史空白。"""
    msg = (reason or '用例未执行').strip()
    try:
        rid = db.create_run_history(
            case_id,
            'error',
            round(max(0.0, float(duration or 0.0)), 2),
            msg,
            '',
            '',
        )
        uat_logger.info('已记录未执行用例历史 case_id=%s run_id=%s', case_id, rid)
        return rid
    except Exception as exc:
        uat_logger.warning('记录未执行用例历史失败 case_id=%s: %s', case_id, exc)
        return None


class _UserUiRunGuard:
    """同一用户 Web 用例运行串行槽（with 块内 return 也会释放锁）。"""

    def __init__(self, user_id: int, label: str = ''):
        self.user_id = user_id
        self.label = label or ''
        self._lock = _get_user_ui_run_lock(user_id)

    def __enter__(self):
        self._lock.acquire(blocking=True)
        uat_logger.info(
            '🔒 [USER_RUN] 用户 %s 进入串行执行槽%s',
            self.user_id,
            f' ({self.label})' if self.label else '',
        )
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            self._lock.release()
        except RuntimeError:
            pass
        uat_logger.info(
            '🔓 [USER_RUN] 用户 %s 释放串行执行槽%s',
            self.user_id,
            f' ({self.label})' if self.label else '',
        )
        return False


def _env_flag_true(name: str) -> bool:
    return (os.environ.get(name) or "").strip().lower() in ("1", "true", "yes", "on")


def _on_case_execution_failure(payload: dict) -> None:
    """执行线程内调用：可选本地向量记忆与自动缺陷（不依赖 Flask 上下文）。"""
    uid = payload.get("user_id")
    tid = payload.get("tenant_id")
    case_id = payload.get("case_id")
    case_name = (payload.get("case_name") or "").strip() or "未命名用例"
    project_id = payload.get("project_id")
    err = (payload.get("error") or "").strip()
    trigger = (payload.get("trigger") or "").strip() or "unknown"

    try:
        import ai_memory_store

        if uid is not None and ai_memory_store.memory_enabled():
            lines = [f"[{trigger}] 用例执行失败: {case_name}", f"case_id={case_id}"]
            if err:
                lines.append(f"error: {err[:3500]}")
            text = "\n".join(lines)
            meta = {
                "case_id": case_id,
                "project_id": project_id,
                "run_history_id": payload.get("run_history_id"),
                "trigger": trigger,
            }
            if payload.get("schedule_id") is not None:
                meta["schedule_id"] = payload.get("schedule_id")
            ai_memory_store.ingest(
                int(uid),
                "case_failure",
                text,
                tenant_id=tid if isinstance(tid, int) else None,
                meta=meta,
            )
    except Exception as exc:
        uat_logger.warning(f"[FAILURE_HOOK] 向量记忆写入跳过: {exc}")

    if not _env_flag_true("AUTO_DEFECT_ON_CASE_FAILURE"):
        return
    if uid is None or project_id is None:
        return
    try:
        db = Database()
        title = f"[自动-{trigger}] {case_name} 执行失败"
        desc_parts = [f"触发来源: {trigger}", f"case_id: {case_id}"]
        if err:
            desc_parts.append(f"错误摘要: {err[:2000]}")
        desc = "\n".join(desc_parts)
        db.create_defect(
            int(project_id),
            title[:200],
            int(uid),
            description=desc[:8000],
            severity="medium",
            priority="medium",
            case_id=int(case_id) if case_id is not None else None,
            run_history_id=payload.get("run_history_id"),
            error_message=(err or "")[:4000],
            status="open",
        )
    except Exception as exc:
        uat_logger.warning(f"[FAILURE_HOOK] 自动创建缺陷失败: {exc}")


def _make_execution_context(trigger: str, extra: dict | None = None) -> ExecutionContext:
    uid = current_user.id if current_user.is_authenticated else None
    tid = None
    if uid is not None:
        try:
            tid = Database().get_user_tenant_id(int(uid))
        except Exception:
            tid = None
    return ExecutionContext(
        user_id=uid,
        tenant_id=tid,
        trigger=trigger,
        on_case_failure=_on_case_execution_failure,
        extra=dict(extra or {}),
    )


def _make_batch_execution_context(trigger: str, data: dict | None = None) -> ExecutionContext:
    """批量执行上下文：含运行时变量池与会话复用开关。"""
    payload = dict(data or {})
    ctx = _make_execution_context(trigger, extra=payload)
    if payload.get("reuse_session") is False:
        ctx.reuse_session = False
    if payload.get("skip_duplicate_login_for_business") is False:
        ctx.skip_duplicate_login_for_business = False
    return ctx


def _force_stop_browser_async():
    """异步强制停止浏览器，避免阻塞停止接口响应。"""
    # 先重置状态（清空 browser/page 引用，设置 Event 信号）
    # 这样执行线程的下一步浏览器断连检测能立即生效
    try:
        force_reset_execution_state()
    except Exception:
        pass
    # 再尝试关闭浏览器进程（可能被 worker 阻塞，但不影响断连检测）
    try:
        sync_close_browser()
    except Exception:
        pass

def _parse_api_bool(val, default: bool = True) -> bool:
    """解析 JSON 中的布尔或字符串开关（避免对 True 调用 .strip() 导致 500）。"""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() not in ("0", "false", "no", "off", "")


# 统一的API错误处理装饰器
def api_error_handler(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            uat_logger.log_exception(f"API Error in {func.__name__}", e)
            detail = str(e)
            if _is_production_env() and os.environ.get("API_DETAILED_ERRORS", "").strip().lower() not in (
                "1",
                "true",
                "yes",
            ):
                detail = "服务器内部错误，请稍后重试或联系管理员"
            return jsonify({"success": False, "error": detail}), 500
    return wrapper


def _safe_response_payload_for_api_log(rv):
    """供 log_api_request 使用：勿对 SSE 等流式响应调用 get_json（可能阻塞或抛错）。"""
    if rv is None:
        return None
    try:
        ct = (rv.headers.get("Content-Type") or getattr(rv, "mimetype", None) or "").lower()
    except Exception:
        ct = ""
    if "event-stream" in ct or "text/event-stream" in ct:
        return {"_type": "sse"}
    try:
        return rv.get_json(silent=True)
    except Exception:
        return None


# API请求日志装饰器
def log_api_request(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        # 记录请求，处理没有请求体的情况
        try:
            request_data = (
                request.get_json(silent=True) if request.method in ["POST", "PUT", "PATCH"] else None
            )
        except Exception:
            # 如果解析JSON失败（如请求体为空），使用None
            request_data = None
        uat_logger.log_api_request(func.__name__, request.method, request_data)
        # 执行函数
        response = func(*args, **kwargs)
        # 记录响应
        try:
            if isinstance(response, tuple):
                rv, status = response[0], response[1]
                uat_logger.log_api_response(
                    func.__name__, status, _safe_response_payload_for_api_log(rv)
                )
            else:
                sc = getattr(response, "status_code", None) or 200
                uat_logger.log_api_response(
                    func.__name__, sc, _safe_response_payload_for_api_log(response)
                )
        except Exception:
            # 如果响应无法解析为JSON，记录基本信息
            if isinstance(response, tuple):
                status_code = response[1]
            else:
                status_code = getattr(response, "status_code", None) or 200
            uat_logger.log_api_response(func.__name__, status_code, None)
        return response
    return wrapper


def _init_cors(flask_app: Flask) -> None:
    """跨域：生产默认不启用宽松 CORS；需跨域前端时设置 FLASK_CORS_ORIGINS（逗号分隔）。"""
    raw = os.environ.get("FLASK_CORS_ORIGINS", "").strip()
    if raw:
        origins = [x.strip() for x in raw.split(",") if x.strip()]
        if origins:
            CORS(flask_app, resources={r"/*": {"origins": origins}})
        return
    if os.environ.get("FLASK_ENV", "").lower() == "production" or os.environ.get("APP_ENV", "").lower() == "production":
        return
    CORS(flask_app, resources={r"/*": {"origins": "*"}})


try:
    from install_paths import resource_root as _testory_resource_root

    _rr = _testory_resource_root()
    app = Flask(
        __name__,
        template_folder=str(_rr / "templates"),
        static_folder=str(_rr / "static"),
    )
except ImportError:
    app = Flask(__name__)
_init_cors(app)
# Session 加密密钥：与 docker-compose 中 SECRET_KEY 一致；未设置时每次启动随机（会话在重启后失效）
_secret_key = (os.environ.get("SECRET_KEY") or os.environ.get("FLASK_SECRET_KEY") or "").strip()
if _secret_key:
    app.secret_key = _secret_key
else:
    app.secret_key = secrets.token_hex(32)
    uat_logger.warning(
        "未设置环境变量 SECRET_KEY 或 FLASK_SECRET_KEY，已使用临时随机密钥；"
        "生产环境请务必设置固定密钥，否则重启后会话将全部失效。"
    )

# 模板热重载：DEBUG 开启时一定重载；否则在非 production 下也默认重载，避免改 ai_test.html 等必须反复杀进程。
# （仅 .py 逻辑仍建议开 FLASK_DEBUG=1 以启用代码重载，或改后手动重启。）
_flask_debug = os.environ.get("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")
_is_prod = os.environ.get("APP_ENV", "").lower() == "production" or os.environ.get("FLASK_ENV", "").lower() == "production"
if _flask_debug or not _is_prod:
    app.config["TEMPLATES_AUTO_RELOAD"] = True


@app.after_request
def _prevent_stale_html_cache(response):
    """开发/联调时避免浏览器长期使用带侧栏的旧 HTML 缓存。"""
    if _is_prod:
        return response
    ctype = (response.content_type or "").split(";")[0].strip().lower()
    if ctype == "text/html":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response


app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "").strip().lower() in (
    "1",
    "true",
    "yes",
)
app.config["PERMANENT_SESSION_LIFETIME"] = datetime.timedelta(days=365)
app.config["REMEMBER_COOKIE_DURATION"] = datetime.timedelta(days=365)
app.config["REMEMBER_COOKIE_HTTPONLY"] = True
app.config["REMEMBER_COOKIE_SAMESITE"] = "Lax"

_max_upload_mb = int(os.environ.get("MAX_UPLOAD_MB", "50") or "50")
app.config["MAX_CONTENT_LENGTH"] = max(1, _max_upload_mb) * 1024 * 1024

if os.environ.get("TRUST_X_FORWARDED", "").strip().lower() in ("1", "true", "yes"):
    from werkzeug.middleware.proxy_fix import ProxyFix

    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)


@app.after_request
def _security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "geolocation=(), microphone=(), camera=()",
    )
    return response


@app.before_request
def _sync_playwright_browser_from_session():
    """将当前登录用户的浏览器偏好同步到 Playwright worker（与 session / PLAYWRIGHT_BROWSER 一致）。"""
    try:
        if not current_user.is_authenticated:
            return
    except Exception:
        return
    path = request.path or ""
    if not path.startswith("/api/"):
        return
    if path.startswith("/api/auth/") or path.startswith("/api/health"):
        return
    try:
        sess_val = session.get("playwright_browser_engine")
        if sess_val:
            automation.set_browser_engine(str(sess_val).strip().lower())
        else:
            automation.set_browser_engine(None)
    except Exception:
        pass


@app.errorhandler(413)
def _request_entity_too_large(_e):
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": "请求体或上传文件超过大小限制"}), 413
    return "Payload Too Large", 413


# ==================== AI模型路由与云端安全网关 ====================
_CLOUD_LLM_ENDPOINT = os.environ.get('CLOUD_LLM_ENDPOINT', '').strip()
_CLOUD_LLM_API_KEY = os.environ.get('CLOUD_LLM_API_KEY', '').strip()
_cloud_llm_gateway = None
_AI_MODEL_NAME_RE = re.compile(r'^[a-zA-Z0-9][a-zA-Z0-9._:/+\-]{0,199}$')
_AI_PROFILE_MODEL_ID_RE = re.compile(r'^[^\s]{1,220}$')


def _validate_ai_model_name(name: str) -> tuple:
    n = (name or '').strip()
    if not n:
        return False, 'model_name不能为空'
    if len(n) > 200:
        return False, '模型名称过长（最多200字符）'
    if not _AI_MODEL_NAME_RE.match(n):
        return False, '模型名称格式无效：需以字母或数字开头，仅允许字母、数字及 . _ : / + -'
    return True, n


def _get_cloud_llm_gateway():
    global _cloud_llm_gateway
    if _cloud_llm_gateway is None:
        if not _CLOUD_LLM_ENDPOINT or not _CLOUD_LLM_API_KEY:
            return None
        _cloud_llm_gateway = CloudLLMGateway(
            endpoint=_CLOUD_LLM_ENDPOINT,
            api_key=_CLOUD_LLM_API_KEY,
            timeout=45,
        )
    return _cloud_llm_gateway


def _route_ai_model(task_type: str) -> dict:
    """
    双模型策略:
    - 本地开源模型: 高频、低复杂任务
    - 云端大模型: 脚本修复/复杂报错分析（强制脱敏后上云）
    """
    local_light_tasks = {'intent_classification', 'report_summary', 'log_tagging'}
    local_mid_tasks = {'test_case_generation', 'test_step_generation', 'base_fix_suggestion'}
    cloud_tasks = {'script_repair', 'complex_error_analysis'}

    if task_type in local_light_tasks:
        return {'provider': 'local', 'model': 'qwen2-1.5b-or-qwen1.8b'}
    if task_type in local_mid_tasks:
        return {'provider': 'local', 'model': 'llama3-8b-instruct'}
    if task_type in cloud_tasks:
        return {'provider': 'cloud', 'model': 'cloud_llm'}
    return {'provider': 'local', 'model': _get_active_local_model()}


def _default_ai_model_config() -> dict:
    local_mid = os.environ.get('LOCAL_LLM_MODEL_MID', 'llama3:8b-instruct').strip()
    local_light = os.environ.get('LOCAL_LLM_MODEL_LIGHT', 'qwen2:1.5b').strip()
    models = [m for m in [local_mid, local_light, 'qwen:1.8b-chat'] if m]
    dedup_models = []
    for m in models:
        if m not in dedup_models:
            dedup_models.append(m)
    return {
        'active_local_model': local_mid,
        'local_models': dedup_models,
    }


def _load_ai_provider_catalog() -> dict:
    return load_ai_provider_catalog_dict()


def _catalog_provider_meta(provider_id: str) -> dict:
    for p in _load_ai_provider_catalog().get('providers') or []:
        if isinstance(p, dict) and p.get('id') == provider_id:
            return p
    return {}


def _infer_ai_provider_simple(base_url: str = '', api_key: str = '') -> str:
    from ai_provider_infer import infer_provider_from_simple_config

    catalog = _load_ai_provider_catalog()
    return infer_provider_from_simple_config(
        base_url,
        api_key,
        catalog.get('providers') if isinstance(catalog, dict) else [],
    )


def _normalize_profile_base_url(base_url: str, provider: str = '') -> str:
    """修正用户误填的完整 chat/completions 地址，避免重复拼接路径。"""
    bu = (base_url or '').strip()
    if not bu:
        cmeta = _catalog_provider_meta(provider) if provider else {}
        return (cmeta.get('default_base_url') or '').strip()
    low = bu.lower().rstrip('/')
    for suffix in ('/v1/chat/completions', '/chat/completions'):
        if low.endswith(suffix):
            bu = bu[: len(bu) - len(suffix)].rstrip('/')
            low = bu.lower().rstrip('/')
    if low.endswith('/v1/v1'):
        bu = bu[: -len('/v1')]
    return bu.rstrip('/')


def _migrate_v1_config_to_v2(raw: dict, defaults: dict) -> dict:
    models = raw.get('local_models')
    if not isinstance(models, list) or not models:
        models = list(defaults.get('local_models') or [])
    clean_models = []
    for m in models:
        m = (str(m) if m is not None else '').strip()
        if m and m not in clean_models:
            clean_models.append(m)
    active_name = (raw.get('active_local_model') or '').strip()
    profiles = []
    active_id = ''
    for m in clean_models:
        pid = str(uuid.uuid4())
        profiles.append({
            'id': pid,
            'provider': 'ollama',
            'api_style': 'ollama',
            'model_type': 'test_case_generation',
            'model_id': m,
            'label': m,
            'api_key': '',
            'base_url': '',
        })
        if m == active_name:
            active_id = pid
    if not active_id and profiles:
        active_id = profiles[0]['id']
    return {'version': 2, 'active_profile_id': active_id, 'profiles': profiles}


def _normalize_v2_config(raw: dict) -> dict:
    profiles_in = raw.get('profiles')
    if not isinstance(profiles_in, list):
        profiles_in = []
    profiles = []
    seen = set()
    for p in profiles_in:
        if not isinstance(p, dict):
            continue
        pid = (p.get('id') or '').strip() or str(uuid.uuid4())
        if pid in seen:
            continue
        seen.add(pid)
        provider = (p.get('provider') or 'ollama').strip()
        bu_raw = (p.get('base_url') or '').strip() if isinstance(p.get('base_url'), str) else ''
        profiles.append({
            'id': pid,
            'provider': provider,
            'api_style': (p.get('api_style') or 'ollama').strip(),
            'model_type': (p.get('model_type') or 'test_case_generation').strip(),
            'model_id': (p.get('model_id') or '').strip(),
            'label': (p.get('label') or '').strip(),
            'api_key': p.get('api_key') if isinstance(p.get('api_key'), str) else '',
            'base_url': _normalize_profile_base_url(bu_raw, provider),
            'group_id': (p.get('group_id') or '').strip() if isinstance(p.get('group_id'), str) else '',
        })
    aid = (raw.get('active_profile_id') or '').strip()
    if aid and not any(x.get('id') == aid for x in profiles):
        aid = profiles[0]['id'] if profiles else ''
    if not aid and profiles:
        aid = profiles[0]['id']
    return {'version': 2, 'active_profile_id': aid, 'profiles': profiles}


def _load_ai_model_config() -> dict:
    """
    读取 ai_model_registry.json。v1 仅含 local_models，将迁移为 v2 profiles。
    """
    with _ai_model_cfg_lock:
        defaults = _default_ai_model_config()
        cfg_path = ai_model_registry_path()
        if not cfg_path.is_file():
            return _migrate_v1_config_to_v2({}, defaults)
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            if not isinstance(raw, dict):
                return _migrate_v1_config_to_v2({}, defaults)
            if int(raw.get('version') or 0) >= 2 or 'profiles' in raw:
                cfg = _normalize_v2_config(raw)
                dirty = False
                raw_profiles = raw.get('profiles') if isinstance(raw.get('profiles'), list) else []
                by_id = {p.get('id'): p for p in raw_profiles if isinstance(p, dict) and p.get('id')}
                for p in cfg.get('profiles') or []:
                    old = by_id.get(p.get('id')) or {}
                    old_bu = (old.get('base_url') or '').strip() if isinstance(old.get('base_url'), str) else ''
                    if old_bu != (p.get('base_url') or ''):
                        dirty = True
                        break
                if dirty:
                    to_write = dict(cfg)
                    to_write['version'] = 2
                    cfg_path.parent.mkdir(parents=True, exist_ok=True)
                    with open(cfg_path, 'w', encoding='utf-8') as wf:
                        json.dump(to_write, wf, ensure_ascii=False, indent=2)
                return cfg
            return _migrate_v1_config_to_v2(raw, defaults)
        except Exception:
            return _migrate_v1_config_to_v2({}, defaults)


def _save_ai_model_config(cfg: dict) -> None:
    with _ai_model_cfg_lock:
        if 'profiles' in cfg or int(cfg.get('version') or 0) >= 2:
            to_write = _normalize_v2_config(cfg)
            to_write['version'] = 2
        else:
            to_write = cfg
        cfg_path = ai_model_registry_path()
        cfg_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cfg_path, 'w', encoding='utf-8') as f:
            json.dump(to_write, f, ensure_ascii=False, indent=2)


def _mask_profile_for_api(p: dict) -> dict:
    if not isinstance(p, dict):
        return {}
    key = p.get('api_key')
    has_key = bool((key or '').strip()) if isinstance(key, str) else False
    out = {k: v for k, v in p.items() if k != 'api_key'}
    out['has_api_key'] = has_key
    return out


def _resolve_inference_profile(selected: str) -> tuple:
    """
    返回 (profile 或 None, 纯 Ollama 模型名字符串)。
    选中云端 profile 时 profile 非空；仅本地 Ollama 字符串模式时 profile 为空、第二项为模型名。
    """
    cfg = _load_ai_model_config()
    profiles = cfg.get('profiles') or []
    sel = (selected or '').strip()
    if not profiles:
        return (None, sel or os.environ.get('LOCAL_LLM_MODEL_MID', 'llama3:8b-instruct'))
    if not sel:
        sel = (cfg.get('active_profile_id') or '').strip()
    for p in profiles:
        if p.get('id') == sel:
            return (p, '')
    if sel:
        for p in profiles:
            if p.get('provider') == 'ollama' and p.get('model_id') == sel:
                return (p, '')
        return (None, sel)
    return (None, os.environ.get('LOCAL_LLM_MODEL_MID', 'llama3:8b-instruct'))


def _get_active_local_model() -> str:
    """当前默认：优先 active_profile_id，兼容旧版 active_local_model / 环境变量。"""
    cfg = _load_ai_model_config()
    profiles = cfg.get('profiles') or []
    if profiles:
        aid = (cfg.get('active_profile_id') or '').strip()
        if aid:
            return aid
        return profiles[0].get('id') or ''
    return (cfg.get('active_local_model') or '').strip() or os.environ.get('LOCAL_LLM_MODEL_MID', 'llama3:8b-instruct')


def _ai_memory_context_block(
    user_id: int,
    goal: str,
    probe_url: str = "",
    project_name: str = "",
) -> str:
    """LOCAL_MEMORY_ENABLE=1 时从 SQLite 向量记忆检索相似片段，拼入本地 LLM 提示。"""
    try:
        from ai_memory_store import build_query_for_case, format_memory_block, memory_enabled, search

        if not memory_enabled():
            return ""
        _db = Database()
        tid = _db.get_user_tenant_id(user_id)
        q = build_query_for_case(goal, probe_url=probe_url, project_name=project_name)
        hits = search(user_id, q, tenant_id=tid)
        return format_memory_block(hits)
    except Exception as e:
        uat_logger.debug("ai memory search skipped: %s", e)
        return ""


def _ai_build_dom_pack(snap, embed_remote: bool = False) -> str:
    """LOCAL_AI_DOM_PACK=1 时生成分区 DOM 包；内嵌画布远程会话时不拉本地页 a11y（避免串页）。"""
    if not isinstance(snap, dict) or not snap:
        return ""
    try:
        from ai_page_probe import dom_context_pack, dom_context_pack_enabled

        if not dom_context_pack_enabled():
            return ""
        a11y = ""
        if not embed_remote and (os.environ.get("LOCAL_AI_DOM_A11Y", "1").strip().lower() not in ("0", "false", "no")):
            if sync_automation_session_usable():
                try:
                    a11y = sync_get_accessibility_outline_text(48) or ""
                except Exception as e:
                    uat_logger.debug("accessibility outline skipped: %s", e)
        return dom_context_pack(snap, a11y_outline=a11y) or ""
    except Exception as e:
        uat_logger.debug("dom context pack skipped: %s", e)
        return ""


def _ai_str(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _norm_captcha_max_attempts(raw, *, for_verify: bool = True) -> Optional[int]:
    """verify 步骤最大自动验证次数：1–20；空/0 表示使用全局 CAPTCHA_SOLVE_RETRY。"""
    if not for_verify:
        return None
    if raw is None or raw == '':
        return None
    try:
        n = int(raw)
    except (TypeError, ValueError):
        return None
    if n < 1:
        return None
    return max(1, min(n, 20))


def _norm_click_repeat_count(raw) -> int:
    """点击步骤连续执行次数：1–99，非法或空视为 1。"""
    if raw is None or raw == '':
        return 1
    try:
        n = int(raw)
    except (TypeError, ValueError):
        return 1
    if n < 1:
        n = 1
    if n > 99:
        n = 99
    return n


def _ai_step_to_db_kwargs(step: dict, case_id: int, step_order: int) -> dict:
    """将 AI 步骤转为 create_test_step 参数；navigate 的 URL 写入 url 与 input_value。"""
    action = _ai_str(step.get("action"))
    st = _ai_str(step.get("selector_type"))
    sv = _ai_str(step.get("selector_value"))
    iv = _ai_str(step.get("input_value"))
    desc = _ai_str(step.get("description"))
    url_col = ""
    if action == "navigate":
        url_col = iv or sv
        st, sv = "", ""
        iv = url_col
    lc = step.get("locator_candidates")
    if lc is not None and not isinstance(lc, str):
        try:
            lc = json.dumps(lc, ensure_ascii=False)
        except Exception:
            lc = ""
    elif lc is None:
        lc = ""
    else:
        lc = str(lc).strip()
    crc = 1
    if action == "click":
        crc = _norm_click_repeat_count(step.get("click_repeat_count"))
    cmp_out = "equals"
    if action == "assert":
        from auth_batch_helpers import normalize_assert_compare_type

        cmp_out = normalize_assert_compare_type(
            _ai_str(step.get("compare_type")) or "text_contains",
            selector_value=sv,
            input_value=iv,
        )
    layer = _ai_str(step.get("automation_layer")).lower() or "web"
    if layer not in ("web", "desktop", "android"):
        layer = "web"
    strategy = _ai_str(step.get("strategy")) or _ai_str(step.get("selector_type")) or "accessibility_id"
    if layer == "android" and not _ai_str(step.get("selector_type")):
        st = strategy
    ms = step.get("mobile_spec")
    if ms is not None and not isinstance(ms, str):
        try:
            ms = json.dumps(ms, ensure_ascii=False)
        except Exception:
            ms = ""
    elif ms is None:
        ms = ""
    else:
        ms = str(ms).strip()
    return {
        "case_id": case_id,
        "action": action,
        "selector_type": st,
        "selector_value": sv,
        "input_value": iv,
        "description": desc,
        "step_order": step_order,
        "page_name": "",
        "swipe_x": "",
        "swipe_y": "",
        "url": url_col,
        "enter_iframe": False,
        "iframe_selector": "",
        "compare_type": cmp_out,
        "locator_candidates": lc,
        "click_repeat_count": crc,
        "automation_layer": layer,
        "mobile_spec": ms,
    }

# ==================== Flask-Login 初始化 ====================
login_manager = LoginManager(app)
login_manager.login_view = 'login_page'
login_manager.login_message = '请先登录'


@login_manager.unauthorized_handler
def _api_unauthorized():
    """API 未登录时返回 JSON，避免前端 fetch().json() 解析到登录页 HTML。"""
    if (request.path or "").startswith("/api/"):
        return jsonify({"success": False, "error": "未登录，请重新登录"}), 401
    return redirect(url_for("login_page", next=request.path))


@app.errorhandler(404)
def _http_not_found(e):
    """API 路径不存在时返回 JSON，避免前端解析到 HTML 404 页。"""
    if (request.path or "").startswith("/api/"):
        return jsonify({
            "success": False,
            "error": "接口不存在，请确认已重启后端服务（python app.py）",
        }), 404
    return e


class UserModel(UserMixin):
    """Flask-Login 用户模型"""
    def __init__(self, user_data: dict):
        self.id = user_data['id']
        self.username = user_data['username']
        self.role = user_data['role']
        self.is_active_flag = bool(user_data.get('is_active', 1))

    def get_id(self):
        return str(self.id)

    @property
    def is_active(self):
        return self.is_active_flag

@login_manager.user_loader
def load_user(user_id):
    _db = Database()
    user_data = _db.get_user_by_id(int(user_id))
    if user_data:
        return UserModel(user_data)
    return None

def role_required(*roles):
    """角色权限装饰器"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({'success': False, 'error': '未登录'}), 401
            if current_user.role not in roles:
                return jsonify({'success': False, 'error': '权限不足'}), 403
            return func(*args, **kwargs)
        return wrapper
    return decorator

def token_or_login_required(func):
    """支持 Bearer Token 或 Flask-Login Session 两种认证方式（用于 Webhook/CI）"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            token_val = auth_header[7:]
            _db = Database()
            token_info = _db.get_token_by_value(token_val)
            if not token_info:
                return jsonify({'success': False, 'error': '无效的 API Token'}), 401
            import datetime
            if token_info.get('expires_at'):
                try:
                    exp = datetime.datetime.strptime(token_info['expires_at'], '%Y-%m-%d %H:%M:%S')
                    if exp < datetime.datetime.now():
                        return jsonify({'success': False, 'error': 'API Token 已过期'}), 401
                except Exception:
                    pass
            return func(*args, **kwargs)
        if current_user.is_authenticated:
            return func(*args, **kwargs)
        return jsonify({'success': False, 'error': '未认证'}), 401
    return wrapper


def project_access_required(min_role='viewer'):
    """项目访问权限装饰器 - 检查用户是否有指定项目的访问权限"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({'success': False, 'error': '未登录'}), 401

            # 从参数中获取 project_id
            project_id = kwargs.get('project_id') or (
                request.view_args.get('project_id') if request.view_args else None
            )
            if not project_id:
                project_id = request.args.get('project_id', type=int)
            if not project_id:
                _body = request.get_json(silent=True) or {}
                pid = _body.get('project_id')
                if pid is not None:
                    try:
                        project_id = int(pid)
                    except (TypeError, ValueError):
                        project_id = None

            if project_id:
                _db = Database()
                if not _db.check_project_access(current_user.id, project_id, min_role):
                    return jsonify({'success': False, 'error': '无权限访问此项目'}), 403

            return func(*args, **kwargs)
        return wrapper
    return decorator


def audit_log(action: str, target_type: str):
    """审计日志装饰器 - 记录用户操作"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)

            # 异步记录审计日志（不阻塞主流程）
            try:
                if current_user.is_authenticated:
                    _db = Database()
                    target_id = None
                    # 尝试从 kwargs 或请求中获取目标ID
                    if 'id' in kwargs:
                        target_id = kwargs['id']
                    elif request.view_args and 'id' in request.view_args:
                        target_id = request.view_args['id']

                    details = None
                    if request.is_json:
                        try:
                            details = json.dumps(request.get_json(silent=True) or {}, ensure_ascii=False)
                        except (TypeError, ValueError):
                            pass

                    _db.add_audit_log(
                        user_id=current_user.id,
                        username=current_user.username,
                        action=action,
                        target_type=target_type,
                        target_id=target_id,
                        details=details,
                        ip_address=request.remote_addr
                    )
            except Exception as e:
                uat_logger.error(f"记录审计日志失败: {e}")

            return result
        return wrapper
    return decorator


def _wants_license_gate_html_redirect() -> bool:
    """浏览器直接打开页面（非 /api、非 XHR/fetch JSON）时跳转 License 页。"""
    if request.path.startswith('/api/'):
        return False
    if request.method not in ('GET', 'HEAD'):
        return False
    # fetch / XHR
    if (request.headers.get('X-Requested-With') or '').lower() == 'xmlhttprequest':
        return False
    mode = (request.headers.get('Sec-Fetch-Mode') or '').lower()
    if mode in ('cors', 'same-origin', 'no-cors'):
        return False
    if mode == 'navigate':
        return True
    accept = (request.headers.get('Accept') or 'text/html').lower()
    # 明确只要 JSON
    if accept.startswith('application/json'):
        return False
    return 'text/html' in accept


def feature_required(feature_name: str):
    """功能可用性检查装饰器 - 检查某功能是否在当前 License 中可用。

    - ``/api/*`` 与 XHR/fetch：返回 JSON 403（``error_code=LICENSE_FEATURE_REQUIRED``）
    - 浏览器打开页面（如 /audit-logs、/sso-settings）：重定向到 ``/license?gate=...&denied=1``
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            if not license_manager.check_feature_available(feature_name):
                gate = license_manager.describe_feature_gate(feature_name)
                error = license_manager.build_feature_denied_message(feature_name)
                upgrade_url = '/license?gate=' + str(feature_name) + '&denied=1'
                if _wants_license_gate_html_redirect():
                    from flask import redirect

                    return redirect(upgrade_url)
                return jsonify({
                    'success': False,
                    'ok': False,
                    'error': error,
                    'error_code': 'LICENSE_FEATURE_REQUIRED',
                    'feature': feature_name,
                    'gate': gate,
                    'upgrade_url': upgrade_url,
                }), 403
            return func(*args, **kwargs)
        return wrapper
    return decorator


def check_limit(limit_type: str, get_current_value_func=None):
    """限制检查装饰器 - 检查是否超出使用限制"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            current_value = 0
            if get_current_value_func:
                try:
                    current_value = get_current_value_func()
                except Exception:
                    current_value = 0

            result = license_manager.check_limit(limit_type, current_value)
            if not result['allowed']:
                return jsonify({
                    'success': False,
                    'error': result['message'],
                    'limit': result['limit'],
                    'current': result['current']
                }), 403

            return func(*args, **kwargs)
        return wrapper
    return decorator

# 初始化数据库
db = Database()

register_deployment_hooks(app, Database, UserModel)
init_server_instance(Database)

try:
    from mobile_routes import register_mobile_routes

    register_mobile_routes(
        app,
        api_error_handler=api_error_handler,
        log_api_request=log_api_request,
        role_required=role_required,
    )
except ImportError:
    pass

try:
    from ai_modules.explore.hub_routes import explore_bp

    app.register_blueprint(explore_bp)
except ImportError:
    pass

# 主页路由
@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/create_project')
@login_required
def create_project_landing():
    """与首页「新建项目」链接一致；支持新标签页打开、收藏夹与直链，避免 404。"""
    return redirect(f"{url_for('index')}?openCreateProject=1")

# ==================== 用户认证路由 ====================

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/register')
def register_page():
    mode = (request.args.get('mode') or '').strip().lower()
    if mode == 'email':
        return render_template('register_email.html')
    return render_template('register.html')

@app.route('/forgot-password')
def forgot_password_page():
    mode = (request.args.get('mode') or '').strip().lower()
    if mode == 'email':
        return render_template('forgot_password_email.html')
    return render_template('forgot_password.html')

@app.route('/profile')
@login_required
def profile_page():
    return render_template('profile.html')

@app.route('/settings')
@login_required
def settings_center_page():
    return render_template('settings.html')


@app.route('/plugin-market')
@login_required
def plugin_market_page():
    return render_template('plugin_market.html')


@app.route('/api/plugin-market/catalog', methods=['GET'])
@login_required
@api_error_handler
def api_plugin_market_catalog():
    from web_capture.plugin_market import get_plugin_catalog

    origin = (request.host_url or '').rstrip('/')
    return jsonify({'success': True, 'plugins': get_plugin_catalog(platform_origin=origin)})


@app.route('/api/plugin-market/install', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_plugin_market_install():
    data = request.get_json(silent=True) or {}
    from web_capture.plugin_market import install_plugin

    plugin_id = (data.get('plugin_id') or data.get('id') or '').strip()
    if not plugin_id and data.get('browser'):
        plugin_id = f"web-capture-{(data.get('browser') or '').strip().lower()}"
    result = install_plugin(plugin_id)
    code = 200 if result.get('success') or result.get('async') else 400
    return jsonify(result), code


@app.route('/api/plugin-market/install/job/<job_id>', methods=['GET'])
@login_required
@api_error_handler
def api_plugin_market_install_job(job_id):
    from plugin_install_jobs import get_job

    job = get_job(job_id)
    if not job:
        return jsonify({'ok': False, 'error': '安装任务不存在或已过期'}), 404
    return jsonify({'ok': True, 'job': job})


@app.route('/api/plugin-market/install/active', methods=['GET'])
@login_required
@api_error_handler
def api_plugin_market_install_active():
    from plugin_install_jobs import list_active_jobs

    return jsonify({'success': True, 'jobs': list_active_jobs()})


@app.route('/api/plugin-market/status', methods=['GET'])
@login_required
@api_error_handler
def api_plugin_market_status():
    from web_capture.plugin_market import get_plugin_catalog, get_preferred_browser_for_capture

    origin = (request.host_url or '').rstrip('/')
    plugins = get_plugin_catalog(platform_origin=origin)
    installed = [p for p in plugins if p.get('installed')]
    return jsonify({
        'success': True,
        'plugins': plugins,
        'preferred_browser': get_preferred_browser_for_capture(),
        'has_capture_plugin': bool(installed),
    })


@app.route('/api/plugin-market/prepare', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_plugin_market_prepare():
    """兼容旧接口：转为一键安装。"""
    data = request.get_json(silent=True) or {}
    from web_capture.plugin_market import install_plugin

    browser = (data.get('browser') or 'chrome').strip().lower()
    result = install_plugin(f"web-capture-{browser}")
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/schedules')
@login_required
@feature_required('schedule')
def schedules_page():
    return render_template('schedules.html')

@app.route('/notifications')
@login_required
@role_required('admin')
def notifications_page():
    return render_template('notifications.html')

def _basic_email_format(email: str) -> bool:
    if not email or '@' not in email:
        return False
    local, domain = email.rsplit('@', 1)
    if not local or not domain or '.' not in domain:
        return False
    return True


def _allow_local_auth() -> bool:
    """本机桌面 / 单机部署允许免 SMTP 注册与找回密钥重置（零外部成本）。"""
    try:
        from deployment_config import is_client_mode, is_standalone_mode, is_local_standalone_desktop

        return bool(is_client_mode() or is_standalone_mode() or is_local_standalone_desktop())
    except Exception:
        return True


def _generate_recovery_key() -> str:
    import secrets

    # 形如 XXXX-XXXX-XXXX-XXXX，便于手抄/保管
    raw = secrets.token_hex(8).upper()
    return "-".join(raw[i : i + 4] for i in range(0, 16, 4))


def _normalize_recovery_key(raw: str) -> str:
    """兼容带/不带横线的找回密钥输入。"""
    s = (raw or "").strip().upper().replace(" ", "").replace("-", "")
    if len(s) == 16 and all(c.isalnum() for c in s):
        return "-".join(s[i : i + 4] for i in range(0, 16, 4))
    return (raw or "").strip().upper().replace(" ", "")


@app.route('/api/auth/register/local', methods=['POST'])
@api_error_handler
def api_register_local():
    """免邮箱/SMTP 本地注册：用户名 + 密码，返回一次性找回密钥。"""
    if not _allow_local_auth():
        return jsonify({'success': False, 'error': '当前部署模式不支持本地简易注册'}), 403

    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    password = (body.get('password') or '').strip()
    confirm_password = (body.get('confirm_password') or '').strip()

    if not username:
        return jsonify({'success': False, 'error': '用户名不能为空'}), 400
    if not password:
        return jsonify({'success': False, 'error': '密码不能为空'}), 400
    if len(username) < 2 or len(username) > 64:
        return jsonify({'success': False, 'error': '用户名长度须为 2-64 个字符'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'error': f'密码长度不能少于 6 位（当前长度：{len(password)}）'}), 400
    if password != confirm_password:
        return jsonify({'success': False, 'error': '两次输入的密码不一致'}), 400

    _db = Database()
    if _db.get_user_by_username(username):
        return jsonify({'success': False, 'error': '用户名已被注册'}), 409

    recovery_key = _generate_recovery_key()
    role = 'admin' if _db.count_users() == 0 else 'tester'
    user_id = _db.create_user(
        username,
        generate_password_hash(password),
        email=None,
        role=role,
        recovery_key_hash=generate_password_hash(recovery_key),
    )
    if user_id is None:
        return jsonify({'success': False, 'error': '注册失败，请稍后重试'}), 500

    user_data = _db.get_user_by_id(user_id)
    login_user(UserModel(user_data))
    _db.update_user_last_login(user_id)

    uat_logger.info("本地简易注册: %s (role=%s)", username, role)
    try:
        from auth_audit import ACTION_REGISTER, record_auth_audit

        record_auth_audit(
            action=ACTION_REGISTER,
            username=username,
            user_id=user_id,
            ip_address=request.remote_addr,
            details={'method': 'local_simple', 'role': role},
            db=_db,
        )
    except Exception:
        pass
    return jsonify({
        'success': True,
        'message': '注册成功',
        'user': {'id': user_id, 'username': username, 'role': role},
        'role': role,
        'recovery_key': recovery_key,
    })


@app.route('/api/auth/forgot-password/recovery-reset', methods=['POST'])
@api_error_handler
def api_forgot_password_recovery_reset():
    """用注册时的找回密钥重置密码（无需 SMTP）。"""
    if not _allow_local_auth():
        return jsonify({'success': False, 'error': '当前部署模式不支持密钥找回'}), 403

    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    recovery_key = _normalize_recovery_key(body.get('recovery_key') or '')
    new_password = (body.get('new_password') or '').strip()
    confirm_password = (body.get('confirm_password') or '').strip()

    if not username or not recovery_key or not new_password:
        return jsonify({'success': False, 'error': '请填写用户名、找回密钥和新密码'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'error': '密码长度不能少于 6 位'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'error': '两次输入的密码不一致'}), 400

    _db = Database()
    user_data = _db.get_user_by_username(username)
    if not user_data:
        return jsonify({'success': False, 'error': '用户名或找回密钥不正确'}), 400
    key_hash = user_data.get('recovery_key_hash')
    if not key_hash or not check_password_hash(key_hash, recovery_key):
        return jsonify({'success': False, 'error': '用户名或找回密钥不正确'}), 400

    # 重置密码后轮换找回密钥，旧密钥失效
    new_recovery_key = _generate_recovery_key()
    ok = _db.update_user(
        user_data['id'],
        password_hash=generate_password_hash(new_password),
        recovery_key_hash=generate_password_hash(new_recovery_key),
    )
    if not ok:
        return jsonify({'success': False, 'error': '重置失败，请稍后重试'}), 500

    uat_logger.info("用户 %s 已通过找回密钥重置密码", username)
    try:
        from auth_audit import ACTION_PASSWORD_RESET, record_auth_audit

        record_auth_audit(
            action=ACTION_PASSWORD_RESET,
            username=username,
            user_id=user_data['id'],
            ip_address=request.remote_addr,
            details={'method': 'recovery_key'},
            db=_db,
        )
    except Exception:
        pass
    return jsonify({
        'success': True,
        'message': '密码已重置，请使用新密码登录，并妥善保存新的找回密钥',
        'recovery_key': new_recovery_key,
    })


@app.route('/api/auth/register/send-code', methods=['POST'])
@api_error_handler
def api_register_send_code():
    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    email = (body.get('email') or '').strip()
    password = (body.get('password') or '').strip()
    confirm_password = (body.get('confirm_password') or '').strip()
    smtp_host = (body.get('smtp_host') or '').strip()
    smtp_port = body.get('smtp_port', 587)
    smtp_username = (body.get('smtp_username') or '').strip()
    smtp_password = (body.get('smtp_password') or '').strip()
    smtp_use_tls = int(body.get('smtp_use_tls', 1))

    if not username:
        return jsonify({'success': False, 'error': '用户名不能为空'}), 400
    if not email:
        return jsonify({'success': False, 'error': '邮箱不能为空'}), 400
    if not password:
        return jsonify({'success': False, 'error': '密码不能为空'}), 400
    if len(username) < 2 or len(username) > 64:
        return jsonify({'success': False, 'error': '用户名长度须为 2-64 个字符'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'error': f'密码长度不能少于 6 位（当前长度：{len(password)}）'}), 400
    if password != confirm_password:
        return jsonify({'success': False, 'error': '两次输入的密码不一致'}), 400
    if not _basic_email_format(email):
        return jsonify({'success': False, 'error': '邮箱格式不正确'}), 400
    if not smtp_host or not smtp_username or not smtp_password:
        return jsonify({'success': False, 'error': '请填写完整的 SMTP 配置'}), 400

    _db = Database()
    if _db.get_user_by_username(username):
        return jsonify({'success': False, 'error': '用户名已被注册'}), 409
    if _db.get_user_by_email(email):
        return jsonify({'success': False, 'error': '邮箱已被注册'}), 409

    smtp_config = {
        'host': smtp_host,
        'port': int(smtp_port),
        'username': smtp_username,
        'password': smtp_password,
        'use_tls': smtp_use_tls,
        'sender_email': smtp_username,
    }
    result = send_verify_code(email, smtp_config, 'register')
    return jsonify(result), 200 if result['success'] else 400


@app.route('/api/auth/register/confirm', methods=['POST'])
@api_error_handler
def api_register_confirm():
    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    email = (body.get('email') or '').strip()
    password = (body.get('password') or '').strip()
    code = (body.get('code') or '').strip()
    smtp_host = (body.get('smtp_host') or '').strip()
    smtp_port = body.get('smtp_port', 587)
    smtp_username = (body.get('smtp_username') or '').strip()
    smtp_password = (body.get('smtp_password') or '').strip()
    smtp_use_tls = int(body.get('smtp_use_tls', 1))

    if not username or not password or not email or not code:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    verify_result = verify_code(email, code, 'register')
    if not verify_result['success']:
        return jsonify(verify_result), 400

    _db = Database()
    if _db.get_user_by_username(username):
        return jsonify({'success': False, 'error': '用户名已被注册'}), 409
    if _db.get_user_by_email(email):
        return jsonify({'success': False, 'error': '邮箱已被注册'}), 409

    role = 'admin' if _db.count_users() == 0 else 'tester'
    user_id = _db.create_user(username, generate_password_hash(password), email=email, role=role)
    if user_id is None:
        return jsonify({'success': False, 'error': '注册失败，请稍后重试'}), 500

    _db.save_user_smtp_config(user_id, email, smtp_host, int(smtp_port),
                               smtp_username, smtp_password, smtp_use_tls)

    user_data = _db.get_user_by_id(user_id)
    login_user(UserModel(user_data))
    _db.update_user_last_login(user_id)

    uat_logger.info(f"新用户注册: {username} (role={role}, email={email})")
    return jsonify({
        'success': True,
        'message': '注册成功',
        'user': {'id': user_id, 'username': username, 'role': role},
        'role': role,
    })


@app.route('/api/auth/forgot-password/send-code', methods=['POST'])
@api_error_handler
@log_api_request
def api_forgot_password_send_code():
    body = request.get_json(silent=True) or {}
    email = (body.get('email') or '').strip().lower()

    if not email or not _basic_email_format(email):
        uat_logger.info("找回密码: 邮箱格式无效=%s", email)
        return jsonify({'success': True, 'message': '若邮箱已注册，验证码已发送'})

    _db = Database()
    user_data = _db.get_user_by_email(email)
    if not user_data:
        uat_logger.info("找回密码: 邮箱未注册=%s", email)
        return jsonify({'success': True, 'message': '若邮箱已注册，验证码已发送'})

    smtp_data = _db.get_user_smtp_config_by_email(email)
    if not smtp_data:
        smtp_data = _db.get_user_smtp_config_by_user_id(user_data['id'])
    if not smtp_data:
        uat_logger.warning(
            "找回密码: 用户 %s (id=%s) 缺少 SMTP 配置，无法发送验证码",
            user_data.get('username', ''), user_data.get('id'),
        )
        return jsonify({
            'success': False,
            'error': '该账号未保存 SMTP 邮件配置，无法发送验证码。请联系管理员。',
        }), 400

    smtp_config = {
        'host': smtp_data['host'],
        'port': smtp_data['port'],
        'username': smtp_data['username'],
        'password': smtp_data['password'],
        'use_tls': smtp_data['use_tls'],
        'sender_email': smtp_data['username'],
    }
    uat_logger.info(
        "找回密码发送验证码: to=%s host=%s port=%s sender=%s use_tls=%s user=%s",
        email, smtp_config['host'], smtp_config['port'],
        smtp_config['sender_email'], smtp_config['use_tls'],
        user_data.get('username', ''),
    )
    is_self_send = (smtp_config['sender_email'].lower() == email.lower())
    if is_self_send:
        uat_logger.warning(
            "找回密码：发件人和收件人相同 (%s)，部分邮箱（163/126/QQ）可能静默丢弃此类邮件",
            email,
        )
    result = send_verify_code(email, smtp_config, 'reset_password')
    if not result.get('success'):
        uat_logger.warning(f"找回密码邮件发送失败: {email} - {result.get('message')}")
        return jsonify({
            'success': False,
            'error': result.get('message', '邮件发送失败，请稍后重试'),
        }), 500
    hint = ""
    if is_self_send:
        hint = "（注意：163/126等邮箱可能不收自己发给自己的邮件，如未收到请检查垃圾箱或使用其他邮箱）"
    return jsonify({'success': True, 'message': '验证码已发送，请查收邮箱' + hint})


@app.route('/api/auth/forgot-password/reset', methods=['POST'])
@api_error_handler
@log_api_request
def api_forgot_password_reset():
    body = request.get_json(silent=True) or {}
    email = (body.get('email') or '').strip().lower()
    code = (body.get('code') or '').strip()
    new_password = (body.get('new_password') or '').strip()
    confirm_password = (body.get('confirm_password') or '').strip()

    if not email:
        return jsonify({'success': False, 'error': '邮箱不能为空'}), 400
    if not code:
        return jsonify({'success': False, 'error': '验证码不能为空'}), 400
    if not new_password:
        return jsonify({'success': False, 'error': '新密码不能为空'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'error': f'密码长度不能少于 6 位（当前长度：{len(new_password)}）'}), 400
    if new_password != confirm_password:
        return jsonify({'success': False, 'error': '两次输入的密码不一致'}), 400

    verify_result = verify_code(email, code, 'reset_password')
    if not verify_result['success']:
        return jsonify(verify_result), 400

    _db = Database()
    user_data = _db.get_user_by_email(email)
    if not user_data:
        return jsonify({'success': False, 'error': '用户不存在'}), 404

    _db.update_user(user_data['id'], password_hash=generate_password_hash(new_password))
    uat_logger.info(f"用户 {user_data['username']} 已通过邮箱找回密码")
    return jsonify({'success': True, 'message': '密码已重置，请使用新密码登录'})


@app.route('/api/auth/smtp-test', methods=['POST'])
@api_error_handler
def api_smtp_test():
    body = request.get_json(silent=True) or {}
    host = (body.get('host') or '').strip()
    port = body.get('port', 587)
    username = (body.get('username') or '').strip()
    password = (body.get('password') or '').strip()
    use_tls = int(body.get('use_tls', 1))
    to_email = (body.get('to_email') or '').strip()

    if not host:
        return jsonify({'success': False, 'message': '请填写 SMTP 服务器地址'}), 400
    if not username:
        return jsonify({'success': False, 'message': '请填写邮箱地址'}), 400
    if not password:
        return jsonify({'success': False, 'message': '请填写 SMTP 授权码'}), 400
    if not to_email:
        return jsonify({'success': False, 'message': '请填写测试收件邮箱'}), 400

    smtp_config = {'host': host, 'port': int(port), 'username': username,
                   'password': password, 'use_tls': use_tls, 'sender_email': username}
    result = send_verify_code(to_email, smtp_config, 'register')
    return jsonify(result), 200 if result['success'] else 400


@app.route('/api/auth/login', methods=['POST'])
@api_error_handler
def api_login():
    ip = _login_client_ip()
    blocked, retry_after = _login_is_rate_limited(ip)
    if blocked:
        resp = jsonify({"success": False, "error": "登录尝试过于频繁，请稍后再试"})
        resp.headers["Retry-After"] = str(retry_after)
        return resp, 429

    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400
    _db = Database()
    user_data = _db.get_user_by_username(username)
    if not user_data or not check_password_hash(user_data['password_hash'], password):
        _login_record_failure(ip)
        try:
            from auth_audit import ACTION_LOGIN_FAILURE, record_auth_audit

            record_auth_audit(
                action=ACTION_LOGIN_FAILURE,
                username=username,
                user_id=0,
                ip_address=ip,
                details={'method': 'password', 'reason': 'bad_credentials'},
                db=_db,
            )
        except Exception:
            pass
        return jsonify({'success': False, 'error': '用户名或密码错误'}), 401
    if not user_data.get('is_active', 1):
        try:
            from auth_audit import ACTION_LOGIN_FAILURE, record_auth_audit

            record_auth_audit(
                action=ACTION_LOGIN_FAILURE,
                username=username,
                user_id=user_data.get('id'),
                ip_address=ip,
                details={'method': 'password', 'reason': 'disabled'},
                db=_db,
            )
        except Exception:
            pass
        return jsonify({'success': False, 'error': '账号已被禁用'}), 403
    _login_clear_failures(ip)
    user = UserModel(user_data)
    remember = data.get('remember_me', False)
    if remember:
        session.permanent = True
    login_user(user, remember=remember)
    _db.update_user_last_login(user_data['id'])
    uat_logger.info(f"用户 {username} 登录成功")
    try:
        from auth_audit import ACTION_LOGIN_SUCCESS, record_auth_audit

        record_auth_audit(
            action=ACTION_LOGIN_SUCCESS,
            username=username,
            user_id=user_data['id'],
            ip_address=ip,
            details={'method': 'password', 'remember': bool(remember)},
            db=_db,
        )
    except Exception:
        pass
    try:
        from client_config_store import get_team_server_url
        from deployment_config import is_client_mode
        from platform_sync import sync_product_user

        sync_product_user(
            user_data["id"],
            username,
            email=user_data.get("email") or "",
            team_server_url=get_team_server_url() if is_client_mode() else "",
            license_type=license_manager.get_current_license().license_type,
        )
    except Exception:
        pass
    return jsonify({'success': True, 'user': {'id': user_data['id'], 'username': username, 'role': user_data['role']}})

@app.route('/api/auth/logout', methods=['POST'])
@login_required
def api_logout():
    username = current_user.username
    uid = getattr(current_user, 'id', None)
    try:
        from auth_audit import ACTION_LOGOUT, record_auth_audit

        record_auth_audit(
            action=ACTION_LOGOUT,
            username=username,
            user_id=uid,
            ip_address=request.remote_addr,
            details={'method': 'session'},
        )
    except Exception:
        pass
    logout_user()
    uat_logger.info(f"用户 {username} 已注销")
    return jsonify({'success': True})

@app.route('/api/auth/me', methods=['GET'])
@login_required
def api_me():
    """获取当前用户信息（包含权限和License信息）"""
    # 获取当前 License 信息
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    
    # 获取今日使用统计
    _db = Database()
    today_stats = _db.get_user_usage_stats(current_user.id)
    
    return jsonify({
        'success': True, 
        'user': {
            'id': current_user.id, 
            'username': current_user.username, 
            'role': current_user.role
        },
        'deployment': {
            'mode': __import__('deployment_config').get_deployment_mode().value,
            'team_server_url': __import__('client_config_store').get_team_server_url() if is_client_mode() else '',
        },
        'license': {
            'type': license_info.license_type,
            # 导航门控用 effective_features（含档位目录合并 / 开源试用解锁）
            'features': limits.get('effective_features') or limits.get('features') or [],
            'certificate_features': limits.get('features') or [],
            'effective_features': limits.get('effective_features') or [],
            'limits': {
                'max_projects': limits['max_projects'],
                'max_cases_per_project': limits['max_cases_per_project'],
                'max_executions_per_day': limits['max_executions_per_day']
            }
        },
        'usage': {
            'today_executions': today_stats.get('execution_count', 0) if today_stats else 0,
            'today_created_cases': today_stats.get('created_cases', 0) if today_stats else 0
        }
    })

@app.route('/api/auth/me', methods=['PUT'])
@login_required
@api_error_handler
def api_update_me():
    """更新当前用户信息"""
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip() or None

    if email and '@' not in email:
        return jsonify({'success': False, 'error': '邮箱格式不正确'}), 400

    _db = Database()
    _db.update_user(current_user.id, email=email)
    uat_logger.info(f"用户 {current_user.username} 更新了个人信息")
    return jsonify({'success': True})

@app.route('/api/auth/me/smtp', methods=['GET'])
@login_required
def api_get_my_smtp():
    _db = Database()
    smtp = _db.get_user_smtp_config_by_user_id(current_user.id)
    if not smtp:
        return jsonify({'success': True, 'smtp': None})
    return jsonify({'success': True, 'smtp': {
        'host': smtp['host'],
        'port': smtp['port'],
        'username': smtp['username'],
        'use_tls': smtp['use_tls'],
    }})

@app.route('/api/auth/me/smtp', methods=['PUT'])
@login_required
@api_error_handler
def api_update_my_smtp():
    data = request.get_json(silent=True) or {}
    host = (data.get('host') or '').strip()
    port = int(data.get('port', 587))
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    use_tls = int(data.get('use_tls', 1))

    if not host or not username or not password:
        return jsonify({'success': False, 'error': '请填写完整的 SMTP 配置'}), 400
    if port <= 0 or port > 65535:
        return jsonify({'success': False, 'error': '无效的端口号'}), 400

    _db = Database()
    user_data = _db.get_user_by_id(current_user.id)
    email = (user_data.get('email') if user_data else None) or username
    _db.save_user_smtp_config(current_user.id, email, host, port, username, password, use_tls)
    uat_logger.info(f"用户 {current_user.username} 更新了 SMTP 配置: host={host} port={port}")
    return jsonify({'success': True, 'message': 'SMTP 配置已保存'})

@app.route('/api/auth/change_password', methods=['POST'])
@login_required
@api_error_handler
def api_change_password():
    data = request.get_json(silent=True) or {}
    old_password = (data.get('old_password') or '').strip()
    new_password = (data.get('new_password') or '').strip()
    if not old_password:
        return jsonify({'success': False, 'error': '原密码不能为空'}), 400
    if not new_password:
        return jsonify({'success': False, 'error': '新密码不能为空'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'error': f'新密码长度不能少于 6 位（当前长度：{len(new_password)}）'}), 400
    _db = Database()
    user_data = _db.get_user_by_id(current_user.id)
    if not check_password_hash(user_data['password_hash'], old_password):
        return jsonify({'success': False, 'error': '原密码错误'}), 401
    _db.update_user(current_user.id, password_hash=generate_password_hash(new_password))
    uat_logger.info(f"用户 {current_user.username} 修改了密码")
    logout_user()
    return jsonify({'success': True, 'require_relogin': True, 'message': '密码已修改，请重新登录'})

# ==================== 用户管理API（仅管理员） ====================

@app.route('/api/users', methods=['GET'])
@login_required
@role_required('admin')
def api_get_users():
    _db = Database()
    users = _db.get_all_users()
    return jsonify({'success': True, 'users': users})

@app.route('/api/users', methods=['POST'])
@login_required
@role_required('admin')
def api_create_user():
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    email = data.get('email', '').strip() or None
    role = data.get('role', 'tester')
    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400
    if role not in ('admin', 'tester', 'viewer'):
        return jsonify({'success': False, 'error': '无效的角色，可选: admin/tester/viewer'}), 400
    _db = Database()
    user_id = _db.create_user(username, generate_password_hash(password), email, role)
    if user_id is None:
        return jsonify({'success': False, 'error': '用户名或邮箱已存在'}), 409
    return jsonify({'success': True, 'user_id': user_id})

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@app.route('/api/users/<int:user_id>/update', methods=['POST'])
@login_required
@role_required('admin')
def api_update_user(user_id):
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'success': False, 'error': '用户名不能为空'}), 400
    role = data.get('role', 'tester')
    if role not in ('admin', 'tester', 'viewer'):
        return jsonify({'success': False, 'error': '无效的角色，可选: admin/tester/viewer'}), 400
    update_kwargs = {'username': username, 'role': role}
    if 'email' in data:
        email_raw = data.get('email')
        if isinstance(email_raw, str):
            update_kwargs['email'] = email_raw.strip() or None
        else:
            update_kwargs['email'] = email_raw
    if 'is_active' in data:
        update_kwargs['is_active'] = data.get('is_active')
    _db = Database()
    success = _db.update_user(user_id, **update_kwargs)
    if not success:
        return jsonify({'success': False, 'error': '更新失败：用户名或邮箱已被占用，或用户不存在'}), 409
    return jsonify({'success': True})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@app.route('/api/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def api_delete_user(user_id):
    if user_id == current_user.id:
        return jsonify({'success': False, 'error': '不能删除当前登录的账号'}), 400
    _db = Database()
    success = _db.delete_user(user_id)
    if not success:
        return jsonify({'success': False, 'error': '删除失败：用户不存在或仍存在无法自动解除的关联数据'}), 400
    return jsonify({'success': True})

@app.route('/create_case_v2')
@login_required
def create_case_v2():
    """旧版独立建用例页；模板已移除，跳转到项目列表（从项目进入「用例管理」创建用例）。"""
    return redirect(url_for('list_projects'))

@app.route('/ai-hub')
@login_required
def ai_hub_page():
    """AI 中心：用例设计 / 自主测试 / 自愈优化入口。"""
    return render_template('ai_hub.html')


@app.route('/api/ai/stats', methods=['GET'])
@login_required
@api_error_handler
def api_ai_stats():
    """AI 中心首页统计数据（真实数据）。"""
    _db = Database()
    stats = _db.get_ai_stats()
    return jsonify({
        "success": True,
        "ai_generated_cases": stats["ai_generated_cases"],
        "efficiency_boost": stats["efficiency_boost"],
        "coverage_boost": stats["coverage_boost"],
    })


@app.route('/ai-design')
@login_required
def ai_design_page():
    """AI 用例设计：需求文档 → 场景 → 批量生成。"""
    return render_template('ai_design.html')


@app.route('/ai-heal')
@login_required
def ai_heal_page():
    """AI 自愈优化：定位修复、失败诊断、步骤助手入口。"""
    return render_template('ai_heal.html')


@app.route('/cross-end')
@login_required
def cross_end_page():
    """跨端联动编排：API + Web + Mobile + Desktop 端到端测试。"""
    return render_template('cross_end.html')


@app.route('/api/ai/hub/heal/analyze-steps', methods=['POST'])
@login_required
@api_error_handler
def api_ai_hub_heal_analyze_steps():
    data = request.get_json(silent=True) or {}
    steps = data.get('steps') if isinstance(data.get('steps'), list) else []
    from ai_modules.hub_routes import hub_heal_analyze_steps

    return jsonify(hub_heal_analyze_steps(steps))


@app.route('/api/ai/hub/heal/diagnose-text', methods=['POST'])
@login_required
@api_error_handler
def api_ai_hub_heal_diagnose_text():
    data = request.get_json(silent=True) or {}
    from ai_modules.hub_routes import hub_heal_diagnose_text

    msg = (data.get('error_message') or data.get('exception_message') or '').strip()
    if not msg:
        return jsonify({'success': False, 'error': 'error_message 不能为空'}), 400
    out = hub_heal_diagnose_text(
        msg,
        step_summary=(data.get('step_summary') or '').strip(),
        url=(data.get('url') or '').strip(),
    )
    return jsonify(out)


@app.route('/api/ai/hub/design/preview', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_hub_design_preview():
    """从需求生成用例草案（不写库）。"""
    from ai_modules.hub_routes import hub_design_preview

    out, code = hub_design_preview(
        request,
        resolve_profile_fn=_resolve_inference_profile,
        get_active_model_fn=_get_active_local_model,
    )
    return jsonify(out), code


@app.route('/api/ai/hub/design/save', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@audit_log('CREATE_CASE', 'case')
def api_ai_hub_design_save():
    """将确认的 AI 设计草案保存到项目。"""
    from ai_modules.hub_routes import hub_design_save

    _db = Database()
    out, code = hub_design_save(
        request,
        db=_db,
        user_id=current_user.id,
        check_project_access_fn=_db.check_project_access,
    )
    if code == 200 and isinstance(out, dict):
        log_ai_plan_to_audit(
            current_user.id,
            current_user.username,
            'AI_DESIGN_SAVE_BATCH',
            {
                'batch_id': out.get('batch_id'),
                'count': out.get('count'),
                'created_case_ids': out.get('created_case_ids'),
            },
            request.remote_addr,
        )
    return jsonify(out), code


@app.route('/api/ai/hub/analyze/frontend-code', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_hub_analyze_frontend_code():
    """解析前端源码：识别复杂 UI 组件/标签与稳定定位线索（不落库）。"""
    from ai_modules.code_intel.ui_agent import analyze_frontend_ui

    data = request.get_json(silent=True) or {}
    snippets = data.get('file_snippets') or data.get('files') or {}
    if isinstance(snippets, list):
        # [{path, content}, ...]
        mapped = {}
        for item in snippets:
            if isinstance(item, dict) and item.get('path'):
                mapped[str(item['path'])] = item.get('content') or item.get('code') or ''
        snippets = mapped
    if not isinstance(snippets, dict):
        return jsonify({'ok': False, 'success': False, 'error': 'file_snippets 须为 {path: code}'}), 400
    if not snippets and not data.get('diff'):
        return jsonify({'ok': False, 'success': False, 'error': '请提供 file_snippets 或 diff'}), 400
    out = analyze_frontend_ui(file_snippets=snippets, diff=_ai_str(data.get('diff')))
    return jsonify({'ok': True, 'success': True, **out})


@app.route('/api/ai/hub/generate/from-frontend-code', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_hub_generate_from_frontend_code():
    """从前端代码精准识别 UI 组件，按自动化测试知识生成可靠用例草案。

    Body: project_id, file_snippets{path:code}, diff?, base_url?, git_sha?,
          extra_requirements?, save(bool), use_llm(bool)
    """
    from ai_modules.code_intel.ui_agent import generate_and_optionally_save
    from ai_modules.code_intel.policy import resolve_use_llm

    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    if project_id is None:
        return jsonify({'ok': False, 'success': False, 'error': 'project_id 必填'}), 400
    try:
        project_id = int(project_id)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'success': False, 'error': 'project_id 无效'}), 400

    _db = Database()
    try:
        if not _db.check_project_access(current_user.id, project_id):
            return jsonify({'ok': False, 'success': False, 'error': '无权访问项目'}), 403
    except Exception:
        pass

    snippets = data.get('file_snippets') or data.get('files') or {}
    if isinstance(snippets, list):
        mapped = {}
        for item in snippets:
            if isinstance(item, dict) and item.get('path'):
                mapped[str(item['path'])] = item.get('content') or item.get('code') or ''
        snippets = mapped
    if not isinstance(snippets, dict):
        snippets = {}
    if not snippets and not data.get('diff'):
        return jsonify({'ok': False, 'success': False, 'error': '请提供 file_snippets 或 diff'}), 400

    out = generate_and_optionally_save(
        _db,
        project_id=project_id,
        file_snippets=snippets,
        diff=data.get('diff') or '',
        base_url=_ai_str(data.get('base_url')),
        git_sha=_ai_str(data.get('git_sha') or data.get('commit')),
        extra_requirements=_ai_str(data.get('extra_requirements') or data.get('requirements')),
        use_llm=resolve_use_llm(data.get('use_llm')),
        save=bool(data.get('save') or data.get('persist')),
        user_id=int(current_user.id or 0),
    )
    if out.get('created_case_ids'):
        try:
            log_ai_plan_to_audit(
                current_user.id,
                current_user.username,
                'AI_FRONTEND_CODE_GEN',
                {
                    'project_id': project_id,
                    'count': out.get('count'),
                    'created_case_ids': out.get('created_case_ids'),
                    'meta': out.get('meta'),
                },
                request.remote_addr,
            )
        except Exception:
            pass
    return jsonify(out)


@app.route('/api/ai/hub/generate/cases-from-requirements', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_hub_generate_cases_from_requirements():
    """
    已弃用：仅返回草案预览。请改用 /api/ai/hub/design/preview 与 /api/ai/hub/design/save。
    """
    from ai_modules.hub_routes import hub_design_preview

    out, code = hub_design_preview(
        request,
        resolve_profile_fn=_resolve_inference_profile,
        get_active_model_fn=_get_active_local_model,
    )
    if isinstance(out, dict):
        out['deprecated'] = True
        out['hint'] = '请使用 /api/ai/hub/design/preview 生成草案，/api/ai/hub/design/save 保存到项目。'
    return jsonify(out), code


@app.route('/api/ai/hub/generate/api-from-scenarios', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@audit_log('CREATE_CASE', 'case')
def api_ai_hub_generate_api_from_scenarios():
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id不能为空'}), 400
    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403
    doc = data.get('document') if isinstance(data.get('document'), dict) else {}
    scenarios = data.get('scenarios')
    if not isinstance(scenarios, list):
        scenarios = doc.get('scenarios') if isinstance(doc.get('scenarios'), list) else []
    if not scenarios:
        return jsonify({'success': False, 'error': 'scenarios 为空'}), 400
    try:
        max_n = int(data.get('max_scenarios') or 20)
    except (TypeError, ValueError):
        max_n = 20
    max_n = max(1, min(max_n, 40))
    from ai_modules.generate.api_from_doc import batch_api_cases_from_scenarios

    return jsonify(
        batch_api_cases_from_scenarios(
            _db,
            project_id=project_id,
            scenarios=scenarios,
            max_scenarios=max_n,
            user_id=current_user.id,
        )
    )


@app.route('/api/ai/hub/generate/mobile-from-scenarios', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@audit_log('CREATE_CASE', 'case')
def api_ai_hub_generate_mobile_from_scenarios():
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id不能为空'}), 400
    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403
    doc = data.get('document') if isinstance(data.get('document'), dict) else {}
    scenarios = data.get('scenarios')
    if not isinstance(scenarios, list):
        scenarios = doc.get('scenarios') if isinstance(doc.get('scenarios'), list) else []
    if not scenarios:
        return jsonify({'success': False, 'error': 'scenarios 为空'}), 400
    selected_model = (data.get('model') or '').strip() or _get_active_local_model()
    profile, legacy_model = _resolve_inference_profile(selected_model)
    project_name = (data.get('project_name') or '').strip()
    try:
        max_n = int(data.get('max_scenarios') or 15)
    except (TypeError, ValueError):
        max_n = 15
    max_n = max(1, min(max_n, 30))
    from ai_modules.generate.mobile_from_doc import batch_mobile_cases_from_scenarios

    def _audit(plan):
        log_ai_plan_to_audit(
            current_user.id,
            current_user.username,
            'AI_PLAN_SCENARIO_BATCH_MOBILE',
            plan,
            request.remote_addr,
        )

    return jsonify(
        batch_mobile_cases_from_scenarios(
            _db,
            project_id=project_id,
            scenarios=scenarios,
            project_name=project_name,
            profile=profile,
            legacy_model=legacy_model,
            memory_context_fn=_ai_memory_context_block,
            user_id=current_user.id,
            max_scenarios=max_n,
            fill_steps_fn=local_ai_service._fill_missing_step_payloads,
            normalize_fn=apply_step_normalization_to_plan,
            step_to_db_kwargs_fn=_ai_step_to_db_kwargs,
            audit_fn=_audit,
        )
    )


@app.route('/api/ai/hub/cross-platform/scenarios', methods=['GET', 'POST'])
@login_required
@api_error_handler
def api_ai_hub_cross_platform_scenarios():
    from ai_modules.execute.orchestrator import (
        list_cross_platform_scenarios,
        save_cross_platform_scenario,
    )

    if request.method == 'GET':
        return jsonify({'success': True, 'scenarios': list_cross_platform_scenarios()})
    data = request.get_json(silent=True) or {}
    return jsonify(save_cross_platform_scenario(data))


@app.route('/api/ai/hub/cross-platform/scenarios/<scenario_id>', methods=['DELETE'])
@login_required
@api_error_handler
def api_ai_hub_cross_platform_scenario_delete(scenario_id: str):
    from ai_modules.execute.orchestrator import delete_cross_platform_scenario

    return jsonify(delete_cross_platform_scenario(scenario_id))


@app.route('/api/ai/hub/cross-platform/execute', methods=['POST'])
@login_required
@api_error_handler
def api_ai_hub_cross_platform_execute():
    data = request.get_json(silent=True) or {}
    sid = (data.get('scenario_id') or '').strip()
    if not sid:
        return jsonify({'success': False, 'error': 'scenario_id 必填'}), 400
    from ai_modules.execute.orchestrator import execute_cross_platform_scenario

    out = execute_cross_platform_scenario(sid, user_id=str(current_user.id))
    if out.get('lock') == 'busy':
        return jsonify(out), 409
    if out.get('lock') == 'unavailable':
        return jsonify(out), 503
    code = 501 if out.get('status') == 'not_implemented' else (200 if out.get('success') else 400)
    return jsonify(out), code


# AI 测试工作台（本地模型 + 内嵌预览：与后台 Playwright 自动化会话经 /api/navigate 同步）
@app.route('/ai-test')
@login_required
def ai_test_page():
    """AI 生成测试步骤；内嵌区与后台 Playwright 会话同步（可选远程画布为另一进程）。"""
    from ai_chat_tool_loop import ai_chat_tools_enabled
    from agent_gateway_client import agent_gateway_configured
    from hermes_config import hermes_cdp_attached

    resp = make_response(
        render_template(
            'ai_test.html',
            ai_chat_tools_env_enabled=ai_chat_tools_enabled(),
            hermes_gateway_configured=agent_gateway_configured(),
            openclaw_gateway_configured=agent_gateway_configured(),
            hermes_cdp_attached=hermes_cdp_attached(),
        )
    )
    # 用于核对浏览器是否命中本仓库模板（与页内 #aiTestBuildMarker 文案一致）
    resp.headers['X-AI-Test-Template'] = 'playwright-ui-dedup-2026-04-24'
    return resp


@app.route('/mobile-testing')
@login_required
def mobile_testing_page():
    """Android 移动端测试：配对/用例；Agent 与 /ai-test 同一大脑（一脑多端双手）。"""
    from ai_chat_tool_loop import ai_chat_tools_enabled
    from agent_gateway_client import agent_gateway_configured

    hermes_cfg = agent_gateway_configured()
    template_kw = dict(
        mobile_disabled=False,
        mobile_disabled_reason='',
        ai_chat_tools_env_enabled=ai_chat_tools_enabled(),
        hermes_gateway_configured=hermes_cfg,
        openclaw_gateway_configured=hermes_cfg,
    )
    try:
        from mobile_env_config import mobile_enabled, mobile_runtime_unavailable_reason

        if not mobile_enabled():
            return render_template(
                'mobile_testing.html',
                mobile_disabled=True,
                mobile_disabled_reason=mobile_runtime_unavailable_reason()
                or '请在 .env 中设置 ENABLE_MOBILE=1',
                ai_chat_tools_env_enabled=ai_chat_tools_enabled(),
                hermes_gateway_configured=hermes_cfg,
                openclaw_gateway_configured=hermes_cfg,
            )
    except ImportError:
        return render_template(
            'mobile_testing.html',
            mobile_disabled=True,
            mobile_disabled_reason='移动端模块未安装',
            ai_chat_tools_env_enabled=ai_chat_tools_enabled(),
            hermes_gateway_configured=hermes_cfg,
            openclaw_gateway_configured=hermes_cfg,
        )
    return render_template('mobile_testing.html', **template_kw)


# 项目管理页面
@app.route('/list_projects')
@login_required
def list_projects():
    return render_template('list_projects.html')


def _validate_and_fix_url(url: str) -> tuple:
    """校验并修备URL。返回 (fixed_url, error_msg)。
    fixed_url 为 None 表示应跳过； error_msg 不为 None 表示应报错。
    """
    import re
    if not url or not url.strip():
        return None, None  # 空 URL 跳过

    # 兼容全角冒号等输入，避免 about：blank 等变体漏判
    url = url.strip().replace('：', ':')

    # 跳过无意义占位符URL
    skip_patterns = [
        'example.com', '0.0.0.0', '0.0.0.1', '127.0.0.1',
        'localhost', 'about:blank', 'about:newtab',
    ]
    for pat in skip_patterns:
        if pat in url.lower():
            uat_logger.warning(f"检测到占位符URL ({url})，跳过初始导航")
            return None, None  # 跳过

    # 自动添加协议
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url

    # 判断 URL 格式是否合法
    # 🔥 修复：IP 地址正则表达式增加 100-199 范围的匹配 (1[0-9]{2})
    # IP 第一段 (1-255): (25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9][0-9]|[1-9])
    # IP 中间/最后段 (0-255): (25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])
    url_re = re.compile(
        r'^https?://'
        r'(([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}'
        r'|((25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9][0-9]|[1-9])\.){1}'
        r'((25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])\.){2}'
        r'(25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9]))'
        r'(:\d+)?(/.*)?$'
    )
    if not url_re.match(url):
        return None, f"无效的URL地址: {url}，请检查用例的测试URL配置"

    return url, None


def _get_first_valid_step_url(steps) -> tuple:
    """从步骤中提取首个有效URL。返回 (fixed_url, source_desc)。"""
    if not steps:
        return None, None

    for step in steps:
        action = (step.get('action') or '').strip().lower()
        candidates = []

        if step.get('url'):
            candidates.append(('step.url', step.get('url')))
        if action == 'navigate' and step.get('input_value'):
            candidates.append(('navigate.input_value', step.get('input_value')))

        for source, raw in candidates:
            fixed_url, url_err = _validate_and_fix_url(raw)
            if fixed_url:
                return fixed_url, source
            if url_err:
                uat_logger.warning(f"步骤URL无效，已跳过（{source}）: {url_err}")

    return None, None


def _resolve_case_navigation_url(case: dict = None, case_id: int = None, steps: list = None, fallback_url: str = None) -> tuple:
    """统一解析运行/录制前导航URL。优先级：case.url > step.url/navigate.input_value > fallback_url"""
    local_case = case or {}
    if not local_case and case_id:
        local_case = db.get_test_case(case_id) or {}

    case_url = local_case.get('url')
    if case_url:
        fixed_url, url_err = _validate_and_fix_url(case_url)
        if fixed_url:
            return fixed_url, 'case.url'
        if url_err:
            uat_logger.warning(f"用例URL无效，继续尝试步骤URL: {url_err}")

    local_steps = steps
    if local_steps is None and case_id:
        local_steps = db.get_case_steps(case_id)
    step_url, step_source = _get_first_valid_step_url(local_steps or [])
    if step_url:
        return step_url, step_source

    if fallback_url:
        fixed_url, url_err = _validate_and_fix_url(fallback_url)
        if fixed_url:
            return fixed_url, 'request.url'
        if url_err:
            return None, f"无效的URL地址: {fallback_url}"

    return None, None


def _resolve_case_navigation_url_for_data_row(
    db,
    case: dict,
    case_id: int,
    steps: list,
    row_data: dict,
    project_id,
    resolve_with_row_fn,
):
    """数据驱动每行：与单用例相同的 URL 优先级，并对地址做项目/用例变量与 {{row.*}} 替换。"""
    local_case = case or {}
    raw_case_url = (local_case.get('url') or '').strip()
    if raw_case_url:
        resolved = resolve_with_row_fn(
            db.resolve_variables(raw_case_url, project_id=project_id, case_id=case_id),
            row_data,
        )
        fixed_url, url_err = _validate_and_fix_url((resolved or '').strip())
        if fixed_url:
            return fixed_url, 'case.url'
        if url_err:
            uat_logger.warning(f'[数据驱动] 用例 URL 无效，继续尝试步骤: {url_err}')

    for step in steps or []:
        action = (step.get('action') or '').strip().lower()
        candidates = []
        if step.get('url'):
            candidates.append(('step.url', step.get('url')))
        if action == 'navigate' and step.get('input_value'):
            candidates.append(('navigate.input_value', step.get('input_value')))
        for source, raw in candidates:
            resolved = resolve_with_row_fn(
                db.resolve_variables(raw or '', project_id=project_id, case_id=case_id),
                row_data,
            )
            fixed_url, url_err = _validate_and_fix_url((resolved or '').strip())
            if fixed_url:
                return fixed_url, source
            if url_err:
                uat_logger.warning(f'[数据驱动] 步骤 URL 无效已跳过（{source}）: {url_err}')

    return None, None


# 测试用例管理页面（新版本）
@app.route('/list_cases_v2/<int:project_id>')
@login_required
@project_access_required(min_role='viewer')
def list_cases_v2(project_id):
    return render_template('list_cases_v2.html', project_id=project_id)

# 测试步骤管理页面
@app.route('/list_steps')
@login_required
def list_steps():
    case_id = request.args.get('case_id', type=int)
    if case_id:
        c = db.get_test_case_v2(case_id)
        if c and _app_case_type(c) == 'api':
            from flask import redirect, url_for
            pid = c.get('project_id') or ''
            q = f"?project_id={pid}" if pid else ""
            return redirect(url_for('api_case_detail_page', case_id=case_id) + q)
    return render_template('list_steps.html')

# API: 创建测试用例
@app.route('/api/create_case', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
@audit_log('CREATE_CASE', 'case')
def api_create_case():
    data = request.get_json(silent=True) or {}
    name = data.get('name', '')
    description = data.get('description', '')
    target_url = data.get('target_url', '')
    
    if not name:
        return jsonify({'error': '用例名称不能为空'}), 400
    
    case_id = db.create_test_case(name, description, target_url)
    return jsonify({'success': True, 'case_id': case_id})

# API: 获取所有测试用例
@app.route('/api/test_cases', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_test_cases():
    cases = db.get_all_test_cases()
    return jsonify({'cases': cases})

# API: 获取单个测试用例
@app.route('/api/test_case/<int:case_id>', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_test_case(case_id):
    case = db.get_test_case(case_id)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404
    return jsonify({'test_case': case})

# API: 更新测试用例
@app.route('/api/test_case/<int:case_id>', methods=['PUT'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
@audit_log('UPDATE_CASE', 'case')
def api_update_test_case(case_id):
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    description = data.get('description')
    target_url = data.get('target_url')
    
    success = db.update_test_case(case_id, name, description, target_url)
    
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '更新测试用例失败'}), 400

# API: 删除测试用例
@app.route('/api/test_case/<int:case_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
@audit_log('DELETE_CASE', 'case')
def api_delete_test_case(case_id):
    success = db.delete_test_case(case_id)
    
    if success:
        try:
            db.prune_orphan_run_history()
        except Exception:
            pass
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '删除测试用例失败'}), 400

# API: 执行多个测试用例
@app.route('/api/execute_multiple_cases', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
@audit_log('EXECUTE_CASES', 'execution')
def api_execute_multiple_cases():
    data = request.get_json(silent=True) or {}
    case_ids = data.get('case_ids', [])
    
    if not case_ids:
        return jsonify({'success': False, 'error': '缺少测试用例ID列表参数'}), 400
    
    if not isinstance(case_ids, list):
        return jsonify({'success': False, 'error': 'case_ids参数必须是数组'}), 400
    
    # ===== 执行次数限制检查 =====
    user_id = current_user.id
    _db = Database()
    
    # 获取当前 License 信息
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    is_free_user = license_info.license_type == LicenseType.FREE.value
    
    # 获取今日已执行次数
    today_stats = _db.get_user_usage_stats(user_id)
    current_count = today_stats.get('execution_count', 0) if today_stats else 0
    
    # 获取每日执行限制（-1表示无限制）
    daily_limit = limits.get('max_executions_per_day', -1)
    
    # 免费版每日限制10次
    if is_free_user:
        DAILY_LIMIT = 10
        
        if current_count >= DAILY_LIMIT:
            return jsonify({
                'success': False,
                'error': f'今日执行次数已达上限（{DAILY_LIMIT}次）。请升级至团队版解除限制。',
                'limit_reached': True,
                'current_count': current_count,
                'daily_limit': DAILY_LIMIT,
                'upgrade_url': '/upgrade'
            }), 403
        
        # 检查是否会超出限制
        if current_count + len(case_ids) > DAILY_LIMIT:
            remaining = DAILY_LIMIT - current_count
            return jsonify({
                'success': False,
                'error': f'本次执行将超出今日限制。剩余次数: {remaining}，本次请求: {len(case_ids)}个用例。',
                'limit_exceeded': True,
                'current_count': current_count,
                'daily_limit': DAILY_LIMIT,
                'remaining': remaining,
                'upgrade_url': '/upgrade'
            }), 403
    elif daily_limit > 0 and current_count >= daily_limit:
        # 非免费版但有每日限制的情况
        return jsonify({
            'success': False,
            'error': f'今日执行次数已达上限（{daily_limit}次）。请联系管理员。',
            'limit_reached': True,
            'current_count': current_count,
            'daily_limit': daily_limit
        }), 403
    
    # 记录执行次数（所有版本都记录，用于统计和显示）
    new_count = _db.increment_execution_count(user_id)
    uat_logger.info(f"📊 [LICENSE] 用户 {user_id} 今日执行次数: {new_count}")
    
    # 添加调试信息
    uat_logger.info(f"📥 [API_ENTRY] 接收到的执行顺序: {case_ids}")
    uat_logger.info(f"📥 [API_ENTRY] 用例数量: {len(case_ids)}")
    
    uat_logger.info(f"开始执行多个测试用例，共 {len(case_ids)} 个用例")
    with _case_run_lock:
        _case_run_jobs[user_id] = {
            'active': True,
            'cancel_requested': False,
            'run_mode': 'multiple_cases',
            'case_id': None,
            'case_name': '批量执行',
            'total_steps': len(case_ids),
            'completed_steps': 0,
            'current_step_order': 0,
            'current_action': 'batch_run',
            'message': f'正在批量执行 {len(case_ids)} 个用例',
            'started_at': time.time(),
        }
    
    results = None
    
    try:
        with _UserUiRunGuard(user_id, f'batch x{len(case_ids)}'):
            uat_logger.info(f"🚀 [API] 开始执行 {len(case_ids)} 个测试用例")
            thread_db = Database()

            def _should_stop_batch():
                with _case_run_lock:
                    return bool(_case_run_jobs.get(user_id, {}).get('cancel_requested'))

            try:
                uat_logger.info(f"🚀 [API] 在主线程中同步执行测试用例序列: {case_ids}")
                results = sync_execute_multiple_test_cases(
                    case_ids,
                    thread_db,
                    should_stop=_should_stop_batch,
                    execution_context=_make_batch_execution_context("ui", data),
                )
                if results.get('error') and not results.get('case_results'):
                    from auth_batch_helpers import record_cases_run_rejected

                    reject_reason = str(results.get('error') or '批量执行未启动')
                    record_cases_run_rejected(thread_db, case_ids, reject_reason)
                    results['case_results'] = [
                        {
                            'case_id': cid,
                            'case_name': (thread_db.get_test_case_v2(cid) or {}).get('name') or '未知',
                            'status': 'error',
                            'error': reject_reason,
                        }
                        for cid in case_ids
                    ]
                uat_logger.info(f"✅ [API] 多个测试用例同步执行完成")
            except Exception as e:
                uat_logger.error(f"❌ [API] 执行测试用例时发生异常: {str(e)}")
                results = {
                    "total_cases": len(case_ids),
                    "successful_cases": 0,
                    "failed_cases": len(case_ids),
                    "case_results": [
                        {
                            "case_id": case_id,
                            "case_name": "未知",
                            "status": "error",
                            "error": f"执行出错: {str(e)}"
                        } for case_id in case_ids
                    ]
                }
                try:
                    from auth_batch_helpers import record_cases_run_rejected

                    record_cases_run_rejected(thread_db, case_ids, str(e))
                except Exception:
                    pass

            uat_logger.info(
                f"多个测试用例执行完成，成功: {results['successful_cases']}, 失败: {results['failed_cases']}"
            )
            _case_job_update(
                user_id,
                completed_steps=len(case_ids),
                current_step_order=len(case_ids),
                message=(
                    f"批量执行完成：成功 {results.get('successful_cases', 0)}，"
                    f"失败 {results.get('failed_cases', 0)}"
                ),
            )
    except Exception as e:
        uat_logger.error(f"执行测试用例时出错: {e}")
        results = {
            "total_cases": len(case_ids),
            "successful_cases": 0,
            "failed_cases": len(case_ids),
            "case_results": [
                {
                    "case_id": case_id,
                    "case_name": "未知",
                    "status": "error",
                    "error": f"系统错误: {str(e)}"
                } for case_id in case_ids
            ]
        }
        try:
            from auth_batch_helpers import record_cases_run_rejected

            record_cases_run_rejected(Database(), case_ids, str(e))
        except Exception:
            pass
    finally:
        # 确保浏览器资源清理，无论测试用例执行结果如何
        try:
            uat_logger.info("🔧 [API_CLEANUP] 开始清理浏览器资源...")
            # 调用同步关闭浏览器函数确保资源释放
            sync_close_browser()
            uat_logger.info("✅ [API_CLEANUP] 浏览器资源清理完成")
        except Exception as close_error:
            uat_logger.warning(f"⚠️ [API_CLEANUP] 清理浏览器时出现警告: {close_error}")
        with _case_run_lock:
            job = _case_run_jobs.get(user_id)
            if job:
                job['active'] = False
                job['finished_at'] = time.time()
                job['duration'] = round(time.time() - job.get('started_at', time.time()), 2)
    
    with _case_run_lock:
        stopped = bool(_case_run_jobs.get(user_id, {}).get('cancel_requested'))
    if results is None:
        results = {
            'total_cases': len(case_ids),
            'successful_cases': 0,
            'failed_cases': len(case_ids),
            'case_results': [],
            'error': '批量执行未返回结果（服务可能已中断，请重启后重试）',
        }
        try:
            from auth_batch_helpers import record_cases_run_rejected

            record_cases_run_rejected(_db, case_ids, results['error'])
        except Exception:
            pass
    batch_ok = not results.get('error') or bool(results.get('case_results'))
    response_data = {
        'success': batch_ok,
        'results': results,
        'stopped': stopped,
    }
    if not batch_ok and results.get('error'):
        response_data['error'] = results.get('error')
    return jsonify(response_data)

# API: 导航到指定URL
@app.route('/api/navigate', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_navigate():
    data = request.get_json(silent=True) or {}
    url = data.get('url', '')
    iframe_selector = data.get('iframe_selector', '')
    
    if not url:
        return jsonify({'error': 'URL不能为空'}), 400

    force_main = (data.get('force_main_playwright') or data.get('force_main') or '').strip().lower() in (
        '1',
        'true',
        'yes',
        'on',
    )
    if (
        not force_main
        and embedded_gateway_enabled()
        and not _ai_allow_main_playwright_fallback()
    ):
        return jsonify({
            'success': True,
            'skipped': True,
            'reason': 'embedded_canvas_preferred',
            'message': '已配置内置画布网关，未启动主 Playwright。请使用 AI 测试页画布或 embedded 会话 API 导航。',
        })

    sync_navigate_to(url, iframe_selector=iframe_selector)
    return jsonify({'success': True})

# API: 执行滚动操作
@app.route('/api/scroll', methods=['POST'])
@api_error_handler
@log_api_request
def api_scroll():
    data = request.get_json(silent=True) or {}
    direction = data.get('direction', 'down')
    pixels = data.get('pixels', 500)
    iframe_selector = data.get('iframe_selector', '')
    
    sync_scroll_page(direction, pixels, iframe_selector=iframe_selector)
    return jsonify({'success': True})

# API: 提取元素文本
@app.route('/api/extract_element_text', methods=['POST'])
@api_error_handler
@log_api_request
def api_extract_element_text():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    selector_type = data.get('selector_type', 'css')
    
    if selector == 'body':
        text = sync_get_page_text()
    else:
        text = sync_extract_element_text(selector, selector_type)
    return jsonify({'success': True, 'text': text})

# API: 提取元素JSON数据
@app.route('/api/extract_element_json', methods=['POST'])
@api_error_handler
@log_api_request
def api_extract_element_json():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    selector_type = data.get('selector_type', 'css')
    
    if not selector:
        return jsonify({'success': False, 'error': '选择器不能为空'}), 400
    
    json_data = sync_extract_element_json(selector, selector_type)
    return jsonify({'success': True, 'json': json_data})

# API: 获取页面标题
@app.route('/api/page_title', methods=['GET'])
@api_error_handler
@log_api_request
def api_page_title():
    title = sync_get_page_title()
    return jsonify({'success': True, 'title': title})

# API: 获取当前URL
@app.route('/api/current_url', methods=['GET'])
@api_error_handler
@log_api_request
def api_current_url():
    url = sync_get_current_url()
    return jsonify({'success': True, 'url': url})

# API: 获取页面上所有链接
@app.route('/api/links', methods=['GET'])
@api_error_handler
@log_api_request
def api_links():
    links = sync_get_all_links()
    return jsonify({'success': True, 'links': links})

def _legacy_visual_picker_removed():
    """旧版网页拾取 API 已收敛至统一元素捕获。"""
    return jsonify({
        'success': False,
        'error': '该接口已移除，请使用 /api/element-picker/start|stop|status',
        'deprecated': True,
    }), 410


# API: 启动可视化选择（已废弃）
@app.route('/api/start_visual_selection', methods=['POST'])
@api_error_handler
@log_api_request
def api_start_visual_selection():
    return _legacy_visual_picker_removed()


# API: 停止可视化选择（已废弃）
@app.route('/api/stop_visual_selection', methods=['POST'])
@api_error_handler
@log_api_request
def api_stop_visual_selection():
    return _legacy_visual_picker_removed()


# API: 检查选择的元素（已废弃）
@app.route('/api/check_selected_element', methods=['GET'])
@api_error_handler
@log_api_request
def api_check_selected_element():
    return _legacy_visual_picker_removed()

# API: 提取元素数据
@app.route('/api/extract_element_data', methods=['POST'])
@api_error_handler
@log_api_request
def api_extract_element_data():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    element_data = sync_extract_element_data(selector)
    return jsonify({'success': True, 'data': element_data})

# API: 获取页面数据
@app.route('/api/page_data', methods=['GET'])
@api_error_handler
@log_api_request
def api_page_data():
    page_data = automation.get_page_data()
    return jsonify({'success': True, 'data': page_data})

# API: 分析页面内容
@app.route('/api/analyze_content', methods=['POST'])
@api_error_handler
@log_api_request
def api_analyze_content():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', 'body')
    
    analysis = automation.analyze_page_content(selector)
    return jsonify({'success': True, 'analysis': analysis})


@app.route('/api/ai/models', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_ai_models():
    cfg = _load_ai_model_config()
    ollama_models, ollama_error = local_ai_service.list_installed_models()
    profiles = [_mask_profile_for_api(p) for p in (cfg.get('profiles') or [])]
    local_names = [p.get('model_id') for p in (cfg.get('profiles') or []) if p.get('provider') == 'ollama' and p.get('model_id')]
    catalog = _load_ai_provider_catalog()
    providers = catalog.get('providers') if isinstance(catalog, dict) else []
    try:
        from ai_config_paths import ai_provider_catalog_source

        catalog_path = ai_provider_catalog_source()
    except ImportError:
        catalog_path = str(ai_provider_catalog_path())
    return jsonify({
        'success': True,
        'version': cfg.get('version') or 2,
        'active_profile_id': cfg.get('active_profile_id'),
        'profiles': profiles,
        'active_local_model': cfg.get('active_local_model'),
        'local_models': local_names or cfg.get('local_models', []),
        'provider_catalog': catalog,
        'provider_catalog_count': len(providers) if isinstance(providers, list) else 0,
        'provider_catalog_path': catalog_path,
        'ollama_base_url': local_ai_service.base_url,
        'ollama_models': ollama_models,
        'ollama_error': ollama_error,
    })


@app.route('/api/ai/vision/readiness', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_ai_vision_readiness():
    """面向用户的视觉自动化就绪状态（画布、网关、本地视觉模型）。"""
    from vision_platform_readiness import check_vision_automation_readiness

    embedded_sid = (request.args.get('embedded_session_id') or '').strip()
    payload = check_vision_automation_readiness(
        user_id=current_user.id,
        embedded_session_id=embedded_sid,
    )
    return jsonify({'success': True, **payload})


@app.route('/api/ai/vision/replay/<run_id>/')
@app.route('/api/ai/vision/replay/<run_id>/index.html')
@login_required
def api_ai_vision_replay_html(run_id: str):
    """打开单次测试步骤回放页（HTML）。"""
    from flask import abort, send_from_directory
    from vision_step_report import replay_index_path

    path = replay_index_path(run_id)
    if not path:
        abort(404)
    return send_from_directory(str(path.parent), path.name)


@app.route('/api/ai/models/ollama', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_ollama_models():
    ollama_models, ollama_error = local_ai_service.list_installed_models()
    return jsonify({
        'success': True,
        'ollama_base_url': local_ai_service.base_url,
        'ollama_models': ollama_models,
        'ollama_error': ollama_error,
    })


def _validate_profile_model_id(mid: str) -> tuple:
    m = (mid or '').strip()
    if not m:
        return False, 'model_id不能为空'
    if len(m) > 220:
        return False, 'model_id过长'
    if not _AI_PROFILE_MODEL_ID_RE.match(m):
        return False, 'model_id格式无效（勿含空白）'
    return True, m


@app.route('/api/ai/models', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_add_ai_model():
    data = request.get_json(silent=True) or {}
    if (data.get('provider') or '').strip() or (data.get('model_id') or '').strip():
        return _api_add_ai_profile(data)
    ok, model_name_or_err = _validate_ai_model_name(data.get('model_name') or '')
    if not ok:
        return jsonify({'success': False, 'error': model_name_or_err}), 400
    model_name = model_name_or_err
    verify_ollama = bool(data.get('verify_ollama'))
    if verify_ollama:
        installed, err = local_ai_service.list_installed_models()
        if err:
            return jsonify({
                'success': False,
                'error': f'无法校验 Ollama：{err}',
                'hint': '可关闭「添加前校验」后重试，或检查 LOCAL_LLM_BASE_URL 与 ollama serve',
            }), 503
        if model_name not in installed:
            return jsonify({
                'success': False,
                'error': f'Ollama 中未找到模型「{model_name}」，请先执行 ollama pull',
                'ollama_models': installed,
            }), 400
    cfg = _load_ai_model_config()
    profiles = list(cfg.get('profiles') or [])
    meta = _catalog_provider_meta('ollama')
    pid = str(uuid.uuid4())
    profiles.append({
        'id': pid,
        'provider': 'ollama',
        'api_style': meta.get('api_style') or 'ollama',
        'model_type': 'test_case_generation',
        'model_id': model_name,
        'label': model_name,
        'api_key': '',
        'base_url': '',
    })
    cfg['profiles'] = profiles
    cfg['version'] = 2
    cfg['active_profile_id'] = pid
    _save_ai_model_config(cfg)
    return jsonify({
        'success': True,
        'active_profile_id': cfg.get('active_profile_id'),
        'profiles': [_mask_profile_for_api(p) for p in profiles],
    })


def _api_add_ai_profile(data: dict):
    from ai_multi_provider import normalize_api_key

    provider = (data.get('provider') or '').strip()
    base_url_early = (data.get('base_url') or '').strip() if isinstance(data.get('base_url'), str) else ''
    api_key_early = normalize_api_key(data.get('api_key') if isinstance(data.get('api_key'), str) else '')
    if not provider:
        provider = _infer_ai_provider_simple(base_url_early, api_key_early)
    cmeta = _catalog_provider_meta(provider)
    if not cmeta:
        return jsonify({'success': False, 'error': f'未知提供商: {provider}'}), 400
    api_style = (data.get('api_style') or cmeta.get('api_style') or '').strip()
    if not api_style:
        return jsonify({'success': False, 'error': '无法解析 api_style'}), 400
    ok, model_id = _validate_profile_model_id(data.get('model_id') or '')
    if not ok:
        return jsonify({'success': False, 'error': model_id}), 400
    model_type = (data.get('model_type') or 'test_case_generation').strip()
    label = (data.get('label') or '').strip() or model_id
    api_key = api_key_early
    base_url = (data.get('base_url') or '').strip() if isinstance(data.get('base_url'), str) else ''
    base_url = _normalize_profile_base_url(base_url, provider)
    group_id = (data.get('group_id') or '').strip() if isinstance(data.get('group_id'), str) else ''
    requires_key = bool(cmeta.get('requires_api_key'))
    if provider == 'custom_openai' and not base_url:
        return jsonify({'success': False, 'error': '请填写 API Base URL（代理/第三方网关地址）'}), 400
    if requires_key and provider != 'ollama' and not api_key:
        return jsonify({'success': False, 'error': '请填写 API 密钥'}), 400
    verify_ollama = bool(data.get('verify_ollama'))
    if provider == 'ollama' and verify_ollama:
        installed, err = local_ai_service.list_installed_models()
        if err:
            return jsonify({
                'success': False,
                'error': f'无法校验 Ollama：{err}',
                'hint': '可关闭「添加前校验」后重试',
            }), 503
        if model_id not in installed:
            return jsonify({
                'success': False,
                'error': f'Ollama 中未找到模型「{model_id}」',
                'ollama_models': installed,
            }), 400
    cfg = _load_ai_model_config()
    profiles = list(cfg.get('profiles') or [])
    pid = str(uuid.uuid4())
    profiles.append({
        'id': pid,
        'provider': provider,
        'api_style': api_style,
        'model_type': model_type,
        'model_id': model_id,
        'label': label,
        'api_key': api_key,
        'base_url': base_url,
        'group_id': group_id,
    })
    cfg['profiles'] = profiles
    cfg['version'] = 2
    cfg['active_profile_id'] = pid
    _save_ai_model_config(cfg)
    return jsonify({
        'success': True,
        'active_profile_id': cfg.get('active_profile_id'),
        'profiles': [_mask_profile_for_api(p) for p in profiles],
    })


@app.route('/api/ai/models/active', methods=['PUT', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_set_active_ai_model():
    data = request.get_json(silent=True) or {}
    profile_id = (data.get('profile_id') or '').strip()
    cfg = _load_ai_model_config()
    profiles = cfg.get('profiles') or []
    if profile_id:
        if not any(p.get('id') == profile_id for p in profiles):
            return jsonify({'success': False, 'error': '未找到该模型配置'}), 404
        cfg['active_profile_id'] = profile_id
        cfg['version'] = 2
        _save_ai_model_config(cfg)
        hermes_info: dict = {}
        try:
            from hermes_service_bootstrap import ensure_hermes_llm_current

            # 切换引擎 = 写入 Hermes .env + config.yaml model；若智能体在跑则重启加载
            hermes_info = ensure_hermes_llm_current(restart_if_stale=True)
        except Exception as e:
            hermes_info = {'error': str(e)[:160]}
        active = next((p for p in profiles if p.get('id') == profile_id), None)
        return jsonify({
            'success': True,
            'active_profile_id': profile_id,
            'active_label': (active or {}).get('label') or (active or {}).get('model_id') or '',
            'active_model_id': (active or {}).get('model_id') or '',
            'hermes': hermes_info,
            'hermes_model': (hermes_info or {}).get('env_model') or (hermes_info or {}).get('synced_model') or '',
            'hermes_base_url': (hermes_info or {}).get('env_base_url') or '',
            'profiles': [_mask_profile_for_api(p) for p in profiles],
        })
    ok, model_name_or_err = _validate_ai_model_name(data.get('model_name') or '')
    if not ok:
        return jsonify({'success': False, 'error': model_name_or_err}), 400
    model_name = model_name_or_err
    for p in profiles:
        if p.get('provider') == 'ollama' and p.get('model_id') == model_name:
            cfg['active_profile_id'] = p.get('id')
            cfg['version'] = 2
            _save_ai_model_config(cfg)
            hermes_info = {}
            try:
                from hermes_service_bootstrap import ensure_hermes_llm_current

                hermes_info = ensure_hermes_llm_current(restart_if_stale=True)
            except Exception as e:
                hermes_info = {'error': str(e)[:160]}
            return jsonify({
                'success': True,
                'active_profile_id': cfg.get('active_profile_id'),
                'hermes': hermes_info,
                'profiles': [_mask_profile_for_api(x) for x in profiles],
            })
    return jsonify({'success': False, 'error': '请使用 profile_id 或已注册的 Ollama 模型名'}), 400


@app.route('/api/ai/models', methods=['DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_delete_ai_model():
    data = request.get_json(silent=True) or {}
    profile_id = (data.get('profile_id') or '').strip()
    cfg = _load_ai_model_config()
    profiles = list(cfg.get('profiles') or [])
    if profile_id:
        if not any(p.get('id') == profile_id for p in profiles):
            return jsonify({'success': False, 'error': '未找到该模型配置'}), 404
        profiles = [p for p in profiles if p.get('id') != profile_id]
        cfg['profiles'] = profiles
        cfg['version'] = 2
        if (cfg.get('active_profile_id') or '').strip() == profile_id:
            cfg['active_profile_id'] = profiles[0]['id'] if profiles else ''
        _save_ai_model_config(cfg)
        return jsonify({
            'success': True,
            'active_profile_id': cfg.get('active_profile_id'),
            'profiles': [_mask_profile_for_api(p) for p in profiles],
        })
    ok, model_name_or_err = _validate_ai_model_name(data.get('model_name') or '')
    if not ok:
        return jsonify({'success': False, 'error': model_name_or_err}), 400
    model_name = model_name_or_err
    removed = False
    new_profiles = []
    for p in profiles:
        if p.get('provider') == 'ollama' and p.get('model_id') == model_name:
            removed = True
            continue
        new_profiles.append(p)
    if not removed:
        return jsonify({'success': False, 'error': '该平台未注册此模型'}), 404
    cfg['profiles'] = new_profiles
    cfg['version'] = 2
    if (cfg.get('active_profile_id') or '').strip() and not any(
        p.get('id') == cfg.get('active_profile_id') for p in new_profiles
    ):
        cfg['active_profile_id'] = new_profiles[0]['id'] if new_profiles else ''
    _save_ai_model_config(cfg)
    return jsonify({
        'success': True,
        'active_profile_id': cfg.get('active_profile_id'),
        'profiles': [_mask_profile_for_api(p) for p in new_profiles],
    })


@app.route('/api/ai/models/verify', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_verify_ai_model_profile():
    """校验模型配置是否可连通（发送极短测试请求）。"""
    data = request.get_json(silent=True) or {}
    explicit_provider = (data.get('provider') or '').strip()
    provider = explicit_provider
    model_id = (data.get('model_id') or '').strip()
    profile_id = (data.get('profile_id') or '').strip()
    api_key = data.get('api_key')
    base_url = (data.get('base_url') or '').strip()

    if profile_id:
        cfg = _load_ai_model_config()
        p = next((x for x in (cfg.get('profiles') or []) if x.get('id') == profile_id), None)
        if not p:
            return jsonify({'success': False, 'error': '未找到该模型配置'}), 404
        if not model_id:
            model_id = (p.get('model_id') or '').strip()
        if not isinstance(api_key, str) or not api_key.strip():
            api_key = p.get('api_key') or ''
        if not base_url:
            base_url = (p.get('base_url') or '').strip()
    if not model_id:
        return jsonify({'success': False, 'error': 'model_id 不能为空'}), 400
    from ai_multi_provider import dispatch_chat, normalize_api_key

    key_norm = normalize_api_key(api_key if isinstance(api_key, str) else '')
    if not explicit_provider:
        provider = _infer_ai_provider_simple(base_url, key_norm)
    if provider == 'ollama' and not key_norm and profile_id:
        cfg_key = _load_ai_model_config()
        p_key = next((x for x in (cfg_key.get('profiles') or []) if x.get('id') == profile_id), None)
        if isinstance(p_key, dict) and (p_key.get('provider') or '').strip() == 'ollama':
            key_norm = ''
    if provider != 'ollama' and not key_norm:
        return jsonify({'success': False, 'error': '请先填写有效的 API 密钥（连接测试不会使用已保存的密钥，除非在编辑已有配置且留空时才会读取库内密钥）'}), 400
    if provider == 'minimax' and key_norm.lower().startswith('tp-'):
        return jsonify({
            'success': False,
            'error': '当前 Key 为第三方/代理格式（tp-），不能用于 MiniMax 官方地址',
            'hint': '请改选提供商「第三方 / 代理（OpenAI 兼容）」，Base URL 填代理商地址（如 https://token-plan-cn.xiaomimimo.com/v1），model 填代理商文档中的名称。',
        }), 400
    if key_norm.lower().startswith('tp-') and provider != 'custom_openai':
        return jsonify({
            'success': False,
            'error': '检测到代理 Key（tp-），请使用提供商「第三方 / 代理（OpenAI 兼容）」',
            'hint': 'Base URL 与 model 必须以代理商文档为准，不要选 MiniMax/OpenAI 等官方提供商。',
        }), 400
    if provider == 'custom_openai' and not base_url:
        return jsonify({'success': False, 'error': '请填写 API Base URL（代理/第三方网关地址）'}), 400
    cmeta = _catalog_provider_meta(provider)
    if not cmeta and provider:
        return jsonify({'success': False, 'error': f'未知提供商: {provider}'}), 400
    base_url = _normalize_profile_base_url(base_url, provider)
    group_id = (data.get('group_id') or '').strip()
    if profile_id and not group_id:
        p_gid = next((x for x in (_load_ai_model_config().get('profiles') or []) if x.get('id') == profile_id), None)
        if isinstance(p_gid, dict):
            group_id = (p_gid.get('group_id') or '').strip()
    profile = {
        'provider': provider,
        'api_style': (cmeta.get('api_style') or 'openai_compatible'),
        'model_id': model_id,
        'api_key': key_norm,
        'base_url': base_url,
        'group_id': group_id,
    }
    try:

        reply = dispatch_chat('Reply with JSON only: {"ok":true}', profile, local_ai_service)
        ok = bool(reply and str(reply).strip())
        return jsonify({
            'success': True,
            'reachable': ok,
            'message': '连接成功，模型可响应' if ok else '已连接但返回为空',
            'normalized_base_url': base_url,
        })
    except ValueError as e:
        err = str(e)
        hint = ''
        low = err.lower()
        if 'not supported model' in low or 'param incorrect' in low:
            popular = cmeta.get('popular_models') or []
            bu_low = base_url.lower()
            if 'xiaomimimo.com' in bu_low or provider == 'xiaomi_mimo_token':
                hint = (
                    '该 Base URL 为小米 MiMo Token Plan（不是 MiniMax）。'
                    ' model 请改为 mimo-v2.5-pro（推荐）、mimo-v2.5 或 mimo-v2-flash；'
                    '勿使用 MiniMax-M3、abab7-preview 等名称。'
                    ' 也可在添加模型时直接选提供商「小米 MiMo Token Plan」。'
                )
            elif provider == 'custom_openai':
                hint = (
                    'API 已连通，但 model_id 不被该代理地址支持。'
                    + (f' 若文档有推荐型号可尝试：{", ".join(popular[:3])}' if popular else '')
                    + ' 请以代理商/网关文档中的 model 名称为准。'
                )
            else:
                hint = (
                    'API 已连通，但 model_id 不被该 Base URL 支持。'
                    + (f' 请尝试：{", ".join(popular[:3])}' if popular else '')
                    + '；Base URL 应填根地址，不要填带 /chat/completions 的完整路径。'
                )
        elif provider == 'minimax' and ('1004' in err or 'authorized_error' in low or 'api secret key' in low):
            hint = (
                'MiniMax 401：请确认使用的是开放平台「订阅 Key（sk-cp- 开头）」或「按量计费 API Key」，'
                '且与 Base URL 地域一致（国内 api.minimaxi.com，国际 api.minimax.io）。'
                '以 tp- 等开头的第三方代理 Key 不能用于官方地址；若仅有代理 Key，请改填代理提供的 Base URL。'
            )
        return jsonify({'success': False, 'error': err, 'hint': hint, 'normalized_base_url': base_url}), 400
    except Exception as e:
        uat_logger.exception('verify ai model failed')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ai/models/profile', methods=['PUT', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_update_ai_profile():
    """更新已有模型配置（换 Key、改 Base URL、同厂商下换模型名等），无需删除后重建。"""
    data = request.get_json(silent=True) or {}
    profile_id = (data.get('profile_id') or '').strip()
    if not profile_id:
        return jsonify({'success': False, 'error': '缺少 profile_id'}), 400
    cfg = _load_ai_model_config()
    profiles = list(cfg.get('profiles') or [])
    idx = next((i for i, p in enumerate(profiles) if p.get('id') == profile_id), -1)
    if idx < 0:
        return jsonify({'success': False, 'error': '未找到该模型配置'}), 404
    p = dict(profiles[idx])
    provider = (p.get('provider') or '').strip()
    cmeta = _catalog_provider_meta(provider)
    if not cmeta:
        return jsonify({'success': False, 'error': f'未知提供商: {provider}'}), 400

    if 'label' in data:
        lab = _ai_str(data.get('label'))
        if lab:
            p['label'] = lab

    if 'model_type' in data:
        mt = _ai_str(data.get('model_type'))
        if mt:
            p['model_type'] = mt

    if 'model_id' in data:
        ok, model_id = _validate_profile_model_id(data.get('model_id') or '')
        if not ok:
            return jsonify({'success': False, 'error': model_id}), 400
        p['model_id'] = model_id

    if 'base_url' in data:
        bu = data.get('base_url')
        bu = bu.strip() if isinstance(bu, str) else ''
        p['base_url'] = _normalize_profile_base_url(bu, provider)

    from ai_multi_provider import normalize_api_key

    api_key_in = data.get('api_key')
    if isinstance(api_key_in, str) and api_key_in.strip():
        p['api_key'] = normalize_api_key(api_key_in)

    final_base = (p.get('base_url') or '').strip()
    final_key = normalize_api_key(p.get('api_key') if isinstance(p.get('api_key'), str) else '')
    inferred_provider = _infer_ai_provider_simple(final_base, final_key)
    inferred_meta = _catalog_provider_meta(inferred_provider)
    if inferred_meta:
        provider = inferred_provider
        p['provider'] = inferred_provider
        p['api_style'] = inferred_meta.get('api_style') or p.get('api_style') or 'openai_compatible'
        if final_base:
            p['base_url'] = _normalize_profile_base_url(final_base, inferred_provider)
        cmeta = inferred_meta

    if 'group_id' in data:
        p['group_id'] = (data.get('group_id') or '').strip() if isinstance(data.get('group_id'), str) else ''
    elif bool(cmeta.get('requires_api_key')) and provider != 'ollama':
        existing = (p.get('api_key') or '').strip() if isinstance(p.get('api_key'), str) else ''
        if not existing:
            return jsonify({'success': False, 'error': '该提供商需要 API 密钥，请填写新的密钥'}), 400

    profiles[idx] = p
    cfg['profiles'] = profiles
    cfg['version'] = 2
    _save_ai_model_config(cfg)
    return jsonify({
        'success': True,
        'active_profile_id': cfg.get('active_profile_id'),
        'profiles': [_mask_profile_for_api(x) for x in profiles],
    })


# --- AI 后台任务：SQLite 持久化存储（支持进程重启恢复） ---
from ai_job_store import get_job_store


def _ai_bg_prune():
    get_job_store().prune()


def _ai_url_probe_disabled() -> bool:
    v = (os.environ.get("LOCAL_AI_SKIP_URL_PROBE") or "").strip().lower()
    return v in ("1", "true", "yes", "on")


def _ai_allow_main_playwright_fallback() -> bool:
    """为 1 时：即使已配置 embedded 网关，也允许回退主 Playwright（可能弹出浏览器窗口）。"""
    return (os.environ.get("AI_ALLOW_MAIN_PLAYWRIGHT_FALLBACK") or "0").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _ai_should_open_main_playwright(*, embedded_sid: str) -> bool:
    """内置画布网关已配置时，默认不拉起主 Playwright（除非 AI_ALLOW_MAIN_PLAYWRIGHT_FALLBACK=1）。"""
    if (embedded_sid or "").strip():
        return False
    if embedded_gateway_enabled() and not _ai_allow_main_playwright_fallback():
        return False
    return True


def _ai_main_session_dom_pack(target_page_url: str, *, strict: bool):
    """
    在主 Playwright 会话中打开 URL 并抓取可交互结构，供本地模型绑定真实 DOM。
    strict=True：任一步失败则返回 (None,..., error)；strict=False：失败时记录日志并返回错误文案供告警（仍可继续生成）。
    """
    from ai_page_probe import probe_registry_from_interactive_snapshot

    url = (target_page_url or "").strip()
    if not url:
        e = "目标 URL 为空"
        return None, None, None, None, (e if strict else "")

    try:
        sync_start_browser(headless=True)
    except Exception:
        pass
    if not sync_automation_session_usable():
        e = "主浏览器未就绪，无法打开页面抓取结构。请确认 Playwright 已安装，或先在界面中启动主浏览器。"
        if strict:
            return None, None, None, None, e
        uat_logger.warning("LOCAL_AI url probe skipped: %s", e)
        return None, None, None, None, e

    try:
        sync_navigate_to(url, ai_probe=True)
    except Exception as ex:
        e = f"导航失败：{ex}"
        if strict:
            return None, None, None, None, e
        uat_logger.warning("LOCAL_AI url probe: %s", e)
        return None, None, None, None, e

    try:
        snap = sync_get_interactive_page_snapshot(150)
    except Exception as ex:
        e = f"获取页面可交互结构失败：{ex}"
        if strict:
            return None, None, None, None, e
        uat_logger.warning("LOCAL_AI url probe: %s", e)
        return None, None, None, None, e

    ps, pr, pu = probe_registry_from_interactive_snapshot(snap)
    probe_url = (pu or url).strip() or None
    return snap, ps or None, pr if pr else None, probe_url, ""


def _attach_vision_replay_to_execution(execution: dict) -> None:
    """脚本执行完成后附加回放链接（主 Playwright 与内置画布共用）。"""
    if not isinstance(execution, dict):
        return
    try:
        from vision_step_report import pop_last_replay_meta

        meta = pop_last_replay_meta()
        if meta:
            execution["vision_replay"] = meta
    except Exception:
        pass


def _embedded_fetch_screenshot_png(user_id: int, embedded_sid: str):
    """从内置画布网关拉取当前视口 PNG。"""
    import base64

    j, err = embedded_gateway_json(
        "GET",
        f"/internal/session/{embedded_sid}/screenshot",
        user_id=user_id,
        timeout_sec=30.0,
    )
    if err or not isinstance(j, dict) or not j.get("success"):
        return None
    raw = (j.get("data") or "").strip()
    if not raw:
        return None
    try:
        return base64.b64decode(raw)
    except Exception:
        return None


def _ai_embedded_run_script_steps_sequentially(user_id: int, embedded_sid: str, script_steps: list) -> list:
    """
    在远程画布会话中逐步调用网关 run-steps（每步单独请求），与 gateway-stream 一致，
    便于 CDP screencast 在步骤间隙刷新 JPEG，实现「边看边跑」。
  同时记录测试回放（截图 + 中文步骤名）。
    """
    import time

    from vision_step_report import VisionReplaySession, vision_replay_enabled

    replay_sess = VisionReplaySession.start() if vision_replay_enabled() else None
    results = []
    for idx, st in enumerate(script_steps, start=1):
        step_t0 = time.perf_counter()
        step_status = "success"
        step_msg = ""
        png = None
        j, err = embedded_gateway_json(
            "POST",
            f"/internal/session/{embedded_sid}/run-steps",
            user_id=user_id,
            body={"steps": [st]},
            timeout_sec=180.0,
        )
        if err or not isinstance(j, dict) or not j.get("success"):
            step_status = "error"
            step_msg = str(err or (j or {}).get("detail") or "远程执行请求失败")
            results.append(
                {
                    "status": "error",
                    "step": st,
                    "error": step_msg,
                }
            )
        else:
            rs0 = (j.get("results") or [{}])[0]
            if not rs0.get("ok"):
                step_status = "error"
                step_msg = str(rs0.get("error") or "远程步骤失败")
                results.append(
                    {
                        "status": "error",
                        "step": st,
                        "error": step_msg,
                    }
                )
            else:
                results.append({"status": "success", "step": st})
        if replay_sess:
            png = _embedded_fetch_screenshot_png(user_id, embedded_sid)
            try:
                replay_sess.record(
                    idx,
                    st if isinstance(st, dict) else {},
                    step_status,
                    step_msg,
                    png,
                    int((time.perf_counter() - step_t0) * 1000),
                )
            except Exception:
                pass
        if step_status == "error":
            break
    if replay_sess:
        try:
            replay_sess.finalize()
        except Exception:
            pass
    return results


def _merge_ai_locator_resolution(plan_dict, snap_data, norm_warnings):
    """有页面快照时：启发式/LLM 选择器修复；随后对 assert 做回放页面探测校正。"""
    if not isinstance(plan_dict, dict) or not isinstance(plan_dict.get("steps"), list):
        return plan_dict, norm_warnings
    try:
        from ai_step_normalization import infer_plan_platform_type

        if infer_plan_platform_type(plan_dict) in ("desktop", "android"):
            return plan_dict, norm_warnings
    except Exception:
        pass
    try:
        registry = None
        if snap_data:
            from ai_page_probe import probe_registry_from_interactive_snapshot

            _, registry, _ = probe_registry_from_interactive_snapshot(snap_data)
        if not registry:
            from ai_page_probe import fetch_page_controls_bundle, resolve_steps_probe_url

            probe_u = resolve_steps_probe_url(
                plan_dict.get("steps") or [],
                (plan_dict.get("case_url") or "").strip(),
            )
            if probe_u.startswith(("http://", "https://")):
                _, _pe, registry = fetch_page_controls_bundle(probe_u)
        if registry:
            from ai_page_probe import heuristic_repair_plan_selectors_from_registry

            repaired, hw = heuristic_repair_plan_selectors_from_registry(
                plan_dict["steps"], registry
            )
            plan_dict["steps"] = repaired
            if hw:
                norm_warnings = list(norm_warnings or []) + list(hw)
        if snap_data:
            from ai_locator_resolution import ai_locator_resolve_enabled, resolve_plan_steps_locators_with_snapshot

            if ai_locator_resolve_enabled():
                new_steps, lw = resolve_plan_steps_locators_with_snapshot(
                    plan_dict["steps"], snap_data
                )
                plan_dict["steps"] = new_steps
                if lw:
                    norm_warnings = list(norm_warnings or []) + list(lw)

        from ai_page_probe import apply_ai_assert_grounding_to_plan

        plan_dict, norm_warnings = apply_ai_assert_grounding_to_plan(plan_dict, norm_warnings)
    except Exception as e:
        uat_logger.warning("AI locator merge skipped: %s", e)
    return plan_dict, norm_warnings


def _ai_desktop_planning_snapshot() -> tuple:
    """返回 (snapshot_text, warning) 供桌面用例规划。"""
    try:
        from desktop_discovery import desktop_runtime_snapshot, discovery_available

        if not discovery_available():
            return "", "桌面自动化当前仅支持 Windows"
        snap = desktop_runtime_snapshot(include_catalog=False)
        wins = snap.get("windows") or []
        lines = ["当前可见窗口（生成 attach_window / click 步骤时请引用 title 或 hwnd）："]
        n = 0
        for w in wins:
            title = (w.get("title") or "").strip()
            if not title:
                continue
            n += 1
            if n > 35:
                break
            lines.append(
                f"  [{n}] title={title!r} hwnd={w.get('hwnd')} pid={w.get('pid')}"
            )
        if n == 0:
            lines.append("  （暂无可见窗口标题）")
        return "\n".join(lines), ""
    except Exception as e:
        return "", f"无法获取桌面快照：{e}"


def _ai_execute_desktop_plan_steps(steps: list) -> list:
    """在本地执行桌面层步骤（AI 运行模式）。"""
    from desktop_run_context import reset_desktop_run_context
    from step_executor import enrich_execution_step, sync_execute_step_by_layer

    reset_desktop_run_context()
    results = []
    for raw in steps or []:
        if not isinstance(raw, dict):
            continue
        step = enrich_execution_step(dict(raw))
        if (step.get("automation_layer") or "").strip().lower() != "desktop":
            step["automation_layer"] = "desktop"
        try:
            row = sync_execute_step_by_layer(step)
            results.append({
                "ok": True,
                "action": step.get("action"),
                "status": row.get("status") if isinstance(row, dict) else "success",
                "description": step.get("description") or "",
            })
        except Exception as exc:
            results.append({
                "ok": False,
                "action": step.get("action"),
                "error": str(exc),
                "description": step.get("description") or "",
            })
            break
    return results


def _execute_ai_task_plan(data: dict, user_id: int, username: str, remote_addr):
    """规划逻辑（供同步 API 与后台线程共用）。返回体可含 _http 表示建议 HTTP 状态码。"""
    from ai_page_probe import probe_registry_from_interactive_snapshot

    data = data or {}
    task_type = (data.get('task_type') or 'test_case_generation').strip()
    platform_type = (data.get('platform_type') or 'web').strip().lower()
    goal = (data.get('goal') or '').strip()
    project_name = (data.get('project_name') or '').strip()
    selected_model = (data.get('model') or '').strip() or _get_active_local_model()
    profile, legacy_model = _resolve_inference_profile(selected_model)

    if not goal:
        return {'success': False, 'error': 'goal不能为空', '_http': 400}

    route = _route_ai_model(task_type)
    if route['provider'] != 'local':
        return {
            'success': False,
            'error': '该任务需走云端分析接口，请调用 /api/ai/task/cloud-analyze',
            '_http': 400,
        }

    embedded_sid = (data.get('embedded_session_id') or data.get('remote_session_id') or '').strip()
    target_page_url = (data.get('target_page_url') or '').strip()
    execution_mode = (data.get('execution_mode') or '').strip().lower()
    run_execute = execution_mode in ('run', 'run_and_record', 'execute')

    page_snapshot = None
    probe_registry = None
    probe_url = (target_page_url or None) if not embedded_sid else None
    snap_data = None
    dom_probe_warning = ""

    if embedded_sid:
        if not embedded_gateway_enabled():
            return {'success': False, 'error': '远程 Chromium 网关未配置', '_http': 503}
        j, err = embedded_gateway_json(
            'GET',
            f'/internal/session/{embedded_sid}/inspect',
            user_id=user_id,
        )
        if not j or not j.get('success'):
            detail = (j or {}).get('detail')
            return {
                'success': False,
                'error': str(err or detail or '无法获取远程页结构，请确认远程画布已连接'),
                '_http': 502,
            }
        snap = j.get('data') or {}
        snap_data = snap
        ps, pr, pu = probe_registry_from_interactive_snapshot(snap)
        page_snapshot = ps or None
        probe_registry = pr if pr else None
        probe_url = (pu or target_page_url).strip() or None
    elif run_execute and platform_type not in ('android', 'desktop'):
        if not target_page_url:
            return {
                'success': False,
                'error': '「运行」模式需要填写目标页面 URL。',
                '_http': 400,
            }
        if not _ai_should_open_main_playwright(embedded_sid=embedded_sid):
            return {
                'success': False,
                'error': (
                    '「运行」需要已连接的内置画布会话（embedded_session_id）。'
                    '请先在 AI 测试页建立画布实时画面；平台会在启动时自动拉起 Browser Runtime。'
                    '（Testory AI (Hermes) 与 Browser Runtime 是不同服务。）'
                    '若确需回退主 Playwright，请在 .env 设置 AI_ALLOW_MAIN_PLAYWRIGHT_FALLBACK=1。'
                ),
                '_http': 400,
            }
        snap_data, page_snapshot, probe_registry, probe_url, err = _ai_main_session_dom_pack(
            target_page_url, strict=True
        )
        if err:
            code = 400 if ("主浏览器未就绪" in err or "目标 URL 为空" in err) else 500
            return {'success': False, 'error': err, '_http': code}
    elif platform_type == 'desktop':
        dsnap, dwarn = _ai_desktop_planning_snapshot()
        if dsnap:
            page_snapshot = dsnap
        if dwarn:
            dom_probe_warning = dwarn
    elif (target_page_url or "").strip() and not embedded_sid and not _ai_url_probe_disabled() and platform_type not in ('android', 'desktop'):
        # 「仅规划」也抓取 LIVE：只要填写了目标 URL，就为主会话打开该页并探测，避免模型凭空写选择器。
        if _ai_should_open_main_playwright(embedded_sid=embedded_sid):
            snap_data, page_snapshot, probe_registry, probe_url, err = _ai_main_session_dom_pack(
                target_page_url, strict=False
            )
            if err:
                dom_probe_warning = (
                    "未能抓取 LIVE 页面结构，选择器将未经页面探测约束："
                    f"{err} "
                    "（可配置 Playwright / 主浏览器，或使用「运行生成」强制探测。设置 LOCAL_AI_SKIP_URL_PROBE=1 可关闭自动打开浏览器。）"
                )
        elif embedded_gateway_enabled():
            from ai_page_probe import fetch_page_controls_bundle

            summary, probe_err, headless_registry = fetch_page_controls_bundle(target_page_url)
            if headless_registry:
                page_snapshot = summary or page_snapshot
                probe_registry = headless_registry
                probe_url = (target_page_url or "").strip() or None
            elif probe_err:
                dom_probe_warning = (
                    "未能获取页面实时结构，选择器将未经页面探测约束："
                    f"{probe_err}（请确认目标 URL 可访问。）"
                )

    dpack = _ai_build_dom_pack(snap_data, embed_remote=bool(embedded_sid)) if snap_data else ""
    mem_ctx = _ai_memory_context_block(
        user_id, goal, probe_url=probe_url or target_page_url or "", project_name=project_name
    )
    try:
        generated = local_ai_service.generate_case_and_steps(
            goal,
            project_name,
            model=legacy_model,
            profile=profile,
            page_snapshot=page_snapshot,
            probe_registry=probe_registry,
            probe_url=probe_url,
            memory_context=mem_ctx or None,
            dom_context_pack=dpack or None,
            platform_type=platform_type,
        )
    except ValueError as e:
        return {
            'success': False,
            'error': str(e),
            'hint': '可先执行: ollama serve，并确认模型已拉取。',
            '_http': 503,
        }
    except Exception as e:
        uat_logger.exception('ai plan execute failed')
        return {'success': False, 'error': str(e), '_http': 500}

    generated, norm_warnings = apply_step_normalization_to_plan(generated)
    generated, norm_warnings = _merge_ai_locator_resolution(generated, snap_data, norm_warnings)
    if dom_probe_warning:
        norm_warnings = [dom_probe_warning] + list(norm_warnings or [])
    log_ai_plan_to_audit(
        user_id,
        username,
        'AI_PLAN_GENERATE',
        generated,
        remote_addr,
    )
    out: dict = {
        'success': True,
        'provider': route['provider'],
        'model': generated.get('meta', {}).get('model') or route['model'],
        'plan': generated,
        'warnings': norm_warnings,
    }
    if run_execute:
        try:
            if platform_type == 'desktop':
                from desktop_runtime import desktop_runtime_available, desktop_runtime_unavailable_reason

                if not desktop_runtime_available():
                    out["execution"] = {
                        "ran": False,
                        "skipped_reason": desktop_runtime_unavailable_reason() or "桌面运行时不可用",
                        "results": [],
                    }
                else:
                    exec_results = _ai_execute_desktop_plan_steps(generated.get('steps') or [])
                    out["execution"] = {"ran": True, "results": exec_results, "via": "desktop"}
                    _attach_vision_replay_to_execution(out["execution"])
            else:
                script_steps = ai_plan_steps_to_playwright_script_steps(generated.get('steps') or [])
                gate_url = (target_page_url or "").strip() or (generated.get("case_url") or "").strip()
                if (
                    gate_url.startswith(("http://", "https://"))
                    and script_steps
                    and script_steps[0].get("action") != "navigate"
                ):
                    script_steps = [
                        {"action": "navigate", "url": gate_url, "description": "（运行）先打开目标页"},
                    ] + script_steps
                max_n = int(os.environ.get("AI_AGENT_GATEWAY_MAX_STEPS", "40") or 40)
                script_steps = script_steps[: max(1, max_n)]
                if not script_steps:
                    out["execution"] = {
                        "ran": False,
                        "skipped_reason": "无可映射为脚本的原子步骤（请检查 navigate/选择器等）。",
                        "results": [],
                    }
                elif embedded_sid:
                    if not embedded_gateway_enabled():
                        out["execution"] = {
                            "ran": False,
                            "skipped_reason": "远程 Chromium 网关未配置，无法在内置画布会话中执行。",
                            "results": [],
                        }
                    else:
                        exec_results = _ai_embedded_run_script_steps_sequentially(
                            user_id, embedded_sid, script_steps
                        )
                        out["execution"] = {"ran": True, "results": exec_results, "via": "embedded"}
                        _attach_vision_replay_to_execution(out["execution"])
                elif embedded_gateway_enabled() and not _ai_allow_main_playwright_fallback():
                    out["execution"] = {
                        "ran": False,
                        "skipped_reason": (
                            "未连接内置画布会话，已跳过主 Playwright 执行。"
                            "请先建立画布或设置 AI_ALLOW_MAIN_PLAYWRIGHT_FALLBACK=1。"
                        ),
                        "results": [],
                    }
                else:
                    exec_results = sync_execute_script_steps(script_steps)
                    out["execution"] = {"ran": True, "results": exec_results, "via": "main_playwright"}
                    _attach_vision_replay_to_execution(out["execution"])
        except Exception as e:
            uat_logger.exception("ai run-and-record execution")
            out["execution"] = {"ran": True, "error": str(e), "results": []}
    return out


def _execute_ai_task_chat(data: dict, user_id: int, username: str, remote_addr, *, abort_event=None):
    """多轮优化逻辑（供同步 API 与后台线程共用）。"""
    from ai_page_probe import probe_registry_from_interactive_snapshot

    data = data or {}
    message = (data.get('message') or '').strip()
    project_name = (data.get('project_name') or '').strip()
    current_plan = data.get('current_plan') or {}
    history = data.get('history') or []
    selected_model = (data.get('model') or '').strip() or _get_active_local_model()
    profile, legacy_model = _resolve_inference_profile(selected_model)
    if not message:
        return {'success': False, 'error': 'message不能为空', '_http': 400}

    route = _route_ai_model('test_case_generation')
    from ai_chat_tool_loop import ai_chat_tools_enabled, profile_supports_ai_chat_tools

    _ai_chat_tools_on = ai_chat_tools_enabled() and profile_supports_ai_chat_tools(profile, legacy_model)
    if route['provider'] != 'local' and not _ai_chat_tools_on:
        return {'success': False, 'error': '当前仅支持本地模型对话', '_http': 400}

    embedded_sid = (data.get('embedded_session_id') or data.get('remote_session_id') or '').strip()
    platform_type = (data.get('platform_type') or 'web').strip().lower()
    target_page_url = (data.get('target_page_url') or '').strip()
    page_snapshot = None
    probe_registry = None
    probe_url = (target_page_url or None) if not embedded_sid else None
    snap_data = None
    chat_dom_probe_warning = ""

    if embedded_sid and platform_type not in ('android', 'desktop'):
        if not embedded_gateway_enabled():
            return {'success': False, 'error': '远程 Chromium 网关未配置', '_http': 503}
        j, err = embedded_gateway_json(
            'GET',
            f'/internal/session/{embedded_sid}/inspect',
            user_id=user_id,
        )
        if not j or not j.get('success'):
            detail = (j or {}).get('detail')
            return {
                'success': False,
                'error': str(err or detail or '无法获取远程页结构'),
                '_http': 502,
            }
        snap = j.get('data') or {}
        snap_data = snap
        ps, pr, pu = probe_registry_from_interactive_snapshot(snap)
        page_snapshot = ps or None
        probe_registry = pr if pr else None
        probe_url = (pu or target_page_url).strip() or None

    if platform_type == 'desktop':
        dsnap, dwarn = _ai_desktop_planning_snapshot()
        if dsnap:
            page_snapshot = dsnap
        if dwarn:
            chat_dom_probe_warning = dwarn
    elif platform_type != 'android':
        url_for_probe = (target_page_url or "").strip()
        if not url_for_probe and isinstance(current_plan, dict):
            url_for_probe = (current_plan.get("case_url") or "").strip()
        if snap_data is None and url_for_probe and not embedded_sid and not _ai_url_probe_disabled():
            if _ai_should_open_main_playwright(embedded_sid=embedded_sid):
                snap_data, page_snapshot, probe_registry, probe_url, err = _ai_main_session_dom_pack(
                    url_for_probe, strict=False
                )
                if err:
                    chat_dom_probe_warning = (
                        "未能抓取 LIVE 页面结构（对话优化）："
                        f"{err} "
                        "（选择器约束可能较弱；请填写目标 URL 或确认主浏览器可用。）"
                    )
            else:
                from ai_page_probe import fetch_page_controls_bundle

                summary, probe_err, headless_registry = fetch_page_controls_bundle(url_for_probe)
                if headless_registry:
                    page_snapshot = summary or page_snapshot
                    probe_registry = headless_registry
                    probe_url = url_for_probe
                elif probe_err:
                    chat_dom_probe_warning = (
                        "未能获取页面实时结构，选择器/断言约束可能较弱："
                        f"{probe_err}（请确认用例 URL 可访问。）"
                    )

    dpack = _ai_build_dom_pack(snap_data, embed_remote=bool(embedded_sid)) if snap_data else ""
    mem_ctx = _ai_memory_context_block(
        user_id, message, probe_url=probe_url or target_page_url or "", project_name=project_name
    )
    interaction_context = None
    _ic: dict = {}
    if data.get('focus_step_index') is not None and str(data.get('focus_step_index', '')).strip() != '':
        _ic['focus_step_index'] = data.get('focus_step_index')
    _fi = data.get('focus_step_indices')
    if isinstance(_fi, list) and _fi:
        _ic['focus_step_indices'] = _fi
    _sel = (data.get('browser_selection_text') or data.get('selection_text') or '').strip()
    if _sel:
        _ic['browser_selection_text'] = _sel
    _kind = (data.get('action_kind') or data.get('intent') or '').strip()
    if _kind:
        _ic['action_kind'] = _kind
    _rm = (data.get('response_mode') or '').strip().lower()
    if _rm in ('delta', 'full'):
        _ic['response_mode'] = _rm
    elif _kind in ('optimize_step', 'merge_steps', 'assert_from_selection'):
        _ic['response_mode'] = 'full'
    elif isinstance(current_plan, dict) and (current_plan.get('steps') or []):
        _ic.setdefault('response_mode', 'full')
    if _ic:
        interaction_context = _ic

    tool_meta_extra = None
    try:
        if _ai_chat_tools_on:
            from ai_chat_tool_loop import ChatToolLoopParams, run_ai_chat_with_tools

            _ctp = ChatToolLoopParams(
                message=message,
                project_name=project_name,
                current_plan=current_plan if isinstance(current_plan, dict) else {},
                history=history,
                profile=profile,
                legacy_model=legacy_model,
                page_snapshot=page_snapshot,
                probe_registry=probe_registry,
                probe_url=probe_url,
                memory_context=mem_ctx or None,
                dom_context_pack=dpack or None,
                interaction_context=interaction_context,
                test_scope=(data.get('test_scope') or data.get('scope') or '').strip() or None,
                embedded_session_id=embedded_sid or None,
                platform_type=platform_type,
            )
            try:
                generated, _, tool_meta_extra = run_ai_chat_with_tools(
                    local_ai_service=local_ai_service,
                    params=_ctp,
                    abort_event=abort_event,
                )
            except (ValueError, TypeError, RuntimeError) as loop_err:
                uat_logger.warning('AI chat tool loop failed, fallback to refine: %s', loop_err)
                tool_meta_extra = {'fallback': 'refine_after_loop_error', 'error': str(loop_err)}
                generated = local_ai_service.refine_case_and_steps(
                    user_message=message,
                    project_name=project_name,
                    current_plan=current_plan if isinstance(current_plan, dict) else {},
                    history=history if isinstance(history, list) else [],
                    model=legacy_model,
                    profile=profile,
                    page_snapshot=page_snapshot,
                    probe_registry=probe_registry,
                    probe_url=probe_url,
                    memory_context=mem_ctx or None,
                    dom_context_pack=dpack or None,
                    interaction_context=interaction_context,
                )
        else:
            generated = local_ai_service.refine_case_and_steps(
                user_message=message,
                project_name=project_name,
                current_plan=current_plan if isinstance(current_plan, dict) else {},
                history=history if isinstance(history, list) else [],
                model=legacy_model,
                profile=profile,
                page_snapshot=page_snapshot,
                probe_registry=probe_registry,
                probe_url=probe_url,
                memory_context=mem_ctx or None,
                dom_context_pack=dpack or None,
                interaction_context=interaction_context,
            )
    except InterruptedError:
        raise
    except ValueError as e:
        msg = str(e)
        low = msg.lower()
        hint = (
            '可先执行 ollama serve，并确认 ollama pull 过所选模型。'
            '若仅在「改步骤」时报错而列表正常，多半是推理接口慢或模型不合适，而非「没启动」。'
        )
        if 'read timed out' in low or 'timed out' in low:
            hint = (
                '本次为推理超时：界面能列出模型仅代表 Ollama 在线；改写步骤会发送长文本，耗时可远大于列表。'
                '建议①在 /ai-test 选用纯文本 instruct 模型（尽量避免名称含 -vl 的视觉模型）；'
                '②终端执行 ollama run <模型名> 预热；③设置环境变量 LOCAL_LLM_TIMEOUT_CHAT（仅影响生成/对话 POST）'
                '或 LOCAL_LLM_TIMEOUT 后重启 HuFirst（未设置时 /api/chat 默认读超时 600 秒）；④确认机器算力与显存足够。'
            )
        return {
            'success': False,
            'error': msg,
            'hint': hint,
            '_http': 503,
        }
    except Exception as e:
        uat_logger.exception('ai chat execute failed')
        return {'success': False, 'error': str(e), '_http': 500}

    generated, norm_warnings = apply_step_normalization_to_plan(generated)
    generated, norm_warnings = _merge_ai_locator_resolution(generated, snap_data, norm_warnings)
    if chat_dom_probe_warning:
        norm_warnings = [chat_dom_probe_warning] + list(norm_warnings or [])
    log_ai_plan_to_audit(
        user_id,
        username,
        'AI_PLAN_REFINE',
        generated,
        remote_addr,
    )
    out = {
        'success': True,
        'provider': 'local',
        'model': generated.get('meta', {}).get('model') or route['model'],
        'plan': generated,
        'warnings': norm_warnings,
    }
    if tool_meta_extra:
        out['chat_tools'] = tool_meta_extra
    return out


_AI_JOB_ABORT_EVENTS: Dict[str, threading.Event] = {}


def _start_ai_bg_job_thread(job_id: str, kind: str, data: dict, user_id: int, username: str, remote_addr):
    data_copy = dict(data)
    store = get_job_store()
    abort_event = threading.Event()
    _AI_JOB_ABORT_EVENTS[job_id] = abort_event

    def _runner():
        out = {'success': False, 'error': 'unknown', '_http': 500}
        try:
            if kind == 'plan':
                out = _execute_ai_task_plan(data_copy, user_id, username, remote_addr)
            elif kind == 'chat':
                out = _execute_ai_task_chat(data_copy, user_id, username, remote_addr, abort_event=abort_event)
            else:
                out = {'success': False, 'error': 'unknown job kind', '_http': 400}
        except InterruptedError as ie:
            # 超时/死循环已被包装成 InterruptedError 时，保留原文，勿一律改成「用户取消」
            out = {
                'success': False,
                'error': str(ie).strip() or '任务已被用户取消',
                '_http': 499,
            }
        except Exception as ex:
            uat_logger.exception('ai bg job %s', job_id)
            out = {'success': False, 'error': str(ex), '_http': 500}
        finally:
            _AI_JOB_ABORT_EVENTS.pop(job_id, None)
        if store.is_cancelled(job_id):
            store.set_cancelled(job_id)
            return
        body = {k: v for k, v in out.items() if k != '_http'}
        http_status = int(out.get('_http', 200))
        store.set_result(job_id, body, http_status)

    threading.Thread(target=_runner, daemon=True, name='ai-bg-' + job_id[:8]).start()


@app.route('/api/ai/task/plan', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_task_plan():
    """
    AI任务规划入口（本地模型真实推理）。
    """
    data = request.get_json(silent=True) or {}
    out = _execute_ai_task_plan(
        data,
        current_user.id,
        current_user.username,
        request.remote_addr,
    )
    code = int(out.get('_http', 200))
    body = {k: v for k, v in out.items() if k != '_http'}
    return jsonify(body), code


@app.route('/api/ai/task/cloud-analyze', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_task_cloud_analyze():
    """
    云端高复杂任务入口（脚本修复/复杂报错分析）。
    铁律：任何上云内容必须先经过自动全量脱敏。
    """
    data = request.get_json(silent=True) or {}
    task_type = (data.get('task_type') or '').strip()
    payload = data.get('payload')

    if task_type not in {'script_repair', 'complex_error_analysis'}:
        return jsonify({'success': False, 'error': '仅支持 script_repair / complex_error_analysis'}), 400
    if payload is None:
        return jsonify({'success': False, 'error': 'payload不能为空'}), 400

    route = _route_ai_model(task_type)
    if route['provider'] != 'cloud':
        return jsonify({'success': False, 'error': '该任务不需要云端模型'}), 400

    gateway = _get_cloud_llm_gateway()
    if gateway is None:
        return jsonify({
            'success': False,
            'error': '云端模型未配置，请设置 CLOUD_LLM_ENDPOINT 和 CLOUD_LLM_API_KEY'
        }), 400

    result = gateway.call({
        'task_type': task_type,
        'payload': payload,
        'request_meta': {
            'requested_by': current_user.username,
            'requested_at': beijing_now_iso(),
        }
    })
    try:
        from ai_memory_store import ingest_repair_case, memory_enabled

        if memory_enabled():
            _dbm = Database()
            tidm = _dbm.get_user_tenant_id(current_user.id)
            ingest_repair_case(
                current_user.id,
                task_type,
                payload,
                cloud_result=result,
                tenant_id=tidm,
            )
    except Exception as e:
        uat_logger.debug("ai memory ingest skipped: %s", e)

    return jsonify({
        'success': True,
        'provider': route['provider'],
        'model': route['model'],
        'result': result
    })


@app.route('/api/ai/memory/ingest', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_memory_ingest():
    """手工写入用户习惯/备注，进入本地向量记忆（需 LOCAL_MEMORY_ENABLE=1 且已拉取 LOCAL_EMBED_MODEL）。"""
    data = request.get_json(silent=True) or {}
    kind = (data.get('kind') or 'habit').strip()[:64]
    text = (data.get('text') or data.get('source_text') or '').strip()
    meta = data.get('meta')
    if not text or len(text) < 4:
        return jsonify({'success': False, 'error': 'text 过短或为空'}), 400
    try:
        from ai_memory_store import ingest, memory_enabled

        if not memory_enabled():
            return jsonify({'success': False, 'error': '未开启：设置 LOCAL_MEMORY_ENABLE=1'}), 400
        _db = Database()
        tid = _db.get_user_tenant_id(current_user.id)
        mid = ingest(
            current_user.id,
            kind,
            text,
            tenant_id=tid,
            meta=meta if isinstance(meta, dict) else None,
        )
        return jsonify({'success': True, 'id': mid})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 502


@app.route('/api/ai/scenarios/from-requirements', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_structured_scenarios():
    """
    从需求说明书文本生成半结构化测试场景（无 UI 定位），便于评审后再生成自动化步骤。
    支持 JSON body：requirements_text / extra_context / model；
    或 multipart：字段 file=说明书（.txt/.md/.pdf/.docx 等）。
    """
    requirements_text = ''
    extra_context = ''
    selected_model = ''
    file_warns: List[str] = []

    if request.files and request.files.get('file'):
        f = request.files['file']
        raw = f.read() or b''
        fname = (f.filename or 'upload.txt').strip()
        from requirements_document_extract import extract_text_from_bytes

        requirements_text, file_warns = extract_text_from_bytes(fname, raw)
        requirements_text = (requirements_text or '').strip()
    else:
        data = request.get_json(silent=True) or {}
        requirements_text = (data.get('requirements_text') or data.get('text') or '').strip()
        extra_context = (data.get('extra_context') or '').strip()
        selected_model = (data.get('model') or '').strip()

    if not requirements_text:
        return jsonify({'success': False, 'error': 'requirements_text 为空或未上传可解析文件'}), 400

    mid = selected_model or _get_active_local_model()
    profile, _legacy = _resolve_inference_profile(mid)
    route = _route_ai_model('test_case_generation')
    try:
        from ai_structured_scenarios import (
            generate_structured_scenarios_from_requirements,
            generate_structured_scenarios_from_requirements_chunked,
            structured_scenarios_chunk_size,
        )

        if len(requirements_text) > int(structured_scenarios_chunk_size() * 1.15):
            doc, w0 = generate_structured_scenarios_from_requirements_chunked(
                requirements_text, profile, extra_context=extra_context
            )
        else:
            doc, w0 = generate_structured_scenarios_from_requirements(
                requirements_text, profile, extra_context=extra_context
            )
    except Exception as e:
        uat_logger.exception('structured scenarios failed')
        return jsonify({'success': False, 'error': str(e)}), 500

    warns = list(file_warns) + list(w0)
    return jsonify(
        {
            'success': True,
            'provider': route.get('provider'),
            'model': route.get('model'),
            'document': doc,
            'warnings': warns,
        }
    )


@app.route('/api/ai/locator/resolve-preview', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_locator_resolve_preview():
    """打开 URL、抓取交互快照并对 steps 执行「执行前定位解析」（不写入数据库）。"""
    data = request.get_json(silent=True) or {}
    url = (data.get('url') or '').strip()
    steps = data.get('steps')
    if not url:
        return jsonify({'success': False, 'error': 'url 不能为空'}), 400
    if not isinstance(steps, list):
        return jsonify({'success': False, 'error': 'steps 须为数组'}), 400

    snap_data, _ps, _pr, _pu, err = _ai_main_session_dom_pack(url, strict=True)
    if err:
        code = 400 if ('目标 URL' in err or '主浏览器未就绪' in err) else 502
        return jsonify({'success': False, 'error': err}), code

    try:
        from ai_locator_resolution import resolve_plan_steps_locators_with_snapshot

        resolved, warns = resolve_plan_steps_locators_with_snapshot(steps, snap_data, force=True)
    except Exception as e:
        uat_logger.exception('locator resolve-preview failed')
        return jsonify({'success': False, 'error': str(e)}), 500

    meta = {}
    if isinstance(snap_data, dict):
        meta = {
            'title': snap_data.get('title'),
            'url': snap_data.get('url'),
            'item_count': len(snap_data.get('items') or []),
        }

    return jsonify({'success': True, 'steps': resolved, 'warnings': warns, 'snapshot_meta': meta})


def _normalize_locator_candidates_for_db(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, str):
        return val
    try:
        return json.dumps(val, ensure_ascii=False)
    except (TypeError, ValueError):
        return str(val)


@app.route('/api/cases/<int:case_id>/ai/locator-resolve-save', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_case_ai_locator_resolve_save(case_id: int):
    """
    对用例已有步骤：打开解析 URL、抓取 DOM 快照，将定位解析结果写回 test_steps（selector / locator_candidates）。
    """
    _db = Database()
    case = _db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404
    if _app_case_type(case) != 'ui':
        return jsonify({'success': False, 'error': '仅支持 Web 用例'}), 400
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403

    steps_db = _db.get_case_steps(case_id)
    steps_db = sorted(steps_db, key=lambda x: int(x.get('step_order') or 0))
    if not steps_db:
        return jsonify({'success': False, 'error': '该用例没有步骤'}), 400

    nav_url, _src = _resolve_case_navigation_url(case=case, steps=steps_db)
    if not nav_url:
        return jsonify({'success': False, 'error': '无法解析用于探测的 URL（请设置用例 URL 或步骤中的 navigate/url）'}), 400

    snap_data, _ps, _pr, _pu, err = _ai_main_session_dom_pack(nav_url, strict=True)
    if err:
        code = 400 if ('目标 URL' in err or '主浏览器未就绪' in err) else 502
        return jsonify({'success': False, 'error': err}), code

    plan = []
    for s in steps_db:
        plan.append(
            {
                'action': s.get('action'),
                'selector_type': s.get('selector_type'),
                'selector_value': s.get('selector_value'),
                'input_value': s.get('input_value'),
                'description': s.get('description'),
                'url': s.get('url'),
            }
        )

    try:
        from ai_locator_resolution import resolve_plan_steps_locators_with_snapshot

        resolved, warns = resolve_plan_steps_locators_with_snapshot(plan, snap_data, force=True)
    except Exception as e:
        uat_logger.exception('locator resolve-save failed')
        return jsonify({'success': False, 'error': str(e)}), 500

    updated = 0
    for i, s in enumerate(steps_db):
        if i >= len(resolved) or not isinstance(resolved[i], dict):
            break
        r = resolved[i]
        lc_new = _normalize_locator_candidates_for_db(r.get('locator_candidates'))
        lc_old = _normalize_locator_candidates_for_db(s.get('locator_candidates'))
        sv_new = (r.get('selector_value') or '').strip()
        sv_old = (s.get('selector_value') or '').strip()
        st_new = (r.get('selector_type') or 'css').strip()
        st_old = (s.get('selector_type') or 'css').strip()
        if sv_new == sv_old and st_new == st_old and (lc_new or '') == (lc_old or ''):
            continue
        kwargs = {'selector_type': st_new, 'selector_value': sv_new}
        if lc_new is not None:
            kwargs['locator_candidates'] = lc_new
        if _db.update_test_step(int(s['id']), **kwargs):
            updated += 1

    return jsonify({'success': True, 'case_id': case_id, 'updated_steps': updated, 'warnings': warns})


@app.route('/api/cases/<int:case_id>/steps/bulk-patch', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_case_steps_bulk_patch(case_id: int):
    """批量按 step id 更新字段（id 必须属于该 case_id）。"""
    data = request.get_json(silent=True) or {}
    items = data.get('steps')
    if not isinstance(items, list) or not items:
        return jsonify({'success': False, 'error': 'steps 须为非空数组'}), 400

    _db = Database()
    case = _db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403

    allowed = {
        'action',
        'selector_type',
        'selector_value',
        'input_value',
        'description',
        'step_order',
        'enter_iframe',
        'iframe_selector',
        'compare_type',
        'locator_candidates',
        'click_repeat_count',
        'api_spec',
        'url',
    }

    patched = 0
    errors: List[str] = []
    for it in items[:500]:
        if not isinstance(it, dict):
            continue
        try:
            sid = int(it.get('id'))
        except (TypeError, ValueError):
            errors.append('跳过：缺少有效 id')
            continue
        row = _db.get_test_step(sid)
        if not row or int(row.get('case_id') or 0) != int(case_id):
            errors.append(f'步骤 {sid} 不属于该用例')
            continue
        kwargs: Dict[str, Any] = {}
        for k in allowed:
            if k not in it:
                continue
            v = it[k]
            if k == 'locator_candidates' and isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            if k == 'api_spec' and isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            kwargs[k] = v
        if not kwargs:
            continue
        if _db.update_test_step(sid, **kwargs):
            patched += 1

    return jsonify({'success': True, 'case_id': case_id, 'patched': patched, 'errors': errors})


def _ai_batch_web_cases_internal(
    *,
    project_id,
    scenarios: List[dict],
    user_id: int,
    username: str,
    remote_addr: str,
    project_name: str = '',
    base_url: str = '',
    selected_model: str = '',
    max_n: int = 20,
) -> Dict[str, Any]:
    """将结构化场景批量转为 Web 用例并落库（供路由与 Hub 复用）。"""
    if not project_id:
        return {'success': False, 'error': 'project_id不能为空', '_http': 400}
    if not scenarios:
        return {'success': False, 'error': 'scenarios 为空', '_http': 400}

    mid = (selected_model or '').strip() or _get_active_local_model()
    profile, legacy_model = _resolve_inference_profile(mid)
    max_n = max(1, min(int(max_n or 20), 40))

    _db = Database()
    if not _db.check_project_access(user_id, project_id, 'editor'):
        return {'success': False, 'error': '无权限在此项目创建用例', '_http': 403}

    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    created_ids: List[int] = []
    all_warnings: List[str] = []

    for sc in scenarios[:max_n]:
        if not isinstance(sc, dict):
            continue
        sid = _ai_str(sc.get('id')) or f"auto_{len(created_ids)+1}"
        title = _ai_str(sc.get('title')) or f"场景 {sid}"
        pre = sc.get('preconditions') or []
        hs = sc.get('high_level_steps') or []
        er = sc.get('expected_results') or []
        if isinstance(pre, str):
            pre = [pre]
        if isinstance(hs, str):
            hs = [hs]
        if isinstance(er, str):
            er = [er]
        pre_t = "\n".join(str(x) for x in pre if str(x).strip())
        hs_t = "\n".join(f"- {x}" for x in hs if str(x).strip())
        er_t = "\n".join(str(x) for x in er if str(x).strip())
        goal = (
            f"测试场景：{title}\n"
            f"场景标识：{sid}\n"
            + (f"前置条件：\n{pre_t}\n\n" if pre_t else "")
            + (f"步骤概要：\n{hs_t}\n\n" if hs_t else "")
            + (f"期望：\n{er_t}" if er_t else "")
        ).strip()
        desc = f"[REQ:{sid}] {title}"
        if len(desc) > 3900:
            desc = desc[:3897] + "..."

        current_case_count = _db.get_project_case_count(project_id)
        if limits['max_cases_per_project'] != -1 and current_case_count >= limits['max_cases_per_project']:
            all_warnings.append(f"已达项目用例上限，停止在 {len(created_ids)} 条")
            break

        mem_ctx = _ai_memory_context_block(user_id, goal, probe_url=base_url, project_name=project_name)
        try:
            generated = local_ai_service.generate_case_and_steps(
                goal,
                project_name,
                model=legacy_model,
                profile=profile,
                memory_context=mem_ctx or None,
            )
        except ValueError as e:
            all_warnings.append(f"场景 {sid} 跳过：{e}")
            continue

        if license_info.license_type == LicenseType.FREE.value:
            _db.increment_created_cases(user_id)

        if base_url and not (generated.get('case_url') or '').strip():
            generated['case_url'] = base_url

        local_ai_service._fill_missing_step_payloads(
            generated.get('steps') or [],
            goal,
            _ai_str(generated.get('case_url')),
            None,
        )
        generated, norm_warns = apply_step_normalization_to_plan(generated)
        all_warnings.extend(norm_warns)

        case_id = _db.create_test_case_v2(
            project_id,
            (title[:200] if title else f"AI-{sid}")[:200],
            generated.get('case_url', ''),
            desc,
            pre_t[:2000] if pre_t else generated.get('precondition', ''),
            er_t[:2000] if er_t else generated.get('expected_result', ''),
        )
        steps = generated.get('steps') or []
        for idx, step in enumerate(steps, start=1):
            _db.create_test_step(**_ai_step_to_db_kwargs(step, case_id, idx))
        log_ai_plan_to_audit(
            user_id,
            username,
            'AI_PLAN_SCENARIO_BATCH',
            {**generated, 'scenario_id': sid},
            remote_addr,
        )
        created_ids.append(case_id)

    return {
        'success': True,
        'created_case_ids': created_ids,
        'count': len(created_ids),
        'warnings': all_warnings,
    }


@app.route('/api/ai/cases/from-scenarios-batch', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_CASE', 'case')
def api_ai_cases_from_scenarios_batch():
    """
    将结构化场景列表批量转为 Web 用例并落库；用例 description 前缀 [REQ:场景id] 便于追溯。
    Body: project_id, document 或 scenarios, model?, project_name?, base_url? , max_scenarios? (默认20)
    """
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    project_name = (data.get('project_name') or '').strip()
    base_url = (data.get('base_url') or '').strip()
    selected_model = (data.get('model') or '').strip()

    try:
        max_n = int(data.get('max_scenarios') or 20)
    except (TypeError, ValueError):
        max_n = 20

    doc = data.get('document') if isinstance(data.get('document'), dict) else {}
    scenarios = data.get('scenarios')
    if not isinstance(scenarios, list):
        scenarios = doc.get('scenarios') if isinstance(doc.get('scenarios'), list) else []

    out = _ai_batch_web_cases_internal(
        project_id=project_id,
        scenarios=scenarios,
        user_id=current_user.id,
        username=current_user.username,
        remote_addr=request.remote_addr,
        project_name=project_name,
        base_url=base_url,
        selected_model=selected_model,
        max_n=max_n,
    )
    code = int(out.pop('_http', 200))
    return jsonify(out), code


@app.route('/api/ai/import/api-spec/preview', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_import_api_spec_preview():
    """解析 OpenAPI / Postman JSON，返回可导入的请求列表预览。"""
    base_url = ''
    content = ''
    if request.files and request.files.get('file'):
        f = request.files['file']
        content = (f.read() or b'').decode('utf-8', errors='replace')
    else:
        data = request.get_json(silent=True) or {}
        content = (data.get('content') or data.get('text') or '').strip()
        base_url = (data.get('base_url') or data.get('server_url') or '').strip()

    from api_doc_import import detect_and_parse_api_doc

    kind, items, warns = detect_and_parse_api_doc(content, base_url_override=base_url)
    if kind == 'unknown':
        return jsonify({'success': False, 'error': (warns or ['无法解析'])[0], 'warnings': warns}), 400
    return jsonify({'success': True, 'kind': kind, 'items': items[:500], 'warnings': warns})


@app.route('/api/ai/import/api-spec/commit', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_CASE', 'case')
def api_ai_import_api_spec_commit():
    """将预览项写入一个接口用例（多 api_request 步骤）。"""
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    case_name = (data.get('case_name') or '').strip() or '导入的接口用例'
    items = data.get('items')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    if not isinstance(items, list) or not items:
        return jsonify({'success': False, 'error': 'items 须为非空数组'}), 400

    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403

    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    current_case_count = _db.get_project_case_count(project_id)
    if limits['max_cases_per_project'] != -1 and current_case_count >= limits['max_cases_per_project']:
        return jsonify({'success': False, 'error': '已达到项目用例数量限制', 'limit_reached': True}), 403
    if license_info.license_type == LicenseType.FREE.value:
        _db.increment_created_cases(current_user.id)

    desc = (data.get('description') or '').strip()[:4000]
    case_id = _db.create_test_case_v2(
        project_id,
        case_name[:200],
        '',
        desc,
        '',
        '',
        case_type='api',
    )
    n = 0
    for i, it in enumerate(items[:300], start=1):
        if not isinstance(it, dict):
            continue
        spec = it.get('api_spec')
        if isinstance(spec, str):
            try:
                spec = json.loads(spec)
            except json.JSONDecodeError:
                continue
        if not isinstance(spec, dict):
            continue
        api_json = json.dumps(spec, ensure_ascii=False)
        desc_step = (it.get('description') or it.get('name') or f"请求{i}")[:2000]
        _db.create_test_step(
            case_id,
            'api_request',
            '',
            '',
            '',
            desc_step,
            i,
            api_spec=api_json,
        )
        n += 1

    if n == 0:
        try:
            _db.delete_test_case(case_id)
        except Exception:
            pass
        return jsonify({'success': False, 'error': '没有可写入的有效 api_spec 项'}), 400

    return jsonify({'success': True, 'case_id': case_id, 'steps_created': n})


@app.route('/api/ai/cases/import-ui-plan', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cases_import_ui_plan():
    """
    将外部 Agent（Hermes / 工具循环）产出的步骤 JSON 写入用例：可选 replace 清空原步骤后写入。
    Body: case_id, steps[], replace? (默认 false 为追加)
    """
    data = request.get_json(silent=True) or {}
    case_id = data.get('case_id')
    steps = data.get('steps') or []
    replace = bool(data.get('replace'))
    if not case_id:
        return jsonify({'success': False, 'error': 'case_id 不能为空'}), 400
    if not isinstance(steps, list) or not steps:
        return jsonify({'success': False, 'error': 'steps 须为非空数组'}), 400

    case = db.get_test_case_v2(int(case_id))
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404
    _db = Database()
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403
    if _app_case_type(case) != 'ui':
        return jsonify({'success': False, 'error': '仅支持 Web 用例'}), 400

    goal_hint = _ai_str(data.get('goal')) or _ai_str(case.get('name')) or ''
    local_ai_service._fill_missing_step_payloads(
        steps,
        goal_hint,
        _ai_str(case.get('url')),
        None,
    )
    clean_steps, warnings = dedupe_and_validate_ai_steps(steps)
    clean_steps = [s for s in clean_steps if (s.get('action') or '').strip().lower() != 'api_request']
    if not clean_steps:
        return jsonify({'success': False, 'error': '没有可写入的 Web 步骤', 'warnings': warnings}), 400

    if replace:
        _db.delete_case_steps(int(case_id))
        start_order = 0
    else:
        old_steps, _t = _db.get_case_steps_paginated(int(case_id), 1, 2000)
        start_order = max([int(s.get('step_order') or 0) for s in old_steps] or [0])

    created = 0
    for idx, step in enumerate(clean_steps, start=1):
        _db.create_test_step(**_ai_step_to_db_kwargs(step, int(case_id), start_order + idx))
        created += 1

    return jsonify({'success': True, 'case_id': int(case_id), 'steps_created': created, 'warnings': warnings})


@app.route('/api/ai/skills', methods=['GET'])
@login_required
@api_error_handler
def api_ai_skills_list():
    """列出 Hermes Skills（UAT_DATA_DIR/hermes/skills）。"""
    from ai_hermes_skills import list_skills

    return jsonify({'success': True, 'skills': list_skills()})


@app.route('/api/ai/skills/export-from-plan', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_skills_export_from_plan():
    """将 AI 计划导出为 Hermes SKILL.md。"""
    from ai_hermes_skills import export_plan_to_skill

    data = request.get_json(silent=True) or {}
    plan = data.get('plan') or data.get('current_plan') or {}
    if not isinstance(plan, dict) or not plan.get('steps'):
        return jsonify({'success': False, 'error': 'plan.steps 不能为空'}), 400
    skill_name = _ai_str(data.get('skill_name')) or _ai_str(plan.get('case_name')) or 'test-flow'
    module_hint = _ai_str(data.get('module_hint')) or skill_name
    env_notes = _ai_str(data.get('environment_notes'))
    path, summary = export_plan_to_skill(
        plan,
        skill_name=skill_name,
        module_hint=module_hint,
        environment_notes=env_notes,
    )
    return jsonify({'success': True, 'skill': summary, 'path': str(path)})


@app.route('/api/ai/skills/promote-from-run', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_skills_promote_from_run():
    """Phase B：从成功的 AgentTeams / 跨端 plan 沉淀 Skill 草稿（失败默认拒绝）。"""
    from ai_modules.skills.promote_from_run import (
        list_promoted_skills,
        promote_agent_run,
        promote_plan_to_skill_draft,
    )

    data = request.get_json(silent=True) or {}
    if request.args.get('list') in ('1', 'true') or data.get('list'):
        return jsonify({'ok': True, 'skills': list_promoted_skills(limit=int(data.get('limit') or 50))})

    force = bool(data.get('force'))
    skill_name = _ai_str(data.get('skill_name'))
    agent_run_id = _ai_str(data.get('agent_run_id') or data.get('run_id'))
    if agent_run_id:
        from ai_modules.agent_teams.test_run_state import load_run

        st = load_run(agent_run_id)
        if not st:
            return jsonify({'ok': False, 'error': 'agent run 不存在'}), 404
        path, meta = promote_agent_run(st, skill_name=skill_name, force=force)
        code = 200 if meta.get('ok') else 400
        return jsonify(meta if meta.get('ok') else {**meta, 'ok': False}), code

    plan = data.get('plan') if isinstance(data.get('plan'), dict) else {}
    success = data.get('success')
    if success is None:
        success = True
    path, meta = promote_plan_to_skill_draft(
        plan,
        skill_name=skill_name,
        source=_ai_str(data.get('source')) or 'api',
        run_id=_ai_str(data.get('history_run_id')),
        evidence_level=_ai_str(data.get('evidence_level')),
        force=force,
        success=bool(success),
    )
    code = 200 if meta.get('ok') else 400
    return jsonify(meta), code


@app.route('/api/ai/skills/promoted', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_skills_promoted_list():
    from ai_modules.skills.promote_from_run import list_promoted_skills

    try:
        limit = int(request.args.get('limit') or 50)
    except (TypeError, ValueError):
        limit = 50
    return jsonify({'ok': True, 'skills': list_promoted_skills(limit=limit)})


@app.route('/api/ai/skills/apply-to-case', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_skills_apply_to_case():
    """从 Skill 导入步骤到用例或返回合并后的 plan。"""
    from ai_hermes_skills import apply_skill_to_plan

    data = request.get_json(silent=True) or {}
    skill_id = _ai_str(data.get('skill_id'))
    if not skill_id:
        return jsonify({'success': False, 'error': 'skill_id 不能为空'}), 400
    case_id = data.get('case_id')
    base = data.get('current_plan') if isinstance(data.get('current_plan'), dict) else {}
    try:
        plan, warnings = apply_skill_to_plan(skill_id, base_plan=base)
    except (FileNotFoundError, ValueError) as e:
        return jsonify({'success': False, 'error': str(e)}), 400

    if case_id:
        steps = plan.get('steps') or []
        if not steps:
            return jsonify({'success': False, 'error': 'Skill 中无步骤'}), 400
        case = db.get_test_case_v2(int(case_id))
        if not case:
            return jsonify({'success': False, 'error': '测试用例不存在'}), 404
        _db = Database()
        if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
            return jsonify({'success': False, 'error': '无权限修改此用例'}), 403
        replace = bool(data.get('replace', True))
        if replace:
            _db.delete_case_steps(int(case_id))
            start_order = 0
        else:
            old_steps, _t = _db.get_case_steps_paginated(int(case_id), 1, 2000)
            start_order = max([int(s.get('step_order') or 0) for s in old_steps] or [0])
        created = 0
        for idx, step in enumerate(steps, start=1):
            _db.create_test_step(**_ai_step_to_db_kwargs(step, int(case_id), start_order + idx))
            created += 1
        return jsonify({'success': True, 'case_id': int(case_id), 'steps_created': created, 'warnings': warnings})

    return jsonify({'success': True, 'plan': plan, 'warnings': warnings})


@app.route('/api/ai/skills/update', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_skills_update():
    """对话式更新 Skill（UI 变更 / 自愈）。"""
    from ai_hermes_skills import request_hermes_skill_update

    data = request.get_json(silent=True) or {}
    skill_id = _ai_str(data.get('skill_id'))
    message = _ai_str(data.get('message'))
    if not skill_id or not message:
        return jsonify({'success': False, 'error': 'skill_id 与 message 不能为空'}), 400
    result = request_hermes_skill_update(
        skill_id,
        message,
        failure_context=data.get('failure_context') if isinstance(data.get('failure_context'), dict) else None,
    )
    return jsonify({'success': True, 'hermes_response': result})


# ----------------------------------------------------------------------
# 跨端联动编排 API
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/decompose', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cross_end_decompose():
    """将自然语言场景描述分解为 CrossEndPlan。"""
    from ai_modules.plan.plan_decomposer import CrossEndPlanDecomposer

    data = request.get_json(silent=True) or {}
    desc = _ai_str(data.get('description') or data.get('desc') or '')
    if not desc:
        return jsonify({'ok': False, 'error': 'description 不能为空'}), 400
    project_id = data.get('project_id')

    decomposer = CrossEndPlanDecomposer()
    try:
        result = decomposer.decompose_sync(desc)
        response = {
            'ok': result.get('ok', False),
            'plan': result.get('plan'),
            'warnings': result.get('warnings', []),
        }
        if not result.get('ok'):
            warnings = result.get('warnings', [])
            response['error'] = warnings[0] if warnings else '场景分解失败，请检查LLM服务是否运行'
        return jsonify(response)
    except Exception as exc:
        uat_logger.exception('cross-end decompose failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/ai/desktop/preflight', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_ai_desktop_preflight():
    """Desktop 主路径预检：不可用返回 ok=false + DESKTOP_NO_SESSION（不假绿）。"""
    from ai_modules.execute.desktop_preflight import check_desktop_preflight

    pre = check_desktop_preflight()
    status = 200 if pre.get('ok') else 503
    return jsonify({'ok': bool(pre.get('ok')), 'preflight': pre}), status


@app.route('/api/ai/cross-end/desktop-mainpath-plan', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cross_end_desktop_mainpath_plan():
    """返回 Windows 记事本桌面主路径标准计划（可直接 execute）。"""
    from ai_modules.execute.desktop_preflight import (
        build_notepad_mainpath_plan,
        check_desktop_preflight,
    )

    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id') or request.args.get('project_id')
    plan = build_notepad_mainpath_plan(project_id=project_id)
    pre = check_desktop_preflight()
    return jsonify({
        'ok': True,
        'plan': plan,
        'preflight': pre,
        'ready_to_run': bool(pre.get('ok')),
        'hint': (
            None
            if pre.get('ok')
            else (pre.get('error') or '桌面会话不可用，执行将诚实失败 DESKTOP_NO_SESSION')
        ),
    })


@app.route('/api/ai/cross-end/otp-demo-plan', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cross_end_otp_demo_plan():
    """可选示例：加载 demos 下 OTP 剧本（非运行时唯一模板；对话 Agent 不依赖本接口）。"""
    from pathlib import Path

    path = Path(__file__).resolve().parent / 'demos' / 'cross_end' / 'desktop_mobile_otp_plan.json'
    if not path.is_file():
        return jsonify({'ok': False, 'error': '示例剧本文件不存在'}), 404
    try:
        plan = json.loads(path.read_text(encoding='utf-8'))
    except Exception as e:
        return jsonify({'ok': False, 'error': f'读取失败: {e}'}), 500
    return jsonify({
        'ok': True,
        'plan': plan,
        'hint': '示例剧本，可改 variables_defaults 后执行；非唯一支持场景',
    })


@app.route('/api/ai/cross-end/erp-desktop-plan', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cross_end_erp_desktop_plan():
    """ERP 桌面样例计划：Fake ERP 或 DESKTOP_APP_ALIASES（@erp）↔ API 种子订单号断言。"""
    from ai_modules.execute.desktop_preflight import check_desktop_preflight
    from ai_modules.execute.erp_desktop_sample import build_erp_desktop_sample_plan

    data = request.get_json(silent=True) or {}
    q = request.args
    project_id = data.get('project_id') or q.get('project_id')
    order_id = data.get('order_id') or q.get('order_id') or 'ORD-DEMO-404'
    launch_mode = (
        data.get('launch_mode')
        or data.get('mode')
        or q.get('launch_mode')
        or q.get('mode')
        or 'fake'
    )
    alias = data.get('alias') or q.get('alias') or 'erp'
    window_title_re = data.get('window_title_re') or q.get('window_title_re')
    plan = build_erp_desktop_sample_plan(
        order_id=order_id,
        project_id=project_id,
        launch_mode=str(launch_mode),
        alias=str(alias),
        window_title_re=window_title_re,
    )
    pre = check_desktop_preflight()
    alias_err = (plan.get('meta') or {}).get('alias_error')
    ready = bool(pre.get('ok')) and not alias_err
    hint = None
    if alias_err:
        hint = alias_err
    elif not pre.get('ok'):
        hint = pre.get('error') or '桌面会话不可用，执行将诚实失败 DESKTOP_NO_SESSION'
    return jsonify({
        'ok': True,
        'plan': plan,
        'preflight': pre,
        'ready_to_run': ready,
        'alias_error': alias_err,
        'hint': hint,
    })


@app.route('/api/desktop/aliases', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_aliases_list():
    """列出桌面应用别名（catalog + 用户文件 + env）。"""
    from desktop_env_config import load_app_alias_specs, load_user_alias_specs, user_aliases_path

    specs = load_app_alias_specs()
    user = load_user_alias_specs()
    return jsonify({
        'ok': True,
        'aliases': specs,
        'user_aliases': user,
        'user_file': str(user_aliases_path()),
        'hint': '真实客户 ERP 请 PUT /api/desktop/aliases/<alias> 持久化 path/args/window_title_re',
    })


@app.route('/api/desktop/aliases/<alias>', methods=['PUT', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_alias_upsert(alias):
    """保存客户 ERP 等别名到 data/desktop_aliases.json（不改 .env）。"""
    from desktop_env_config import probe_app_alias, save_user_alias

    data = request.get_json(silent=True) or {}
    path = _ai_str(data.get('path') or data.get('exe') or '')
    if not path:
        return jsonify({'ok': False, 'error': 'path 不能为空', 'error_code': 'ALIAS_PATH_REQUIRED'}), 400
    raw_args = data.get('args') or data.get('arguments') or []
    if isinstance(raw_args, str):
        args = [raw_args] if raw_args.strip() else []
    elif isinstance(raw_args, list):
        args = [str(a) for a in raw_args]
    else:
        args = []
    title_re = _ai_str(data.get('window_title_re') or data.get('title_re') or '')
    try:
        entry = save_user_alias(alias, path=path, args=args, window_title_re=title_re)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    probe = probe_app_alias(alias, order_id=_ai_str(data.get('order_id') or ''))
    return jsonify({'ok': True, 'alias': entry, 'probe': probe})


@app.route('/api/desktop/aliases/<alias>/probe', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_alias_probe(alias):
    """探测别名 path 是否存在（不启动）。"""
    from desktop_env_config import probe_app_alias

    data = request.get_json(silent=True) or {}
    order_id = _ai_str(data.get('order_id') or request.args.get('order_id') or '')
    probe = probe_app_alias(alias, order_id=order_id)
    code = 200 if probe.get('ok') else 404
    return jsonify({'ok': bool(probe.get('ok')), **probe}), code



# ----------------------------------------------------------------------
# CI/CD 跨端计划触发 API
# ----------------------------------------------------------------------


@app.route('/api/ci/cross-end/runs', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_cross_end_runs():
    """CI 触发跨端计划执行。"""
    from ai_modules.execute.cross_end_async import create_run, start_run_thread
    from ai_modules.execute.orchestrator import execute_cross_end_plan

    data = request.get_json(silent=True) or {}
    plan = data.get('plan')
    if not isinstance(plan, dict) or not plan.get('stages'):
        return jsonify({'ok': False, 'error': 'plan.stages 不能为空'}), 400

    build_id = str(data.get('build_id') or data.get('pipeline_id') or '').strip()
    git_sha = str(data.get('git_sha') or data.get('commit') or '').strip()
    branch = str(data.get('branch') or '').strip()
    callback_url = str(data.get('callback_url') or data.get('webhook_url') or '').strip()
    project_id = data.get('project_id')
    want_async = bool(data.get('async') or str(request.args.get('async') or '').lower() in ('1', 'true'))
    uid = current_user.id if current_user.is_authenticated else None

    if want_async:
        rec = create_run(plan, user_id=str(uid or ''), project_id=project_id, trigger_source='ci')
        run_id = rec['run_id']

        def _ci_worker():
            try:
                result = execute_cross_end_plan(plan, user_id=str(uid or ''), project_id=project_id, trigger_source='ci')
                from ai_modules.execute.cross_end_async import _patch
                status = 'success' if result.get('success') is True else 'failed'
                _patch(run_id, status=status, result=result, error=result.get('error'), error_code=result.get('error_code'))
                if callback_url:
                    _deliver_ci_callback(callback_url, run_id, result)
            except Exception as e:
                from ai_modules.execute.cross_end_async import _patch
                _patch(run_id, status='failed', error=str(e)[:300], error_code='CI_CROSS_END_ERROR')

        t = threading.Thread(target=_ci_worker, name=f'ci-cross-end-{run_id}', daemon=True)
        t.start()
        return jsonify({'ok': True, 'accepted': True, 'async': True, 'run_id': run_id, 'status': 'queued', 'poll_url': f'/api/ci/cross-end/runs/{run_id}', 'build_id': build_id or None}), 202

    result = execute_cross_end_plan(plan, user_id=str(uid or ''), project_id=project_id, trigger_source='ci')
    run_id = f'ci-ce-{uuid.uuid4().hex[:10]}'
    status = 'success' if result.get('success') is True else 'failed'
    if callback_url:
        _deliver_ci_callback(callback_url, run_id, result)
    return jsonify({'ok': True, 'async': False, 'run_id': run_id, 'status': status, 'success': bool(result.get('success')), 'gate_passed': bool(result.get('gate_passed')), 'result': result, 'build_id': build_id or None})


@app.route('/api/ci/cross-end/runs/<run_id>', methods=['GET'])
@token_or_login_required
def api_ci_cross_end_run_get(run_id):
    """查询 CI 跨端执行状态。"""
    from ai_modules.execute.cross_end_async import get_run
    rec = get_run(run_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    terminal = rec.get('status') in ('success', 'failed')
    return jsonify({'ok': True, 'run_id': rec.get('run_id'), 'status': rec.get('status'), 'terminal': terminal, 'success': bool(rec.get('status') == 'success'), 'error': rec.get('error'), 'error_code': rec.get('error_code'), 'result': rec.get('result')})


@app.route('/api/ci/cross-end/runs/<run_id>/junit.xml', methods=['GET'])
@token_or_login_required
def api_ci_cross_end_run_junit(run_id):
    """导出 CI 跨端执行结果为 JUnit XML。"""
    from ai_modules.execute.cross_end_async import get_run
    rec = get_run(run_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    result = rec.get('result') or {}
    stage_results = result.get('stage_results') or []
    elapsed = result.get('total_elapsed_ms', 0)
    elapsed_sec = elapsed / 1000 if elapsed else 0
    fail_count = sum(1 for s in stage_results if not s.get('ok_assert'))
    lines = ['<?xml version="1.0" encoding="UTF-8"?>']
    lines.append('<testsuites tests="{}" failures="{}" time="{:.2f}">'.format(len(stage_results), fail_count, elapsed_sec))
    lines.append('  <testsuite name="cross-end-{}" tests="{}">'.format(run_id, len(stage_results)))
    for sr in stage_results:
        sid = sr.get('stage_id', 'unknown')
        layer = sr.get('layer', '')
        t = (sr.get('elapsed_ms') or 0) / 1000
        lines.append('    <testcase name="{}" classname="cross-end.{}" time="{:.2f}">'.format(sid, layer, t))
        if not sr.get('ok_assert'):
            err = (sr.get('error') or 'unknown').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            ec = sr.get('error_code', '')
            lines.append('      <failure message="{}" type="{}">{}</failure>'.format(err, ec, err))
        lines.append('    </testcase>')
    lines.append('  </testsuite>')
    lines.append('</testsuites>')
    xml_text = '\n'.join(lines)
    return Response(xml_text, mimetype='application/xml', headers={'Content-Disposition': 'attachment; filename="testory-cross-end-{}-junit.xml"'.format(run_id)})


def _deliver_ci_callback(callback_url, run_id, result):
    """向 CI 回调 URL 发送执行结果。"""
    if not callback_url:
        return
    import urllib.request
    payload = json.dumps({'run_id': run_id, 'success': bool(result.get('success')), 'status': 'success' if result.get('success') else 'failed', 'gate_passed': bool(result.get('gate_passed')), 'error': result.get('error'), 'stage_count': len(result.get('stage_results') or [])}, ensure_ascii=False).encode('utf-8')
    try:
        req = urllib.request.Request(callback_url, data=payload, headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=15):
            pass
    except Exception:
        pass



# ----------------------------------------------------------------------
# 性能监控与报告 API
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/performance/<run_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_performance(run_id):
    """获取指定执行的性能报告。"""
    from ai_modules.execute.performance_monitor import get_performance_report
    report = get_performance_report(run_id)
    if not report:
        return jsonify({'ok': False, 'error': '性能数据不存在'}), 404
    return jsonify({'ok': True, 'performance': report})


@app.route('/api/ai/cross-end/performance/<run_id>/bottleneck', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_bottleneck(run_id):
    """瓶颈分析。"""
    from ai_modules.execute.performance_monitor import analyze_bottlenecks
    analysis = analyze_bottlenecks(run_id)
    if analysis.get('error'):
        return jsonify({'ok': False, **analysis}), 404
    return jsonify({'ok': True, 'analysis': analysis})


@app.route('/api/ai/cross-end/performance/trends', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_trends():
    """性能趋势数据。"""
    from ai_modules.execute.performance_monitor import get_performance_trends
    scenario = request.args.get('scenario', '')
    limit = request.args.get('limit', 10, type=int)
    trends = get_performance_trends(scenario=scenario, limit=limit)
    return jsonify({'ok': True, 'trends': trends, 'count': len(trends)})


@app.route('/api/ci/cross-end/runs/<run_id>/junit-enhanced.xml', methods=['GET'])
@token_or_login_required
def api_ci_cross_end_run_junit_enhanced(run_id):
    """导出增强版 JUnit XML（含性能指标）。"""
    from ai_modules.execute.cross_end_async import get_run
    from ai_modules.execute.performance_monitor import generate_enhanced_junit
    rec = get_run(run_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    result = rec.get('result') or {}
    xml_text = generate_enhanced_junit(run_id, result)
    return Response(xml_text, mimetype='application/xml', headers={'Content-Disposition': 'attachment; filename="testory-cross-end-{}-junit-enhanced.xml"'.format(run_id)})


# ----------------------------------------------------------------------
# 场景模板库 API
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/templates', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_templates():
    """列出场景模板。"""
    from ai_modules.execute.scenario_templates import list_templates
    industry = request.args.get('industry', '')
    tag = request.args.get('tag', '')
    return jsonify({'ok': True, 'templates': list_templates(industry=industry, tag=tag)})


@app.route('/api/ai/cross-end/templates/<template_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_template_get(template_id):
    """获取模板详情。"""
    from ai_modules.execute.scenario_templates import get_template
    tpl = get_template(template_id)
    if not tpl:
        return jsonify({'ok': False, 'error': '模板不存在'}), 404
    return jsonify({'ok': True, 'template': tpl})


@app.route('/api/ai/cross-end/templates/<template_id>/instantiate', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_template_instantiate(template_id):
    """实例化模板（填充参数生成 plan）。"""
    from ai_modules.execute.scenario_templates import instantiate_template
    data = request.get_json(silent=True) or {}
    result = instantiate_template(
        template_id,
        data.get('parameters') or {},
        scenario_name=data.get('scenario_name', ''),
    )
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/ai/cross-end/templates/custom', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_template_save():
    """保存自定义模板。"""
    from ai_modules.execute.scenario_templates import save_custom_template
    data = request.get_json(silent=True) or {}
    result = save_custom_template(data)
    return jsonify(result)


@app.route('/api/ai/cross-end/templates/custom/<template_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_template_delete(template_id):
    """删除自定义模板。"""
    from ai_modules.execute.scenario_templates import delete_custom_template
    result = delete_custom_template(template_id)
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


# ----------------------------------------------------------------------
# 场景版本管理 API
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/scenario/<scenario_id>/versions', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_versions(scenario_id):
    """获取场景版本历史。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    mgr = ScenarioVersionManager()
    limit = request.args.get('limit', 20, type=int)
    history = mgr.get_history(scenario_id, limit=limit)
    return jsonify({'ok': True, 'scenario_id': scenario_id, 'versions': history})


@app.route('/api/ai/cross-end/scenario/<scenario_id>/versions/<int:version>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_version_get(scenario_id, version):
    """获取指定版本的完整 plan。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    mgr = ScenarioVersionManager()
    ver = mgr.get_version(scenario_id, version)
    if not ver:
        return jsonify({'ok': False, 'error': f'版本 {version} 不存在'}), 404
    return jsonify({'ok': True, 'version': ver})


@app.route('/api/ai/cross-end/scenario/<scenario_id>/rollback', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_rollback(scenario_id):
    """回滚到指定版本。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    data = request.get_json(silent=True) or {}
    version = data.get('version')
    if not version:
        return jsonify({'ok': False, 'error': 'version 必填'}), 400
    mgr = ScenarioVersionManager()
    result = mgr.rollback(scenario_id, int(version))
    return jsonify(result)


@app.route('/api/ai/cross-end/scenario/<scenario_id>/diff', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_diff(scenario_id):
    """对比两个版本的差异。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    v1 = request.args.get('v1', type=int)
    v2 = request.args.get('v2', type=int)
    if v1 is None or v2 is None:
        return jsonify({'ok': False, 'error': 'v1 和 v2 参数必填'}), 400
    mgr = ScenarioVersionManager()
    diff = mgr.diff(scenario_id, v1, v2)
    return jsonify({'ok': True, 'diff': diff, 'v1': v1, 'v2': v2})


@app.route('/api/ai/cross-end/scenario/<scenario_id>/export', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_export(scenario_id):
    """导出场景（指定版本或最新版本）。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    version = request.args.get('version', type=int)
    mgr = ScenarioVersionManager()
    result = mgr.export_version(scenario_id, version)
    if not result.get('success'):
        return jsonify(result), 404
    return jsonify(result)


@app.route('/api/ai/cross-end/scenario/<scenario_id>/import', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_scenario_import(scenario_id):
    """从导出的 JSON 导入场景版本。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    data = request.get_json(silent=True) or {}
    message = data.get('message', '导入')
    author = str(getattr(current_user, 'username', '') or current_user.id if current_user.is_authenticated else '')
    mgr = ScenarioVersionManager()
    result = mgr.import_version(scenario_id, data, message=message, author=author)
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/ai/cross-end/scenarios/all', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_all_versioned_scenarios():
    """列出所有有版本历史的场景。"""
    from ai_modules.execute.scenario_version_manager import ScenarioVersionManager
    mgr = ScenarioVersionManager()
    return jsonify({'ok': True, 'scenarios': mgr.list_all_scenarios()})


# ----------------------------------------------------------------------
# iOS 设备管理 API
# ----------------------------------------------------------------------


@app.route('/api/mobile/ios/devices', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ios_devices():
    """列出已连接的 iOS 设备。"""
    from mobile_engine.device.ios_device import check_ios_preflight
    pre = check_ios_preflight()
    return jsonify({'ok': True, **pre})


@app.route('/api/mobile/ios/preflight', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ios_preflight():
    """iOS 环境预检。"""
    from mobile_engine.device.ios_device import check_ios_preflight
    pre = check_ios_preflight()
    status = 200 if pre.get('idb_available') else 503
    return jsonify({'ok': pre.get('idb_available', False), 'preflight': pre}), status


@app.route('/api/mobile/ios/simulators', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ios_simulators():
    """列出 iOS 模拟器。"""
    from mobile_engine.device.ios_device import IOSDeviceManager
    sims = IOSDeviceManager.list_simulators()
    return jsonify({'ok': True, 'simulators': sims, 'count': len(sims)})


@app.route('/api/mobile/ios/simulators/<udid>/boot', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ios_simulator_boot(udid):
    """启动模拟器。"""
    from mobile_engine.device.ios_device import IOSDeviceManager
    ok, msg = IOSDeviceManager.boot_simulator(udid)
    return jsonify({'ok': ok, 'message': msg, 'udid': udid})


# ----------------------------------------------------------------------
# 多设备并行调度 API
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/multi-device/discover', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_multi_device_discover():
    """发现当前可用设备。"""
    from ai_modules.execute.multi_device_scheduler import _discover_available_devices
    platform = request.args.get('platform', 'android')
    max_dev = request.args.get('max_devices', type=int, default=0)
    devices = _discover_available_devices(platform_filter=platform, max_devices=max_dev)
    return jsonify({'ok': True, 'devices': devices, 'count': len(devices)})


@app.route('/api/ai/cross-end/multi-device/test', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_multi_device_test():
    """在多台设备上并行执行 steps。"""
    from ai_modules.execute.multi_device_scheduler import execute_multi_device_stage, multi_device_summary
    data = request.get_json(silent=True) or {}
    steps = data.get('steps') or []
    if not isinstance(steps, list) or not steps:
        return jsonify({'ok': False, 'error': 'steps 不能为空'}), 400
    stage = {'id': data.get('stage_id') or 'multi-device-test', 'layer': 'mobile', 'steps': steps, 'parallel_devices': data.get('parallel_devices') or True}
    if isinstance(data.get('devices'), list):
        stage['devices'] = data['devices']
    if data.get('max_devices'):
        stage.setdefault('parallel_devices', {})
        if isinstance(stage['parallel_devices'], dict):
            stage['parallel_devices']['max_devices'] = data['max_devices']
    result, extracted = execute_multi_device_stage(stage)
    result['multi_device_summary'] = multi_device_summary(result)
    return jsonify({'ok': True, 'result': result, 'extracted': extracted})


# ----------------------------------------------------------------------
# 可视化调试：时间线 API
# ----------------------------------------------------------------------


@app.route('/cross-end/debug', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
def cross_end_debug_page():
    """跨端调试面板页面。"""
    return render_template('cross_end_debug.html')


@app.route('/api/ai/cross-end/timeline', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_timeline_list():
    """列出活跃的跨端执行时间线。"""
    from ai_modules.execute.timeline_tracker import list_trackers
    status = request.args.get('status', '')
    limit = request.args.get('limit', 20, type=int)
    trackers = list_trackers(status=status, limit=limit)
    return jsonify({'ok': True, 'trackers': trackers})


@app.route('/api/ai/cross-end/timeline/<run_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_timeline_get(run_id):
    """获取指定执行的完整时间线。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    return jsonify({'ok': True, **tracker.to_dict()})


@app.route('/api/ai/cross-end/timeline/<run_id>/events', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
def api_cross_end_timeline_sse(run_id):
    """SSE 实时推送时间线事件。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404

    def generate():
        q = []
        tracker.add_sse_queue(q)
        try:
            for ev in tracker.get_events_since():
                yield 'data: {}\n\n'.format(json.dumps(ev, ensure_ascii=False))
            while True:
                if not q:
                    time.sleep(0.3)
                    if tracker.status in ('success', 'failed'):
                        yield 'data: {}\n\n'.format(json.dumps({'kind': 'done', 'status': tracker.status}))
                        break
                    continue
                while q:
                    payload = q.pop(0)
                    yield 'data: {}\n\n'.format(payload)
        finally:
            tracker.remove_sse_queue(q)

    return Response(stream_with_context(generate()), mimetype='text/event-stream', headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/ai/cross-end/timeline/<run_id>/stage/<stage_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_timeline_stage(run_id, stage_id):
    """获取指定阶段的详细时间线。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    stage = tracker.get_stage(stage_id)
    if not stage:
        return jsonify({'ok': False, 'error': '阶段不存在'}), 404
    return jsonify({'ok': True, 'stage': stage})


@app.route('/api/ai/cross-end/timeline/<run_id>/variables', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_timeline_vars(run_id):
    """获取变量流转历史。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    data = tracker.to_dict()
    return jsonify({'ok': True, 'variables': data.get('variables', {}), 'var_history': data.get('var_history', [])})


@app.route('/api/ai/cross-end/timeline/<run_id>/breakpoints', methods=['GET', 'POST', 'DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_breakpoints(run_id):
    """管理断点。GET=列出，POST=设置，DELETE=清除。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404

    if request.method == 'GET':
        return jsonify({'ok': True, 'breakpoints': tracker.get_breakpoints()})

    if request.method == 'DELETE':
        tracker.clear_breakpoints()
        return jsonify({'ok': True, 'message': '已清除所有断点'})

    data = request.get_json(silent=True) or {}
    stage_id = data.get('stage_id', '').strip()
    if not stage_id:
        return jsonify({'ok': False, 'error': 'stage_id 必填'}), 400
    tracker.set_breakpoint(
        stage_id,
        condition=str(data.get('condition', '')),
        enabled=bool(data.get('enabled', True)),
    )
    return jsonify({'ok': True, 'stage_id': stage_id})


@app.route('/api/ai/cross-end/timeline/<run_id>/step-mode', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_step_mode(run_id):
    """启用/禁用单步模式。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get('enabled', True))
    tracker.set_step_mode(enabled)
    return jsonify({'ok': True, 'step_mode': enabled})


@app.route('/api/ai/cross-end/timeline/<run_id>/resume', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_resume(run_id):
    """恢复执行（从断点/单步暂停中恢复）。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    tracker.resume()
    return jsonify({'ok': True, 'message': '已恢复执行'})


@app.route('/api/ai/cross-end/timeline/<run_id>/var-diff/<stage_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_var_diff(run_id, stage_id):
    """获取阶段执行前后的变量差异。"""
    from ai_modules.execute.timeline_tracker import get_tracker
    tracker = get_tracker(run_id)
    if not tracker:
        return jsonify({'ok': False, 'error': '时间线不存在'}), 404
    diff = tracker.get_var_diff(stage_id)
    return jsonify({'ok': True, 'diff': diff, 'stage_id': stage_id})


@app.route('/api/ai/cross-end/debug/panel', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_cross_end_debug_panel():
    """返回调试面板数据。"""
    from ai_modules.execute.timeline_tracker import list_trackers
    from agent_capability_registry import snapshot_capabilities
    active = list_trackers(status='running', limit=10)
    recent = list_trackers(limit=10)
    caps = snapshot_capabilities()
    return jsonify({'ok': True, 'active_executions': active, 'recent_executions': recent, 'capabilities': caps.get('capabilities', {}), 'available_skills': caps.get('available_skills', [])})

# ----------------------------------------------------------------------
# 对话测试 API（保留原区块分隔由后续路由承接）
# ----------------------------------------------------------------------


@app.route('/api/ai/cross-end/execute', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_cross_end_execute():
    """执行跨端计划。

    body.async=true 时立即返回 run_id，后台执行（企业运营：HITL 等待时可同页 resume）。
    """
    from ai_modules.execute.orchestrator import execute_cross_end_plan

    data = request.get_json(silent=True) or {}
    plan = data.get('plan')
    if not isinstance(plan, dict):
        return jsonify({'ok': False, 'error': 'plan 不能为空'}), 400
    if not plan.get('stages'):
        return jsonify({'ok': False, 'error': 'plan.stages 不能为空'}), 400
    project_id = data.get('project_id')
    want_async = bool(
        data.get('async')
        or data.get('async_mode')
        or str(request.args.get('async') or '').lower() in ('1', 'true', 'yes')
    )

    if want_async:
        from ai_modules.execute.cross_end_async import create_run, start_run_thread

        rec = create_run(
            plan,
            user_id=str(current_user.id),
            project_id=project_id,
            trigger_source='ui-async',
        )
        start_run_thread(
            rec['run_id'],
            plan,
            user_id=str(current_user.id),
            project_id=project_id,
            trigger_source='ui-async',
        )
        return jsonify({
            'ok': True,
            'async': True,
            'run_id': rec['run_id'],
            'status': rec['status'],
            'message': '已异步启动；请轮询 /api/ai/cross-end/runs/<run_id>，HITL/审批见运营面板',
        })

    try:
        result = execute_cross_end_plan(
            plan,
            user_id=str(current_user.id),
            project_id=project_id,
            trigger_source='ui',
        )
        try:
            from ai_modules.plan.user_facing_errors import enrich_result_with_user_hint
            enrich_result_with_user_hint(result)
        except Exception:
            pass
        if result.get('lock') == 'busy':
            return jsonify({
                'ok': False,
                'error': result.get('error') or '本机已有自动化任务在执行',
                'user_hint': result.get('user_hint'),
                'lock': 'busy',
                'error_code': result.get('error_code'),
                'result': result,
            }), 409
        if result.get('lock') == 'unavailable':
            return jsonify({
                'ok': False,
                'error': result.get('error') or 'execution_lock 不可用',
                'user_hint': result.get('user_hint'),
                'lock': 'unavailable',
                'error_code': result.get('error_code'),
                'result': result,
            }), 503
        # 业务失败仍返回 200 + result，由前端根据 success/user_hint 展示
        return jsonify({
            'ok': True,
            'result': result,
            'user_hint': result.get('user_hint'),
            'error_code': result.get('error_code'),
        })
    except Exception as exc:
        uat_logger.exception('cross-end execute failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/ai/cross-end/runs/<run_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_cross_end_run_get(run_id):
    """查询异步跨端执行状态。"""
    from ai_modules.execute.cross_end_async import get_run

    rec = get_run(run_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    return jsonify({'ok': True, **rec})


@app.route('/api/ai/ops/gates', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_ops_gates():
    """运营门禁：HITL waiting + RiskGuard pending。"""
    from ai_modules.execute.cross_end_async import list_ops_gates

    blob = list_ops_gates(user_id=str(current_user.id))
    return jsonify({'ok': True, **blob})


@app.route('/api/ai/risk/approve', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_risk_approve():
    """批准 L2 审批；返回 token，可写入 plan.approvals 后重试。"""
    from ai_modules.security.risk_guard import approve_risk

    data = request.get_json(silent=True) or {}
    aid = str(data.get('approval_id') or '').strip()
    if not aid:
        return jsonify({'ok': False, 'error': 'approval_id 必填'}), 400
    ok, token_or_err = approve_risk(
        aid,
        token=str(data.get('token') or '').strip(),
        approver=str(getattr(current_user, 'username', None) or current_user.id),
    )
    if not ok:
        return jsonify({'ok': False, 'error': token_or_err or '批准失败'}), 400
    return jsonify({
        'ok': True,
        'approval_id': aid,
        'token': token_or_err,
        'message': '已批准；请将 token 写入 plan.approvals[stage_id] 后重新执行',
    })


@app.route('/api/ai/risk/deny', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_risk_deny():
    from ai_modules.security.risk_guard import deny_risk

    data = request.get_json(silent=True) or {}
    aid = str(data.get('approval_id') or '').strip()
    if not aid:
        return jsonify({'ok': False, 'error': 'approval_id 必填'}), 400
    ok = deny_risk(
        aid,
        reason=str(data.get('reason') or '运营拒绝'),
        denier=str(getattr(current_user, 'username', None) or current_user.id),
    )
    if not ok:
        return jsonify({'ok': False, 'error': '审批记录不存在'}), 404
    return jsonify({'ok': True, 'approval_id': aid, 'message': '已拒绝'})


@app.route('/api/ai/cross-end/scenario', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_cross_end_scenarios():
    """列出或保存跨端场景（scenario_id 为字符串 UUID）。"""
    from ai_modules.execute.orchestrator import (
        list_cross_platform_scenarios,
        save_cross_platform_scenario,
        scenario_store_info,
    )

    if request.method == 'GET':
        project_id = request.args.get('project_id', type=int)
        all_scenarios = list_cross_platform_scenarios()
        if project_id is not None:
            all_scenarios = [
                s for s in all_scenarios
                if s.get('project_id') == project_id or str(s.get('project_id') or '') == str(project_id)
            ]
        return jsonify({
            'ok': True,
            'scenarios': all_scenarios,
            'store': scenario_store_info(),
        })

    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    if not project_id:
        return jsonify({'ok': False, 'error': 'project_id 不能为空'}), 400
    payload = dict(data)
    payload.setdefault('project_id', project_id)
    # name 保留在顶层；save 内会同步到 plan
    result = save_cross_platform_scenario(payload)
    if not result.get('success'):
        return jsonify({'ok': False, 'error': result.get('error') or '保存失败'}), 400
    sc = result.get('scenario') or {}
    sid = result.get('scenario_id') or sc.get('scenario_id')
    return jsonify({
        'ok': True,
        'scenario_id': sid,
        'id': sid,
        'scenario': sc,
        'store': scenario_store_info(),
    })


@app.route('/api/ai/cross-end/scenario/<scenario_id>', methods=['GET', 'DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_cross_end_scenario(scenario_id):
    """获取或删除单个跨端场景（字符串 scenario_id，兼容历史数字字符串）。"""
    from ai_modules.execute.orchestrator import get_cross_platform_scenario, delete_cross_platform_scenario

    sid = (scenario_id or '').strip()
    if not sid:
        return jsonify({'ok': False, 'error': 'scenario_id 不能为空'}), 400

    if request.method == 'GET':
        scenario = get_cross_platform_scenario(sid)
        if not scenario:
            return jsonify({'ok': False, 'error': '场景不存在'}), 404
        return jsonify({'ok': True, 'scenario': scenario, 'scenario_id': sid})

    result = delete_cross_platform_scenario(sid)
    if not result.get('success'):
        return jsonify({'ok': False, 'error': result.get('error') or '删除失败'}), 404
    return jsonify({'ok': True, 'scenario_id': result.get('scenario_id') or sid})


# ----------------------------------------------------------------------
# Agent Teams（Phase A：TestRunState + Planner/Executor/Verifier）
# ----------------------------------------------------------------------


@app.route('/api/ai/agent-teams/spec', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_spec():
    """加载 testory-cross-end-qa-team Spec（本地控制面，可映射 AgentTeams）。"""
    from ai_modules.agent_teams import load_team_spec

    return jsonify({'ok': True, 'spec': load_team_spec()})


@app.route('/api/ai/agent-teams/runs', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_agent_teams_runs():
    """GET 列出近期 run；POST 启动三角色闭环（同步）。"""
    from ai_modules.agent_teams import load_team_spec, run_cross_end_qa_team
    from ai_modules.agent_teams.test_run_state import list_run_ids, load_run

    if request.method == 'GET':
        ids = list_run_ids(limit=int(request.args.get('limit') or 50))
        runs = []
        for rid in ids:
            st = load_run(rid)
            if st:
                runs.append({
                    'run_id': st.run_id,
                    'status': st.status,
                    'goal': st.goal,
                    'created_at': st.created_at,
                    'finished_at': st.finished_at,
                    'agents_seen': st.agent_kinds_seen(),
                })
        return jsonify({'ok': True, 'runs': runs, 'spec': load_team_spec()})

    data = request.get_json(silent=True) or {}
    description = _ai_str(data.get('description') or data.get('goal') or data.get('desc') or '')
    plan = data.get('plan') if isinstance(data.get('plan'), dict) else None
    if not description and not (plan and plan.get('stages')):
        return jsonify({
            'ok': False,
            'error': '请提供 description 或带 stages 的 plan',
            'user_hint': '多 Agent 闭环需要自然语言目标，或直接传入已分解的跨端 plan。',
        }), 400

    try:
        state = run_cross_end_qa_team(
            description=description,
            plan=plan,
            user_id=str(current_user.id),
            idempotency_key=_ai_str(data.get('idempotency_key') or ''),
            run_id=_ai_str(data.get('run_id') or ''),
            project_id=data.get('project_id'),
        )
    except Exception as exc:
        uat_logger.exception('agent-teams run failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500

    passed = state.status == 'success'
    hist_ev = None
    for e in reversed(state.events or []):
        if (e or {}).get('message') == '已写入运行历史':
            hist_ev = e
            break
    return jsonify({
        'ok': True,
        'passed': passed,
        'run_id': state.run_id,
        'status': state.status,
        'report': state.report,
        'state': state.to_dict(),
        'run_history_id': ((hist_ev or {}).get('payload') or {}).get('run_history_id'),
        'user_hint': (
            '多 Agent 验证通过'
            if passed
            else (
                (state.report or {}).get('reason')
                or (state.errors[-1] if state.errors else '验证未通过')
            )
        ),
    })


@app.route('/api/ai/agent-teams/runs/<run_id>', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_run(run_id):
    """读取 TestRunState。"""
    from ai_modules.agent_teams.test_run_state import load_run

    rid = (run_id or '').strip()
    if not rid:
        return jsonify({'ok': False, 'error': 'run_id 不能为空'}), 400
    st = load_run(rid)
    if not st:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    return jsonify({'ok': True, 'run_id': rid, 'state': st.to_dict(), 'report': st.report})


@app.route('/api/ai/agent-teams/runs/<run_id>/report', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_report(run_id):
    """Verifier 报告（report.json 语义）。"""
    from ai_modules.agent_teams.test_run_state import load_run

    rid = (run_id or '').strip()
    st = load_run(rid)
    if not st:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    if not st.report:
        return jsonify({'ok': False, 'error': '报告尚未生成'}), 404
    return jsonify({'ok': True, 'run_id': rid, 'report': st.report, 'status': st.status})


@app.route('/api/ai/agent-teams/runs/<run_id>/promote-skill', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_promote_skill(run_id):
    """成功 AgentTeams 运行 → Skill 草稿。"""
    from ai_modules.agent_teams.test_run_state import load_run
    from ai_modules.skills.promote_from_run import promote_agent_run

    st = load_run((run_id or '').strip())
    if not st:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    data = request.get_json(silent=True) or {}
    _path, meta = promote_agent_run(
        st,
        skill_name=_ai_str(data.get('skill_name')),
        force=bool(data.get('force')),
    )
    return jsonify(meta), (200 if meta.get('ok') else 400)


@app.route('/api/enterprise/farm/nodes', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_nodes():
    """执行农场节点列表 / 登记。"""
    from ai_modules.enterprise.execution_farm import farm_summary, register_node

    if request.method == 'GET':
        return jsonify({'ok': True, **farm_summary()})
    data = request.get_json(silent=True) or {}
    try:
        node = register_node(
            name=_ai_str(data.get('name')),
            base_url=_ai_str(data.get('base_url') or data.get('url')),
            capabilities=data.get('capabilities') if isinstance(data.get('capabilities'), list) else None,
            node_id=_ai_str(data.get('node_id')),
        )
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    return jsonify({'ok': True, 'node': node})


@app.route('/api/enterprise/farm/nodes/<node_id>/probe', methods=['POST', 'GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_probe(node_id):
    from ai_modules.enterprise.execution_farm import probe_node

    result = probe_node(node_id)
    if result.get('error_code') == 'NODE_NOT_FOUND':
        return jsonify(result), 404
    # 不可达返回 200 + ok=False，避免与「节点不存在」混淆
    return jsonify(result), 200


@app.route('/api/enterprise/farm/nodes/<node_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'project_manager')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_remove(node_id):
    from ai_modules.enterprise.execution_farm import remove_node

    ok = remove_node(node_id)
    return jsonify({'ok': ok}), (200 if ok else 404)


@app.route('/api/enterprise/farm/dispatch-readiness', methods=['GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_dispatch_readiness():
    """调度就绪检查：前置齐备 ≠ 并行用例已通过。"""
    from ai_modules.enterprise.execution_farm import dispatch_readiness

    return jsonify(dispatch_readiness())


@app.route('/api/enterprise/farm/dispatch-hint', methods=['POST', 'GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_dispatch_hint():
    """输出 remote 环境建议；不自动改 .env。"""
    from ai_modules.enterprise.execution_farm import dispatch_hint

    data = request.get_json(silent=True) or {}
    node_id = _ai_str(data.get('node_id') or request.args.get('node_id'))
    capability = _ai_str(data.get('capability') or request.args.get('capability') or 'desktop')
    result = dispatch_hint(node_id=node_id, capability=capability)
    code = 200 if result.get('ok') else (404 if result.get('error_code') == 'NODE_NOT_FOUND' else 400)
    return jsonify(result), code


@app.route('/api/enterprise/farm/resolve-gateway', methods=['GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_resolve_gateway():
    """解析 Gateway URL（env 优先；DESKTOP_FARM_GATEWAY=1 时可用农场节点）。"""
    from ai_modules.enterprise.gateway_resolve import resolve_desktop_gateway

    return jsonify(resolve_desktop_gateway())


@app.route('/api/enterprise/gateway/live-probe', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_gateway_live_probe():
    """MCP/农场：探活 Gateway；可选 wait 一步。health 成功 ≠ 用例通过。"""
    from testory_mcp.gateway_live import mcp_live_demo

    data = request.get_json(silent=True) or {}
    try_step = bool(data.get('try_step') or request.args.get('try_step') in ('1', 'true', 'yes'))
    return jsonify(mcp_live_demo(try_step=try_step))


@app.route('/api/enterprise/farm/jobs', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_jobs():
    """农场任务队列：入队 ≠ 业务用例通过。"""
    from ai_modules.enterprise.farm_jobs import enqueue_job, jobs_summary, list_jobs

    if request.method == 'GET':
        try:
            limit = int(request.args.get('limit') or 50)
        except (TypeError, ValueError):
            limit = 50
        return jsonify({'ok': True, **jobs_summary(), 'jobs': list_jobs(limit=limit)})
    data = request.get_json(silent=True) or {}
    result = enqueue_job(
        job_type=_ai_str(data.get('job_type') or data.get('type')),
        node_id=_ai_str(data.get('node_id')),
        payload=data.get('payload') if isinstance(data.get('payload'), dict) else None,
        auto_run=bool(data.get('auto_run')),
    )
    code = 200 if result.get('ok') or result.get('job') else 400
    if result.get('error_code') in ('JOB_TYPE_REQUIRED', 'JOB_TYPE_UNSUPPORTED'):
        code = 400
    return jsonify(result), code


@app.route('/api/enterprise/farm/jobs/<job_id>/run', methods=['POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_job_run(job_id):
    from ai_modules.enterprise.farm_jobs import run_job

    result = run_job(job_id)
    code = 200
    if result.get('error_code') == 'JOB_NOT_FOUND':
        code = 404
    return jsonify(result), code


@app.route('/api/enterprise/farm/jobs/<job_id>', methods=['GET', 'DELETE'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_job_one(job_id):
    from ai_modules.enterprise.farm_jobs import cancel_job, get_job

    if request.method == 'GET':
        job = get_job(job_id)
        if not job:
            return jsonify({'ok': False, 'error': '任务不存在'}), 404
        return jsonify({'ok': True, 'job': job, 'case_pass_claimed': False})
    result = cancel_job(job_id)
    code = 200 if result.get('ok') else (404 if result.get('error_code') == 'JOB_NOT_FOUND' else 400)
    return jsonify(result), code


@app.route('/api/enterprise/farm/fanout-probe', methods=['POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_fanout_probe():
    """多节点 probe fan-out；all_nodes_reachable ≠ 并行用例通过。"""
    from ai_modules.enterprise.farm_batch import run_probe_fanout

    data = request.get_json(silent=True) or {}
    auto_run = data.get('auto_run')
    if auto_run is None:
        auto_run = True
    result = run_probe_fanout(auto_run=bool(auto_run))
    code = 400 if result.get('error_code') == 'NO_NODES' else 200
    return jsonify(result), code


@app.route('/api/enterprise/farm/jobs/drain', methods=['POST'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@feature_required('parallel_execution')
@api_error_handler
def api_enterprise_farm_jobs_drain():
    """消化 queued 农场作业（Worker 一批）；不宣称用例通过。"""
    from ai_modules.enterprise.farm_worker import drain_queued_jobs

    data = request.get_json(silent=True) or {}
    try:
        limit = int(data.get('limit') or request.args.get('limit') or 20)
    except (TypeError, ValueError):
        limit = 20
    return jsonify(drain_queued_jobs(limit=limit))


@app.route('/api/enterprise/sla-evidence', methods=['GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@api_error_handler
def api_enterprise_sla_evidence():
    """SLA 证据样本摘要；sla_claim 恒 false。"""
    from ai_modules.enterprise.sla_evidence import summarize_sla_evidence

    try:
        limit = int(request.args.get('limit') or 200)
    except (TypeError, ValueError):
        limit = 200
    return jsonify(summarize_sla_evidence(limit=limit))


@app.route('/api/enterprise/sla-alerts', methods=['GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@api_error_handler
def api_enterprise_sla_alerts():
    """SLA 阈值告警（运维提示）；sla_met 恒 false。"""
    from ai_modules.enterprise.sla_alerts import evaluate_sla_alerts

    try:
        limit = int(request.args.get('limit') or 200)
    except (TypeError, ValueError):
        limit = 200
    return jsonify(evaluate_sla_alerts(limit=limit))


@app.route('/api/enterprise/sla-alerts/webhook', methods=['POST'])
@login_required
@role_required('admin', 'project_manager')
@api_error_handler
def api_enterprise_sla_alerts_webhook():
    """可选推送 SLA 告警到 SLA_ALERT_WEBHOOK_URL；不构成达标判定。"""
    from ai_modules.enterprise.sla_webhook import maybe_post_sla_webhook

    data = request.get_json(silent=True) or {}
    force = bool(data.get('force') or request.args.get('force') in ('1', 'true', 'yes'))
    return jsonify(maybe_post_sla_webhook(force=force))


@app.route('/api/ai/agent-teams/sdk-runtime/probe', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_sdk_runtime_probe():
    """官方 SDK 运行时探测；未安装返回 SDK_NOT_INSTALLED。"""
    from ai_modules.agent_teams.sdk_runtime import try_official_sdk_runtime
    from ai_modules.agent_teams.test_run_state import load_run

    data = request.get_json(silent=True) or {}
    run_id = _ai_str(data.get('run_id') or request.args.get('run_id'))
    st = load_run(run_id) if run_id else None
    result = try_official_sdk_runtime(st)
    return jsonify(result), 200


@app.route('/api/enterprise/ops-readiness', methods=['GET'])
@login_required
@role_required('admin', 'project_manager', 'test_lead')
@api_error_handler
def api_enterprise_ops_readiness():
    """企业运营就绪清单（非 SLA 达标证明）。"""
    from ai_modules.enterprise.readiness import enterprise_ops_readiness

    return jsonify(enterprise_ops_readiness())


@app.route('/api/ai/agent-teams/runs/<run_id>/sdk-events', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_sdk_events(run_id):
    """将本地 TestRunState 映射为 SDK 风格事件（官方 SDK 可选）。"""
    from ai_modules.agent_teams.sdk_bridge import adapt_local_run_to_sdk_events
    from ai_modules.agent_teams.test_run_state import load_run

    st = load_run((run_id or '').strip())
    if not st:
        return jsonify({'ok': False, 'error': 'run 不存在'}), 404
    payload = adapt_local_run_to_sdk_events(st)
    return jsonify({'ok': True, **payload})


@app.route('/api/config-registry/info', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_config_registry_info():
    """R19：本地 Spec/Skill 配置注册中心（Nacos 叙事轻量替代）。"""
    from ai_modules.config_registry import registry_info, seed_from_builtin_team_spec

    if request.args.get('seed') in ('1', 'true', 'yes'):
        seed_from_builtin_team_spec()
    return jsonify({'ok': True, **registry_info()})


@app.route('/api/ai/incident-memory/search', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_incident_memory_search():
    """R15：IncidentMemory / Runbook 轻量检索（建议不判绿）。"""
    from ai_modules.memory.incident_memory import search_incidents, search_runbooks

    data = request.get_json(silent=True) or {}
    q = _ai_str(data.get('q') or data.get('query') or request.args.get('q') or request.args.get('query') or '')
    kind = _ai_str(data.get('kind') or request.args.get('kind') or '')
    try:
        limit = int(data.get('limit') or request.args.get('limit') or 5)
    except (TypeError, ValueError):
        limit = 5
    if kind == 'runbook':
        hits = search_runbooks(q, limit=limit)
    else:
        hits = search_incidents(q, limit=limit, kind=kind or None)
    return jsonify({
        'ok': True,
        'query': q,
        'kind': kind or 'all',
        'hits': hits,
        'count': len(hits),
        'disclaimer': '检索命中仅为排障建议，不得据此自动判绿',
    })


@app.route('/api/ai/agent-teams/runs/<run_id>/trace', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_ai_agent_teams_trace(run_id):
    """导出多 Agent 运行 Trace 证据包（JSON 元数据或 ZIP）。"""
    from ai_modules.execute.trace_pack import export_trace_pack
    from flask import send_file

    rid = (run_id or '').strip()
    if not rid:
        return jsonify({'ok': False, 'error': 'run_id 不能为空'}), 400
    fmt = (request.args.get('format') or 'json').strip().lower()
    want_zip = fmt in ('zip', 'download')
    exported = export_trace_pack(agent_run_id=rid, make_zip=want_zip)
    if want_zip:
        zp = exported.get('zip_path')
        if not zp or not os.path.isfile(zp):
            return jsonify({
                'ok': False,
                'error': exported.get('error') or 'ZIP 生成失败',
                'error_code': exported.get('error_code'),
                'manifest': exported.get('manifest'),
            }), 404 if exported.get('error_code') == 'TRACE_INCOMPLETE' else 500
        return send_file(
            zp,
            mimetype='application/zip',
            as_attachment=True,
            download_name=exported.get('download_name') or f'{rid}.zip',
        )
    code = 200 if exported.get('ok') else (404 if exported.get('error_code') == 'TRACE_INCOMPLETE' else 400)
    return jsonify(exported), code


@app.route('/api/ai/trace-packs/export', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_trace_pack_export():
    """通用 Trace 证据包导出：audit_id / run_history_id / agent_run_id 三选一。"""
    from ai_modules.execute.trace_pack import export_trace_pack
    from flask import send_file

    data = request.get_json(silent=True) or {}
    q = request.args
    audit_id = _ai_str(data.get('audit_id') or q.get('audit_id') or '')
    agent_run_id = _ai_str(data.get('agent_run_id') or q.get('agent_run_id') or '')
    rh_raw = data.get('run_history_id') if data.get('run_history_id') is not None else q.get('run_history_id')
    run_history_id = None
    if rh_raw is not None and str(rh_raw).strip() != '':
        try:
            run_history_id = int(rh_raw)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'run_history_id 必须是整数'}), 400
    fmt = _ai_str(data.get('format') or q.get('format') or 'json').lower()
    want_zip = fmt in ('zip', 'download')
    exported = export_trace_pack(
        audit_id=audit_id,
        run_history_id=run_history_id,
        agent_run_id=agent_run_id,
        make_zip=want_zip,
    )
    if want_zip:
        zp = exported.get('zip_path')
        if not zp or not os.path.isfile(zp):
            return jsonify({
                'ok': False,
                'error': exported.get('error') or 'ZIP 生成失败',
                'error_code': exported.get('error_code'),
                'manifest': exported.get('manifest'),
            }), 400
        return send_file(
            zp,
            mimetype='application/zip',
            as_attachment=True,
            download_name=exported.get('download_name') or 'trace_pack.zip',
        )
    code = 200 if exported.get('ok') else 400
    return jsonify(exported), code


# ----------------------------------------------------------------------
# 对话测试 API
# ----------------------------------------------------------------------


@app.route('/api/ai/dialog/test', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_dialog_test():
    """执行移动端对话测试。"""
    from ai_modules.execute.dialog_tester import DialogTester, DialogPersona

    data = request.get_json(silent=True) or {}
    persona_type = _ai_str(data.get('persona') or 'professional')
    messages = data.get('messages') if isinstance(data.get('messages'), list) else []
    platform = _ai_str(data.get('platform') or 'mobile')

    if not messages:
        return jsonify({'ok': False, 'error': 'messages 不能为空'}), 400

    tester = DialogTester(persona=DialogPersona(persona_type), platform=platform)
    try:
        result = tester.run(messages)
        return jsonify({'ok': True, 'result': result})
    except Exception as exc:
        uat_logger.exception('dialog test failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500


# ----------------------------------------------------------------------
# 自愈批量扫描与验证 API
# ----------------------------------------------------------------------


@app.route('/api/ai/heal/capabilities', methods=['GET'])
@login_required
@api_error_handler
def api_ai_heal_capabilities():
    """Self-heal Hub 能力矩阵（诚实：Desktop 无运行时自愈）。"""
    from ai_modules.optimize.self_heal import heal_capability_matrix, summarize_heal_claim

    layer = (request.args.get('layer') or '').strip() or None
    matrix = heal_capability_matrix()
    claim = summarize_heal_claim(layer=layer) if layer else None
    return jsonify({
        'ok': True,
        'matrix': matrix,
        'claim': claim,
        'marketing_claim_allowed': False,
    })


@app.route('/api/ai/heal/batch-scan', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_heal_batch_scan():
    """批量扫描项目用例，检测需要自愈的步骤。"""
    from ai_modules.optimize.self_heal import batch_scan_project

    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    if not project_id:
        return jsonify({'ok': False, 'error': 'project_id 不能为空'}), 400

    try:
        result = batch_scan_project(project_id)
        return jsonify({'ok': True, 'result': result})
    except Exception as exc:
        uat_logger.exception('batch scan failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/ai/heal/verify', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_heal_verify():
    """验证并应用自愈后的步骤。"""
    from ai_modules.optimize.heal_verifier import batch_verify_and_apply

    data = request.get_json(silent=True) or {}
    healed_steps = data.get('healed_steps') if isinstance(data.get('healed_steps'), list) else []
    if not healed_steps:
        return jsonify({'ok': False, 'error': 'healed_steps 不能为空'}), 400

    try:
        result = batch_verify_and_apply(healed_steps)
        return jsonify({'ok': True, 'result': result})
    except Exception as exc:
        uat_logger.exception('heal verify failed')
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/ai/heal/history', methods=['GET'])
@login_required
@api_error_handler
def api_ai_heal_history():
    """获取自愈审计历史。"""
    from mobile_engine.element_repo.element_repository import ElementRepository

    project_id = request.args.get('project_id', type=int)
    alias = request.args.get('alias')
    limit = request.args.get('limit', 50, type=int)

    if not project_id:
        return jsonify({'ok': False, 'error': 'project_id 不能为空'}), 400

    repo = ElementRepository()
    history = repo.get_heal_history(project_id, alias=alias, limit=limit)
    stats = repo.get_heal_stats(project_id)
    return jsonify({'ok': True, 'history': history, 'stats': stats})


@app.route('/api/ai/llm/readiness', methods=['GET'])
@login_required
@api_error_handler
def api_ai_llm_readiness():
    """混合 LLM 就绪检测（Ollama 优先，否则云端 profile）。"""
    from ai_llm_readiness import assess_llm_readiness

    return jsonify({'success': True, **assess_llm_readiness(local_ai_service=local_ai_service)})


@app.route('/api/ai/llm/wizard-dismiss', methods=['POST'])
@login_required
@api_error_handler
def api_ai_llm_wizard_dismiss():
    os.environ['AI_LLM_WIZARD_DISMISSED'] = '1'
    return jsonify({'success': True})


@app.route('/api/ai/skills/record-success', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_skills_record_success():
    """记录一次成功执行，可选自动导出 Skill。"""
    from hermes_skill_loop import record_execution_success

    data = request.get_json(silent=True) or {}
    plan = data.get('plan') if isinstance(data.get('plan'), dict) else {}
    if not plan.get('steps'):
        return jsonify({'success': False, 'error': 'plan.steps 不能为空'}), 400
    result = record_execution_success(
        plan,
        case_url=_ai_str(data.get('case_url') or plan.get('case_url')),
        instruction=_ai_str(data.get('instruction') or data.get('message')),
        outcome=_ai_str(data.get('outcome')) or 'ok',
    )
    return jsonify({'success': True, **result})


@app.route('/api/ai/diagnostics/failure-bundle', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_failure_diag_bundle():
    """
    在当前主 Playwright 页面采集失败诊断包，并可选生成结构化缺陷草稿（LLM）。
    Body: { "failed_step": {...}, "exception_message": "..." }
    """
    data = request.get_json(silent=True) or {}
    failed_step = data.get('failed_step') if isinstance(data.get('failed_step'), dict) else {}
    exception_message = (data.get('exception_message') or data.get('error') or '').strip()

    try:
        from playwright_automation import sync_automation_session_usable, sync_gather_failure_signals

        if not sync_automation_session_usable():
            return jsonify({'success': False, 'error': '主浏览器未就绪，无法采集页面信号'}), 503
        signals = sync_gather_failure_signals()
    except Exception as e:
        uat_logger.warning('failure-bundle signals: %s', e)
        return jsonify({'success': False, 'error': str(e)}), 500

    try:
        from execution_diag_bundle import (
            build_failure_bundle,
            classify_failure_with_llm,
            merge_bundle_and_draft,
        )

        bundle = build_failure_bundle(failed_step, exception_message, signals)
        draft, dwarns = classify_failure_with_llm(bundle, force=False)
        out = merge_bundle_and_draft(bundle, draft)
        out['warnings'] = dwarns
        return jsonify({'success': True, 'bundle': out})
    except Exception as e:
        uat_logger.exception('failure-bundle compose failed')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ai/cases/generate-and-save', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_CASE', 'case')
def api_ai_generate_case_and_save():
    """
    使用本地模型生成测试用例与步骤，并直接保存到项目。
    """
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    goal = (data.get('goal') or '').strip()
    project_name = (data.get('project_name') or '').strip()
    case_name_override = (data.get('case_name_override') or data.get('session_title') or '').strip()
    selected_model = (data.get('model') or '').strip() or _get_active_local_model()
    client_plan = data.get('plan')
    profile, legacy_model = _resolve_inference_profile(selected_model)

    if not project_id:
        return jsonify({'success': False, 'error': 'project_id不能为空'}), 400

    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403

    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    current_case_count = _db.get_project_case_count(project_id)
    if limits['max_cases_per_project'] != -1 and current_case_count >= limits['max_cases_per_project']:
        return jsonify({
            'success': False,
            'error': f'已达到项目用例数量限制（{limits["max_cases_per_project"]}个）。请升级至团队版或企业版以提升配额。',
            'limit_reached': True,
            'current_count': current_case_count,
            'limit': limits['max_cases_per_project'],
            'upgrade_url': '/license'
        }), 403
    if license_info.license_type == LicenseType.FREE.value:
        _db.increment_created_cases(current_user.id)

    if isinstance(client_plan, dict) and isinstance(client_plan.get('steps'), list) and client_plan.get('steps'):
        goal = goal or _ai_str(client_plan.get('case_name')) or 'AI用例'
        generated = {
            'case_name': client_plan.get('case_name') or goal,
            'case_url': client_plan.get('case_url') or '',
            'description': client_plan.get('description') or '',
            'precondition': client_plan.get('precondition') or '',
            'expected_result': client_plan.get('expected_result') or '',
            'steps': list(client_plan.get('steps') or []),
            'meta': client_plan.get('meta') or {},
        }
        if case_name_override:
            generated['case_name'] = case_name_override
    else:
        if not goal:
            return jsonify({'success': False, 'error': 'goal不能为空'}), 400
        mem_ctx = _ai_memory_context_block(current_user.id, goal, probe_url="", project_name=project_name)
        try:
            generated = local_ai_service.generate_case_and_steps(
                goal,
                project_name,
                model=legacy_model,
                profile=profile,
                memory_context=mem_ctx or None,
            )
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': str(e),
                'hint': '可先执行: ollama serve，并确认模型已拉取。'
            }), 503
        if case_name_override:
            generated['case_name'] = case_name_override
    local_ai_service._fill_missing_step_payloads(
        generated.get('steps') or [],
        goal or '',
        _ai_str(generated.get('case_url')),
        None,
    )
    generated, warnings = apply_step_normalization_to_plan(generated)
    from ai_page_probe import apply_ai_assert_grounding_to_plan

    generated, ground_warns = apply_ai_assert_grounding_to_plan(generated, warnings)
    warnings = ground_warns
    log_ai_plan_to_audit(
        current_user.id,
        current_user.username,
        'AI_PLAN_GENERATE',
        generated,
        request.remote_addr,
    )
    case_id = db.create_test_case_v2(
        project_id,
        generated.get('case_name', ''),
        generated.get('case_url', ''),
        generated.get('description', ''),
        generated.get('precondition', ''),
        generated.get('expected_result', ''),
    )

    created_steps = 0
    steps = generated.get('steps') or []
    for idx, step in enumerate(steps, start=1):
        db.create_test_step(**_ai_step_to_db_kwargs(step, case_id, idx))
        created_steps += 1

    return jsonify({
        'success': True,
        'case_id': case_id,
        'case_name': generated.get('case_name', ''),
        'steps_created': created_steps,
        'generated': generated,
        'warnings': warnings,
    })


@app.route('/api/ai/task/chat', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_task_chat():
    """
    多轮AI对话生成/优化测试用例步骤（本地模型）。

    可选 body 字段（与左栏/浮层/右键菜单对接，全平台保持同一套语义）：
    - focus_step_index: 用户聚焦的步骤序号（1-based，与 current_plan.steps 顺序一致）
    - focus_step_indices: 多选步骤，如合并步骤、批量优化
    - browser_selection_text / selection_text: 内嵌浏览器划词，用于「断言见划词内容」
    - action_kind / intent: 如 optimize_step、merge_steps、assert_from_selection
    """
    data = request.get_json(silent=True) or {}
    out = _execute_ai_task_chat(
        data,
        current_user.id,
        current_user.username,
        request.remote_addr,
    )
    code = int(out.get('_http', 200))
    body = {k: v for k, v in out.items() if k != '_http'}
    return jsonify(body), code


@app.route('/api/ai/studio/steps-assist', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_studio_steps_assist():
    """
    编排入口：若请求体未带 current_plan 但含 case_id，则从数据库装载步骤后再调用本地模型 chat；
    append_after_chat 为真且 chat 成功时，将返回的 plan.steps 追加到该用例末尾（权限与 /api/ai/cases/append-steps 一致）。
    """
    from assistant_gateway import merge_case_into_chat_payload

    data = request.get_json(silent=True) or {}
    err = merge_case_into_chat_payload(db, data)
    if err:
        return jsonify({
            'success': False,
            'error': err,
            'hint': '请提供有效的 case_id，或自行传入完整 current_plan（与步骤页 AI 浮层一致）。',
        }), 400
    out = _execute_ai_task_chat(
        data,
        current_user.id,
        current_user.username,
        request.remote_addr,
    )
    code = int(out.get('_http', 200))
    body = {k: v for k, v in out.items() if k != '_http'}
    if data.get('append_after_chat') and body.get('success'):
        plan = body.get('plan') or {}
        steps = plan.get('steps') or []
        cid = data.get('case_id')
        if isinstance(steps, list) and steps and cid is not None:
            try:
                case_id_int = int(cid)
            except (TypeError, ValueError):
                case_id_int = 0
            if case_id_int > 0:
                case = db.get_test_case_v2(case_id_int)
                if case:
                    _db = Database()
                    if case.get('project_id') and not _db.check_project_access(
                        current_user.id, case['project_id'], 'editor'
                    ):
                        body['append'] = {'success': False, 'error': '无权限修改此用例'}
                    else:
                        old_steps, _total = db.get_case_steps_paginated(case_id_int, 1, 1000)
                        max_order = 0
                        if old_steps:
                            max_order = max(int(s.get('step_order') or 0) for s in old_steps)
                        goal_hint = _ai_str(data.get('goal')) or _ai_str(case.get('name')) or _ai_str(
                            case.get('description')
                        )
                        local_ai_service._fill_missing_step_payloads(
                            steps,
                            goal_hint,
                            _ai_str(case.get('url')),
                            None,
                        )
                        clean_steps, append_warnings = dedupe_and_validate_ai_steps(steps)
                        created_steps = 0
                        for idx, step in enumerate(clean_steps, start=1):
                            db.create_test_step(
                                **_ai_step_to_db_kwargs(step, case_id_int, max_order + idx)
                            )
                            created_steps += 1
                        body['append'] = {
                            'success': True,
                            'steps_created': created_steps,
                            'warnings': append_warnings,
                        }
    return jsonify(body), code


def _agent_gateway_sse_line(obj: dict) -> str:
    return f"data: {json.dumps(obj, ensure_ascii=False)}\n\n"


@app.route('/api/ai/agent/gateway-stream', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_agent_gateway_stream():
    """
    对话式网关（MVP）：解析左栏自然语言中的「安全子集」意图，通过 text/event-stream 推送进度。
    若请求携带有效的 embedded_session_id 且内嵌网关已配置，则在 **远程画布同一会话** 的 Chromium 中逐步执行（与 WebSocket 画面一致）；
    否则在 HuFirst **后台 Playwright 自动化会话**中执行（与「打开」、/api/navigate、左侧 iframe 同步）。
    若预览首步不是 navigate 但请求携带了目标 URL（或 plan.case_url），网关会自动插入一步导航再执行。
    复杂规划仍请用「生成用例 / 优化」或 /api/ai/task/chat。
    """
    from agent_intent import parse_agent_intent

    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    plan = data.get('plan') if isinstance(data.get('plan'), dict) else {}
    steps = plan.get('steps') if isinstance(plan.get('steps'), list) else []
    has_steps = len(steps) > 0
    gate_url = (data.get('target_page_url') or '').strip()
    if not gate_url:
        gate_url = (plan.get('case_url') or '').strip()
    embedded_sid = (data.get('embedded_session_id') or data.get('remote_session_id') or '').strip()
    use_embedded = bool(embedded_sid and embedded_gateway_enabled())
    exec_uid = int(current_user.id)
    cdp_ws_client = (data.get('cdp_ws_url') or '').strip()

    if (not message) and bool(data.get('force_execute')) and has_steps:
        kind, meta = 'execute_plan', {}
    else:
        kind, meta = parse_agent_intent(message, has_steps)

    def generate():
        if use_embedded and cdp_ws_client:
            try:
                from hermes_config import sync_hermes_cdp_endpoint

                if sync_hermes_cdp_endpoint(cdp_ws_client, restart_gateway=True):
                    yield _agent_gateway_sse_line({"t": "log", "message": "已同步画布 CDP 到 Hermes。"})
            except Exception:
                pass
        yield _agent_gateway_sse_line({"t": "intent", "kind": kind, "meta": meta})
        if kind == "none":
            yield _agent_gateway_sse_line(
                {
                    "t": "error",
                    "message": "未识别为可执行命令。示例：「执行当前预览」「打开 https://example.com」「探索登录流程」",
                }
            )
            yield _agent_gateway_sse_line({"t": "end"})
            return

        if kind == "hermes_explore":
            from agent_gateway_client import agent_gateway_configured, get_agent_gateway_client
            from hermes_config import hermes_cdp_attached, sync_hermes_cdp_endpoint

            if not agent_gateway_configured():
                yield _agent_gateway_sse_line(
                    {"t": "error", "message": "Hermes Agent 未配置，无法执行自然语言探索。"}
                )
                yield _agent_gateway_sse_line({"t": "end"})
                return
            if use_embedded and not hermes_cdp_attached():
                j_cdp, err_cdp = embedded_gateway_json(
                    "GET",
                    f"/internal/session/{embedded_sid}",
                    user_id=exec_uid,
                )
                cdp_ws = ""
                if isinstance(j_cdp, dict):
                    cdp_ws = (j_cdp.get("cdp_browser_ws") or "").strip()
                if cdp_ws:
                    sync_hermes_cdp_endpoint(cdp_ws, restart_gateway=True)
                    yield _agent_gateway_sse_line({"t": "log", "message": "已同步画布 CDP 到 Hermes。"})
                elif not hermes_cdp_attached():
                    yield _agent_gateway_sse_line(
                        {
                            "t": "error",
                            "message": "画布 CDP 未就绪，无法在同一浏览器中探索。请先连接实时画面。",
                        }
                    )
                    yield _agent_gateway_sse_line({"t": "end"})
                    return
            yield _agent_gateway_sse_line({"t": "log", "message": "正在通过 Testory AI (Hermes) 执行探索…"})
            client = get_agent_gateway_client()
            result = ""
            try:
                from hermes_skill_hints import build_explore_instruction

                explore_instr = build_explore_instruction(
                    meta.get("message") or message,
                    {"platform": (data.get("platform") or meta.get("platform") or "auto")},
                )
            except Exception:
                explore_instr = meta.get("message") or message
            try:
                for ev_kind, ev_payload in client.execute_user_instruction_stream(explore_instr):
                    if ev_kind == "trace":
                        msg = (ev_payload or {}).get("message") or (ev_payload or {}).get("stage") or "…"
                        yield _agent_gateway_sse_line(
                            {"t": "hermes_trace", "message": str(msg)[:400], **(ev_payload or {})}
                        )
                    elif ev_kind == "delta":
                        piece = (ev_payload or {}).get("text") or ""
                        if piece:
                            yield _agent_gateway_sse_line({"t": "hermes_delta", "text": piece[:800]})
                    elif ev_kind == "error":
                        yield _agent_gateway_sse_line(
                            {"t": "error", "message": (ev_payload or {}).get("error") or "Hermes 失败"}
                        )
                        yield _agent_gateway_sse_line({"t": "end"})
                        return
                    elif ev_kind == "result":
                        result = (ev_payload or {}).get("content") or ""
            except Exception:
                result = client.execute_user_instruction(explore_instr)
            yield _agent_gateway_sse_line({"t": "hermes_result", "content": (result or "")[:48000]})
            try:
                from hermes_skill_loop import record_execution_success

                loop_out = record_execution_success(
                    plan if isinstance(plan, dict) else {},
                    case_url=gate_url,
                    instruction=message,
                    outcome="hermes_explore",
                )
                if loop_out.get("auto_exported"):
                    sk = loop_out.get("skill") or {}
                    yield _agent_gateway_sse_line(
                        {
                            "t": "skill_learned",
                            "message": f"已自动保存为 Skill: {sk.get('id') or ''}",
                            "skill": sk,
                        }
                    )
            except Exception:
                pass
            yield _agent_gateway_sse_line({"t": "end"})
            return

        embed_exec = use_embedded and kind == "execute_plan"
        embed_nav = use_embedded and kind == "navigate_url"
        need_main_playwright = not (embed_exec or embed_nav)
        if need_main_playwright and embedded_gateway_enabled() and not _ai_allow_main_playwright_fallback():
            yield _agent_gateway_sse_line(
                {
                    "t": "error",
                    "message": (
                        "未携带有效的 embedded_session_id，且已配置内置画布网关，"
                        "不会启动主 Playwright。请先在 AI 测试页连接画布后再执行。"
                    ),
                }
            )
            yield _agent_gateway_sse_line({"t": "end"})
            return

        if need_main_playwright:
            try:
                headless = resolve_playwright_headless(True)
                sync_start_browser(headless=headless)
            except Exception as e:
                yield _agent_gateway_sse_line({"t": "error", "message": f"启动浏览器失败: {e}"})
                yield _agent_gateway_sse_line({"t": "end"})
                return
            if not sync_automation_session_usable():
                yield _agent_gateway_sse_line(
                    {
                        "t": "error",
                        "message": (
                            "后台 **Playwright 自动化会话** 未就绪。请先在 AI 测试页点击「打开」同步页面，"
                            "或在部署环境完成 Playwright 内核安装（如：python -m playwright install）。"
                            "（未连接远程画布时，识别并执行依赖此会话。）"
                        ),
                    }
                )
                yield _agent_gateway_sse_line({"t": "end"})
                return

        if embed_exec:
            yield _agent_gateway_sse_line(
                {
                    "t": "log",
                    "message": "已在远程画布会话中执行步骤（与当前画布 WebSocket 为同一 Chromium）。",
                }
            )

        if kind == "navigate_url":
            url = (meta.get("url") or "").strip()
            if not url.startswith(("http://", "https://")):
                yield _agent_gateway_sse_line({"t": "error", "message": "仅支持 http(s) URL"})
                yield _agent_gateway_sse_line({"t": "end"})
                return
            yield _agent_gateway_sse_line({"t": "log", "message": f"导航 → {url}"})
            try:
                if embed_nav:
                    j, err = embedded_gateway_json(
                        "POST",
                        f"/internal/session/{embedded_sid}/run-steps",
                        user_id=exec_uid,
                        body={"steps": [{"action": "navigate", "url": url, "description": "gateway navigate"}]},
                        timeout_sec=120.0,
                    )
                    if err or not isinstance(j, dict) or not j.get("success"):
                        yield _agent_gateway_sse_line(
                            {"t": "error", "message": str(err or (j or {}).get("detail") or "远程导航失败")}
                        )
                    else:
                        rs = (j.get("results") or [{}])[0]
                        if not rs.get("ok"):
                            yield _agent_gateway_sse_line(
                                {"t": "error", "message": str(rs.get("error") or "远程导航失败")}
                            )
                        else:
                            yield _agent_gateway_sse_line({"t": "done", "message": "导航完成", "url": url})
                else:
                    sync_navigate_to(url, ai_probe=True)
                    yield _agent_gateway_sse_line({"t": "done", "message": "导航完成", "url": url})
            except Exception as e:
                yield _agent_gateway_sse_line({"t": "error", "message": str(e)})
            yield _agent_gateway_sse_line({"t": "end"})
            return

        # execute_plan
        script_steps = ai_plan_steps_to_playwright_script_steps(steps)
        if not script_steps:
            msg = '当前预览没有可映射为脚本的原子步骤'
            if has_steps:
                msg += '（常见原因：click/input/verify 缺少 selector，或 navigate 缺少 URL）。请检查左栏预览或重新生成。'
            else:
                msg += '（预览步骤为空）。'
            yield _agent_gateway_sse_line({"t": "error", "message": msg})
            yield _agent_gateway_sse_line({"t": "end"})
            return
        if (
            gate_url.startswith(('http://', 'https://'))
            and script_steps
            and script_steps[0].get('action') != 'navigate'
        ):
            script_steps = [
                {'action': 'navigate', 'url': gate_url, 'description': '（网关）先打开目标页'},
            ] + script_steps
            yield _agent_gateway_sse_line(
                {'t': 'log', 'message': f'预览首步非 navigate，已根据目标 URL 插入导航：{gate_url}'}
            )
        max_n = int(os.environ.get('AI_AGENT_GATEWAY_MAX_STEPS', '40') or 40)
        script_steps = script_steps[: max(1, max_n)]
        total = len(script_steps)
        yield _agent_gateway_sse_line({"t": "log", "message": f"共 {total} 步，开始逐步执行…"})
        all_ok = True
        for i, st in enumerate(script_steps, start=1):
            yield _agent_gateway_sse_line(
                {"t": "step_begin", "index": i, "total": total, "action": st.get("action"), "step": st}
            )
            try:
                if embed_exec:
                    j, err = embedded_gateway_json(
                        "POST",
                        f"/internal/session/{embedded_sid}/run-steps",
                        user_id=exec_uid,
                        body={"steps": [st]},
                        timeout_sec=180.0,
                    )
                    if err or not isinstance(j, dict) or not j.get("success"):
                        raise RuntimeError(str(err or (j or {}).get("detail") or "远程执行请求失败"))
                    rs0 = (j.get("results") or [{}])[0]
                    if not rs0.get("ok"):
                        raise RuntimeError(str(rs0.get("error") or "远程步骤失败"))
                else:
                    sync_execute_script_steps([st])
                yield _agent_gateway_sse_line({"t": "step_ok", "index": i, "total": total})
            except Exception as e:
                yield _agent_gateway_sse_line({"t": "step_err", "index": i, "total": total, "error": str(e)})
                all_ok = False
                break
        if all_ok and isinstance(plan, dict) and plan.get('steps'):
            try:
                from hermes_skill_loop import record_execution_success

                loop_out = record_execution_success(
                    plan,
                    case_url=gate_url,
                    instruction=message,
                    outcome='execute_plan_ok',
                )
                if loop_out.get('auto_exported'):
                    sk = loop_out.get('skill') or {}
                    yield _agent_gateway_sse_line(
                        {
                            't': 'skill_learned',
                            'message': f"已自动保存为 Skill: {sk.get('id') or ''}",
                            'skill': sk,
                            'success_count': loop_out.get('success_count'),
                        }
                    )
                elif loop_out.get('suggest_export'):
                    yield _agent_gateway_sse_line(
                        {
                            't': 'skill_suggest',
                            'message': '执行成功。可将此流程保存为 Hermes Skill 供下次复用。',
                            'success_count': loop_out.get('success_count'),
                        }
                    )
            except Exception:
                pass
        yield _agent_gateway_sse_line({"t": "done", "message": "执行序列结束"})
        yield _agent_gateway_sse_line({"t": "end"})

    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers=headers,
    )


@app.route('/api/ai/task/plan-async', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_task_plan_async():
    """提交规划任务，立即返回 job_id；推理在后台线程继续（见 GET /api/ai/task/job/<id>）。"""
    data = request.get_json(silent=True) or {}
    goal = (data.get('goal') or '').strip()
    if not goal:
        return jsonify({'success': False, 'error': 'goal不能为空'}), 400
    task_type = (data.get('task_type') or 'test_case_generation').strip()
    route = _route_ai_model(task_type)
    if route['provider'] != 'local':
        return jsonify({
            'success': False,
            'error': '该任务需走云端分析接口，请调用 /api/ai/task/cloud-analyze',
        }), 400
    store = get_job_store()
    store.prune()
    job_id = store.create(current_user.id, 'plan')
    _start_ai_bg_job_thread(
        job_id,
        'plan',
        data,
        current_user.id,
        current_user.username,
        request.remote_addr,
    )
    return jsonify({'success': True, 'job_id': job_id}), 202


@app.route('/api/ai/task/chat-async', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_ai_task_chat_async():
    """提交优化任务，立即返回 job_id；推理在后台线程继续。"""
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'success': False, 'error': 'message不能为空'}), 400
    route = _route_ai_model('test_case_generation')
    if route['provider'] != 'local':
        return jsonify({'success': False, 'error': '当前仅支持本地模型对话'}), 400
    store = get_job_store()
    store.prune()
    job_id = store.create(current_user.id, 'chat')
    _start_ai_bg_job_thread(
        job_id,
        'chat',
        data,
        current_user.id,
        current_user.username,
        request.remote_addr,
    )
    return jsonify({'success': True, 'job_id': job_id}), 202


@app.route('/api/ai/task/chat-stream', methods=['POST'])
@login_required
def api_ai_task_chat_stream():
    """SSE 流式 AI 对话：实时推送 LLM 推理过程和 tool calling 进度。"""
    import json as _json
    from flask import Response, stream_with_context

    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'success': False, 'error': 'message不能为空'}), 400

    # 解析参数（复用 chat-async 的参数结构）
    project_name = (data.get('project_name') or '').strip()
    current_plan = data.get('current_plan') or {}
    history = data.get('history') or []
    selected_model = (data.get('model') or '').strip()
    embedded_sid = (data.get('embedded_session_id') or '').strip()
    platform_type = (data.get('platform_type') or 'web').strip()
    interaction_context = data.get('interaction_context')

    # gateway-stream 同样走统一路由，避免桌面话术却挂网页工具
    try:
        from agent_intent import resolve_task_route

        _gs_route = resolve_task_route(message, ui_platform=platform_type)
        if _gs_route.needs_automation and _gs_route.platform in ("web", "desktop", "android"):
            platform_type = _gs_route.platform
    except Exception:
        pass

    abort_event = threading.Event()

    def generate():
        try:
            from ai_local_inference import local_ai_service
            from ai_chat_tool_loop import ChatToolLoopParams, run_ai_chat_with_tools_stream, ai_chat_tools_enabled, profile_supports_ai_chat_tools
            from ai_page_probe import probe_registry_from_interactive_snapshot

            profile, legacy_model = _resolve_inference_profile(selected_model)

            # 构建 probe registry
            probe_reg = None
            probe_url = (data.get('target_page_url') or '').strip()
            if probe_url:
                try:
                    probe_reg = probe_registry_from_interactive_snapshot(probe_url)
                except Exception:
                    pass

            _ctp = ChatToolLoopParams(
                message=message,
                project_name=project_name,
                current_plan=current_plan if isinstance(current_plan, dict) else {},
                history=history if isinstance(history, list) else [],
                legacy_model=legacy_model,
                profile=profile,
                page_snapshot=data.get('page_snapshot') or '',
                probe_registry=probe_reg,
                probe_url=probe_url or None,
                memory_context=data.get('memory_context') or '',
                dom_context_pack=data.get('dom_context_pack') or '',
                interaction_context=interaction_context,
                test_scope=data.get('test_scope') or '',
                embedded_session_id=embedded_sid or None,
                platform_type=platform_type,
            )

            for evt_type, evt_data in run_ai_chat_with_tools_stream(
                local_ai_service=local_ai_service,
                params=_ctp,
                abort_event=abort_event,
            ):
                sse_data = _json.dumps({"type": evt_type, "data": evt_data}, ensure_ascii=False, default=str)
                yield f"data: {sse_data}\n\n"

        except Exception as e:
            sse_data = _json.dumps({"type": "error", "data": str(e)}, ensure_ascii=False)
            yield f"data: {sse_data}\n\n"

        yield "data: {\"type\": \"end\", \"data\": null}\n\n"

    headers = {
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
        'Connection': 'keep-alive',
    }
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers=headers,
    )


@app.route('/api/ai/task/job/<job_id>', methods=['GET'])
@login_required
@api_error_handler
def api_ai_task_job_status(job_id):
    """查询后台 AI 任务状态；完成后 result 与同步接口 JSON 一致。"""
    rec = get_job_store().get(job_id)
    if not rec or rec.get('user_id') != current_user.id:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    st = rec.get('status', 'running')
    if st == 'running':
        return jsonify({'success': True, 'status': 'running'})
    if st == 'cancelled':
        return jsonify({'success': True, 'status': 'cancelled'})
    if st == 'error':
        return jsonify({
            'success': True,
            'status': 'error',
            'error': rec.get('error'),
            'result': rec.get('result'),
        })
    return jsonify({'success': True, 'status': 'done', 'result': rec.get('result')})


@app.route('/api/ai/task/job/<job_id>/cancel', methods=['POST'])
@login_required
@api_error_handler
def api_ai_task_job_cancel(job_id):
    """标记取消：中断后台线程的重试循环并设置取消标志。"""
    store = get_job_store()
    rec = store.get(job_id)
    if not rec or rec.get('user_id') != current_user.id:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    abort_evt = _AI_JOB_ABORT_EVENTS.get(job_id)
    if abort_evt:
        abort_evt.set()
    store.set_cancelled(job_id)
    return jsonify({'success': True})


@app.route('/api/ai/task/execute', methods=['POST'])
@login_required
@api_error_handler
def api_ai_task_execute():
    """AI 任务执行（SSE 流式输出，复用 ai_chat_tool_loop 多轮工具循环）"""
    data = request.get_json(silent=True) or {}
    task = data.get('task', '')
    project_id = data.get('project_id')
    platform = (data.get('platform') or 'auto').strip().lower() or 'auto'
    engine = (data.get('engine') or '').strip()  # profile_id from UI
    # 起始 URL 一律从用户任务原文解析（前端已无独立 URL 输入框）
    try:
        from agent_intent import extract_task_url

        url = extract_task_url(task) or ''
    except Exception:
        url = ''
    if not url:
        # 兼容 API 调用方仍传 url/case_url
        explicit = (data.get('url') or data.get('case_url') or '').strip()
        if explicit:
            try:
                from agent_intent import extract_task_url as _etu

                url = _etu(explicit) or (
                    explicit if explicit.lower().startswith(('http://', 'https://')) else ''
                )
            except Exception:
                url = explicit if explicit.lower().startswith(('http://', 'https://')) else ''
    enable_vision = data.get('enable_vision', False)
    timeout = data.get('timeout', 120)
    # 默认关闭：日常操控完成后立刻结束，避免卡在「优化测试用例」直到超时
    generate_case_after_run = bool(
        data.get("generate_case_after_run")
        or data.get("refine_after_run")
        or data.get("generate_test_case")
    )

    if not task:
        return jsonify({'success': False, 'error': '请输入任务描述'}), 400

    # 统一路由：chat / web / desktop / android（消息信号可纠正错误的 UI 平台）
    try:
        from agent_intent import resolve_task_route

        _route = resolve_task_route(task, ui_platform=platform)
        if _route.platform in ("web", "desktop", "android") and _route.platform != platform:
            uat_logger.info(
                "task route override platform %s -> %s reason=%s task=%s",
                platform,
                _route.platform,
                _route.reason,
                (task or "")[:80],
            )
        if _route.needs_automation and _route.platform in ("web", "desktop", "android"):
            platform = _route.platform
    except Exception as _route_ex:
        uat_logger.debug("resolve_task_route failed: %s", _route_ex)
        _route = None

    # 执行前：UI 所选引擎必须成为本次（及全局）生效模型，避免「下拉显示 A、后台仍是 B」
    # 桌面外层 windows_* 不依赖 Hermes 进程热加载；跳过可能长达 ~25s 的 restart，避免 SSE 前空白
    _desk_skip_hermes_restart = bool(
        platform == "desktop"
        or (
            _route is not None
            and (
                getattr(_route, "needs_desktop_tools", False)
                or getattr(_route, "platform", "") == "desktop"
            )
        )
    )
    if engine:
        try:
            from ai_multi_provider import get_llm_profile_by_id, set_active_llm_profile_id

            if get_llm_profile_by_id(engine):
                set_active_llm_profile_id(engine)
                if not _desk_skip_hermes_restart:
                    try:
                        from hermes_service_bootstrap import ensure_hermes_llm_current

                        ensure_hermes_llm_current(restart_if_stale=True)
                    except Exception:
                        try:
                            from hermes_config import sync_platform_llm_credentials_to_hermes_env

                            sync_platform_llm_credentials_to_hermes_env()
                        except Exception:
                            pass
                else:
                    # 桌面路径：仅轻量同步凭证到 env，绝不在进流前重启 Gateway
                    try:
                        from hermes_config import sync_platform_llm_credentials_to_hermes_env

                        sync_platform_llm_credentials_to_hermes_env()
                    except Exception:
                        pass
        except Exception as _eng_ex:
            uat_logger.debug("apply engine profile failed: %s", _eng_ex)

    # 必须在进入生成器前捕获：无 stream_with_context 时 yield 后 current_user 会变成 None
    # → AttributeError: 'NoneType' object has no attribute 'id'
    try:
        _exec_user_id = str(current_user.id)
    except Exception:
        _exec_user_id = ""
    if not _exec_user_id:
        return jsonify({'success': False, 'error': '用户未登录'}), 401

    # 设置超时
    old_timeout = os.environ.get("HERMES_GATEWAY_TIMEOUT", "")
    os.environ["HERMES_GATEWAY_TIMEOUT"] = str(max(30, int(timeout or 120)))

    import uuid as _uuid
    job_id = _uuid.uuid4().hex

    def _friendly_sse_error(err: object) -> str:
        try:
            from hermes_gateway_client import _friendly_corrupt_msg, _is_corrupt_session_error

            s = err if isinstance(err, str) else str(err)
            if _is_corrupt_session_error(s) or (
                "NoneType" in s and "attribute 'id'" in s
            ):
                return _friendly_corrupt_msg(s)
            return s[:400]
        except Exception:
            return str(err)[:400]

    def _gen():
        import json as _json
        import threading
        import time as _time

        def send(event_type, **kwargs):
            kwargs['type'] = event_type
            return 'data: ' + _json.dumps(kwargs, ensure_ascii=False) + '\n\n'

        abort_event = threading.Event()
        user_id = _exec_user_id
        task_timeout_sec = max(30, int(timeout or 120))
        deadline_ts = _time.time() + float(task_timeout_sec)

        def _timeout_watchdog():
            # 左侧「超时时间」生效：到期后置位 abort，打断工具循环
            remain = max(0.5, deadline_ts - _time.time())
            if abort_event.wait(timeout=remain):
                return
            # 标记为超时（而非用户取消），便于错误文案
            setattr(abort_event, "_timed_out", True)
            setattr(abort_event, "_abort_reason", "timeout")
            abort_event.set()

        threading.Thread(target=_timeout_watchdog, daemon=True, name="ai-task-timeout").start()

        with _ai_task_abort_lock:
            _ai_task_abort_events[job_id] = abort_event

        yield send('job_started', job_id=job_id, timeout_sec=task_timeout_sec)

        try:
            # 立刻推送，避免用户感觉「点了没反应」
            yield send('think', text='已收到任务，正在解析意图…', status='running')

            from agent_intent import message_needs_automation, resolve_task_route

            # 勿在本生成器内给闭包名 platform 赋值，否则整函数变成局部变量 → UnboundLocalError
            run_platform = (platform or "auto").strip().lower() or "auto"
            task_route = None
            needs_automation = message_needs_automation(task)
            try:
                task_route = resolve_task_route(task, ui_platform=run_platform)
                needs_automation = bool(task_route.needs_automation)
                if task_route.needs_automation and task_route.platform in (
                    "web",
                    "desktop",
                    "android",
                ):
                    run_platform = task_route.platform
            except Exception:
                task_route = None
            yield send(
                'think',
                text='正在理解您的意图…' if not needs_automation else '正在检查执行环境…',
                status='running',
            )
            if task_route is not None and needs_automation:
                yield send(
                    'think',
                    text=(
                        f"任务路由：{task_route.platform}"
                        f"（UI={task_route.ui_platform}，{task_route.reason}）"
                    ),
                    status='done',
                )

            # 不自动启动：仅检测用户是否已手动点「启动智能体」
            from hermes_gateway_client import HermesGatewayClient
            hermes_client = HermesGatewayClient()
            # 桌面外层工具不依赖 Hermes；缩短探测，避免再卡 1.5s+
            _needs_mobile_await = bool(
                task_route is not None and getattr(task_route, "needs_mobile_await", False)
            )
            _cross_end_cap = bool(
                task_route is not None
                and str(getattr(task_route, "reason", "") or "").startswith("cross_end")
            )
            _desk_only_early = run_platform == "desktop" or (
                task_route is not None and bool(getattr(task_route, "needs_desktop_tools", False))
            )
            _outer_tools_ok_early = _desk_only_early or _needs_mobile_await or _cross_end_cap
            _hermes_probe_sec = 0.4 if _outer_tools_ok_early else 1.5
            hermes_available = hermes_client.is_configured() and hermes_client.health_check(
                timeout_sec=_hermes_probe_sec
            )

            if abort_event.is_set():
                yield send('error', error='任务已取消')
                return

            # 明确要做真实操作时才要求智能体已启动；闲聊/问答可走平台 LLM
            # 桌面外层 windows_* / 跨端 mobile_* 不依赖 Hermes Gateway
            _desk_only = _desk_only_early or (
                task_route is not None and bool(getattr(task_route, "needs_desktop_tools", False))
            )
            _outer_tools_ok = _desk_only or _needs_mobile_await or _cross_end_cap
            if needs_automation and not hermes_available and not _outer_tools_ok:
                yield send('think', text='智能体未启动。请先在左上角点击「启动」后再执行自动化任务。', status='warning')
                yield send('error', error='请先启动智能体后再执行操作类任务')
                return

            if hermes_available:
                yield send('think', text='智能体已就绪', status='done')
            elif _outer_tools_ok and needs_automation:
                if _needs_mobile_await or _cross_end_cap:
                    yield send('think', text='跨端工具就绪（桌面 + 手机 await，无需 Hermes）', status='done')
                else:
                    yield send('think', text='桌面工具就绪（无需 Hermes）', status='done')
            else:
                yield send('think', text='以对话模式回复（未启动自动化）', status='done')

            # 闲聊/问答：直接走平台推理模型，不进工具循环、不碰 Hermes 会话
            if not needs_automation:
                yield send('think', text='正在思考回复…', status='running')
                try:
                    from ai_multi_provider import get_active_llm_profile, dispatch_chat
                    from ai_local_inference import local_ai_service

                    profile = get_active_llm_profile()
                    if not profile:
                        yield send('error', error='请先在左侧选择推理引擎')
                        return
                    chat_prompt = (
                        "你是 Testory 平台的 AI 测试助手。用简洁中文回答用户问题。"
                        "若用户只是闲聊或询问身份/能力，直接回答，不要输出测试用例 JSON，"
                        "不要假装去操作浏览器或桌面。"
                        "注意：你当前处于「纯对话」模式（未进入自动化工具循环）。"
                        "若用户其实想操作微信/桌面/网页，请明确告诉他："
                        "请再发一次任务指令（或确认左栏为桌面/对应平台），不要声称「平台永远无法操作桌面应用」。\n\n"
                        f"用户：{task}"
                    )
                    reply = (dispatch_chat(chat_prompt, profile, local_ai_service) or "").strip()
                    if not reply:
                        yield send('error', error='模型未返回内容，请检查推理引擎配置')
                        return
                    yield send('reply', text=reply)
                    yield send('done', message='已回复')
                except Exception as e:
                    yield send('error', error=_friendly_sse_error(f'对话失败: {e}'))
                return

            # 闲聊已在上方 return；此处为自动化任务 — Hermes 单脑编排（默认关闭关键词桌面早退）
            import os as _os_fp
            _fastpath = _os_fp.environ.get("DESKTOP_NL_FASTPATH", "0").strip().lower() in (
                "1", "true", "yes", "on",
            )
            if _fastpath:
                from agent_intent import message_needs_browser
                from agent_desktop_fastpath import is_desktop_nl_task, execute_desktop_nl

                if is_desktop_nl_task(task) or (needs_automation and not message_needs_browser(task)):
                    yield send('think', text='（调试）桌面快路径 DESKTOP_NL_FASTPATH=1 …', status='running')
                    desk = execute_desktop_nl(task)
                    if desk.get("ok"):
                        display = desk.get("display") or "桌面应用"
                        step_results = desk.get("step_results") or []
                        if step_results:
                            for item in step_results:
                                st = item.get("step") if isinstance(item.get("step"), dict) else {}
                                act = (st.get("action") or "launch_app").strip()
                                if act in ("wait",):
                                    continue
                                yield send(
                                    'action',
                                    action_type=act,
                                    target=st.get("target") or st.get("description") or display,
                                    status='success' if item.get("ok") else 'failed',
                                )
                        else:
                            overall_ok = bool(desk.get("ok")) and not desk.get("partial")
                            for st in desk.get("steps") or []:
                                yield send(
                                    'action',
                                    action_type=st.get("action") or "launch_app",
                                    target=st.get("target") or display,
                                    status='success' if overall_ok else 'failed',
                                )
                        plan = {
                            "case_name": (task[:40] or f"桌面操作-{display}"),
                            "case_url": "",
                            "description": task,
                            "platform": "desktop",
                            "steps": desk.get("steps") or [],
                            "meta": {
                                "source": "desktop_nl",
                                "via": desk.get("via"),
                                "partial": bool(desk.get("partial")),
                            },
                        }
                        reply = desk.get("reply") or f"桌面任务已处理（{display}）。"
                        yield send('reply', text=reply)
                        if desk.get("partial"):
                            yield send('done', message='部分完成', plan=plan, partial=True)
                        else:
                            yield send('done', message='完成', plan=plan)
                    else:
                        # 部分步骤可能已执行：按 step_results 如实展示
                        for item in desk.get("step_results") or []:
                            st = item.get("step") if isinstance(item.get("step"), dict) else {}
                            act = (st.get("action") or "desktop").strip()
                            if act in ("wait",):
                                continue
                            yield send(
                                'action',
                                action_type=act,
                                target=st.get("description") or st.get("target") or act,
                                status='success' if item.get("ok") else 'failed',
                            )
                        yield send('error', error=desk.get("error") or desk.get("reply") or "桌面操作失败")
                    return

            # 预检 + 能力注册表 + 任务上下文总线
            from agent_capability_registry import preflight_for_task, snapshot_capabilities
            from agent_task_context import new_task_context

            # 桌面外层工具：不强制 Hermes；避免重复 health 拖慢首包
            _pf_need_hermes = not _outer_tools_ok
            pf_ok, pf_msg, pf_snap = preflight_for_task(
                task, require_hermes=_pf_need_hermes
            )
            if not pf_ok:
                yield send('error', error=pf_msg or "预检失败")
                return
            if pf_msg:
                yield send('think', text=f'预检提示：{pf_msg}', status='warning')

            caps = (pf_snap or snapshot_capabilities()).get("capabilities") or {}
            caps_summary = "; ".join(
                f"{k}={'on' if (caps.get(k) or {}).get('available') else 'off'}"
                for k in ("hermes", "web", "desktop", "mobile", "api")
            )
            yield send('think', text=f'能力：{caps_summary}', status='done')

            task_ctx = new_task_context(
                active_surface=(
                    "web"
                    if (url or run_platform == "web")
                    else (run_platform or "auto")
                ),
                mobile_udid=((caps.get("mobile") or {}).get("udid") or ""),
                meta={"task": task[:200], "user_id": user_id, "start_url": url or ""},
            )
            yield send(
                'session',
                session_id=task_ctx.session_id,
                context=task_ctx.to_public_dict(),
            )

            browser_ready_holder = {"ready": False}

            def _ensure_browser_before_agent():
                """Web 需要时拉起浏览器；起始 URL 从任务原文解析，禁止依赖已移除的 URL 输入框。"""
                from agent_intent import extract_task_url, message_needs_browser

                needs_br = message_needs_browser(task) or (run_platform == "web")
                if not needs_br:
                    return True, ""
                nav_url = (extract_task_url(task) or url or "").strip()
                if browser_ready_holder["ready"]:
                    from hermes_config import hermes_cdp_attached
                    if hermes_cdp_attached():
                        if nav_url:
                            try:
                                from ai_external_browser_bridge import ensure_browser

                                ensure_browser(headless=False, url=nav_url, browser="edge")
                            except Exception:
                                pass
                        return True, ""
                if not hermes_available:
                    return False, "智能体未启动，无法执行浏览器自动化"
                try:
                    from ai_external_browser_bridge import ensure_browser, is_browser_alive, force_cleanup_browser
                    from hermes_config import hermes_cdp_attached
                    if hermes_cdp_attached() and not is_browser_alive():
                        force_cleanup_browser()
                    if not hermes_cdp_attached() or not is_browser_alive():
                        ok = ensure_browser(headless=False, url=nav_url or "", browser="edge")
                        if not ok:
                            if run_platform == "auto":
                                return True, ""  # 允许 Hermes 改走桌面/其它手
                            return False, "本机浏览器启动失败，请确认已安装 Edge/Chrome"
                    elif nav_url:
                        # 已有会话时仍导航到任务目标 URL，避免卡在 about:blank
                        try:
                            ensure_browser(headless=False, url=nav_url, browser="edge")
                        except Exception:
                            pass
                    browser_ready_holder["ready"] = True
                    if not hermes_cdp_attached():
                        if run_platform == "auto":
                            return True, ""
                        return False, "浏览器 CDP 未连接"
                    return True, ""
                except Exception as ex:
                    if run_platform == "auto":
                        return True, ""
                    return False, f"浏览器桥接失败: {str(ex)[:160]}"

            _time.sleep(0.05)

            if abort_event.is_set():
                yield send('error', error='任务已取消')
                return

            # Hermes 上游鉴权不兼容时：若聊天工具循环可用，仍走 windows_* FC；
            # 仅在工具循环关闭时才回退到本机桌面确定性路径（不再对微信任务抢跑热键方案）。
            try:
                from hermes_config import (
                    hermes_upstream_llm_status,
                    sync_platform_llm_credentials_to_hermes_env,
                )
                from agent_desktop_fastpath import is_desktop_nl_task, execute_desktop_nl
                from ai_chat_tool_loop import ai_chat_tools_enabled

                # 桌面/跨端外层路径已不依赖 Hermes 子进程模型；跳过二次 sync
                if not _outer_tools_ok:
                    sync_platform_llm_credentials_to_hermes_env()
                llm_st = hermes_upstream_llm_status() if not _outer_tools_ok else {"ok": True}
                # MiMo 已证实可用 Bearer，不再因 active_is_xiaomi 判定 Hermes 不可用
                hermes_llm_bad = not bool(llm_st.get("ok"))
                # 微信/桌面 GUI：优先外层 windows_* 工具循环；禁止因「是微信」就绕过 Agent
                use_desktop_first = (
                    is_desktop_nl_task(task)
                    and hermes_llm_bad
                    and not ai_chat_tools_enabled()
                )
                if use_desktop_first:
                    tip = (
                        llm_st.get("message")
                        or "当前引擎与 Hermes 鉴权不兼容且未启用聊天工具，改走平台本机桌面执行…"
                    )
                    yield send('think', text=tip, status='warning')
                    desk = execute_desktop_nl(task)
                    # 仅上报真实执行过的步骤；失败步骤标 failed，未执行的不展示为成功
                    step_results = desk.get("step_results") or []
                    if step_results:
                        for item in step_results:
                            if not isinstance(item, dict):
                                continue
                            st = item.get("step") if isinstance(item.get("step"), dict) else {}
                            act = (st.get("action") or item.get("action") or "desktop").strip()
                            if act in ("wait",):
                                continue
                            tgt = (
                                st.get("description")
                                or st.get("target")
                                or st.get("input_value")
                                or act
                            )
                            st_ok = bool(item.get("ok"))
                            yield send(
                                'action',
                                action_type=act,
                                target=str(tgt)[:120],
                                status='success' if st_ok else 'failed',
                            )
                            yield send(
                                'action_record',
                                action_type=act,
                                target=str(tgt)[:120],
                                status='success' if st_ok else 'failed',
                                has_vision=False,
                            )
                    else:
                        for st in desk.get("steps") or []:
                            if not isinstance(st, dict):
                                continue
                            act = (st.get("action") or "desktop").strip()
                            if act in ("wait",):
                                continue
                            tgt = (
                                st.get("description")
                                or st.get("target")
                                or st.get("input_value")
                                or act
                            )
                            # 无逐步结果时：整体失败则一律标 failed，避免假绿勾
                            overall_ok = bool(desk.get("ok")) and not desk.get("partial")
                            yield send(
                                'action',
                                action_type=act,
                                target=str(tgt)[:120],
                                status='success' if overall_ok else 'failed',
                            )
                            yield send(
                                'action_record',
                                action_type=act,
                                target=str(tgt)[:120],
                                status='success' if overall_ok else 'failed',
                                has_vision=False,
                            )
                    reply = desk.get("reply") or desk.get("error") or "桌面任务已处理"
                    yield send('reply', text=reply)
                    plan = {
                        "case_name": (task[:40] or "桌面操作"),
                        "case_url": "",
                        "description": task,
                        "platform": "desktop",
                        "steps": desk.get("steps") or [],
                        "meta": {
                            "source": "platform_desktop_first",
                            "via": desk.get("via"),
                            "hermes_llm": llm_st.get("reason"),
                            "partial": bool(desk.get("partial") or not desk.get("ok")),
                        },
                    }
                    yield send(
                        'done',
                        message='部分完成' if (desk.get("partial") or not desk.get("ok")) else '完成',
                        plan=plan,
                        partial=bool(desk.get("partial") or not desk.get("ok")),
                        session_id=task_ctx.session_id,
                    )
                    return
            except Exception as _desk_first_ex:
                uat_logger.debug("desktop-first path skipped: %s", _desk_first_ex)

            run_platform_early = run_platform or "auto"
            use_outer_desktop = False
            try:
                from ai_chat_tool_loop import prefer_outer_desktop_tools
                from agent_intent import resolve_task_route

                route2 = resolve_task_route(task, ui_platform=run_platform_early)
                if route2.needs_automation and route2.platform in ("web", "desktop", "android"):
                    run_platform_early = route2.platform
                    run_platform = run_platform_early
                use_outer_desktop = (
                    bool(route2.needs_desktop_tools)
                    or bool(getattr(route2, "needs_mobile_await", False))
                    or str(getattr(route2, "reason", "") or "").startswith("cross_end")
                    or prefer_outer_desktop_tools(
                        platform_type=run_platform_early,
                        message=task,
                    )
                )
            except Exception:
                use_outer_desktop = run_platform_early == "desktop"

            # Step 2: 工具循环（桌面=外层 windows_*；其它=Hermes）
            if url and not use_outer_desktop:
                yield send(
                    'think',
                    text=f'已从任务解析起始 URL：{url}',
                    status='running',
                )
            yield send(
                'think',
                text=(
                    '正在通过桌面工具逐步操控本机…'
                    if use_outer_desktop
                    else '正在通过智能体跨层工具链处理任务…'
                ),
                status='running',
            )

            try:
                from ai_chat_tool_loop import (
                    run_ai_chat_with_tools_stream,
                    ChatToolLoopParams,
                    hermes_execute_allowed,
                )
                from ai_multi_provider import get_active_llm_profile
                from ai_local_inference import local_ai_service

                profile = get_active_llm_profile()

                # 获取项目名称（使用预捕获的 user_id，避免 SSE 后 current_user 失效）
                project_name = "default"
                if project_id:
                    try:
                        projects = get_user_projects(user_id)
                        for p in projects:
                            if str(p.get("id")) == str(project_id):
                                project_name = p.get("name", "default")
                                break
                    except Exception:
                        pass

                current_plan = {
                    "case_name": task[:60],
                    "case_url": url or "",
                    "steps": [],
                }

                _page_snapshot = ""
                _probe_registry = None
                _dom_context_pack = ""

                screen_share_state = _screen_share_states.get(user_id, {})
                allow_screen_tools = bool(
                    screen_share_state.get('enabled', False) or enable_vision
                )

                from ai_action_recorder import ActionRecorder
                recorder = ActionRecorder(vision_enabled=bool(allow_screen_tools), platform=run_platform)

                run_platform = run_platform_early
                # 桌面 NL：外层直接 windows_*（OpenClaw 式）；避免 hermes_execute 嵌套空转
                allow_agent = bool(
                    hermes_available
                    and needs_automation
                    and hermes_execute_allowed(embedded_session_id="", platform_type=run_platform)
                    and not use_outer_desktop
                )
                if use_outer_desktop:
                    run_platform = "desktop"

                _uid = 0
                try:
                    _uid = int(getattr(current_user, "id", 0) or 0)
                except Exception:
                    _uid = 0

                _agent_sid = ""
                try:
                    _agent_sid = str(
                        (data.get("agent_session_id") if isinstance(data, dict) else None)
                        or request.args.get("agent_session_id")
                        or ""
                    ).strip()
                except Exception:
                    _agent_sid = ""

                _hands = {}
                try:
                    from agent_unified_session import snapshot_connected_hands

                    _hands = snapshot_connected_hands(_uid)
                except Exception:
                    _hands = {"phone": False, "desktop": use_outer_desktop, "browser": False}

                # 连接态双手：有桌面则挂桌面工具；有手机则保证 mobile_*
                if _hands.get("desktop"):
                    use_outer_desktop = True
                    run_platform = "desktop"
                elif use_outer_desktop:
                    run_platform = "desktop"

                params = ChatToolLoopParams(
                    message=task,
                    project_name=project_name,
                    current_plan=current_plan,
                    history=[],
                    profile=profile,
                    legacy_model="",
                    page_snapshot=_page_snapshot,
                    probe_registry=_probe_registry,
                    probe_url=url or "",
                    memory_context="",
                    dom_context_pack=_dom_context_pack,
                    interaction_context={
                        "url": url,
                        "platform": run_platform,
                        "enable_vision": allow_screen_tools,
                        "allow_screen_tools": allow_screen_tools,
                        "session_id": task_ctx.session_id,
                        "hands": _hands,
                        "entry": "ai_test",
                    },
                    test_scope=task,
                    embedded_session_id="",
                    platform_type=run_platform,
                    abort_event=abort_event,
                    recorder=recorder,
                    allow_screen_tools=True if use_outer_desktop else allow_screen_tools,
                    allow_desktop_windows_tools=True if use_outer_desktop else (
                        False if run_platform == "web" else None
                    ),
                    deadline_ts=deadline_ts,
                    ensure_browser_before_agent=_ensure_browser_before_agent if allow_agent else None,
                    allow_hermes_execute=allow_agent,
                    task_session_id=task_ctx.session_id,
                    capabilities_summary=caps_summary,
                    generate_case_after_run=generate_case_after_run,
                    allow_refine_test_plan=False,
                    user_id=_uid,
                    agent_session_id=_agent_sid or None,
                    connected_hands=_hands,
                )

                final_plan = None
                need_user_action = None
                hermes_partial = False
                hermes_failed = False
                last_reply_sent = ""
                for evt_type, evt_data in run_ai_chat_with_tools_stream(
                    local_ai_service=local_ai_service,
                    params=params,
                    abort_event=abort_event,
                ):
                    if evt_type == "thinking":
                        yield send('think', text=evt_data.get('content', 'AI 正在推理...'), status='running')

                    elif evt_type == "tool_call_start":
                        tool_name = evt_data.get('tool', '')
                        args_summary = evt_data.get('args_summary', '')
                        if tool_name == "hermes_execute":
                            yield send('action', action_type='hermes_execute', target=args_summary[:80], status='running')
                            yield send('think', text='Hermes 正在跨层执行…', status='running')
                        elif tool_name in ("windows_wait", "get_screen_text", "get_screen_description"):
                            yield send('think', text=f'{tool_name}…', status='running')
                        elif tool_name.startswith("windows_"):
                            yield send(
                                'action',
                                action_type=tool_name,
                                target=str(args_summary)[:80],
                                status='running',
                                quiet_chat=True,
                            )
                            yield send('think', text=f'桌面操作：{tool_name}', status='running')
                        elif tool_name.startswith("mobile_"):
                            yield send(
                                'action',
                                action_type=tool_name,
                                target=str(args_summary)[:80] or '等待手机本机',
                                status='running',
                                quiet_chat=True,
                            )
                            yield send(
                                'think',
                                text=f'手机双手：{tool_name}（enqueue 后等待本机领取）',
                                status='running',
                            )
                        elif tool_name.startswith("desktop_"):
                            yield send(
                                'action',
                                action_type=tool_name,
                                target=str(args_summary)[:80],
                                status='running',
                                quiet_chat=True,
                            )
                            yield send('think', text=f'桌面双手：{tool_name}', status='running')
                        elif tool_name == "refine_test_plan":
                            yield send('think', text='正在优化测试用例', status='running')

                    elif evt_type == "hermes_trace":
                        yield send(
                            'think',
                            text=f"Hermes: {(evt_data.get('message') or '')[:200]}",
                            status='running',
                        )

                    elif evt_type == "tool_call_result":
                        tool_name = evt_data.get('tool', '')
                        result_preview = evt_data.get('result_preview', '')
                        if tool_name == "hermes_execute":
                            # 工具结果里的 corrupt / NoneType.id / 空流 不应标成「执行完成」
                            _bad = False
                            _had_tools = False
                            _stream_empty = False
                            try:
                                from hermes_gateway_client import _is_corrupt_session_error
                                import json as _j_prev
                                _preview_l = (result_preview or "").lower()
                                _stream_empty = "stream_empty_text" in _preview_l
                                _bad = _is_corrupt_session_error(result_preview) or (
                                    '"ok": false' in _preview_l
                                    or '"ok":false' in _preview_l
                                    or _stream_empty
                                    or "auth_fatal" in _preview_l
                                )
                                try:
                                    _pj = _j_prev.loads(result_preview)
                                    if isinstance(_pj, dict):
                                        _had_tools = bool(_pj.get("had_tool_activity"))
                                        if _pj.get("ok") is False and not _pj.get("partial"):
                                            hermes_failed = True
                                except Exception:
                                    pass
                            except Exception:
                                _bad = "NoneType" in (result_preview or "")
                            yield send(
                                'action',
                                action_type='hermes_execute',
                                target='执行失败' if _bad else '执行完成',
                                status='error' if _bad else 'success',
                                result=result_preview[:200],
                            )
                            if _bad:
                                # 空流且无工具 = 失败；有工具但无摘要/部分失败 = 部分完成
                                if _stream_empty and not _had_tools:
                                    hermes_failed = True
                                elif "auth_fatal" in (result_preview or "").lower():
                                    hermes_failed = True
                                else:
                                    hermes_partial = True
                            from agent_hitl import looks_like_hitl_needed, set_need_user_action
                            if looks_like_hitl_needed(result_preview):
                                need_user_action = set_need_user_action(
                                    user_id,
                                    session_id=task_ctx.session_id,
                                    reason="需要人工确认或登录",
                                    hint=result_preview[:200],
                                )
                                task_ctx.request_hitl(need_user_action["reason"], need_user_action.get("hint") or "")
                                yield send('need_user_action', **need_user_action)
                            if (
                                not hermes_failed
                                and (
                                    '"partial": true' in (result_preview or "").lower()
                                    or '"partial":true' in (result_preview or "").lower()
                                    or "未完整" in (result_preview or "")
                                    or "未能" in (result_preview or "")
                                )
                            ):
                                hermes_partial = True
                            if _stream_empty:
                                yield send(
                                    'think',
                                    text=(
                                        'Hermes 空流结束，已禁止再次 hermes_execute'
                                        + ('（判定为失败，不可保存编造用例）' if hermes_failed else '')
                                    ),
                                    status='done',
                                )
                        elif tool_name == "refine_test_plan":
                            yield send('think', text='测试用例已更新', status='done')
                        elif tool_name in ("windows_wait", "get_screen_text", "get_screen_description"):
                            yield send('think', text=f'{tool_name} 完成', status='done')
                        elif tool_name.startswith("windows_"):
                            _ok = True
                            _verified = True
                            _human = ""
                            try:
                                import json as _j
                                _p = _j.loads(result_preview)
                                if isinstance(_p, dict):
                                    if _p.get("success") is False or _p.get("ok") is False:
                                        _ok = False
                                    if _p.get("verified") is False:
                                        _verified = False
                                    cap = _p.get("capture_after") if isinstance(_p.get("capture_after"), dict) else {}
                                    _human = (
                                        _p.get("reply")
                                        or _p.get("matched_title")
                                        or _p.get("matched")
                                        or _p.get("error")
                                        or ""
                                    )
                                    if cap.get("unchanged"):
                                        _verified = False
                                        yield send(
                                            'think',
                                            text='观察：操作后画面无变化（可能未打入目标窗）',
                                            status='warning',
                                        )
                                    elif cap.get("texts_preview"):
                                        preview = " / ".join(
                                            str(x) for x in (cap.get("texts_preview") or [])[:6]
                                        )
                                        yield send(
                                            'think',
                                            text=f'观察摘要：{preview[:160]}',
                                            status='done',
                                        )
                            except Exception:
                                low = (result_preview or "").lower()
                                _ok = '"success": false' not in low and '"success":false' not in low
                            # unchanged / verified=false 一律不得标 success
                            if _ok and _verified:
                                status = 'success'
                            elif not _ok:
                                status = 'failed'
                            else:
                                status = 'warning'
                            yield send(
                                'action',
                                action_type=tool_name,
                                target=str(_human or result_preview)[:80],
                                status=status,
                                result=(_human or result_preview)[:200],
                                quiet_chat=True,
                            )
                        elif tool_name.startswith("mobile_") or tool_name.startswith("desktop_"):
                            _ok = True
                            _human = ""
                            try:
                                import json as _j
                                _p = _j.loads(result_preview)
                                if isinstance(_p, dict):
                                    if _p.get("success") is False or _p.get("ok") is False:
                                        _ok = False
                                    _human = (
                                        _p.get("error")
                                        or _p.get("sms_otp")
                                        or _p.get("job_id")
                                        or _p.get("status")
                                        or ""
                                    )
                                    if _p.get("error_code"):
                                        _human = f"{_human} [{_p.get('error_code')}]".strip()
                            except Exception:
                                low = (result_preview or "").lower()
                                _ok = '"success": false' not in low and '"ok": false' not in low
                            yield send(
                                'action',
                                action_type=tool_name,
                                target=str(_human or result_preview)[:80],
                                status='success' if _ok else 'failed',
                                result=(_human or result_preview)[:200],
                                quiet_chat=True,
                            )
                            yield send(
                                'think',
                                text=(
                                    f'{tool_name} 完成'
                                    if _ok
                                    else f'{tool_name} 失败：{str(_human or result_preview)[:120]}'
                                ),
                                status='done' if _ok else 'error',
                            )

                    elif evt_type == "action_records":
                        for rec in (evt_data if isinstance(evt_data, list) else []):
                            yield send('action_record', **rec)
                            st = (rec.get('status') or '').lower()
                            # 实时用例只收录成功步骤；失败/取消留在「执行动作」
                            if st in ('success', 'ok', 'done', 'completed', 'complete'):
                                yield send(
                                    'case_step',
                                    action=rec.get('action_type', '操作'),
                                    target=rec.get('target', '目标'),
                                    verified=True,
                                )

                    elif evt_type == "vision_start":
                        yield send('vision_start', message=evt_data.get('message', 'AI 正在观察屏幕...'))

                    elif evt_type == "vision_result":
                        yield send('vision_result', text=evt_data.get('text', '')[:300])

                    elif evt_type == "plan_update":
                        plan = evt_data.get('plan', {})
                        final_plan = plan

                    elif evt_type == "reply":
                        reply_text = (evt_data.get("text") if isinstance(evt_data, dict) else str(evt_data)) or ""
                        if reply_text and reply_text.strip() != last_reply_sent.strip():
                            last_reply_sent = reply_text
                            yield send('reply', text=reply_text)
                            from agent_hitl import looks_like_hitl_needed, set_need_user_action
                            if looks_like_hitl_needed(reply_text):
                                need_user_action = set_need_user_action(
                                    user_id,
                                    session_id=task_ctx.session_id,
                                    reason="需要人工确认",
                                    hint=reply_text[:200],
                                )
                                yield send('need_user_action', **need_user_action)

                    elif evt_type == "done":
                        plan = evt_data.get('plan', {}) if isinstance(evt_data, dict) else {}
                        done_meta = evt_data.get('meta') if isinstance(evt_data, dict) else {}
                        if not isinstance(done_meta, dict):
                            done_meta = {}
                        if done_meta.get("hermes_failed") or done_meta.get("failed"):
                            hermes_failed = True
                        if done_meta.get("savable") is False:
                            hermes_failed = hermes_failed or bool(done_meta.get("hermes_failed") or done_meta.get("failed"))
                        if evt_data.get("failed") if isinstance(evt_data, dict) else False:
                            hermes_failed = True
                        if plan and not hermes_failed:
                            final_plan = plan
                        elif hermes_failed:
                            # 失败时丢弃模型编造的 plan，避免弹出「可保存用例」
                            final_plan = None
                        reply_text = (evt_data.get('reply') or "").strip() if isinstance(evt_data, dict) else ""
                        # 避免 reply + done.reply 重复刷同一段说明
                        if reply_text and reply_text.strip() != last_reply_sent.strip():
                            last_reply_sent = reply_text
                            yield send('reply', text=reply_text)

                        # 热路径：ActionRecorder → normalize；失败时不生成可保存 plan
                        norm_warnings = []
                        try:
                            if hermes_failed:
                                final_plan = None
                            elif generate_case_after_run and recorder and getattr(recorder, "records", None):
                                built, norm_warnings = recorder.build_normalized_plan(
                                    case_name=(final_plan or {}).get("case_name") or task[:60],
                                    case_url=url or (final_plan or {}).get("case_url") or "",
                                    instruction=task,
                                )
                                if built.get("steps"):
                                    if final_plan and isinstance(final_plan, dict) and final_plan.get("case_name"):
                                        built["case_name"] = final_plan["case_name"]
                                    built.setdefault("meta", {})
                                    built["meta"]["session_id"] = task_ctx.session_id
                                    built["meta"]["vars"] = dict(task_ctx.vars)
                                    built["meta"]["trace_count"] = len(task_ctx.tool_trace)
                                    final_plan = built
                                    yield send(
                                        'think',
                                        text=f'已将 {len(built["steps"])} 步动作规范化为可维护用例',
                                        status='done',
                                    )
                            elif (
                                generate_case_after_run
                                and not hermes_failed
                                and task_ctx.tool_trace
                                and (not final_plan or not final_plan.get("steps"))
                            ):
                                # 仅非失败时：用真实 tool_trace 生成占位 plan（仍默认不强制可保存）
                                steps = []
                                for tr in task_ctx.tool_trace:
                                    tool_name_tr = (tr.get("tool") or "").strip()
                                    if tool_name_tr in ("hermes_execute", "openclaw_execute"):
                                        continue
                                    steps.append({
                                        "action": "note",
                                        "target": tool_name_tr or "hermes",
                                        "description": tr.get("summary") or "",
                                        "automation_layer": run_platform or "auto",
                                    })
                                if steps:
                                    final_plan = {
                                        "case_name": task[:60],
                                        "case_url": url or "",
                                        "description": task,
                                        "platform": run_platform or "auto",
                                        "steps": steps,
                                        "meta": {
                                            "source": "hermes_trace",
                                            "unsavable": True,
                                            "session_id": task_ctx.session_id,
                                            "vars": dict(task_ctx.vars),
                                        },
                                    }
                        except Exception as ex:
                            uat_logger.debug("hot-path normalize failed: %s", ex)

                        savable = False
                        if (
                            not hermes_failed
                            and final_plan
                            and isinstance(final_plan, dict)
                            and final_plan.get("steps")
                        ):
                            meta_fp = final_plan.get("meta") if isinstance(final_plan.get("meta"), dict) else {}
                            if not meta_fp.get("unsavable") and meta_fp.get("source") != "hermes_trace":
                                savable = True
                                try:
                                    from hermes_skill_loop import record_execution_success
                                    record_execution_success(
                                        final_plan,
                                        case_url=url or final_plan.get('case_url', ''),
                                        instruction=task,
                                        outcome='ok' if not hermes_partial else 'partial',
                                    )
                                except Exception:
                                    pass

                        if hermes_failed:
                            done_msg = '执行失败'
                        elif hermes_partial or need_user_action:
                            done_msg = '部分完成'
                        else:
                            done_msg = '完成'
                        done_payload = {
                            'message': done_msg,
                            'plan': final_plan or {},
                            'session_id': task_ctx.session_id,
                            'partial': bool(not hermes_failed and (hermes_partial or need_user_action)),
                            'failed': bool(hermes_failed),
                            'savable': bool(savable),
                        }
                        if need_user_action:
                            done_payload['need_user_action'] = need_user_action
                        if norm_warnings:
                            done_payload['warnings'] = norm_warnings[:20]
                        yield send('done', **done_payload)

                    elif evt_type == "error":
                        raw_err = evt_data.get("error") if isinstance(evt_data, dict) else evt_data
                        err_text = _friendly_sse_error(raw_err)
                        if getattr(abort_event, "_timed_out", False) or (
                            deadline_ts and _time.time() >= deadline_ts
                        ):
                            err_text = "任务已超过设定的超时时间，已自动停止"
                        elif str(getattr(abort_event, "_abort_reason", "") or "") == "tool_loop":
                            err_text = "智能体因工具死循环已中止（非用户取消）"
                        yield send('error', error=err_text)

            except ImportError as e:
                yield send('think', text='工具循环模块不可用: ' + str(e), status='warning')
                if hermes_available and needs_automation:
                    instruction = task
                    if url:
                        instruction = f"目标URL: {url}\n\n任务: {task}"
                    try:
                        _ensure_browser_before_agent()
                        from hermes_skill_hints import build_explore_instruction
                        instruction = build_explore_instruction(
                            instruction,
                            {
                                "platform": run_platform or "auto",
                                "context_prefix": task_ctx.instruction_prefix(),
                                "capabilities_summary": caps_summary,
                            },
                        )
                        # 降级路径也不传平台 session，避免 Hermes 会话损坏
                        raw = hermes_client.execute_user_instruction(instruction, "")
                        try:
                            parsed = _json.loads(raw) if isinstance(raw, str) else None
                        except Exception:
                            parsed = None
                        if isinstance(parsed, dict) and parsed.get("ok") is False:
                            yield send('error', error=_friendly_sse_error(parsed.get("error") or raw))
                        else:
                            yield send('action', action_type='execute', target='执行完成', status='success', result=str(raw)[:200])
                            yield send('done', message='完成（降级模式）', session_id=task_ctx.session_id)
                    except Exception as ex:
                        yield send('error', error=_friendly_sse_error(ex))
                else:
                    yield send('error', error='对话模块不可用，请检查依赖')

            except Exception as e:
                err_text = _friendly_sse_error(e)
                try:
                    from agent_task_context import reset_task_context
                    reset_task_context(task_ctx.session_id)
                except Exception:
                    pass
                yield send('error', error=err_text)

        except Exception as e:
            yield send('error', error=_friendly_sse_error(e))
        finally:
            with _ai_task_abort_lock:
                _ai_task_abort_events.pop(job_id, None)
            # 恢复超时设置
            if old_timeout:
                os.environ["HERMES_GATEWAY_TIMEOUT"] = old_timeout
            elif "HERMES_GATEWAY_TIMEOUT" in os.environ:
                del os.environ["HERMES_GATEWAY_TIMEOUT"]

    return Response(
        stream_with_context(_gen()),
        mimetype='text/event-stream; charset=utf-8',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
            'Content-Type': 'text/event-stream; charset=utf-8',
        },
    )


@app.route('/api/ai/task/execute/logs', methods=['GET'])
@login_required
@api_error_handler
def api_ai_task_execute_logs():
    return jsonify({'success': True, 'logs': []})


@app.route('/api/ai/task/execute/stop', methods=['POST'])
@login_required
@api_error_handler
def api_ai_task_execute_stop():
    data = request.get_json(silent=True) or {}
    job_id = data.get('job_id')
    hard = bool(data.get('hard'))  # hard=True：连带强制中断 Hermes 在途请求（重启 gateway 过重，仅 abort）

    with _ai_task_abort_lock:
        if job_id and job_id in _ai_task_abort_events:
            _ai_task_abort_events[job_id].set()
        else:
            for evt in list(_ai_task_abort_events.values()):
                try:
                    evt.set()
                except Exception:
                    pass
            if not job_id:
                _ai_task_abort_events.clear()

    # 注意：Hermes /v1/chat/completions 无法从客户端真正取消 socket；
    # abort_event 会在 200ms 内返回，后台 HTTP 线程会在超时后自然结束。
    return jsonify({
        'success': True,
        'message': '已发送停止信号' + ('（硬停止）' if hard else ''),
        'job_id': job_id,
    })

# ── Hermes Gateway 管理 API ──────────────────────────────

@app.route('/api/ai/hermes/status', methods=['GET'])
@login_required
@api_error_handler
def api_ai_hermes_status():
    """获取 Hermes Gateway 运行状态（含启动中状态）"""
    try:
        from hermes_service_bootstrap import get_bootstrap_status
        status = get_bootstrap_status()
    except Exception as e:
        status = {"running": False, "configured": False, "starting": False, "start_error": str(e), "cdp_connected": False}
    return jsonify({"success": True, "status": status})


@app.route('/api/ai/hermes/start', methods=['POST'])
@login_required
@api_error_handler
def api_ai_hermes_start():
    """一键启动 Hermes Gateway（后台异步启动，不阻塞请求）"""
    try:
        from hermes_service_bootstrap import get_bootstrap_status, _force_stale_stopping_unlock
        from hermes_gateway_client import HermesGatewayClient
        client = HermesGatewayClient()
        if not client.is_configured():
            return jsonify({"success": True, "started": False, "message": "Hermes Gateway 未配置"})
        if client.health_check(timeout_sec=1.0):
            try:
                from hermes_service_bootstrap import ensure_hermes_llm_current

                llm_info = ensure_hermes_llm_current(restart_if_stale=True)
            except Exception as e:
                llm_info = {"error": str(e)[:120]}
            return jsonify({
                "success": True,
                "started": True,
                "already_running": True,
                "message": "Hermes Gateway 已在运行",
                "hermes_llm": llm_info,
            })
        # 卡住的 stopping 先清掉，允许立刻再启动
        _force_stale_stopping_unlock()
        cur = get_bootstrap_status()
        if cur.get("stopping"):
            # 停止刚开始：仍允许排队启动（force），避免用户连点被 409 卡死
            # bootstrap 内部会对短暂 stopping 做 force 抢占
            pass
        if cur.get("starting") and not cur.get("start_error") and (cur.get("starting_elapsed_sec") or 0) < 40:
            return jsonify({"success": True, "started": False, "starting": True, "message": "正在启动中…"})
        import threading as _threading
        def _boot():
            try:
                from hermes_service_bootstrap import bootstrap_hermes_services
                bootstrap_hermes_services(force=True, manual=True)
            except Exception as exc:
                try:
                    from hermes_service_bootstrap import _LIFECYCLE_LOCK, _clear_starting_locked
                    with _LIFECYCLE_LOCK:
                        _clear_starting_locked(error=str(exc)[:200], finished=True)
                except Exception:
                    pass
        _threading.Thread(target=_boot, daemon=True, name="hermes-boot").start()
        return jsonify({"success": True, "started": False, "starting": True, "message": "正在启动 Hermes Gateway…"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "started": False}), 500


@app.route('/api/ai/hermes/stop', methods=['POST'])
@login_required
@api_error_handler
def api_ai_hermes_stop():
    """一键停止 Hermes Gateway（含官方 PID、进程树、端口占用与浏览器残留）"""
    try:
        # 先中断进行中的 AI 任务，避免停止后仍有后台工具调用
        with _ai_task_abort_lock:
            for evt in list(_ai_task_abort_events.values()):
                try:
                    evt.set()
                except Exception:
                    pass
            _ai_task_abort_events.clear()

        from hermes_service_bootstrap import stop_hermes_gateway, get_bootstrap_status
        # 同步停止但内部有总超时；finally 保证 clearing stopping
        detail = stop_hermes_gateway() or {}
        status = get_bootstrap_status()
        fully_stopped = bool(detail.get("fully_stopped")) and not status.get("running") and not status.get("stopping")
        message = "Hermes Gateway 已完全停止" if fully_stopped else "已发送停止指令；若仍显示运行中请再点一次停止"
        return jsonify({
            "success": True,
            "message": message,
            "fully_stopped": fully_stopped,
            "stopping": bool(status.get("stopping")),
            "detail": detail,
            "status": status,
        })
    except Exception as e:
        # 异常时也尽量拉一次 status，避免前端永远「停止中」
        try:
            from hermes_service_bootstrap import get_bootstrap_status, _force_stale_stopping_unlock
            _force_stale_stopping_unlock()
            status = get_bootstrap_status()
        except Exception:
            status = {"stopping": False, "running": False}
        return jsonify({"success": False, "error": str(e), "status": status}), 500


@app.route('/api/ai/hermes/cdp-sync', methods=['POST'])
def api_ai_hermes_cdp_sync():
    """接收 Chrome 扩展发送的 CDP WebSocket URL，同步给 Hermes（无需登录，扩展无法携带 cookie）"""
    data = request.get_json(silent=True) or {}
    cdp_ws_url = data.get('cdp_ws_url', '').strip()
    if not cdp_ws_url:
        return jsonify({"success": False, "error": "缺少 CDP WebSocket URL"}), 400

    try:
        from hermes_config import sync_hermes_cdp_endpoint
        sync_hermes_cdp_endpoint(cdp_ws_url)
        return jsonify({"success": True, "message": "CDP 端点已同步给 Hermes"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/ai/actions/to-case', methods=['POST'])
@login_required
@api_error_handler
def api_ai_actions_to_case():
    """将执行动作转换为测试用例，走 ai_step_normalization 全管线，并沉淀为 Hermes Skill"""
    from ai_step_normalization import (
        normalize_ai_step,
        repair_raw_ai_steps_for_platform,
        dedupe_and_validate_ai_steps,
        apply_step_normalization_to_plan,
    )

    data = request.get_json(silent=True) or {}
    actions = data.get('actions', [])
    project_id = data.get('project_id')
    case_name = data.get('case_name', 'AI 生成用例')
    instruction = data.get('instruction', '')
    case_url = data.get('url', '')
    platform = data.get('platform', 'web')

    # Step 1: 原始动作 → 步骤格式，写入 automation_layer 供 normalize_ai_step 读取平台
    raw_steps = []
    for action in actions:
        raw_steps.append({
            'action': action.get('action_type', action.get('type', '操作')),
            'target': action.get('target', action.get('content', '')),
            'input_value': action.get('input_data', action.get('input_value', '')),
            'description': action.get('result', ''),
            'automation_layer': platform,  # normalize_ai_step 从此字段读平台
        })

    # Step 2: 逐条规范化（复用 837 行管线）
    normalized_steps = [normalize_ai_step(s) for s in raw_steps]

    # Step 3: 平台修复（实测：无 platform 参数，就地修改，返回告警列表）
    warnings1 = repair_raw_ai_steps_for_platform(normalized_steps) or []

    # Step 4: 去重 + 验证（实测：platform 为关键字参数）
    clean_steps, warnings2 = dedupe_and_validate_ai_steps(normalized_steps, platform=platform)

    # Step 5: 构造 plan 并应用规范化（实测：无 platform 参数）
    plan = {'case_name': case_name, 'case_url': case_url or '', 'steps': clean_steps}
    plan, warnings3 = apply_step_normalization_to_plan(plan)
    warnings3 = warnings3 or []

    all_warnings = warnings1 + (warnings2 or []) + warnings3

    # Step 6: 选择器恢复（复用 4 层降级，如有 probe_registry）
    try:
        from ai_external_browser_bridge import get_probe_registry
        registry = get_probe_registry()
        if registry and clean_steps:
            from ai_locator_resolution import resolve_plan_steps_locators_with_snapshot
            plan = resolve_plan_steps_locators_with_snapshot(plan, registry)
    except Exception:
        pass

    # Step 7: 保存到数据库
    saved_to_db = False
    case_id = None
    if project_id:
        try:
            from database import get_db_connection
            conn = get_db_connection()
            cursor = conn.cursor()
            import json as _json
            cursor.execute(
                "INSERT INTO test_cases (project_id, name, steps, created_by, created_at) VALUES (?, ?, ?, ?, datetime('now'))",
                (int(project_id), case_name, _json.dumps(clean_steps, ensure_ascii=False), current_user.id)
            )
            conn.commit()
            case_id = cursor.lastrowid
            conn.close()
            saved_to_db = True
        except Exception:
            pass

    # Step 8: 沉淀为 Hermes Skill
    skill_exported = False
    try:
        from hermes_skill_loop import record_execution_success
        record_execution_success(
            plan,
            case_url=case_url,
            instruction=instruction or case_name,
            outcome='ok',
        )
        skill_exported = True
    except Exception:
        pass

    return jsonify({
        'success': True,
        'message': f'动作转用例成功（规范化 {len(clean_steps)} 步）' + ('，已沉淀为 Skill' if skill_exported else ''),
        'case': {
            'name': case_name,
            'project_id': project_id,
            'steps': clean_steps,
        },
        'case_id': case_id,
        'saved_to_db': saved_to_db,
        'skill_exported': skill_exported,
        'warnings': all_warnings,
    })


@app.route('/api/ai/vision/capture', methods=['POST'])
@login_required
@api_error_handler
def api_ai_vision_capture():
    """捕获当前浏览器/屏幕截图。"""
    data = request.get_json(silent=True) or {}
    source = data.get('source', 'browser')  # browser / screen
    try:
        if source == 'browser':
            from ai_external_browser_bridge import capture_screenshot
            png = capture_screenshot()
        else:
            import mss
            from mss.tools import to_png
            with mss.mss() as sct:
                monitor = sct.monitors[1]
                shot = sct.grab(monitor)
                png = to_png(shot.rgb, shot.size)
        if not png:
            return jsonify({'success': False, 'error': '截图失败（无可用画面）'}), 500
        import tempfile, os, uuid
        filename = f"vision_{uuid.uuid4().hex[:8]}.png"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        with open(filepath, 'wb') as f:
            f.write(png)
        return jsonify({'success': True, 'screenshot_path': filepath, 'size': len(png)})
    except ImportError:
        return jsonify({'success': False, 'error': '截图依赖未安装（mss 或 Playwright）'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)[:200]}), 500


@app.route('/api/ai/vision/snapshot', methods=['GET'])
@login_required
@api_error_handler
def api_ai_vision_snapshot():
    """获取最新浏览器截图 + OCR 文本。"""
    try:
        from ai_external_browser_bridge import capture_screenshot
        png = capture_screenshot()
        if not png:
            return jsonify({'success': False, 'error': '无可用截图'}), 404
        ocr_text = ""
        try:
            from ai_vision_local import ocr_region_png
            ocr_text = ocr_region_png(png)  # 实测接受 bytes
        except Exception:
            pass
        return jsonify({'success': True, 'ocr_text': ocr_text[:2000], 'screenshot_size': len(png)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)[:200]}), 500


# 模块级屏幕共享状态（按用户ID隔离）
_screen_share_states: Dict[str, Dict[str, Any]] = {}


@app.route('/api/ai/screen-share/toggle', methods=['POST'])
@login_required
@api_error_handler
def api_ai_screen_share_toggle():
    """开启/关闭屏幕观察工具。开启后 Agent 工具列表包含 get_screen_text / get_screen_description。"""
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get('enabled', False))
    user_id = str(current_user.id)
    interval = int(data.get('interval', 3))
    _screen_share_states[user_id] = {
        'enabled': enabled,
        'interval': interval,
    }

    # 同步到 ai_external_browser_bridge 模块
    try:
        import ai_external_browser_bridge as _bridge
        with _bridge._bridge_lock:
            _bridge._screen_share_active = enabled
            _bridge._screen_share_interval = interval
    except Exception:
        pass

    return jsonify({
        'success': True,
        'enabled': enabled,
        'message': '已启用屏幕观察工具（按需调用）' if enabled else '已关闭屏幕观察工具',
    })


@app.route('/api/ai/agent/capabilities', methods=['GET'])
@login_required
@api_error_handler
def api_ai_agent_capabilities():
    """能力注册表快照（预检 / 调试）。"""
    from agent_capability_registry import snapshot_capabilities

    return jsonify({'success': True, **snapshot_capabilities()})


@app.route('/api/ai/agent/hitl/resume', methods=['POST'])
@login_required
@api_error_handler
def api_ai_agent_hitl_resume():
    """用户完成验证码/登录后继续同一会话。

    支持：
    - 原 AI Agent 路径：按 user_id pending resume
    - 跨端阻塞门禁：按 gate_id / session_id 调用 resume_hitl_gate（不立即清 gate，由 waiter 消费）
    """
    from agent_hitl import (
        mark_user_resumed,
        clear_user_action,
        get_pending,
        resume_hitl_gate,
        get_hitl_gate,
    )
    from agent_task_context import get_task_context

    data = request.get_json(silent=True) or {}
    user_id = str(current_user.id)
    pending = get_pending(user_id)
    gate_id = (
        (data.get('gate_id') or data.get('session_id') or '')
        or ((pending or {}).get('gate_id') or (pending or {}).get('session_id') or '')
    )
    gate_id = str(gate_id or '').strip()

    gate_ok = False
    if gate_id:
        gate_ok = bool(resume_hitl_gate(gate_id))

    user_ok = mark_user_resumed(user_id)
    sid = (data.get('session_id') or (pending or {}).get('session_id') or gate_id or '').strip()
    ctx = get_task_context(sid) if sid else None
    if ctx:
        ctx.clear_hitl()
        ctx.set_var('hitl_resumed_at', __import__('time').time())
        if gate_id:
            ctx.set_var('hitl_gate_id', gate_id)

    # UI pending 可清；gate 由 wait_hitl_gate 消费后再清，避免竞态丢 resume
    clear_user_action(user_id)

    gate_snapshot = get_hitl_gate(gate_id) if gate_id else None
    return jsonify({
        'success': True,
        'session_id': sid,
        'gate_id': gate_id or None,
        'gate_resumed': gate_ok,
        'user_pending_resumed': user_ok,
        'gate_status': (gate_snapshot or {}).get('status') if gate_snapshot else (
            'consumed_or_missing' if gate_ok else None
        ),
        'message': '已确认人工步骤，可继续发送任务或等待跨端编排恢复',
        'context': ctx.to_public_dict() if ctx else None,
    })


@app.route('/api/ai/agent/hitl/status', methods=['GET'])
@login_required
@api_error_handler
def api_ai_agent_hitl_status():
    """查询用户 pending 或指定 gate 状态（跨端轮询/调试）。"""
    from agent_hitl import get_pending, get_hitl_gate, list_hitl_gates

    user_id = str(current_user.id)
    gate_id = (request.args.get('gate_id') or request.args.get('session_id') or '').strip()
    pending = get_pending(user_id)
    gate = get_hitl_gate(gate_id) if gate_id else None
    if gate is None and pending:
        gate = get_hitl_gate(
            str(pending.get('gate_id') or pending.get('session_id') or '')
        )
    waiting = list_hitl_gates(status='waiting')
    return jsonify({
        'success': True,
        'pending': pending,
        'gate': gate,
        'waiting': waiting,
    })


@app.route('/api/ai/agent/hitl/cancel', methods=['POST'])
@login_required
@api_error_handler
def api_ai_agent_hitl_cancel():
    """取消 HITL 门禁（跨端等待将失败，不会假绿）。"""
    from agent_hitl import cancel_hitl_gate, get_pending, clear_user_action

    data = request.get_json(silent=True) or {}
    user_id = str(current_user.id)
    pending = get_pending(user_id)
    gate_id = str(
        data.get('gate_id')
        or data.get('session_id')
        or (pending or {}).get('gate_id')
        or (pending or {}).get('session_id')
        or ''
    ).strip()
    if not gate_id:
        return jsonify({'success': False, 'error': 'gate_id 必填'}), 400
    ok = bool(cancel_hitl_gate(gate_id))
    clear_user_action(user_id)
    return jsonify({
        'success': True,
        'gate_id': gate_id,
        'cancelled': ok,
        'message': '已取消人工确认' if ok else '门禁不存在或已结束',
    })


@app.route('/api/ai/agent/api-http', methods=['POST'])
@login_required
@api_error_handler
def api_ai_agent_api_http():
    """Hermes / 平台共用的临时 HTTP 执行入口。"""
    from agent_api_runner import run_temp_http, run_api_case, summarize_for_agent
    from agent_task_context import get_task_context

    data = request.get_json(silent=True) or {}
    session_id = (data.get('session_id') or '').strip()
    case_id = data.get('case_id')
    if case_id:
        result = run_api_case(int(case_id), db)
    else:
        result = run_temp_http(
            method=data.get('method') or 'GET',
            url=data.get('url') or '',
            headers=data.get('headers') if isinstance(data.get('headers'), dict) else None,
            body=data.get('body'),
            timeout_sec=float(data.get('timeout_sec') or 30),
        )
    ctx = get_task_context(session_id) if session_id else None
    if ctx:
        ctx.append_trace('api_http', summarize_for_agent(result)[:300], ok=bool(result.get('ok')))
        store_as = (data.get('store_as') or '').strip()
        if store_as and result.get('ok'):
            ctx.set_var(store_as, result.get('result') or result)
    return jsonify({'success': bool(result.get('ok')), 'result': result, 'session_id': session_id})


@app.route('/api/ai/cases/append-steps', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_append_steps_to_case():
    """
    将 AI 生成步骤追加到现有用例，不新建用例。
    可选 case_name_override / session_title：同时重命名该用例（需编辑权限）。
    """
    data = request.get_json(silent=True) or {}
    case_id = data.get('case_id')
    steps = data.get('steps') or []
    case_name_override = (data.get('case_name_override') or data.get('session_title') or '').strip()
    if not case_id:
        return jsonify({'success': False, 'error': 'case_id不能为空'}), 400
    if not isinstance(steps, list) or not steps:
        return jsonify({'success': False, 'error': 'steps不能为空'}), 400

    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404

    _db = Database()
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403

    old_steps, _total = db.get_case_steps_paginated(case_id, 1, 1000)
    max_order = 0
    if old_steps:
        max_order = max([int(s.get('step_order') or 0) for s in old_steps] or [0])

    goal_hint = _ai_str(data.get('goal')) or _ai_str(case.get('name')) or _ai_str(case.get('description'))
    local_ai_service._fill_missing_step_payloads(
        steps,
        goal_hint,
        _ai_str(case.get('url')),
        None,
    )
    clean_steps, warnings = dedupe_and_validate_ai_steps(steps)
    ct_ai = _app_case_type(case)
    if ct_ai == "ui":
        clean_steps = [s for s in clean_steps if (s.get("action") or "").strip().lower() != "api_request"]
    if not clean_steps:
        return jsonify(
            {
                "success": False,
                "error": "没有可写入的步骤（Web 用例已忽略接口类步骤）",
                "warnings": warnings,
            }
        ), 400

    created_steps = 0
    for idx, step in enumerate(clean_steps, start=1):
        db.create_test_step(**_ai_step_to_db_kwargs(step, case_id, max_order + idx))
        created_steps += 1

    final_name = case.get('name') or ''
    if case_name_override:
        db.update_test_case_v2(case_id, name=case_name_override)
        final_name = case_name_override

    return jsonify({
        'success': True,
        'case_id': case_id,
        'case_name': final_name,
        'steps_created': created_steps,
        'warnings': warnings,
    })


@app.route('/api/ai/cases/create-from-plan', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_ai_create_case_from_plan():
    """根据 AI 生成的完整 plan 创建新用例并写入步骤。"""
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    plan = data.get('plan') or {}

    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    if not isinstance(plan, dict):
        return jsonify({'success': False, 'error': 'plan 格式错误'}), 400

    # 权限检查
    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403

    # 创建用例（勿传 created_by：create_test_case_v2 无此参数）
    case_name = _ai_str(plan.get('case_name')) or 'AI 生成用例'
    case_url = _ai_str(plan.get('case_url')) or ''
    description = _ai_str(plan.get('description')) or ''
    platform = _ai_str(plan.get('platform')) or 'web'
    if (plan.get('meta') or {}).get('unsavable') or (plan.get('meta') or {}).get('source') == 'hermes_trace':
        return jsonify({'success': False, 'error': '该计划来自失败/占位轨迹，不可保存为用例'}), 400
    steps = plan.get('steps') or []
    if not isinstance(steps, list) or not steps:
        return jsonify({'success': False, 'error': 'plan.steps 为空，无法保存'}), 400

    case_id = db.create_test_case_v2(
        project_id=project_id,
        name=case_name,
        url=case_url,
        description=description,
        platform=platform,
        generated_by_ai=True,
    )

    # 写入步骤
    created_steps = 0
    warnings = []
    if steps:
        goal_hint = case_name
        local_ai_service._fill_missing_step_payloads(
            steps,
            goal_hint,
            case_url,
            None,
        )
        clean_steps, warnings = dedupe_and_validate_ai_steps(steps)
        ct_ai = _app_case_type({'url': case_url, 'platform': platform})
        if ct_ai == "ui":
            clean_steps = [s for s in clean_steps if (s.get("action") or "").strip().lower() != "api_request"]
        for idx, step in enumerate(clean_steps, start=1):
            db.create_test_step(**_ai_step_to_db_kwargs(step, case_id, idx))
            created_steps += 1

    return jsonify({
        'success': True,
        'case_id': case_id,
        'case_name': case_name,
        'steps_created': created_steps,
        'warnings': warnings,
    })

# API: 悬停在元素上
@app.route('/api/hover_element', methods=['POST'])
@api_error_handler
@log_api_request
def api_hover_element():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    selector_type = data.get('selector_type', 'css')
    iframe_selector = data.get('iframe_selector', '')
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_hover_element(selector, selector_type, iframe_selector=iframe_selector)
    return jsonify({'success': True})

# API: 双击元素
@app.route('/api/double_click', methods=['POST'])
@api_error_handler
@log_api_request
def api_double_click():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    selector_type = data.get('selector_type', 'css')
    iframe_selector = data.get('iframe_selector', '')
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_double_click_element(selector, selector_type, iframe_selector=iframe_selector)
    return jsonify({'success': True})

# API: 点击元素
@app.route('/api/click_element', methods=['POST'])
@api_error_handler
@log_api_request
def api_click_element():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_click_element(selector)
    return jsonify({'success': True})



# API: 右键点击元素
@app.route('/api/right_click', methods=['POST'])
@api_error_handler
@log_api_request
def api_right_click():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    selector_type = data.get('selector_type', 'css')
    iframe_selector = data.get('iframe_selector', '')
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_right_click_element(selector, selector_type, iframe_selector=iframe_selector)
    return jsonify({'success': True})

# API: 等待元素出现
@app.route('/api/wait_for_selector', methods=['POST'])
@api_error_handler
@log_api_request
def api_wait_for_selector():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    timeout = data.get('timeout', 30000)
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_wait_for_selector(selector, timeout)
    return jsonify({'success': True})

# API: 等待元素可见
@app.route('/api/wait_for_element_visible', methods=['POST'])
@api_error_handler
@log_api_request
def api_wait_for_element_visible():
    data = request.get_json(silent=True) or {}
    selector = data.get('selector', '')
    timeout = data.get('timeout', 30000)
    
    if not selector:
        return jsonify({'error': '选择器不能为空'}), 400
    
    sync_wait_for_element_visible(selector, timeout)
    return jsonify({'success': True})

# API: 获取页面元素
@app.route('/api/page_elements', methods=['GET'])
@api_error_handler
@log_api_request
def api_page_elements():
    elements = sync_get_page_elements()
    return jsonify({'success': True, 'elements': elements})

# API: 检查是否存在测试用例
@app.route('/api/has_test_cases', methods=['GET'])
@api_error_handler
@log_api_request
def api_has_test_cases():
    cases = db.get_all_test_cases()
    has_cases = len(cases) > 0
    return jsonify({'success': True, 'has_cases': has_cases})

# API: 获取页面截图
@app.route('/api/screenshot', methods=['GET'])
@api_error_handler
@log_api_request
def api_screenshot():
    try:
        from flask import send_file

        # 生成截图文件名
        timestamp = int(time.time())
        filename = f"screenshot_{timestamp}.png"
        filepath = os.path.join(os.getcwd(), filename)
        
        # 保存截图
        sync_take_screenshot(filepath)
        
        # 返回截图文件
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# —— 内置浏览器（与 AI 测试工作台共用 Playwright 会话，PNG 直出供侧栏预览）——

@app.route('/api/browser/frame', methods=['GET'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_frame():
    from flask import send_file

    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动，请先「打开页面」'}), 400
    data = sync_take_screenshot_bytes()
    return send_file(
        io.BytesIO(data),
        mimetype='image/png',
        max_age=0,
        conditional=False,
    )


@app.route('/api/browser/status', methods=['GET'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_status():
    try:
        if not sync_automation_session_usable():
            return jsonify({'success': True, 'ready': False, 'url': '', 'title': ''})
        url = sync_get_current_url()
        title = sync_get_page_title()
        return jsonify({'success': True, 'ready': True, 'url': url, 'title': title})
    except Exception as e:
        uat_logger.debug(f"api_browser_status: {e}")
        return jsonify({'success': True, 'ready': False, 'url': '', 'title': ''})


@app.route('/api/browser/go-back', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_go_back():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    ok = sync_browser_go_back()
    return jsonify({'success': True, 'moved': ok})


@app.route('/api/browser/go-forward', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_go_forward():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    ok = sync_browser_go_forward()
    return jsonify({'success': True, 'moved': ok})


@app.route('/api/browser/reload', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_reload():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    sync_browser_reload()
    return jsonify({'success': True})


@app.route('/api/browser/diagnostics', methods=['GET'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_diagnostics():
    """页面/性能轻量诊断（开发者面板）。"""
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    payload = sync_get_page_diagnostics()
    return jsonify({'success': True, 'data': payload})


@app.route('/api/browser/inspect', methods=['GET'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_inspect():
    """内置浏览器「页面结构」：可交互元素列表 + 建议定位（非完整 F12）。"""
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动，请先打开页面'}), 400
    payload = sync_get_interactive_page_snapshot(150)
    return jsonify({'success': True, 'data': payload})


@app.route('/api/captcha/optional-deps', methods=['GET'])
@login_required
@api_error_handler
def api_captcha_optional_deps():
    """验证码可选组件状态（ddddocr 等，不随主安装包分发）。"""
    from captcha_engine import get_ddddocr_install_info

    return jsonify({'success': True, 'ddddocr': get_ddddocr_install_info()})


@app.route('/api/captcha/optional-deps/install', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_captcha_optional_deps_install():
    """一键安装验证码可选组件（如 ddddocr，约 200MB+）。"""
    data = request.get_json(silent=True) or {}
    component = (data.get('component') or 'ddddocr').strip().lower()
    if component != 'ddddocr':
        return jsonify({'success': False, 'error': f'未知组件: {component}'}), 400
    from captcha_engine import install_ddddocr_subprocess

    result = install_ddddocr_subprocess()
    return jsonify({'success': bool(result.get('success')), **result})


@app.route('/api/desktop/verify-element', methods=['POST'])
@app.route('/api/desktop/verify_element', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_verify_element():
    """校验桌面元素是否可解析（UIA 精准链，不依赖遮挡/Z 序）。"""
    data = request.get_json(silent=True) or {}
    spec, err = _parse_desktop_spec_body(data)
    if err:
        return jsonify({'success': False, 'error': err}), 400
    try:
        from desktop_automation import sync_desktop_verify_element

        result = sync_desktop_verify_element(
            data.get('selector_type') or '',
            data.get('selector_value') or '',
            spec or {},
            data.get('locator_candidates'),
        )
        if result.get('success'):
            return jsonify({'success': True, **result})
        fail_body = {'success': False, 'error': result.get('error') or '校验失败'}
        if result.get('need_relearn'):
            fail_body['need_relearn'] = True
            fail_body['best_score'] = float(result.get('best_score') or 0.0)
        return jsonify(fail_body), 400
    except Exception as e:
        uat_logger.log_exception('api_desktop_verify_element', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/relaunch-with-embed-hooks', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_relaunch_with_embed_hooks():
    """
    无需源码：结束当前前台应用（或指定 exe）并用 Testory 挂钩重启，
    自动注入 --force-renderer-accessibility 与 WebView2 调试参数，便于捕获内部元素。
    """
    data = request.get_json(silent=True) or {}
    try:
        from desktop_embed_launch import (
            relaunch_foreground_with_embed_hooks,
            relaunch_path_with_embed_hooks,
        )

        path = (data.get('path') or data.get('exe') or '').strip()
        kill = bool(data.get('kill', True))
        if path:
            result = relaunch_path_with_embed_hooks(path, terminate_pids=None)
        else:
            result = relaunch_foreground_with_embed_hooks(kill=kill)
        status = 200 if result.get('ok') else 400
        return jsonify({'success': bool(result.get('ok')), **result}), status
    except Exception as e:
        uat_logger.log_exception('api_desktop_relaunch_with_embed_hooks', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/inspect', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_inspect():
    """导出前台窗口 UIA 控件树（调试用）。"""
    data = request.get_json(silent=True) or {}
    max_depth = int(data.get('max_depth') or 4)
    max_nodes = int(data.get('max_nodes') or 120)
    try:
        from desktop_automation import sync_desktop_inspect

        nodes = sync_desktop_inspect(max_depth=max_depth, max_nodes=max_nodes)
        return jsonify({'success': True, 'nodes': nodes, 'count': len(nodes)})
    except Exception as e:
        uat_logger.log_exception('api_desktop_inspect', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/pick', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_pick():
    """根据屏幕坐标捕获元素（UIA/Win32/嵌入式 CDP/OCR 混合）。"""
    data = request.get_json(silent=True) or {}
    try:
        x = int(data.get('x'))
        y = int(data.get('y'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '请提供整数坐标 x, y'}), 400
    try:
        from desktop_visual_picker import build_pick_from_smart_click

        pick = build_pick_from_smart_click(x, y, action=str(data.get('action') or 'click'))
        return jsonify({'success': True, 'pick': pick})
    except Exception as e:
        uat_logger.log_exception('api_desktop_pick', e)
        return jsonify({'success': False, 'error': str(e)}), 500


def _parse_desktop_spec_body(data: dict) -> tuple:
    spec = data.get('desktop_spec') or {}
    if isinstance(spec, str):
        try:
            spec = json.loads(spec) if spec.strip() else {}
        except json.JSONDecodeError:
            return None, 'desktop_spec JSON 无效'
    if not isinstance(spec, dict):
        return None, 'desktop_spec 格式无效'
    return spec, None


@app.route('/api/desktop/picker/start', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_picker_start():
    """启动桌面悬浮拾取/录制器（点选控件自动写入定位，无需手写）。"""
    data = request.get_json(silent=True) or {}
    spec, err = _parse_desktop_spec_body(data)
    if err:
        return jsonify({'success': False, 'error': err}), 400
    try:
        from desktop_picker import desktop_picker_available, sync_start_desktop_picker

        if not desktop_picker_available():
            return jsonify({
                'success': False,
                'error': '桌面框选录制需 Windows + opencv-python + mss',
            }), 400
        record_mode = _parse_api_bool(data.get('record_mode'), default=False)
        if not record_mode:
            mode_s = data.get('mode')
            if isinstance(mode_s, bool):
                record_mode = mode_s
            else:
                record_mode = str(mode_s or '').strip().lower() in (
                    '1',
                    'true',
                    'yes',
                    'record',
                    'recording',
                )
        record_action = (data.get('record_action') or data.get('action') or 'click').strip().lower()
        if record_action not in ('click', 'double_click', 'input', 'verify'):
            record_action = 'click'
        result = sync_start_desktop_picker(
            spec or {},
            record_mode=record_mode,
            record_action=record_action,
            input_value=(data.get('input_value') or '').strip(),
            verify_type=(data.get('verify_type') or 'auto').strip().lower() or 'auto',
        )
        status = 200 if result.get('success') else 400
        return jsonify(result), status
    except Exception as e:
        uat_logger.log_exception('api_desktop_picker_start', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/picker/stop', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_picker_stop():
    """关闭桌面拾取/录制悬浮窗。"""
    try:
        from desktop_picker import sync_stop_desktop_picker

        return jsonify(sync_stop_desktop_picker())
    except Exception as e:
        uat_logger.log_exception('api_desktop_picker_stop', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/picker/status', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_picker_status():
    """轮询拾取结果或录制新增步骤（与 Web check_selected_element 类似）。"""
    try:
        from desktop_picker import sync_get_desktop_picker_status

        consume = (request.args.get('consume') or '').strip().lower() in ('1', 'true', 'yes')
        return jsonify(sync_get_desktop_picker_status(consume_last_pick=consume))
    except Exception as e:
        uat_logger.log_exception('api_desktop_picker_status', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/visual/relearn', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_visual_relearn():
    """匹配失败自学习：用户点击屏幕位置，更新步骤 visual 模板。"""
    data = request.get_json(silent=True) or {}
    step_id = data.get('step_id')
    selector_value = (data.get('selector_value') or '').strip()
    try:
        click_x = int(data.get('x'))
        click_y = int(data.get('y'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '请提供整数坐标 x, y'}), 400
    if not selector_value and step_id:
        step = db.get_test_step(int(step_id))
        if not step:
            return jsonify({'success': False, 'error': '步骤不存在'}), 404
        selector_value = (step.get('selector_value') or '').strip()
    if not selector_value:
        return jsonify({'success': False, 'error': '缺少 selector_value 或 step_id'}), 400
    try:
        from desktop_visual_engine import update_visual_template_at_click

        new_json = update_visual_template_at_click(
            selector_value, click_x, click_y, half_size=int(data.get('half_size') or 24)
        )
        if step_id:
            db.update_test_step(
                int(step_id),
                selector_type='visual',
                selector_value=new_json,
            )
        return jsonify({'success': True, 'selector_value': new_json})
    except Exception as e:
        uat_logger.log_exception('api_desktop_visual_relearn', e)
        return jsonify({'success': False, 'error': str(e)}), 500


def _enrich_desktop_spec_from_case(case_id, spec: dict) -> dict:
    """元素捕获未传 desktop_spec 时，从用例已有桌面步骤推断附着窗口。"""
    if not isinstance(spec, dict):
        spec = {}
    if spec.get('hwnd') or spec.get('process') or spec.get('path') or spec.get('window_title'):
        return spec
    if not case_id:
        return spec
    try:
        cid = int(case_id)
    except (TypeError, ValueError):
        return spec
    for st in db.get_case_steps(cid) or []:
        if (st.get('automation_layer') or 'web').strip().lower() != 'desktop':
            continue
        raw = st.get('desktop_spec')
        if not raw:
            continue
        try:
            parsed = json.loads(raw) if isinstance(raw, str) else raw
        except (json.JSONDecodeError, TypeError):
            continue
        if isinstance(parsed, dict) and parsed:
            return parsed
    return spec


def _resolve_element_picker_web_url(data: dict) -> tuple:
    """解析 Web 拾取导航 URL（与 start_visual_selection 一致）。"""
    case_id = data.get('case_id')
    target_url = (data.get('url') or '').strip()
    fixed_url = None
    if target_url:
        fixed_url, url_err = _validate_and_fix_url(target_url)
        if url_err:
            uat_logger.warning(f"元素捕获传入 URL 无效，将尝试按用例解析: {url_err}")
    if not fixed_url and case_id:
        try:
            parsed_case_id = int(case_id)
        except Exception:
            parsed_case_id = None
        if parsed_case_id:
            case_info = db.get_test_case_v2(parsed_case_id)
            case_steps = db.get_case_steps(parsed_case_id)
            resolved_url, _source = _resolve_case_navigation_url(
                case=case_info, case_id=parsed_case_id, steps=case_steps
            )
            if resolved_url:
                fixed_url = resolved_url
    return fixed_url, None


def _persist_desktop_picker_steps_to_case(case_id: int, steps: list) -> list:
    """将捕获器 session 中的桌面步骤直接写入用例（status 轮询时调用，避免前端 POST 丢失）。"""
    if not case_id or not steps:
        return []
    case_row = db.get_test_case_v2(int(case_id))
    if not case_row:
        return []
    pid = case_row.get('project_id')
    if pid and not db.check_project_access(current_user.id, int(pid), 'editor'):
        return []
    saved_ids = []
    for st in steps:
        if not isinstance(st, dict):
            continue
        action = (st.get('action') or 'click').strip()
        layer = 'desktop'
        step_err = _validate_step_action_for_case(case_row, action, layer)
        if step_err:
            continue
        ds = st.get('desktop_spec') or {}
        if not isinstance(ds, str):
            ds = json.dumps(ds, ensure_ascii=False)
        step_id = db.create_test_step(
            int(case_id),
            action,
            st.get('selector_type') or 'automation_id',
            st.get('selector_value') or '',
            st.get('input_value') or '',
            st.get('description') or '',
            None,
            '',
            '',
            '',
            '',
            False,
            '',
            st.get('compare_type') or 'equals',
            st.get('locator_candidates') or '',
            1,
            '',
            automation_layer=layer,
            desktop_spec=ds or '',
        )
        if step_id:
            saved_ids.append(step_id)
    return saved_ids


@app.route('/api/element-picker/start', methods=['POST'])
@app.route('/api/element_picker/start', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_element_picker_start():
    """统一元素捕获：Windows 桌面 + Web 浏览器。"""
    data = request.get_json(silent=True) or {}
    spec, err = _parse_desktop_spec_body(data)
    if err:
        return jsonify({'success': False, 'error': err}), 400
    spec = _enrich_desktop_spec_from_case(data.get('case_id'), spec or {})
    record_mode = _parse_api_bool(data.get('record_mode'), default=False)
    if not record_mode:
        mode_s = (data.get('mode') or '').strip().lower()
        record_mode = mode_s in ('1', 'true', 'yes', 'record', 'recording')
    web_navigate = _parse_api_bool(data.get('web_navigate'), default=False)
    web_attach_existing = _parse_api_bool(
        data.get('web_attach_existing'), default=False
    )
    capture_channel = (
        (data.get('capture_channel') or data.get('channel') or 'desktop')
        .strip()
        .lower()
    )
    if capture_channel not in ('desktop', 'web'):
        capture_channel = 'desktop'
    enable_web = _parse_api_bool(
        data.get('enable_web'),
        default=capture_channel == 'web' or bool(web_navigate or web_attach_existing),
    )
    web_url = ''
    web_fallback_url = ''
    resolved_url, _ = _resolve_element_picker_web_url(data)
    if capture_channel == 'web':
        web_url = ''
        web_fallback_url = ''
        web_navigate = False
        web_attach_existing = True
    elif web_navigate:
        web_url = resolved_url or ''
    elif enable_web:
        web_fallback_url = resolved_url or ''
    try:
        case_id_raw = data.get('case_id')
        try:
            picker_case_id = int(case_id_raw) if case_id_raw not in (None, '') else None
        except (TypeError, ValueError):
            picker_case_id = None
        from element_picker import sync_start_element_picker

        if capture_channel == 'desktop':
            from desktop_runtime import (
                desktop_runtime_available,
                desktop_runtime_unavailable_reason,
            )

            if sys.platform == "win32" and not desktop_runtime_available():
                reason = desktop_runtime_unavailable_reason()
                if reason:
                    return jsonify(
                        {
                            "success": False,
                            "error": reason,
                            "python_executable": sys.executable,
                        }
                    ), 400

        platform_origin = (request.host_url or '').rstrip('/')
        web_capture_mode = (
            (data.get('web_capture_mode') or data.get('mode') or 'cdp').strip().lower()
        )
        if web_capture_mode not in ('cdp', 'extension', 'legacy_inject'):
            web_capture_mode = 'cdp'
        browser = (data.get('browser') or 'edge').strip().lower()
        start_url = (data.get('start_url') or data.get('url') or resolved_url or '').strip()
        result = sync_start_element_picker(
            desktop_spec=spec or {},
            record_mode=record_mode,
            capture_channel=capture_channel,
            web_url=web_url or '',
            web_fallback_url=web_fallback_url or '',
            enable_web=enable_web,
            web_navigate=web_navigate,
            web_attach_existing=web_attach_existing,
            case_id=picker_case_id,
            platform_origin=platform_origin,
            web_capture_mode=web_capture_mode,
            browser=browser,
            start_url=start_url,
        )
        status = 200 if result.get('success') else 400
        return jsonify(result), status
    except Exception as e:
        uat_logger.log_exception('api_element_picker_start', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/element-picker/stop', methods=['POST'])
@app.route('/api/element_picker/stop', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_element_picker_stop():
    """停止统一元素捕获。"""
    try:
        from element_picker import sync_get_element_picker_status, sync_stop_element_picker

        pre = sync_get_element_picker_status(consume_last_pick=False)
        case_id = int(
            pre.get('case_id')
            or (pre.get('desktop') or {}).get('case_id')
            or 0
        )
        new_steps = list(pre.get('new_recorded_steps') or [])
        if not new_steps:
            desk = pre.get('desktop') or {}
            rec = list(desk.get('recorded_steps') or [])
            sent = int(desk.get('_sent_count') or 0)
            if len(rec) > sent:
                new_steps = rec[sent:]
        result = sync_stop_element_picker()
        if case_id and new_steps:
            saved_ids = _persist_desktop_picker_steps_to_case(case_id, new_steps)
            if saved_ids:
                result['saved_step_ids'] = saved_ids
        return jsonify(result)
    except Exception as e:
        uat_logger.log_exception('api_element_picker_stop', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/element-picker/status', methods=['GET'])
@app.route('/api/element_picker/status', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_element_picker_status():
    """轮询统一捕获状态（桌面 + Web）。"""
    try:
        from element_picker import sync_get_element_picker_status

        consume = (request.args.get('consume') or '').strip().lower() in ('1', 'true', 'yes')
        result = sync_get_element_picker_status(consume_last_pick=consume)
        case_id = int(
            result.get('case_id')
            or (result.get('desktop') or {}).get('case_id')
            or 0
        )
        new_steps = list(result.get('new_recorded_steps') or [])
        if not new_steps:
            desk = result.get('desktop') or {}
            new_steps = list(desk.get('new_recorded_steps') or [])
        if case_id and new_steps:
            saved_ids = _persist_desktop_picker_steps_to_case(case_id, new_steps)
            if saved_ids:
                result['saved_step_ids'] = saved_ids
        if case_id and result.get('picker_closed') and not result.get('saved_step_ids'):
            desk = result.get('desktop') or {}
            flush_all = list(result.get('recorded_steps') or desk.get('recorded_steps') or [])
            sent = int(desk.get('_sent_count') or result.get('_sent_count') or 0)
            if len(flush_all) > sent:
                saved_ids = _persist_desktop_picker_steps_to_case(case_id, flush_all[sent:])
                if saved_ids:
                    result['saved_step_ids'] = saved_ids
        return jsonify(result)
    except Exception as e:
        uat_logger.log_exception('api_element_picker_status', e)
        return jsonify({'success': False, 'error': str(e)}), 500


def _web_dom_picker_cors(resp):
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp


# --- 网页捕获（CDP / 扩展，与 desktop API 分离）---


@app.route('/api/web-capture/session/start', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_session_start():
    data = request.get_json(silent=True) or {}
    from web_capture.session import start_session

    mode = (data.get('mode') or data.get('web_capture_mode') or 'cdp').strip().lower()
    result = start_session(
        mode=mode,
        record_mode=_parse_api_bool(data.get('record_mode'), default=False),
        case_id=int(data.get('case_id') or 0) or None,
        platform_origin=(request.host_url or '').rstrip('/'),
        browser=(data.get('browser') or 'edge').strip().lower(),
        start_url=(data.get('start_url') or data.get('url') or '').strip(),
    )
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/web-capture/browser/launch', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_browser_launch():
    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser

    result = cdp_browser.launch_debug_browser(
        browser=(data.get('browser') or 'edge').strip().lower(),
        port=data.get('port'),
        url=(data.get('url') or '').strip(),
    )
    if result.get('success'):
        cdp_browser.connect_playwright_over_cdp(result.get('debug_port'))
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/web-capture/browser/pages', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_browser_pages():
    from web_capture import cdp_browser

    return jsonify(cdp_browser.list_pages())


@app.route('/api/web-capture/browser/attach', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_browser_attach():
    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser

    result = cdp_browser.attach_page(
        page_index=int(data.get('page_index') or 0),
        target_id=(data.get('target_id') or '').strip(),
    )
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/web-capture/browser/chrome-capture', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_browser_chrome_capture():
    data = request.get_json(silent=True) or {}
    from web_capture.session import capture_browser_chrome, validate_session_id

    session_id = (data.get('session_id') or '').strip()
    if not validate_session_id(session_id):
        return jsonify({'success': False, 'error': '捕获会话无效或已结束'}), 404
    target = (data.get('target') or '').strip().lower()
    result = capture_browser_chrome(session_id, target)
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/web-capture/pick', methods=['POST', 'OPTIONS'])
def api_web_capture_pick():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    from web_capture.session import report_pick

    data = request.get_json(silent=True) or {}
    session_id = (
        data.get('session_id') or data.get('session') or request.args.get('session') or ''
    ).strip()
    payload = data.get('payload') if isinstance(data.get('payload'), dict) else data
    result = report_pick(session_id, payload or {})
    return _web_dom_picker_cors(jsonify(result)), (200 if result.get('success') else 400)


@app.route('/api/web-capture/locator/test', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_locator_test():
    import asyncio

    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser
    from web_capture.validator import evaluate_locator_async

    page = cdp_browser.get_active_page()
    if not page:
        conn = cdp_browser.connect_playwright_over_cdp()
        if not conn.get('success'):
            return jsonify({'success': False, 'error': conn.get('error')}), 400
        page = cdp_browser.get_active_page()
    result = asyncio.run(
        evaluate_locator_async(
            page,
            (data.get('selector_type') or 'css').strip(),
            (data.get('selector_value') or '').strip(),
        )
    )
    return jsonify(result)


@app.route('/api/web-capture/verify', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_verify():
    import asyncio

    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser
    from web_capture.validator import verify_element_async

    page = cdp_browser.get_active_page()
    if not page:
        conn = cdp_browser.connect_playwright_over_cdp()
        if not conn.get('success'):
            return jsonify({'success': False, 'error': conn.get('error')}), 400
        page = cdp_browser.get_active_page()
    result = asyncio.run(
        verify_element_async(
            page,
            (data.get('selector_type') or 'css').strip(),
            (data.get('selector_value') or '').strip(),
        )
    )
    return jsonify(result)


@app.route('/api/web-capture/similar', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_similar():
    import asyncio

    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser
    from web_capture.similar import find_similar_elements_async

    page = cdp_browser.get_active_page()
    if not page:
        return jsonify({'success': False, 'error': '无 CDP 页面'}), 400
    result = asyncio.run(
        find_similar_elements_async(
            page,
            (data.get('selector_type') or 'css').strip(),
            (data.get('selector_value') or '').strip(),
        )
    )
    return jsonify(result)


@app.route('/api/web-capture/highlight', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_highlight():
    data = request.get_json(silent=True) or {}
    from web_capture import cdp_browser
    from web_capture.cdp_picker import arm_picker, inject_all_frames

    page = cdp_browser.get_active_page()
    if not page:
        return jsonify({'success': False, 'error': '无 CDP 页面'}), 400
    sid = (data.get('session_id') or '').strip()
    api_base = (request.host_url or '').rstrip('/')
    inject_all_frames(page, api_base=api_base, session_id=sid)
    return jsonify(arm_picker(page))


@app.route('/api/web-capture/extension/status', methods=['GET'])
@login_required
@api_error_handler
def api_web_capture_extension_status():
    from web_capture.extension_bridge import get_extension_status

    return jsonify(get_extension_status())


@app.route('/api/web-capture/extension/install', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_web_capture_extension_install():
    import subprocess
    import sys

    root = os.path.dirname(os.path.abspath(__file__))
    script = os.path.join(root, 'tools', 'install_web_extension.py')
    if not os.path.isfile(script):
        return jsonify({'success': False, 'error': '安装脚本不存在'}), 404
    try:
        proc = subprocess.run(
            [sys.executable, script, '--prepare'],
            capture_output=True,
            text=True,
            timeout=120,
            cwd=root,
        )
        return jsonify(
            {
                'success': proc.returncode == 0,
                'stdout': proc.stdout[-2000:] if proc.stdout else '',
                'stderr': proc.stderr[-2000:] if proc.stderr else '',
            }
        )
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@app.route('/api/web-capture/highlight.js', methods=['GET'])
def api_web_capture_highlight_js():
    from web_capture.cdp_picker import get_highlight_js
    from web_capture.session import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(
            Response('// invalid session\n', status=404, mimetype='application/javascript')
        )
    api_base = (request.host_url or '').rstrip('/')
    page_only = (request.args.get('page_only') or '').strip() in ('1', 'true', 'yes')
    body = get_highlight_js(api_base, session_id, page_only=page_only)
    return _web_dom_picker_cors(
        Response(body, mimetype='application/javascript; charset=utf-8')
    )


@app.route('/api/web-capture/toolbar.js', methods=['GET'])
def api_web_capture_toolbar_js():
    from web_capture.cdp_picker import get_browser_toolbar_js
    from web_capture.session import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(
            Response('// invalid session\n', status=404, mimetype='application/javascript')
        )
    api_base = (request.host_url or '').rstrip('/')
    body = get_browser_toolbar_js(api_base, session_id)
    return _web_dom_picker_cors(
        Response(body, mimetype='application/javascript; charset=utf-8')
    )


@app.route('/api/web-capture/arm', methods=['POST', 'OPTIONS'])
def api_web_capture_arm():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    data = request.get_json(silent=True) or {}
    from web_capture.session import arm_session, validate_session_id

    session_id = (data.get('session_id') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(jsonify({'success': False, 'error': '捕获会话无效或已结束'})), 404
    api_base = (request.host_url or '').rstrip('/')
    browser = (data.get('browser') or '').strip().lower()
    result = arm_session(session_id, api_base=api_base, browser=browser)
    return _web_dom_picker_cors(jsonify(result)), (200 if result.get('success') else 400)


@app.route('/api/web-capture/disarm', methods=['POST', 'OPTIONS'])
def api_web_capture_disarm():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    data = request.get_json(silent=True) or {}
    from web_capture.session import disarm_session, validate_session_id

    session_id = (data.get('session_id') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(jsonify({'success': False, 'error': '捕获会话无效或已结束'})), 404
    result = disarm_session(session_id)
    return _web_dom_picker_cors(jsonify(result)), (200 if result.get('success') else 400)


@app.route('/api/web-capture/stop', methods=['POST', 'OPTIONS'])
def api_web_capture_stop():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    data = request.get_json(silent=True) or {}
    from web_capture.session import stop_session, validate_session_id

    session_id = (data.get('session_id') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(jsonify({'success': False, 'error': '捕获会话无效或已结束'})), 404
    stop_session(fast=True)
    try:
        from element_picker import sync_stop_element_picker

        sync_stop_element_picker()
    except Exception:
        pass
    return _web_dom_picker_cors(jsonify({'success': True, 'message': '网页捕获已结束'}))


@app.route('/api/web-capture/arm-status', methods=['GET'])
def api_web_capture_arm_status():
    from web_capture.session import validate_session_id, get_session_debug_snapshot

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(jsonify({'success': False, 'armed': False})), 404
    snap = get_session_debug_snapshot()
    return _web_dom_picker_cors(jsonify({'success': True, 'armed': bool(snap.get('armed'))}))


@app.route('/api/web-dom-picker/inject.js', methods=['GET'])
def api_web_dom_picker_inject_js():
    """目标页加载的捕获脚本（凭 session 校验，无需登录 Cookie）。"""
    from pathlib import Path

    from web_dom_picker import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return _web_dom_picker_cors(
            Response('// invalid or expired web dom picker session\n', status=404, mimetype='application/javascript')
        )
    js_path = Path(__file__).resolve().parent / 'static' / 'js' / 'web_dom_picker_inject.js'
    try:
        raw = js_path.read_text(encoding='utf-8')
    except OSError:
        return _web_dom_picker_cors(
            Response('// inject script missing\n', status=500, mimetype='application/javascript')
        )
    api_base = (request.host_url or '').rstrip('/')
    body = raw.replace('__API_BASE__', api_base).replace('__SESSION__', session_id)
    return _web_dom_picker_cors(
        Response(body, mimetype='application/javascript; charset=utf-8')
    )


@app.route('/api/web-dom-picker/pick', methods=['POST', 'OPTIONS'])
def api_web_dom_picker_pick():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    from web_dom_picker import sync_report_web_dom_pick

    data = request.get_json(silent=True) or {}
    session_id = (
        data.get('session_id') or data.get('session') or request.args.get('session') or ''
    ).strip()
    payload = data.get('payload') if isinstance(data.get('payload'), dict) else data
    result = sync_report_web_dom_pick(session_id, payload or {})
    return _web_dom_picker_cors(jsonify(result)), (200 if result.get('success') else 400)


@app.route('/api/web-dom-picker/close', methods=['POST', 'OPTIONS'])
def api_web_dom_picker_close():
    if request.method == 'OPTIONS':
        return _web_dom_picker_cors(Response('', status=204))
    from web_dom_picker import close_web_dom_picker_session

    data = request.get_json(silent=True) or {}
    session_id = (data.get('session') or '').strip()
    result = close_web_dom_picker_session(session_id)
    return _web_dom_picker_cors(jsonify(result)), (200 if result.get('success') else 400)


def _web_capture_proxy_fetch_html(target_url: str, session_id: str, api_base: str) -> tuple:
    """拉取待测页 HTML 并注入捕获脚本（同源代理，供预览 iframe 内高亮拾取）。"""
    import re
    from html import escape as html_escape
    from urllib.parse import urlparse
    from urllib.request import Request, urlopen

    parsed = urlparse(target_url)
    if parsed.scheme not in ('http', 'https'):
        return None, '仅支持 http/https 地址'
    req = Request(
        target_url,
        headers={'User-Agent': 'UAT-WebCapture/1.0', 'Accept': 'text/html,*/*'},
    )
    try:
        with urlopen(req, timeout=25) as resp:
            raw = resp.read()
            ctype = (resp.headers.get('Content-Type') or '').lower()
    except Exception as exc:
        return None, f'无法加载页面: {exc}'

    charset = 'utf-8'
    m = re.search(r'charset=([^\s;]+)', ctype)
    if m:
        charset = m.group(1).strip('\'"')
    try:
        html = raw.decode(charset, errors='replace')
    except LookupError:
        html = raw.decode('utf-8', errors='replace')

    base_href = html_escape(target_url, quote=True)
    api_base = (api_base or '').rstrip('/')
    inject = (
        f'<base href="{base_href}">\n'
        f'<script src="{api_base}/api/web-dom-picker/inject.js?session={html_escape(session_id, quote=True)}"></script>\n'
    )
    low = html.lower()
    if '</head>' in low:
        html = re.sub(r'</head>', inject + '</head>', html, count=1, flags=re.IGNORECASE)
    elif '</body>' in low:
        html = re.sub(r'</body>', inject + '</body>', html, count=1, flags=re.IGNORECASE)
    else:
        html = inject + html
    return html, None


@app.route('/web-capture/workspace')
@login_required
def web_capture_workspace():
    """网页捕获器工作区（独立窗口，内含预览与 DOM 拾取）。"""
    from web_dom_picker import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return render_template(
            'web_capture_workspace.html',
            error='捕获会话无效或已结束。请返回步骤页重新点击「网页捕获」。',
        )
    initial_url = (request.args.get('url') or '').strip()
    case_id = request.args.get('case_id')
    if not initial_url and case_id:
        try:
            initial_url, _ = _resolve_element_picker_web_url({'case_id': case_id})
        except Exception:
            initial_url = ''
    return render_template(
        'web_capture_workspace.html',
        error='',
        session_id=session_id,
        api_base=(request.host_url or '').rstrip('/'),
        initial_url=initial_url or '',
    )


@app.route('/web-capture/shell')
def web_capture_shell():
    """捕获浏览器启动页（独立配置无平台 Cookie，凭 session 校验）。"""
    from web_capture.session import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    api_base = (request.host_url or '').rstrip('/')
    if not validate_session_id(session_id):
        return (
            '<!DOCTYPE html><html><body style="font-family:sans-serif;padding:24px;">'
            '<p>捕获会话无效或已结束，请返回步骤页重新点击「网页捕获」。</p></body></html>',
            404,
        )
    return render_template(
        'web_capture_shell.html',
        session_id=session_id,
        api_base=api_base,
    )


@app.route('/web-capture/toolbar')
@login_required
def web_capture_toolbar():
    """网页元素捕获器（唯一控制窗口，对齐桌面 Tk 工具条）。"""
    from web_capture.session import validate_session_id, get_session_debug_snapshot, get_session_status

    session_id = (request.args.get('session') or '').strip()
    api_base = (request.host_url or '').rstrip('/')
    if not validate_session_id(session_id):
        return render_template(
            'web_capture_toolbar.html',
            error='捕获会话无效或已结束。请返回步骤页重新点击「网页捕获」。',
            session_id='',
            api_base=api_base,
            bookmarklet='',
        )
    snap = get_session_debug_snapshot()
    status = get_session_status()
    return render_template(
        'web_capture_toolbar.html',
        error='',
        session_id=session_id,
        api_base=api_base,
        bookmarklet=status.get('bookmarklet') or '',
        mode=snap.get('mode') or 'extension',
    )


@app.route('/web-capture/workspace-v2')
@login_required
def web_capture_workspace_v2():
    """CDP 网页捕获工作区（调试浏览器 + 状态面板）。"""
    from web_capture.session import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    if not validate_session_id(session_id):
        return render_template(
            'web_capture_workspace_v2.html',
            error='捕获会话无效或已结束。请返回步骤页重新点击「网页捕获」。',
            session_id='',
            api_base='',
            cdp_port=0,
        )
    from web_capture.session import get_session_debug_snapshot

    snap = get_session_debug_snapshot()
    return render_template(
        'web_capture_workspace_v2.html',
        error='',
        session_id=session_id,
        api_base=(request.host_url or '').rstrip('/'),
        cdp_port=int(snap.get('cdp_port') or 0),
        mode=snap.get('mode') or 'cdp',
    )


@app.route('/web-capture/proxy')
def web_capture_proxy():
    """捕获预览代理：注入拾取脚本，使 iframe 内可悬停高亮。"""
    from web_dom_picker import validate_session_id

    session_id = (request.args.get('session') or '').strip()
    target_url = (request.args.get('url') or '').strip()
    if not validate_session_id(session_id):
        return Response('捕获会话无效', status=403)
    if not target_url:
        return Response('缺少 url 参数', status=400)
    fixed, err = _validate_and_fix_url(target_url)
    if err or not fixed:
        return Response(err or 'URL 无效', status=400)
    api_base = (request.host_url or '').rstrip('/')
    html, fetch_err = _web_capture_proxy_fetch_html(fixed, session_id, api_base)
    if fetch_err:
        return Response(fetch_err, status=502, mimetype='text/plain; charset=utf-8')
    return Response(html, mimetype='text/html; charset=utf-8')


@app.route('/api/desktop/config', methods=['GET'])
@login_required
@api_error_handler
def api_desktop_config():
    """返回 .env 中的桌面默认配置，供步骤编辑页预填与提示。"""
    try:
        from desktop_env_config import public_config

        return jsonify({'success': True, **public_config()})
    except ImportError:
        return jsonify({'success': False, 'error': '桌面模块未安装'}), 500


@app.route('/api/desktop/windows', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_windows():
    """枚举本机当前可见顶层窗口（零配置附着窗口）。"""
    from desktop_discovery import discovery_available, list_visible_windows

    if not discovery_available():
        return jsonify({
            'success': False,
            'error': '窗口枚举仅支持 Windows',
        }), 400
    return jsonify({'success': True, 'windows': list_visible_windows()})


@app.route('/api/desktop/processes', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_processes():
    """枚举本机当前运行进程。"""
    from desktop_discovery import discovery_available, list_running_processes

    if not discovery_available():
        return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
    return jsonify({'success': True, 'processes': list_running_processes()})


@app.route('/api/desktop/snapshot', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_snapshot():
    """运行中窗口 + 进程 + 开始菜单应用目录（供步骤编辑点选）。"""
    from desktop_discovery import desktop_runtime_snapshot, discovery_available

    if not discovery_available():
        return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
    refresh = (request.args.get('refresh') or '').strip().lower() in ('1', 'true', 'yes')
    if refresh:
        try:
            from desktop_app_catalog import ensure_catalog_built

            ensure_catalog_built(force=True)
        except Exception:
            pass
    snap = desktop_runtime_snapshot(include_catalog=True)
    return jsonify({'success': True, **snap})


@app.route('/api/desktop/catalog', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_catalog():
    """本机应用目录（开始菜单扫描，含自动别名）。"""
    from desktop_discovery import discovery_available

    if not discovery_available():
        return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
    try:
        from desktop_app_catalog import ensure_catalog_built, list_catalog_apps

        data = ensure_catalog_built(
            force=(request.args.get('refresh') or '').strip().lower() in ('1', 'true', 'yes')
        )
        return jsonify({
            'success': True,
            'apps': list_catalog_apps(),
            'built_at': data.get('built_at'),
            'app_count': len(data.get('apps') or []),
        })
    except Exception as e:
        uat_logger.log_exception('api_desktop_catalog', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/catalog/refresh', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_catalog_refresh():
    """强制重新扫描开始菜单并更新别名库。"""
    from desktop_discovery import discovery_available

    if not discovery_available():
        return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
    try:
        from desktop_app_catalog import ensure_catalog_built
        from desktop_discovery import invalidate_discovery_cache

        invalidate_discovery_cache()
        data = ensure_catalog_built(force=True)
        return jsonify({
            'success': True,
            'app_count': len(data.get('apps') or []),
            'built_at': data.get('built_at'),
        })
    except Exception as e:
        uat_logger.log_exception('api_desktop_catalog_refresh', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/window-spec', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_window_spec():
    """根据 hwnd 生成 attach_window 用的 desktop_spec（写入步骤，不依赖 .env）。"""
    data = request.get_json(silent=True) or {}
    try:
        hwnd = int(data.get('hwnd'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '请提供整数 hwnd'}), 400
    try:
        from desktop_discovery import attachment_spec_for_window, discovery_available

        if not discovery_available():
            return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
        spec, title = attachment_spec_for_window(hwnd)
        return jsonify({
            'success': True,
            'desktop_spec': spec,
            'suggested_input': title,
            'process': spec.get('process') or '',
        })
    except Exception as e:
        uat_logger.log_exception('api_desktop_window_spec', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/resolve-app', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
def api_desktop_resolve_app():
    """按程序名解析本机 exe 路径（PATH / App Paths）。"""
    name = (request.args.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': '请提供 name 参数'}), 400
    try:
        from desktop_discovery import (
            discovery_available,
            format_resolve_error,
            resolve_executable_with_meta,
        )

        if not discovery_available():
            return jsonify({'success': False, 'error': '仅支持 Windows'}), 400
        meta = resolve_executable_with_meta(name)
        if not meta.found:
            return jsonify({
                'success': False,
                'error': format_resolve_error(meta),
                'tried': meta.tried,
                'suggestions': meta.suggestions,
            }), 404
        return jsonify({
            'success': True,
            'path': meta.path,
            'name': name,
            'method': meta.method,
            'tried': meta.tried,
        })
    except Exception as e:
        uat_logger.log_exception('api_desktop_resolve_app', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/desktop/machines', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_desktop_machines():
    """远程桌面 Agent 测试机注册（Phase 3）；本地版 DEPLOYMENT_PROFILE=local 默认禁用。"""
    try:
        from desktop_env_config import is_local_deployment

        if is_local_deployment():
            return jsonify({
                'success': False,
                'error': '本地部署模式不提供远程测试机注册；请在用户本机执行桌面步骤（inprocess）。',
                'deployment_profile': 'local',
            }), 403
    except ImportError:
        pass
    if request.method == 'GET':
        machines = db.list_test_machines()
        return jsonify({'success': True, 'machines': machines})
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    agent_url = (data.get('agent_url') or '').strip().rstrip('/')
    if not name or not agent_url:
        return jsonify({'success': False, 'error': 'name 与 agent_url 必填'}), 400
    mid = db.create_test_machine(
        name,
        agent_url,
        (data.get('os_version') or '').strip(),
        (data.get('secret') or '').strip(),
    )
    return jsonify({'success': True, 'machine_id': mid})


_PLAYWRIGHT_BROWSER_OPTIONS = [
    {"id": "chromium", "label_zh": "Chromium（Playwright 内置）", "label_en": "Chromium (bundled)"},
    {"id": "chrome", "label_zh": "Google Chrome", "label_en": "Google Chrome"},
    {"id": "edge", "label_zh": "Microsoft Edge", "label_en": "Microsoft Edge"},
    {"id": "firefox", "label_zh": "Firefox", "label_en": "Firefox"},
    {"id": "webkit", "label_zh": "WebKit", "label_en": "WebKit"},
]


@app.route('/api/playwright/browser', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_playwright_browser():
    """
    主自动化会话使用的 Playwright 浏览器（Chromium / Chrome / Edge 等）。
    偏好保存在 Flask session；未保存时遵循环境变量 PLAYWRIGHT_BROWSER。
    使用系统 Chrome/Edge 前请在部署环境执行: playwright install chrome / playwright install msedge
    """
    if request.method == 'GET':
        raw = session.get('playwright_browser_engine') or os.environ.get('PLAYWRIGHT_BROWSER') or 'chromium'
        cur = normalize_playwright_browser_name(str(raw))
        return jsonify({
            'success': True,
            'current': cur,
            'options': _PLAYWRIGHT_BROWSER_OPTIONS,
            'env_default': (os.environ.get('PLAYWRIGHT_BROWSER') or '').strip() or 'chromium',
            'running_engine': automation.get_browser_engine(),
        })
    data = request.get_json(silent=True) or {}
    b = (data.get('browser') or data.get('engine') or '').strip()
    if not b:
        return jsonify({'success': False, 'error': 'browser 不能为空'}), 400
    norm = normalize_playwright_browser_name(b)
    session['playwright_browser_engine'] = norm
    session.modified = True
    automation.set_browser_engine(norm)
    return jsonify({
        'success': True,
        'current': norm,
        'hint': '已保存。若自动化浏览器已在运行，下次打开/导航时会切换引擎；也可关闭浏览器窗口后重试。',
    })


@app.route('/api/embedded-browser/status', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_embedded_browser_status():
    """内嵌画布 Chromium 已废弃；保留接口以免旧客户端 404，固定返回不可用。"""
    return jsonify({
        'success': True,
        'enabled': False,
        'deprecated': True,
        'reachable': False,
        'message': '内嵌画布已废弃，请使用本机 Edge/Chrome（/api/web-capture/browser/launch）',
    })


@app.route('/api/embedded-browser/session', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_embedded_browser_session_create():
    """内嵌画布已废弃。"""
    return jsonify({
        'success': False,
        'deprecated': True,
        'error': '内嵌画布 Chromium 已废弃，请改用本机浏览器：POST /api/web-capture/browser/launch',
    }), 410


@app.route("/api/embedded-browser/session/<session_id>/navigate", methods=["POST"])
@login_required
@role_required("admin", "tester", "project_manager", "test_lead")
@api_error_handler
@log_api_request
def api_embedded_browser_session_navigate(session_id: str):
    """内嵌画布已废弃。"""
    return jsonify({
        'success': False,
        'deprecated': True,
        'error': '内嵌画布已废弃，请使用本机浏览器',
    }), 410


@app.route('/api/embedded-browser/session/<session_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_embedded_browser_session_delete(session_id: str):
    """内嵌画布已废弃。"""
    return jsonify({'success': True, 'deprecated': True, 'message': '内嵌画布已废弃，无需删除会话'})


@app.route('/api/embedded-browser/session/<session_id>/inspect', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_embedded_browser_session_inspect(session_id: str):
    """内嵌画布已废弃。"""
    return jsonify({
        'success': False,
        'deprecated': True,
        'error': '内嵌画布已废弃',
    }), 410


@app.route('/api/embedded-browser/session/<session_id>/diagnostics', methods=['GET'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_embedded_browser_session_diagnostics(session_id: str):
    return jsonify({
        'success': False,
        'deprecated': True,
        'error': '内嵌画布已废弃',
    }), 410


@app.route('/api/browser/viewport', methods=['GET'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_viewport():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    v = sync_get_viewport_size()
    return jsonify({'success': True, 'viewport': v})


@app.route('/api/browser/click', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_click():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    data = request.get_json(silent=True) or {}
    try:
        x = float(data.get('x'))
        y = float(data.get('y'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '需要有效的 x, y（视口坐标，与截图像素一致）'}), 400
    btn = (data.get('button') or 'left')
    dbl = bool(data.get('double'))
    try:
        sync_browser_mouse_click(x, y, button=btn, click_count=2 if dbl else 1)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True})


@app.route('/api/browser/wheel', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_wheel():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    data = request.get_json(silent=True) or {}
    try:
        sync_browser_mouse_wheel(float(data.get('delta_x', 0)), float(data.get('delta_y', 0)))
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True})


@app.route('/api/browser/type', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_type_text():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    data = request.get_json(silent=True) or {}
    t = (data.get('text') or '')
    try:
        sync_browser_keyboard_type(t)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True})


@app.route('/api/browser/key', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_key_press():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    data = request.get_json(silent=True) or {}
    k = (data.get('key') or 'Enter')
    try:
        sync_browser_keyboard_press(k)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True})


@app.route('/api/browser/pick', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
def api_browser_pick_element():
    if not sync_automation_session_usable():
        return jsonify({'success': False, 'error': '浏览器未启动'}), 400
    data = request.get_json(silent=True) or {}
    try:
        x = float(data.get('x'))
        y = float(data.get('y'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '需要有效的 x, y'}), 400
    try:
        el = sync_element_info_at_point(x, y)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    return jsonify({'success': True, 'element': el})


@app.route('/api/browser/ai-analyze', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_browser_ai_analyze():
    """根据当前已打开页的可交互结构生成用例（写入侧栏/工作台预览由前端处理）。"""
    data = request.get_json(silent=True) or {}
    embedded_sid = (data.get('embedded_session_id') or data.get('remote_session_id') or '').strip()
    hint = (data.get('hint') or '').strip()
    project_name = (data.get('project_name') or '').strip()
    selected_model = (data.get('model') or '').strip() or _get_active_local_model()
    profile, legacy_model = _resolve_inference_profile(selected_model)

    if embedded_sid:
        if not embedded_gateway_enabled():
            return jsonify({'success': False, 'error': '远程 Chromium 网关未配置'}), 503
        j, err = embedded_gateway_json(
            'GET',
            f'/internal/session/{embedded_sid}/inspect',
            user_id=current_user.id,
        )
        if not j or not j.get('success'):
            detail = (j or {}).get('detail')
            return jsonify({
                'success': False,
                'error': str(err or detail or '远程会话无效或已过期，请重新连接「远程画布」'),
            }), 502
        snap = j.get('data') or {}
        items = snap.get('items') or []
        page_body = {
            'url': snap.get('url', ''),
            'title': snap.get('title', ''),
            'viewport': snap.get('viewport', {}),
            'items': items,
            'source': snap.get('source') or 'embedded_gateway',
        }
    else:
        if not sync_automation_session_usable():
            return jsonify({'success': False, 'error': '浏览器未启动，请先打开页面或使用远程画布'}), 400
        try:
            snap = sync_get_interactive_page_snapshot(120)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
        items = snap.get('items') or []
        page_body = {
            'url': snap.get('url', ''),
            'title': snap.get('title', ''),
            'viewport': snap.get('viewport', {}),
            'items': items,
        }

    ctx_label = (
        '用户已在远程 Chromium（服务端嵌入式网关）中打开的当前页面'
        if embedded_sid
        else '用户已在内置浏览器/主 Playwright 会话中打开的当前页面'
    )
    from ai_page_probe import probe_registry_from_interactive_snapshot

    page_snapshot_txt, probe_registry, probe_pu = probe_registry_from_interactive_snapshot(snap)
    goal_lines = [
        f'请根据以下「{ctx_label}」生成可执行 Web 用例与步骤。',
        f"页面标题: {page_body['title']}",
        f"URL: {page_body['url']}",
    ]
    if hint:
        goal_lines.append('用户补充的测试目标: ' + hint)
    goal = '\n'.join(goal_lines)
    probe_url = (probe_pu or page_body.get('url') or '').strip() or None
    dpack = _ai_build_dom_pack(snap, embed_remote=bool(embedded_sid))
    mem_ctx = _ai_memory_context_block(
        current_user.id, goal, probe_url=probe_url or page_body.get('url') or '', project_name=project_name
    )
    try:
        generated = local_ai_service.generate_case_and_steps(
            goal,
            project_name,
            model=legacy_model,
            profile=profile,
            page_snapshot=page_snapshot_txt or None,
            probe_registry=probe_registry if probe_registry else None,
            probe_url=probe_url,
            memory_context=mem_ctx or None,
            dom_context_pack=dpack or None,
        )
    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'hint': '可先执行: ollama serve，并确认模型已拉取。',
        }), 503
    generated, norm_warnings = apply_step_normalization_to_plan(generated)
    generated, norm_warnings = _merge_ai_locator_resolution(generated, snap, norm_warnings)
    log_ai_plan_to_audit(
        current_user.id,
        current_user.username,
        'AI_PLAN_PAGE_GENERATE',
        generated,
        request.remote_addr,
    )
    return jsonify({
        'success': True,
        'plan': generated,
        'page_snapshot': page_body,
        'warnings': norm_warnings,
    })


# API: 启用元素选择模式（已废弃，转发提示）
@app.route('/api/enable_element_selection', methods=['POST'])
@api_error_handler
@log_api_request
def api_enable_element_selection():
    return _legacy_visual_picker_removed()

# API: 禁用元素选择模式（已废弃）
@app.route('/api/disable_element_selection', methods=['POST'])
@api_error_handler
@log_api_request
def api_disable_element_selection():
    return _legacy_visual_picker_removed()

# API: 获取选中的元素信息（已废弃）
@app.route('/api/get_selected_element', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_selected_element():
    return _legacy_visual_picker_removed()

# API: 从选定元素提取JSON数据
@app.route('/api/extract_json_from_selected_element', methods=['GET'])
@api_error_handler
@log_api_request
def api_extract_json_from_selected_element():
    try:
        json_data = sync_extract_json_from_selected_element()
        return jsonify({'success': True, 'json_data': json_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== 项目管理API ====================

# API: 创建项目
@app.route('/api/projects', methods=['POST'])
@login_required
@role_required('admin', 'tester')
@api_error_handler
@log_api_request
@audit_log('CREATE_PROJECT', 'project')
def api_create_project():
    data = request.get_json(silent=True) or {}
    name = data.get('name', '')
    description = data.get('description', '')

    if not name:
        return jsonify({'error': '项目名称不能为空'}), 400

    # 检查项目数量限制
    _db = Database()
    user_projects = _db.get_user_projects(current_user.id)
    limits = license_manager.get_limits()
    if limits['max_projects'] != -1 and len(user_projects) >= limits['max_projects']:
        return jsonify({
            'success': False,
            'error': f'已达到项目数量限制（{limits["max_projects"]}个）。请升级到企业版。'
        }), 403

    project_id = db.create_project(name, description)

    # 将创建者添加为项目所有者
    _db.add_project_member(project_id, current_user.id, role='owner')

    return jsonify({'success': True, 'project_id': project_id})

# API: 获取所有项目
@app.route('/api/projects', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_projects():
    # 只返回用户有权限的项目
    _db = Database()
    projects = _db.get_user_projects(current_user.id)
    return jsonify({'projects': projects})

# API: 获取单个项目
@app.route('/api/projects/<int:project_id>', methods=['GET'])
@login_required
@project_access_required(min_role='viewer')
@api_error_handler
@log_api_request
def api_get_project(project_id):
    project = db.get_project(project_id)
    if not project:
        return jsonify({'error': '项目不存在'}), 404
    return jsonify({'project': project})

# API: 更新项目
@app.route('/api/projects/<int:project_id>', methods=['PUT'])
@login_required
@project_access_required(min_role='editor')
@api_error_handler
@log_api_request
@audit_log('UPDATE_PROJECT', 'project')
def api_update_project(project_id):
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    description = data.get('description')

    success = db.update_project(project_id, name, description)

    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '更新项目失败'}), 400

# API: 删除项目
@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
@login_required
@project_access_required(min_role='owner')
@api_error_handler
@log_api_request
@audit_log('DELETE_PROJECT', 'project')
def api_delete_project(project_id):
    success = db.delete_project(project_id)

    if success:
        try:
            db.prune_orphan_run_history()
        except Exception:
            pass
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '删除项目失败'}), 400

def _app_case_type(case_dict) -> str:
    if not case_dict:
        return "ui"
    t = (case_dict.get("case_type") or "ui").strip().lower()
    return "api" if t == "api" else "ui"


def _validate_step_action_for_case(case_dict, action: str, automation_layer: str = "web"):
    """服务端步骤写入校验：Web 用例禁止 api_request；接口用例仅允许 api_request；按层校验动作。"""
    if not case_dict:
        return "用例不存在"
    act = (action or "").strip()
    ct = _app_case_type(case_dict)
    if ct == "ui" and act == "api_request":
        return "Web 用例不允许添加接口步骤，请在「接口测试」模块中维护接口用例"
    if ct == "api" and act != "api_request":
        return "接口用例仅允许「接口请求」步骤（api_request）"
    if ct == "ui":
        try:
            from desktop_automation import validate_step_for_layer

            layer = (automation_layer or "web").strip().lower()
            layer_err = validate_step_for_layer(act, layer)
            if layer_err:
                return layer_err
        except ImportError:
            pass
    return None


# API: 获取项目下的所有测试用例
@app.route('/api/projects/<int:project_id>/cases', methods=['GET'])
@login_required
@project_access_required(min_role='viewer')
@api_error_handler
@log_api_request
def api_get_project_cases(project_id):
    raw_ct = (request.args.get("case_type") or "ui").strip().lower()
    case_type = "api" if raw_ct == "api" else "ui"
    unit_id = request.args.get("unit_id")
    page_raw = request.args.get("page", type=int)
    page_size_raw = request.args.get("page_size", type=int)
    if page_raw is not None or page_size_raw is not None:
        page = max(1, page_raw or 1)
        page_size = max(1, min(page_size_raw or 10, 100))
        cases, total = db.get_project_cases_paginated(
            project_id,
            case_type=case_type,
            unit_id=unit_id,
            page=page,
            page_size=page_size,
        )
        total_pages = max(1, (total + page_size - 1) // page_size) if total else 1
        return jsonify({
            'cases': cases,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
        })
    cases = db.get_project_cases(project_id, case_type=case_type, unit_id=unit_id)
    return jsonify({'cases': cases})


@app.route('/api/projects/<int:project_id>/units', methods=['GET', 'POST'])
@login_required
@project_access_required(min_role='viewer')
@api_error_handler
@log_api_request
def api_project_test_units(project_id):
    _db = Database()
    if request.method == 'GET':
        units = _db.get_test_units(project_id)
        ungrouped = _db.get_project_cases(project_id, unit_id='ungrouped')
        return jsonify({
            'success': True,
            'units': units,
            'ungrouped_case_count': len(ungrouped),
        })
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': '单元名称不能为空'}), 400
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限'}), 403
    uid = _db.create_test_unit(
        project_id,
        name,
        description=(data.get('description') or '').strip(),
        sort_order=int(data.get('sort_order') or 0),
    )
    return jsonify({'success': True, 'unit_id': uid})


@app.route('/api/projects/<int:project_id>/units/<int:unit_id>', methods=['PUT', 'DELETE'])
@login_required
@project_access_required(min_role='editor')
@api_error_handler
@log_api_request
def api_project_test_unit_detail(project_id, unit_id):
    _db = Database()
    unit = _db.get_test_unit(unit_id)
    if not unit or int(unit.get('project_id') or 0) != int(project_id):
        return jsonify({'success': False, 'error': '单元不存在'}), 404
    if request.method == 'DELETE':
        case_count = int(unit.get('case_count') or 0)
        if not case_count:
            u = _db.get_test_unit(unit_id)
            if u:
                cases = _db.get_project_cases(project_id, unit_id=unit_id)
                case_count = len(cases)
        ok = _db.delete_test_unit(unit_id)
        return jsonify({'success': ok, 'ungrouped_cases': case_count})
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    if name is not None and not str(name).strip():
        return jsonify({'success': False, 'error': '单元名称不能为空'}), 400
    ok = _db.update_test_unit(
        unit_id,
        name=name,
        description=data.get('description'),
        sort_order=data.get('sort_order'),
    )
    if not ok:
        return jsonify({'success': False, 'error': '更新失败或无可更新字段'}), 400
    return jsonify({'success': ok, 'unit': _db.get_test_unit(unit_id)})


# ==================== 测试用例管理API（新版本） ====================

# API: 创建测试用例（关联到项目）
@app.route('/api/cases', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_CASE', 'case')
def api_create_case_v2():
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    name = data.get('name', '')
    url = data.get('url', '')
    description = data.get('description', '')
    precondition = data.get('precondition', '')
    expected_result = data.get('expected_result', '')
    case_type_raw = (data.get('case_type') or 'ui').strip().lower()
    case_type = 'api' if case_type_raw == 'api' else 'ui'
    platform_raw = (data.get('platform') or 'web').strip().lower()
    if platform_raw in ('android', 'mobile', 'app'):
        platform = 'android'
    elif platform_raw == 'desktop':
        platform = 'desktop'
    else:
        platform = 'web'
    
    if not project_id:
        return jsonify({'error': '项目ID不能为空'}), 400
    if not name:
        return jsonify({'error': '用例名称不能为空'}), 400
    
    # 检查项目访问权限
    _db = Database()
    if not _db.check_project_access(current_user.id, project_id, 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403
    
    # ===== 用例数量限制检查 =====
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    
    # 获取项目当前用例数量（仅 COUNT，避免全量拉取用例与步骤 JOIN）
    current_case_count = _db.get_project_case_count(project_id)
    
    if limits['max_cases_per_project'] != -1 and current_case_count >= limits['max_cases_per_project']:
        return jsonify({
            'success': False,
            'error': f'已达到项目用例数量限制（{limits["max_cases_per_project"]}个）。请升级至团队版或企业版以提升配额。',
            'limit_reached': True,
            'current_count': current_case_count,
            'limit': limits['max_cases_per_project'],
            'upgrade_url': '/license'
        }), 403
    
    # 记录创建用例计数（仅免费版）
    if license_info.license_type == LicenseType.FREE.value:
        _db.increment_created_cases(current_user.id)
    
    unit_id_raw = data.get('unit_id')
    unit_id = int(unit_id_raw) if unit_id_raw not in (None, '', 0, '0') else None

    case_id = db.create_test_case_v2(
        project_id, name, url, description, precondition, expected_result,
        case_type=case_type, platform=platform, unit_id=unit_id,
    )
    return jsonify({'success': True, 'case_id': case_id, 'platform': platform})

# API: 获取测试用例详情（新版本）
@app.route('/api/cases/<int:case_id>', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_get_case_v2(case_id):
    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404
    
    # 检查项目访问权限
    _db = Database()
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'viewer'):
        return jsonify({'success': False, 'error': '无权限访问此用例'}), 403
    
    return jsonify({'test_case': case})

# API: 更新测试用例（新版本）
@app.route('/api/cases/<int:case_id>', methods=['PUT'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('UPDATE_CASE', 'case')
def api_update_case_v2(case_id):
    # 检查用例是否存在并获取所属项目
    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404
    
    # 检查项目访问权限
    _db = Database()
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403
    
    data = request.get_json(silent=True) or {}
    name = data.get('name')
    url = data.get('url')
    description = data.get('description')
    precondition = data.get('precondition')
    expected_result = data.get('expected_result')
    unit_id = data.get('unit_id')
    if 'unit_id' in data:
        if unit_id in (None, '', 0, '0'):
            parsed_unit_id = None
        else:
            try:
                parsed_unit_id = int(unit_id)
            except (TypeError, ValueError):
                parsed_unit_id = None
        success = db.update_test_case_v2(
            case_id, name, url, description, precondition, expected_result, unit_id=parsed_unit_id
        )
    else:
        success = db.update_test_case_v2(
            case_id, name, url, description, precondition, expected_result
        )
    
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '更新测试用例失败'}), 400

# API: 删除测试用例（新版本）
@app.route('/api/cases/<int:case_id>', methods=['DELETE'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('DELETE_CASE', 'case')
def api_delete_case_v2(case_id):
    # 检查用例是否存在并获取所属项目
    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404
    
    # 检查项目访问权限（删除需要editor权限）
    _db = Database()
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'editor'):
        return jsonify({'success': False, 'error': '无权限删除此用例'}), 403
    
    success = db.delete_test_case_v2(case_id)
    
    if success:
        try:
            db.prune_orphan_run_history()
        except Exception:
            pass
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '删除测试用例失败'}), 400

# ==================== 测试步骤管理API ====================

# API: 获取测试用例的所有步骤
@app.route('/api/cases/<int:case_id>/steps', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_case_steps(case_id):
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    
    # 确保参数有效
    page = max(1, page)
    page_size = max(1, min(page_size, 100))  # 限制最大页面大小为100
    
    steps, total = db.get_case_steps_paginated(case_id, page, page_size)
    
    return jsonify({
        'steps': steps,
        'total': total,
        'page': page,
        'page_size': page_size
    })

# API: 获取单个测试步骤详情
@app.route('/api/steps/<int:step_id>', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_step(step_id):
    step = db.get_test_step(step_id)
    if not step:
        return jsonify({'error': '测试步骤不存在'}), 404
    return jsonify({'step': step})

# API: 创建测试步骤
@app.route('/api/steps', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_create_step():
    data = request.get_json(silent=True) or {}
    case_id = data.get('case_id')
    action = data.get('action', '')
    selector_type = data.get('selector_type', '')
    selector_value = data.get('selector_value', '')
    input_value = data.get('input_value', '')
    description = data.get('description', '')
    step_order = data.get('step_order')  # 不设置默认值，让它为None
    page_name = data.get('page_name', '')
    swipe_x = data.get('swipe_x', '')
    swipe_y = data.get('swipe_y', '')
    url = data.get('url', '')
    enter_iframe = data.get('enter_iframe', False)
    iframe_selector = data.get('iframe_selector', '')
    compare_type = data.get('compare_type', 'equals')
    locator_candidates = data.get('locator_candidates', '')
    if locator_candidates is not None and not isinstance(locator_candidates, str):
        locator_candidates = json.dumps(locator_candidates, ensure_ascii=False)
    click_repeat_count = _norm_click_repeat_count(data.get('click_repeat_count')) if action == 'click' else 1
    captcha_max_attempts = _norm_captcha_max_attempts(data.get('captcha_max_attempts')) if action == 'verify' else None
    api_spec = data.get('api_spec', '')
    if api_spec is not None and not isinstance(api_spec, str):
        api_spec = json.dumps(api_spec, ensure_ascii=False)
    automation_layer = (data.get('automation_layer') or 'web').strip().lower()
    desktop_spec = data.get('desktop_spec', '')
    if desktop_spec is not None and not isinstance(desktop_spec, str):
        desktop_spec = json.dumps(desktop_spec, ensure_ascii=False)
    mobile_spec = data.get('mobile_spec', '')
    if mobile_spec is not None and not isinstance(mobile_spec, str):
        mobile_spec = json.dumps(mobile_spec, ensure_ascii=False)
    
    if not case_id:
        return jsonify({'success': False, 'error': '用例ID不能为空'}), 400
    if not action:
        return jsonify({'success': False, 'error': '操作类型不能为空'}), 400

    case_row = db.get_test_case_v2(int(case_id))
    if not case_row:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    pid = case_row.get('project_id')
    if pid and not db.check_project_access(current_user.id, int(pid), 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403
    step_err = _validate_step_action_for_case(case_row, action, automation_layer)
    if step_err:
        return jsonify({'success': False, 'error': step_err}), 422
    
    step_id = db.create_test_step(case_id, action, selector_type, selector_value, 
                                  input_value, description, step_order, page_name,
                                  swipe_x, swipe_y, url, enter_iframe, iframe_selector, compare_type,
                                  locator_candidates or '', click_repeat_count=click_repeat_count,
                                  api_spec=api_spec or '',
                                  automation_layer=automation_layer,
                                  desktop_spec=desktop_spec or '',
                                  mobile_spec=mobile_spec or '',
                                  captcha_max_attempts=captcha_max_attempts)
    return jsonify({'success': True, 'step_id': step_id})

# API: 更新测试步骤
@app.route('/api/steps/<int:step_id>', methods=['PUT'])
@api_error_handler
@log_api_request
def api_update_step(step_id):
    data = request.get_json(silent=True) or {}
    step_row = db.get_test_step(step_id)
    if not step_row:
        return jsonify({'error': '测试步骤不存在'}), 404
    case_row = db.get_test_case_v2(int(step_row['case_id']))
    action = data.get('action')
    eff_action = action if action is not None else step_row.get('action')
    eff_layer = data.get('automation_layer')
    if eff_layer is None:
        eff_layer = step_row.get('automation_layer') or 'web'
    step_err = _validate_step_action_for_case(case_row, eff_action, eff_layer)
    if step_err:
        return jsonify({'error': step_err}), 422
    selector_type = data.get('selector_type')
    selector_value = data.get('selector_value')
    input_value = data.get('input_value')
    description = data.get('description')
    step_order = data.get('step_order')
    enter_iframe = data.get('enter_iframe')
    iframe_selector = data.get('iframe_selector')
    compare_type = data.get('compare_type')
    locator_candidates = data.get('locator_candidates')
    if locator_candidates is not None and not isinstance(locator_candidates, str):
        locator_candidates = json.dumps(locator_candidates, ensure_ascii=False)
    if action is not None:
        if action == 'click':
            click_repeat_count = _norm_click_repeat_count(data.get('click_repeat_count', 1))
        else:
            click_repeat_count = 1
    elif 'click_repeat_count' in data:
        click_repeat_count = _norm_click_repeat_count(data.get('click_repeat_count'))
    else:
        click_repeat_count = None

    if action is not None:
        if action == 'verify':
            captcha_max_attempts = _norm_captcha_max_attempts(data.get('captcha_max_attempts'))
        else:
            captcha_max_attempts = None
    elif 'captcha_max_attempts' in data:
        captcha_max_attempts = _norm_captcha_max_attempts(data.get('captcha_max_attempts'))
    else:
        captcha_max_attempts = None

    api_spec = data.get('api_spec') if 'api_spec' in data else None
    if api_spec is not None and not isinstance(api_spec, str):
        api_spec = json.dumps(api_spec, ensure_ascii=False)
    automation_layer = data.get('automation_layer') if 'automation_layer' in data else None
    desktop_spec = data.get('desktop_spec') if 'desktop_spec' in data else None
    if desktop_spec is not None and not isinstance(desktop_spec, str):
        desktop_spec = json.dumps(desktop_spec, ensure_ascii=False)
    mobile_spec = data.get('mobile_spec') if 'mobile_spec' in data else None
    if mobile_spec is not None and not isinstance(mobile_spec, str):
        mobile_spec = json.dumps(mobile_spec, ensure_ascii=False)
    _sel_eff = (
        (selector_type if selector_type is not None else step_row.get('selector_type') or '')
        .strip()
        .lower()
    )
    if _sel_eff == 'visual':
        automation_layer = 'desktop'

    success = db.update_test_step(step_id, action, selector_type, selector_value,
                                   input_value, description, step_order, enter_iframe, iframe_selector, compare_type,
                                   locator_candidates, click_repeat_count, api_spec=api_spec,
                                   automation_layer=automation_layer, desktop_spec=desktop_spec,
                                   mobile_spec=mobile_spec,
                                   captcha_max_attempts=captcha_max_attempts)
    
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '更新测试步骤失败'}), 400

# API: 删除测试步骤
@app.route('/api/steps/<int:step_id>', methods=['DELETE'])
@login_required
@api_error_handler
@log_api_request
def api_delete_step(step_id):
    step_row = db.get_test_step(step_id)
    if not step_row:
        return jsonify({'success': False, 'error': '测试步骤不存在'}), 404
    case_row = db.get_test_case_v2(int(step_row['case_id']))
    if not case_row:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    pid = case_row.get('project_id')
    if pid and not db.check_project_access(current_user.id, int(pid), 'editor'):
        return jsonify({'success': False, 'error': '无权限删除此步骤'}), 403
    success = db.delete_test_step(step_id)

    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '删除测试步骤失败'}), 400

# API: 删除测试用例的所有步骤
@app.route('/api/cases/<int:case_id>/steps', methods=['DELETE'])
@api_error_handler
@log_api_request
def api_delete_case_steps(case_id):
    success = db.delete_case_steps(case_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '删除测试用例步骤失败'}), 400

# API: 更新测试步骤顺序
@app.route('/api/cases/<int:case_id>/steps/order', methods=['PUT'])
@api_error_handler
@log_api_request
def api_update_step_order(case_id):
    data = request.get_json(silent=True) or {}
    steps = data.get('steps', [])
    
    success = db.update_step_order(case_id, steps)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '更新步骤顺序失败'}), 400


@app.route('/api-testing')
@login_required
def api_testing_page():
    return render_template('api_testing.html')


@app.route('/api-testing/case/<int:case_id>')
@login_required
def api_case_detail_page(case_id):
    case = db.get_test_case_v2(case_id)
    if not case:
        return redirect(url_for('api_testing_page', err='case_not_found'))
    if _app_case_type(case) != 'api':
        return redirect(url_for('api_testing_page', err='not_api_case'))
    pid = case.get('project_id')
    _db = Database()
    if pid and not _db.check_project_access(current_user.id, int(pid), 'viewer'):
        return redirect(url_for('api_testing_page', err='no_project_access'))
    return render_template(
        'api_case_detail.html',
        case_id=case_id,
        project_id=int(pid) if pid else 0,
        case_name=(case.get('name') or '').strip() or ('#' + str(case_id)),
    )


@app.route('/api/cases/<int:case_id>/migrate-api-steps', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_migrate_api_steps(case_id):
    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    _db = Database()
    pid = case.get('project_id')
    if pid and not _db.check_project_access(current_user.id, int(pid), 'editor'):
        return jsonify({'success': False, 'error': '无权限修改此用例'}), 403
    if _app_case_type(case) != 'ui':
        return jsonify({'success': False, 'error': '仅可从 Web 用例迁移接口步骤'}), 400
    data = request.get_json(silent=True) or {}
    tid = data.get('target_api_case_id')
    if tid is not None:
        try:
            tid = int(tid)
        except (TypeError, ValueError):
            tid = None
    tname = (data.get('target_api_case_name') or '').strip() or None
    r = db.migrate_api_steps_from_ui_case(case_id, target_api_case_id=tid, target_api_case_name=tname)
    if r.get('success'):
        return jsonify(r)
    return jsonify(r), 400


@app.route('/api/api-cases/from-ui-case', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_create_api_case_from_ui_case():
    data = request.get_json(silent=True) or {}
    project_id = data.get('project_id')
    source_ui_case_id = data.get('source_ui_case_id')
    migrate_steps = bool(data.get('migrate_api_steps'))
    name = (data.get('name') or '').strip()
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    _db = Database()
    if not _db.check_project_access(current_user.id, int(project_id), 'editor'):
        return jsonify({'success': False, 'error': '无权限在此项目创建用例'}), 403

    limits = license_manager.get_limits()
    lic = license_manager.get_current_license()
    cur_cnt = _db.get_project_case_count(int(project_id))
    if limits['max_cases_per_project'] != -1 and cur_cnt >= limits['max_cases_per_project']:
        return jsonify({
            'success': False,
            'error': f'已达到项目用例数量限制（{limits["max_cases_per_project"]}个）。',
            'limit_reached': True,
        }), 403

    def _bump_free_created():
        if lic.license_type == LicenseType.FREE.value:
            _db.increment_created_cases(current_user.id)

    if not source_ui_case_id:
        if not name:
            return jsonify({'success': False, 'error': '新建接口用例时 name 不能为空'}), 400
        case_id = _db.create_test_case_v2(
            int(project_id),
            name,
            data.get('url', ''),
            data.get('description', ''),
            data.get('precondition', ''),
            data.get('expected_result', ''),
            case_type='api',
        )
        _bump_free_created()
        return jsonify({'success': True, 'case_id': case_id})
    src = _db.get_test_case_v2(int(source_ui_case_id))
    if not src:
        return jsonify({'success': False, 'error': '源 Web 用例不存在'}), 404
    if int(src.get('project_id') or 0) != int(project_id):
        return jsonify({'success': False, 'error': '源用例必须属于所选项目'}), 400
    if _app_case_type(src) != 'ui':
        return jsonify({'success': False, 'error': '源用例必须是 Web 用例'}), 400
    new_name = name or f"{src.get('name') or '用例'} (接口)"
    new_id = _db.create_test_case_v2(
        int(project_id),
        new_name,
        src.get('url') or '',
        src.get('description') or '',
        src.get('precondition') or '',
        src.get('expected_result') or '',
        case_type='api',
    )
    _bump_free_created()
    if migrate_steps:
        m = _db.migrate_api_steps_from_ui_case(int(source_ui_case_id), target_api_case_id=new_id)
        if not m.get('success'):
            try:
                _db.delete_test_case_v2(new_id)
            except Exception:
                pass
            return jsonify({'success': False, 'error': m.get('error', '迁移失败')}), 400
        return jsonify({'success': True, 'case_id': m.get('target_api_case_id', new_id), 'migrated': m})
    return jsonify({'success': True, 'case_id': new_id})


@app.route('/api/api-cases/<int:case_id>/duplicate', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_duplicate_api_case(case_id):
    src = db.get_test_case_v2(case_id)
    if not src:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    pid = src.get('project_id')
    if not pid:
        return jsonify({'success': False, 'error': '用例无归属项目'}), 400
    _db = Database()
    if not _db.check_project_access(current_user.id, int(pid), 'editor'):
        return jsonify({'success': False, 'error': '无权限'}), 403
    if _app_case_type(src) != 'api':
        return jsonify({'success': False, 'error': '仅可复制接口用例'}), 400
    limits = license_manager.get_limits()
    lic = license_manager.get_current_license()
    cur_cnt = _db.get_project_case_count(int(pid))
    if limits['max_cases_per_project'] != -1 and cur_cnt >= limits['max_cases_per_project']:
        return jsonify({
            'success': False,
            'error': f'已达到项目用例数量限制（{limits["max_cases_per_project"]}个）。',
            'limit_reached': True,
        }), 403
    data = request.get_json(silent=True) or {}
    new_name = (data.get('name') or '').strip() or f"{src.get('name') or '接口用例'} 副本"
    nid = _db.duplicate_test_case_with_steps(case_id, new_name, case_type='api')
    if not nid:
        return jsonify({'success': False, 'error': '复制失败'}), 500
    if lic.license_type == LicenseType.FREE.value:
        _db.increment_created_cases(current_user.id)
    return jsonify({'success': True, 'case_id': nid})


@app.route('/api/api-cases/<int:case_id>/run', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_run_api_case(case_id):
    start_time = time.time()
    _db = Database()
    user_id = current_user.id
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    is_free_user = license_info.license_type == LicenseType.FREE.value
    today_stats = _db.get_user_usage_stats(user_id)
    current_count = today_stats.get('execution_count', 0) if today_stats else 0
    daily_limit = limits.get('max_executions_per_day', -1)
    if is_free_user:
        DAILY_LIMIT = 10
        if current_count >= DAILY_LIMIT:
            return jsonify({
                'success': False,
                'error': f'今日执行次数已达上限（{DAILY_LIMIT}次）。请升级至团队版解除限制。',
                'limit_reached': True,
            }), 403
    elif daily_limit > 0 and current_count >= daily_limit:
        return jsonify({
            'success': False,
            'error': f'今日执行次数已达上限（{daily_limit}次）。',
            'limit_reached': True,
        }), 403
    _db.increment_execution_count(user_id)
    case = _db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404
    if case.get('project_id') and not _db.check_project_access(user_id, int(case['project_id']), 'viewer'):
        return jsonify({'success': False, 'error': '无权限执行此用例'}), 403
    if _app_case_type(case) != 'api':
        return jsonify({
            'success': False,
            'error': '该用例不是接口用例，请使用 Web 用例的「运行」按钮执行浏览器自动化。',
        }), 400
    body = request.get_json(silent=True) or {}
    raw_step_ids = body.get('step_ids')
    step_ids_param = None
    if raw_step_ids is not None:
        if not isinstance(raw_step_ids, list):
            return jsonify({'success': False, 'error': 'step_ids 须为非空整数数组'}), 400
        try:
            step_ids_param = [int(x) for x in raw_step_ids]
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'step_ids 须为整数'}), 400
        if not step_ids_param:
            return jsonify({'success': False, 'error': '请至少选择一个要执行的接口步骤'}), 400
        all_steps = _db.get_case_steps(case_id, page=1, page_size=9999)
        valid = {int(s['id']) for s in (all_steps or []) if s.get('id') is not None}
        step_ids_param = list(dict.fromkeys([sid for sid in step_ids_param if sid in valid]))
        if not step_ids_param:
            return jsonify({
                'success': False,
                'error': '所选步骤不属于本用例或不存在，请刷新后重试。',
            }), 400
    payload = sync_run_api_case_for_batch(
        case_id, _db, execution_context=None, step_ids=step_ids_param
    )
    duration = round(time.time() - start_time, 2)
    out = {
        'success': payload.get('status') == 'success',
        'status': payload.get('status'),
        'duration': duration,
        'case_id': case_id,
        'run_history_id': payload.get('run_history_id'),
        'step_results': payload.get('step_results'),
    }
    if payload.get('status') == 'warning':
        out['warning'] = payload.get('warning')
    if payload.get('status') == 'error':
        out['error'] = payload.get('error') or '接口用例执行失败'
    return jsonify(out)


@app.route('/api/api-cases/dry-run-request', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_dry_run_api_request():
    """试发当前表单中的单条 HTTP 规格（不写库），用于弹窗内查看响应与断言结果。"""
    data = request.get_json(silent=True) or {}
    case_id = data.get('case_id')
    if case_id is None:
        return jsonify({'success': False, 'error': 'case_id 不能为空'}), 400
    try:
        case_id = int(case_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'case_id 无效'}), 400

    _db = Database()
    case = _db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    if case.get('project_id') and not _db.check_project_access(
        current_user.id, int(case['project_id']), 'viewer'
    ):
        return jsonify({'success': False, 'error': '无权限'}), 403
    if _app_case_type(case) != 'api':
        return jsonify({'success': False, 'error': '仅支持接口用例'}), 400

    spec = data.get('api_spec')
    if isinstance(spec, str):
        try:
            spec = json.loads(spec)
        except json.JSONDecodeError:
            return jsonify({'success': False, 'error': 'api_spec JSON 无效'}), 400
    if not isinstance(spec, dict):
        return jsonify({'success': False, 'error': 'api_spec 须为对象'}), 400

    project_id = case.get('project_id')

    out = run_api_spec_pipeline(
        spec,
        _db,
        project_id,
        case_id,
        browser_cookie_jar=None,
        persist_extracts=False,
        collect_script_logs=True,
    )

    rt = out.get('response_text') or ''
    if len(rt) > 48000:
        rt = rt[:48000] + '\n…(已截断)'
    # 未拿到 HTTP 响应（连接失败、URL 无效、前置链在发出主请求前失败等）时 success=False，便于前端明确提示
    hard_fail = out.get('status_code') is None and bool(out.get('error'))
    payload = {
        'success': not hard_fail,
        'status_code': out.get('status_code'),
        'response_text': rt,
        'response_json': out.get('response_json'),
        'response_headers': out.get('response_headers') or {},
        'elapsed_ms': out.get('elapsed_ms'),
        'ok_assert': out.get('ok_assert'),
        'assert_message': out.get('assert_message'),
        'error': out.get('error'),
    }
    if out.get('script_logs'):
        payload['script_logs'] = out.get('script_logs')
    return jsonify(payload)


def _effective_step_iframe_selector(automation, db, step, project_id, case_id, row_resolve_fn=None):
    """
    本步操作使用的 iframe 定位串（与 frame_locator 一致）。
    「进入 iframe / 跳出 iframe」为隐式上下文：执行 enter_iframe 步骤后，只要未 exit_iframe，
    且当前步骤未单独勾选「在 iframe 内」+ 选择器列，则默认沿用 automation.current_iframe。
    若步骤单独配置了 iframe 列，则该步优先使用步骤配置（支持 {{变量}}）。
    row_resolve_fn: 数据驱动等对解析结果再套一层 {{row.field}}。
    """
    if bool(step.get('enter_iframe')):
        raw = (step.get('iframe_selector') or '').strip()
        if raw:
            resolved = db.resolve_variables(raw, project_id=project_id, case_id=case_id)
            if row_resolve_fn and resolved:
                resolved = row_resolve_fn(resolved)
            return (resolved or '').strip() or None
    ci = getattr(automation, 'current_iframe', None) or {}
    sel = ci.get('selector')
    return sel if sel else None


def _run_db_step_scroll(input_value: str, iframe_selector=None):
    """按平台存储的 up/down/left/right 像素执行滚动；无有效值时回退为向下 500px。"""
    v = parse_platform_scroll_input_value(input_value)
    dx = v["right"] - v["left"]
    dy = v["down"] - v["up"]
    if dx != 0 or dy != 0:
        sync_scroll_by_delta(dx, dy, iframe_selector=iframe_selector)
    else:
        sync_scroll_page("down", 500, iframe_selector=iframe_selector)


def _run_assert_automation_step(
    step: dict,
    selector_value: str,
    input_value: str,
    selector_type: str,
    iframe_sel,
):
    """执行断言步骤（与单用例运行一致）。文本类断言返回应写入 extracted 的片段，否则返回 None。"""
    from auth_batch_helpers import assert_empty_expected_error, normalize_assert_compare_type

    assert_type = normalize_assert_compare_type(
        step.get('compare_type', 'text_equals'),
        selector_value=selector_value or "",
        input_value=input_value or "",
    )
    expected_value = input_value
    empty_err = assert_empty_expected_error(assert_type, expected_value)
    if empty_err:
        raise Exception(empty_err)
    if assert_type == "vision_contains":
        cond = (step.get("description") or expected_value or "").strip()
        if not cond:
            raise Exception("vision_contains 断言缺少描述/预期")
    uat_logger.info(
        f"执行断言操作: 类型={assert_type}, 选择器={selector_value}, 预期={expected_value}"
    )
    extracted_fragment = None
    try:
        if assert_type in ['page_text_equals', 'page_text_contains', 'page_text_regex']:
            actual_text = (sync_get_page_text() or "").strip()
            if assert_type == 'page_text_equals':
                from auth_batch_helpers import page_text_has_exact_snippet

                if not page_text_has_exact_snippet(actual_text, expected_value or ""):
                    raise Exception(
                        f"整页文本断言失败(equals): 页面未出现与预期完全一致的文案 {expected_value!r}"
                    )
            elif assert_type == 'page_text_contains':
                from ai_page_probe import page_text_matches_assert_expected

                if not page_text_matches_assert_expected(
                    actual_text, expected_value, 'page_text_contains'
                ):
                    raise Exception(
                        f"整页文本断言失败(contains): 页面未包含 {expected_value[:160]!r}"
                    )
            elif assert_type == 'page_text_regex':
                from ai_page_probe import page_text_matches_assert_expected

                if not page_text_matches_assert_expected(
                    actual_text, expected_value or '', 'page_text_regex'
                ):
                    raise Exception(
                        f"整页正则断言失败: pattern={expected_value!r} 实际文本长度={len(actual_text)}"
                    )
            extracted_fragment = actual_text[:500] if actual_text else ""
            uat_logger.info(f"断言成功: {assert_type}")
        elif assert_type in ['text_equals', 'text_contains', 'text_regex']:
            from ai_page_probe import page_text_matches_assert_expected

            assert_wait_ms = int(os.environ.get("UAT_ASSERT_ELEMENT_WAIT_MS", "15000") or 15000)
            try:
                actual_text = sync_extract_element_text(
                    selector_value,
                    selector_type,
                    iframe_selector=iframe_sel,
                    locator_candidates=step.get('locator_candidates') or None,
                    wait_timeout_ms=assert_wait_ms,
                )
            except Exception as elem_ex:
                page_text = (sync_get_page_text() or "").strip()
                if expected_value and page_text_matches_assert_expected(
                    page_text, expected_value, assert_type
                ):
                    uat_logger.warning(
                        "元素断言定位失败，已回退整页文本断言成功: %s", expected_value[:80]
                    )
                    extracted_fragment = page_text[:500]
                    uat_logger.info("断言成功: page_text(回退)")
                    return extracted_fragment
                if assert_type == 'text_regex' or (
                    expected_value and '|' in str(expected_value)
                ):
                    uat_logger.warning(
                        "元素断言超时，尝试 page_text_regex 回退: %s", str(expected_value)[:80]
                    )
                    if page_text_matches_assert_expected(
                        page_text, expected_value, 'page_text_regex'
                    ):
                        extracted_fragment = page_text[:500]
                        uat_logger.info("断言成功: page_text_regex(回退)")
                        return extracted_fragment
                raise elem_ex
            if assert_type == 'text_equals':
                if actual_text != expected_value:
                    raise Exception(
                        f"断言失败: 实际文本 '{actual_text}' 不等于预期 '{expected_value}'"
                    )
            elif assert_type == 'text_contains':
                if expected_value not in actual_text:
                    raise Exception(
                        f"断言失败: 实际文本 '{actual_text}' 不包含预期 '{expected_value}'"
                    )
            elif assert_type == 'text_regex':
                import re

                if not re.search(expected_value, actual_text):
                    raise Exception(
                        f"断言失败: 实际文本 '{actual_text}' 不匹配正则 '{expected_value}'"
                    )
            extracted_fragment = actual_text
            uat_logger.info(f"断言成功: {assert_type}")
        elif assert_type in ['element_exists', 'element_visible']:
            if assert_type == 'element_exists':
                sync_wait_for_selector(selector_value, timeout=5000)
                uat_logger.info(f"断言成功: 元素存在 {selector_value}")
            else:
                sync_wait_for_element_visible(selector_value, timeout=5000)
                uat_logger.info(f"断言成功: 元素可见 {selector_value}")
        elif assert_type == 'element_count':
            actual_count = sync_get_element_count(selector_value, selector_type)
            expected_count = int(expected_value) if expected_value else 0
            operator = step.get('swipe_x', 'equals')
            success = False
            if operator == 'equals':
                success = actual_count == expected_count
            elif operator == 'gt':
                success = actual_count > expected_count
            elif operator == 'lt':
                success = actual_count < expected_count
            elif operator == 'gte':
                success = actual_count >= expected_count
            elif operator == 'lte':
                success = actual_count <= expected_count
            if not success:
                raise Exception(
                    f"断言失败: 实际数量 {actual_count} 不符合预期 {operator} {expected_count}"
                )
            uat_logger.info("断言成功: 元素数量符合预期")
        elif assert_type in ['url_equals', 'url_contains']:
            actual_url = sync_get_current_url()
            if assert_type == 'url_equals':
                if not _url_assert_matches_pa(actual_url, expected_value, 'url_equals'):
                    raise Exception(
                        f"断言失败: 实际URL '{actual_url}' 不等于预期 '{expected_value}'"
                    )
            else:
                if expected_value and not _url_assert_matches_pa(actual_url, expected_value, 'url_contains'):
                    raise Exception(
                        f"断言失败: 实际URL '{actual_url}' 不包含预期 '{expected_value}'"
                    )
            uat_logger.info(f"断言成功: {assert_type}")
        elif assert_type == 'element_attr':
            attr_name = step.get('page_name', '')

            async def get_attr():
                page = await automation.get_page()
                element = await page.query_selector(selector_value)
                if element:
                    return await element.get_attribute(attr_name)
                return None

            actual_attr = worker.execute(get_attr)
            if actual_attr != expected_value:
                raise Exception(
                    f"断言失败: 属性 {attr_name} 实际值 '{actual_attr}' 不等于预期 '{expected_value}'"
                )
            uat_logger.info(f"断言成功: 属性 {attr_name} = {actual_attr}")
        else:
            raise Exception(f"不支持的 assert compare_type: {assert_type}")
    except Exception as assert_error:
        uat_logger.error(f"断言失败: {assert_error}")
        raise
    sync_wait_for_timeout(500)
    return extracted_fragment


def _run_extract_text_automation_step(
    action: str,
    step: dict,
    selector_value: str,
    input_value: str,
    description: str,
    selector_type: str,
    iframe_sel,
    locator_candidates=None,
):
    """extract_text / text_compare 与单用例运行一致。返回 (extracted_text, expected_text)。"""
    current_expected = input_value or description
    expected_text = current_expected
    verify_type = step.get('compare_type', step.get('verify_type', 'equals'))
    extracted_text = ""

    if selector_value:
        try:
            current_extracted = sync_extract_element_text(
                selector_value,
                selector_type,
                iframe_selector=iframe_sel,
                locator_candidates=locator_candidates,
            )
            uat_logger.info(f"提取到文本: {current_extracted[:100]}...")
            extracted_text = current_extracted
        except Exception as extract_error:
            uat_logger.error(f"提取文本失败: {extract_error}")
            raise Exception(f"提取文本失败: {extract_error}") from extract_error

        if expected_text:
            if extracted_text:
                uat_logger.info(
                    f"验证文本 - 提取: {extracted_text[:100]}..., 预期: {expected_text[:100]}..., 验证方式: {verify_type}"
                )
                if verify_type == 'equals':
                    if extracted_text != expected_text:
                        uat_logger.error("文本验证失败: 提取的文本与预期结果不相等")
                        raise Exception("文本验证失败: 提取的文本与预期结果不相等")
                elif verify_type == 'not_equals':
                    if extracted_text == expected_text:
                        uat_logger.error("文本验证失败: 提取的文本与预期结果相等")
                        raise Exception("文本验证失败: 提取的文本与预期结果相等")
                elif verify_type == 'contains':
                    if expected_text not in extracted_text:
                        uat_logger.error("文本验证失败: 提取的文本不包含预期内容")
                        raise Exception("文本验证失败: 提取的文本不包含预期内容")
                elif verify_type == 'partial':
                    if expected_text not in extracted_text:
                        uat_logger.error(
                            "文本验证失败: 提取的文本不包含预期的部分内容"
                        )
                        raise Exception(
                            "文本验证失败: 提取的文本不包含预期的部分内容"
                        )
                uat_logger.info("文本验证成功")
            else:
                if action == 'text_compare':
                    uat_logger.warning("未提取到文本，跳过文本验证")
                else:
                    uat_logger.info("提取文本操作完成（未提取到文本）")
    else:
        try:
            current_extracted = sync_get_page_text()
            uat_logger.info(f"提取到页面文本: {current_extracted[:100]}...")
            extracted_text = current_extracted
        except Exception as extract_error:
            uat_logger.error(f"提取页面文本失败: {extract_error}")
            raise Exception(f"提取页面文本失败: {extract_error}") from extract_error

        if expected_text:
            if extracted_text:
                uat_logger.info(
                    f"验证页面文本 - 提取: {extracted_text[:100]}..., 预期: {expected_text[:100]}..., 验证方式: {verify_type}"
                )
                if verify_type == 'equals':
                    if extracted_text != expected_text:
                        uat_logger.error("页面文本验证失败: 提取的文本与预期结果不相等")
                        raise Exception(
                            "页面文本验证失败: 提取的文本与预期结果不相等"
                        )
                elif verify_type == 'not_equals':
                    if extracted_text == expected_text:
                        uat_logger.error("页面文本验证失败: 提取的文本与预期结果相等")
                        raise Exception(
                            "页面文本验证失败: 提取的文本与预期结果相等"
                        )
                elif verify_type == 'contains':
                    if expected_text not in extracted_text:
                        uat_logger.error(
                            "页面文本验证失败: 提取的文本不包含预期内容"
                        )
                        raise Exception(
                            "页面文本验证失败: 提取的文本不包含预期内容"
                        )
                elif verify_type == 'partial':
                    if expected_text not in extracted_text:
                        uat_logger.error(
                            "页面文本验证失败: 提取的文本不包含预期的部分内容"
                        )
                        raise Exception(
                            "页面文本验证失败: 提取的文本不包含预期的部分内容"
                        )
                uat_logger.info("页面文本验证成功")
            else:
                if action == 'text_compare':
                    uat_logger.warning("未提取到页面文本，跳过文本验证")
                else:
                    uat_logger.error("提取页面文本操作失败：未提取到文本")
                    raise Exception("提取页面文本操作失败：未提取到文本")

    sync_wait_for_timeout(1000)
    return extracted_text, expected_text


# ==================== Playwright Codegen 录制：粘贴导入 ====================


def _playwright_codegen_command(nav_url: str) -> tuple:
    """
    返回 (argv, engine_label)。
    优先使用仓库内 playwright-xpath-fork 构建产物（XPath 优先录制定位），否则回退 pip playwright。
    环境变量 PLAYWRIGHT_XPATH_CODEGEN_CLI：可执行或「node path/to/cli.js ...」前缀，后自动拼接 codegen 参数。
    """
    target_args = ["codegen", nav_url, "--target", "python"]
    app_dir = os.path.dirname(os.path.abspath(__file__))
    fork_cli = os.path.join(
        app_dir,
        "playwright-xpath-fork",
        "packages",
        "playwright-core",
        "cli.js",
    )
    fork_bundle = os.path.join(
        app_dir,
        "playwright-xpath-fork",
        "packages",
        "playwright-core",
        "lib",
        "coreBundle.js",
    )
    # Codegen/Inspector 的静态界面，仅完整 `npm run build` 后才有；缺它会出现 ENOENT index.html
    fork_recorder_index = os.path.join(
        app_dir,
        "playwright-xpath-fork",
        "packages",
        "playwright-core",
        "lib",
        "vite",
        "recorder",
        "index.html",
    )

    env_override = (os.environ.get("PLAYWRIGHT_XPATH_CODEGEN_CLI") or "").strip()
    if env_override:
        try:
            base = shlex.split(env_override, posix=(os.name != "nt"))
        except ValueError:
            base = [env_override]
        return base + target_args, "custom"

    node = shutil.which("node")
    if (
        node
        and os.path.isfile(fork_cli)
        and os.path.isfile(fork_bundle)
        and os.path.isfile(fork_recorder_index)
    ):
        return [node, fork_cli] + target_args, "xpath_fork"

    return [sys.executable, "-m", "playwright"] + target_args, "pypi"


@app.route("/api/cases/<int:case_id>/start-playwright-codegen", methods=["POST"])
@login_required
@api_error_handler
@log_api_request
def api_case_start_playwright_codegen(case_id):
    """在本机新开控制台启动官方 playwright codegen（不写入服务端文件；录制完成后请复制代码使用「从 Codegen 导入」）。"""
    data = request.get_json(silent=True) or {}
    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({"success": False, "error": "测试用例不存在"}), 404

    _db = Database()
    if case.get("project_id") and not _db.check_project_access(
        current_user.id, case["project_id"], "editor"
    ):
        return jsonify({"success": False, "error": "无权限修改此用例"}), 403

    override_url = (data.get("url") or "").strip()
    steps_preview = db.get_case_steps(case_id)
    nav_url, nav_src = _resolve_case_navigation_url(
        case=case,
        case_id=case_id,
        steps=steps_preview,
        fallback_url=override_url or None,
    )
    if not nav_url:
        return jsonify(
            {
                "success": False,
                "error": "无法确定起始地址：请在本用例填写测试 URL，或在请求体中传入 url 字段",
            }
        ), 400

    app_dir = os.path.dirname(os.path.abspath(__file__))
    cmd, engine = _playwright_codegen_command(nav_url)
    try:
        if sys.platform == "win32":
            cf = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
            if not cf:
                cf = 0x00000010
            subprocess.Popen(cmd, creationflags=cf, cwd=app_dir if engine == "xpath_fork" else None)
        else:
            subprocess.Popen(cmd, start_new_session=True, cwd=app_dir if engine == "xpath_fork" else None)
    except Exception as e:
        uat_logger.log_exception("api_case_start_playwright_codegen", e)
        return jsonify(
            {"success": False, "error": f"无法启动 playwright codegen：{e}"}
        ), 500
    uat_logger.info(
        f"Playwright codegen 已启动 engine={engine} case_id={case_id} user={current_user.id} url={nav_url}"
    )
    if engine == "xpath_fork":
        hint = (
            "已使用本仓库 **XPath 优先** 的 Playwright 分叉启动 Codegen（生成代码会尽量用 page.locator('xpath=…') 等）。"
            " 完成后请复制代码，用「从 Codegen 导入」写入用例。"
        )
    elif engine == "pypi":
        hint = (
            "当前使用 pip 安装的 **官方** Playwright，Codegen 仍优先 getByRole 等。"
            " 若要 XPath 优先录制：在本机进入仓库目录 `playwright-xpath-fork` 执行 `npm ci` 与 `npm run build` 后，"
            "再点「启动 Playwright Codegen」；详见「录制教程」。"
            " 完成后复制代码，用「从 Codegen 导入」。"
        )
    else:
        hint = "Codegen 已启动。完成后请复制代码，用「从 Codegen 导入」写入用例。"

    return jsonify(
        {
            "success": True,
            "navigated_to": nav_url,
            "nav_source": nav_src,
            "codegen_engine": engine,
            "hint": hint,
        }
    )


@app.route('/recording-tutorial')
@login_required
def recording_tutorial_page():
    """Playwright Codegen 与 Selenium IDE（.side）粘贴导入说明。"""
    return render_template('recording_tutorial.html')


@app.route('/api/cases/<int:case_id>/import-playwright-codegen', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_import_playwright_codegen(case_id):
    data = request.get_json(silent=True) or {}
    code = (data.get('code') or '').strip()
    replace_existing = bool(data.get('replace_existing'))
    if not code:
        return jsonify({'success': False, 'error': '请粘贴 Playwright Codegen 生成的代码'}), 400

    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404
    if _app_case_type(case) == 'api':
        return jsonify({'success': False, 'error': '接口用例不支持 Codegen 导入，请在接口测试页维护 HTTP 步骤。'}), 400

    _db = Database()
    steps = enrich_steps_with_xpath_priority(steps)
    if not steps:
        return jsonify({
            'success': False,
            'error': '未能解析出有效步骤，请确认粘贴的是 Codegen 窗口中的 Python 或 JS 片段（含 page.goto / click / fill 等）',
            'warnings': warnings,
        }), 400

    if replace_existing:
        db.delete_case_steps(case_id)

    if not db.batch_insert_steps(case_id, steps):
        return jsonify({'success': False, 'error': '写入步骤失败'}), 500

    return jsonify({'success': True, 'imported': len(steps), 'warnings': warnings})


@app.route('/api/cases/<int:case_id>/import-selenium-ide', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_import_selenium_ide(case_id):
    """粘贴 Selenium IDE 保存的 .side 项目 JSON（或其它可解析结构）。"""
    data = request.get_json(silent=True) or {}
    raw = data.get('json') or data.get('side') or data.get('payload')
    replace_existing = bool(data.get('replace_existing'))

    if raw is None:
        return jsonify({'success': False, 'error': '请在请求体中提供 json 字段（.side 文件全文）'}), 400
    if isinstance(raw, dict):
        payload = raw
    elif isinstance(raw, str):
        payload = raw.strip()
        if not payload:
            return jsonify({'success': False, 'error': 'JSON 内容为空'}), 400
    else:
        return jsonify({'success': False, 'error': 'json 字段类型无效'}), 400

    case = db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404
    if _app_case_type(case) == 'api':
        return jsonify({'success': False, 'error': '接口用例不支持 Selenium IDE 导入。'}), 400

    _db = Database()
    if not steps:
        return jsonify({
            'success': False,
            'error': '未能解析出有效步骤。请粘贴完整的 .side 项目（含 tests[].commands），'
                     '教程见「录制教程」页中的 Selenium IDE 一节。',
            'warnings': warnings,
        }), 400

    if replace_existing:
        db.delete_case_steps(case_id)

    if not db.batch_insert_steps(case_id, steps):
        return jsonify({'success': False, 'error': '写入步骤失败'}), 500

    return jsonify({'success': True, 'imported': len(steps), 'warnings': warnings})


# API: 运行测试用例
@app.route('/api/cases/<int:case_id>/run', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_run_case(case_id):
    # 记录开始时间
    start_time = time.time()

    # 初始化数据库连接（修复变量作用域问题）
    db = Database()

    # ===== 执行次数限制检查 =====
    user_id = current_user.id

    # 获取当前 License 信息
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    is_free_user = license_info.license_type == LicenseType.FREE.value

    # 获取今日已执行次数
    today_stats = db.get_user_usage_stats(user_id)
    current_count = today_stats.get('execution_count', 0) if today_stats else 0

    # 获取每日执行限制（-1表示无限制）
    daily_limit = limits.get('max_executions_per_day', -1)

    # 免费版每日限制10次
    if is_free_user:
        DAILY_LIMIT = 10

        if current_count >= DAILY_LIMIT:
            return jsonify({
                'success': False,
                'error': f'今日执行次数已达上限（{DAILY_LIMIT}次）。请升级至团队版解除限制。',
                'limit_reached': True,
                'current_count': current_count,
                'daily_limit': DAILY_LIMIT,
                'upgrade_url': '/upgrade'
            }), 403
    elif daily_limit > 0 and current_count >= daily_limit:
        # 非免费版但有每日限制的情况
        return jsonify({
            'success': False,
            'error': f'今日执行次数已达上限（{daily_limit}次）。请联系管理员。',
            'limit_reached': True,
            'current_count': current_count,
            'daily_limit': daily_limit
        }), 403

    # 记录执行次数（所有版本都记录，用于统计和显示）
    new_count = db.increment_execution_count(user_id)
    uat_logger.info(f"📊 [LICENSE] 用户 {user_id} 今日执行次数: {new_count}")

    # 获取测试用例信息
    case, steps = load_case_and_steps(case_id, db)
    if not case:
        return jsonify({'error': '测试用例不存在'}), 404

    if should_delegate_execution_to_clients() and not getattr(
        __import__('flask').g, 'force_local_run', False
    ):
        job_resp = patch_run_case_for_server(db, case_id, user_id)
        return jsonify(job_resp), 202

    if _app_case_type(case) == 'api':
        return jsonify({
            'success': False,
            'error': '这是接口用例，请在「接口测试」模块中执行，或调用 POST /api/api-cases/<用例ID>/run。',
        }), 400

    # steps 已由 load_case_and_steps 加载
    from auth_batch_helpers import prepare_steps_for_execution

    steps, _runtime_probe_warns = prepare_steps_for_execution(
        steps, (case.get("url") or "").strip()
    )
    for w in _runtime_probe_warns or []:
        uat_logger.warning("运行时 LIVE 步骤修复: %s", w)

    if any((s.get('action') or '').strip() == 'api_request' for s in steps):
        return jsonify({
            'success': False,
            'error': '该 Web 用例仍包含接口测试步骤，请先使用「一键迁移」迁移到接口用例后再运行 AI 自动化测试。',
        }), 400

    if not steps:
        # 用例无步骤，保存一条失败历史并友好提示
        run_id = db.create_run_history(case_id, 'error', 0, '该用例没有步骤，无法执行', '', '')
        uat_logger.warning(f"测试用例 #{case_id} 没有步骤，无法执行")
        return jsonify({'success': False, 'status': 'error', 'duration': 0,
                        'error': '该用例尚未添加任何步骤，请先编辑用例添加步骤后再执行。'})

    try:
        from execution_factory import get_executor_factory

        env_msg = get_executor_factory().validate_case_environment(steps)
        if env_msg and (
            "不支持" in env_msg
            or "需 Windows" in env_msg
            or "缺少依赖" in env_msg
            or "未启用" in env_msg
            or "未安装" in env_msg
        ):
            from desktop_runtime import desktop_runtime_unavailable_reason

            detail = desktop_runtime_unavailable_reason()
            return jsonify({
                'success': False,
                'error': env_msg,
                'error_detail': detail or env_msg,
                'python_executable': sys.executable,
                'platform': sys.platform,
            }), 400
        if env_msg:
            uat_logger.warning("混排运行环境提示: %s", env_msg)
    except ImportError:
        pass

    machine_lock_acquired = False
    playwright_lock_acquired = False
    with _UserUiRunGuard(user_id, f'case #{case_id}'):
        try:
            from execution_lock import ExecutionLockError, acquire as acquire_machine_lock

            machine_lock_acquired = acquire_machine_lock(
                owner=f"case_run:{case_id}:user:{user_id}", timeout_sec=120
            )
            if not machine_lock_acquired:
                reject_msg = '本机已有自动化任务在执行，请稍后再试。'
                _record_run_history_rejected(
                    db, case_id, reject_msg, round(time.time() - start_time, 2)
                )
                return jsonify({
                    'success': False,
                    'error': reject_msg,
                    'lock': 'busy',
                }), 409
        except ExecutionLockError as lock_exc:
            reject_msg = str(lock_exc)
            _record_run_history_rejected(
                db, case_id, reject_msg, round(time.time() - start_time, 2)
            )
            return jsonify({'success': False, 'error': reject_msg, 'lock': 'busy'}), 409
        except ImportError:
            pass

        # 预检：如果上一个任务已被取消但锁仍被持有，短暂等待释放
        if _execution_lock.locked():
            with _case_run_lock:
                prev_job = _case_run_jobs.get(user_id)
                prev_cancelled = prev_job and prev_job.get('cancel_requested')
            if prev_cancelled:
                uat_logger.info("检测到之前被取消的任务锁仍然持有，等待清理...")
                for _ in range(30):  # 最多等待 3 秒
                    if not _execution_lock.locked():
                        break
                    time.sleep(0.1)

        if not _execution_lock.acquire(blocking=True, timeout=10):
            if machine_lock_acquired:
                try:
                    from execution_lock import release as release_machine_lock

                    release_machine_lock()
                except ImportError:
                    pass
            reject_msg = '本机已有自动化任务在执行，请稍后再试。'
            _record_run_history_rejected(
                db, case_id, reject_msg, round(time.time() - start_time, 2)
            )
            return jsonify({
                'success': False,
                'error': reject_msg,
                'lock': 'busy',
            }), 409
        playwright_lock_acquired = True
        # 清除上一次的强制释放信号
        try:
            from playwright_automation import _execution_force_release_requested as _force_rel_event
            _force_rel_event.clear()
        except Exception:
            pass
        set_execution_in_progress(True)

        uat_logger.info(f"开始运行测试用例 #{case_id}: {case['name']}")
        uat_logger.info(f"测试用例共有 {len(steps)} 个步骤")
        with _case_run_lock:
            _case_run_jobs[user_id] = {
                'active': True,
                'cancel_requested': False,
                'case_id': case_id,
                'case_name': case.get('name', ''),
                'total_steps': len(steps),
                'completed_steps': 0,
                'current_step_order': 0,
                'current_action': '',
                'message': '准备执行...',
                'started_at': time.time(),
            }
        try:
            from captcha_engine import set_captcha_status_callback
    
            set_captcha_status_callback(lambda msg: _case_job_update(user_id, message=msg))
        except ImportError:
            pass
    
        try:
            from step_executor import case_steps_include_android, case_steps_include_web
            from mobile_routes import execute_mobile_case
    
            if case_steps_include_android(steps) and not case_steps_include_web(steps):
                try:
                    resp, status = execute_mobile_case(
                        case_id,
                        case,
                        steps,
                        db,
                        user_id,
                        start_time,
                        job_update=lambda **kw: _case_job_update(user_id, **kw),
                        job_cancelled=lambda: _case_run_cancelled(user_id),
                    )
                    with _case_run_lock:
                        if user_id in _case_run_jobs:
                            _case_run_jobs[user_id]["active"] = False
                    if machine_lock_acquired:
                        try:
                            from execution_lock import release as release_machine_lock
    
                            release_machine_lock()
                        except ImportError:
                            pass
                    return resp, status
                except Exception as mobile_early_exc:
                    uat_logger.error("Android 用例执行失败: %s", mobile_early_exc)
                    if machine_lock_acquired:
                        try:
                            from execution_lock import release as release_machine_lock
    
                            release_machine_lock()
                        except ImportError:
                            pass
                    return jsonify({"success": False, "error": str(mobile_early_exc)}), 500
        except ImportError:
            pass
        
        # 提取的文本
        extracted_text = ""
        # 预期结果
        expected_text = ""
        # 截图列表（失败截图路径）
        screenshots = []
        # 步骤结果列表（用于步骤级记录）
        step_results_list = []
        # 浏览器状态标记
        browser_closed_manually = False
        browser_started = False
        
        try:
            # 若用户仍处于元素/桌面拾取会话，执行前关闭拾取 UI，避免全屏遮罩与钩子干扰
            try:
                from element_picker import sync_stop_element_picker
    
                picker_stopped = sync_stop_element_picker()
                if (picker_stopped.get("desktop") or {}).get("was_active"):
                    uat_logger.info("检测到元素捕获仍在运行，执行前已自动关闭")
            except Exception:
                pass
            try:
                from web_dom_picker import get_web_dom_picker_status, sync_stop_web_dom_picker
    
                if get_web_dom_picker_status().get('active'):
                    uat_logger.info("检测到网页 DOM 捕获会话，执行前自动关闭")
                    sync_stop_web_dom_picker()
            except Exception:
                pass
            if bool(getattr(automation, '_selection_mode_active', False)):
                uat_logger.info("检测到旧版 Web 拾取器会话仍在，执行前自动关闭")
                try:
                    sync_disable_element_selection()
                except Exception:
                    pass
                try:
                    automation._selection_mode_active = False
                except Exception:
                    pass
    
            # 🔥 增强浏览器断连检测和自动恢复逻辑
            # 启动浏览器前先检查状态：如果浏览器已断连，先强制重置所有状态
            browser_disconnected = False
            try:
                if automation.browser is not None:
                    try:
                        browser_disconnected = not automation.browser.is_connected()
                    except Exception:
                        # is_connected() 抛出异常说明浏览器对象已失效
                        browser_disconnected = True
                else:
                    # browser 为 None 说明浏览器未启动或已被清理，这是正常状态
                    browser_disconnected = False
            except Exception as e:
                uat_logger.warning(f"⚠️ [浏览器检测] 检测浏览器状态时出错: {e}，假定已断连")
                browser_disconnected = True
            
            if browser_disconnected:
                uat_logger.warning("⚠️ [浏览器恢复] 检测到浏览器已断连，执行前强制重置所有状态")
                force_reset_execution_state()
                uat_logger.info("✅ [浏览器恢复] 状态已重置")
    
            # 纯桌面用例不启动 Playwright，避免先弹出 about:blank 空浏览器窗口
            from step_executor import case_steps_include_web, is_desktop_step
    
            def _ensure_browser_for_web_step() -> None:
                nonlocal browser_started
                if browser_started:
                    # 检测浏览器是否在步骤执行过程中被手动关闭
                    if automation.browser is None:
                        raise Exception("浏览器已关闭，无法继续执行")
                    try:
                        if not automation.browser.is_connected():
                            raise Exception("浏览器连接已断开，无法继续执行")
                    except Exception:
                        if automation.browser is None:
                            raise
                        raise Exception("浏览器已关闭，无法继续执行")
                    return
                sync_start_browser()
                browser_started = True
                initial_nav_url, nav_source = _resolve_case_navigation_url(
                    case=case, case_id=case_id, steps=steps
                )
                try:
                    from auth_batch_helpers import _case_role

                    if _case_role(case) == "login_feature":
                        sync_prepare_fresh_web_session(initial_nav_url or "")
                        if initial_nav_url:
                            uat_logger.info(
                                f"登录功能用例：已重置会话并导航({nav_source}) -> {initial_nav_url}"
                            )
                        return
                except Exception as prep_ex:
                    uat_logger.warning("登录用例会话重置失败，回退常规导航: %s", prep_ex)
                if initial_nav_url:
                    uat_logger.log_automation_step(
                        "navigate", initial_nav_url, f"首次 Web 步骤前导航({nav_source})"
                    )
                    sync_navigate_to(initial_nav_url)
    
            if not case_steps_include_web(steps):
                uat_logger.info("纯桌面/Android 用例，跳过 Playwright 浏览器启动")
    
            # 执行测试步骤
            try:
                try:
                    automation.set_case_run_hint(
                        case_name=(case.get("name") or ""),
                        step_descriptions=[
                            str(s.get("description") or "") for s in (steps or [])
                        ],
                    )
                except Exception:
                    pass
                total_step_count = len(steps)
                for step_index, step in enumerate(steps, start=1):
                    if _case_run_cancelled(user_id):
                        raise Exception("用户已停止执行")
                    # 检查跨线程强制释放信号（用户手动关闭浏览器时触发）
                    try:
                        from playwright_automation import _execution_lock_check_force_release as _check_force_rel
                        if _check_force_rel():
                            raise Exception("用户已停止执行")
                    except ImportError:
                        pass
    
                    action = step.get('action', '')
                    selector_type = step.get('selector_type', 'css')
                    # 变量替换：支持 {{变量名}} 语法
                    selector_value = db.resolve_variables(step.get('selector_value', ''), project_id=case.get('project_id'), case_id=case_id)
                    input_value = db.resolve_variables(step.get('input_value', ''), project_id=case.get('project_id'), case_id=case_id)
                    description = step.get('description', '')
                    # 添加iframe相关字段
                    enter_iframe = step.get('enter_iframe', False)
                    iframe_selector = step.get('iframe_selector', '')
                    iframe_for_step = _effective_step_iframe_selector(
                        automation, db, step, case.get('project_id'), case_id
                    )
                    locator_candidates = step.get('locator_candidates') or None
    
                    step_start_time = time.time()
                    # 🔥 修复：初始化为 error，只有执行成功才改为 success
                    step_status = 'error'
                    step_error = ''
                    step_screenshot = ''
                    
                    uat_logger.log_automation_step(action, selector_value or input_value, description)
                    _case_job_update(
                        user_id,
                        current_step_order=step_index,
                        current_action=action,
                        message=f"正在执行步骤 {step_index}/{total_step_count}: {action}",
                    )
                                                            
                    # 详细的调试日志，跟踪 action 值和执行的方法
                    uat_logger.debug(
                        f"执行步骤：ID={step.get('id')}, Action={action}, SelectorType={selector_type}, "
                        f"SelectorValue={selector_value}, InputValue={input_value}, EnterIframe={enter_iframe}, "
                        f"IframeSelector={iframe_selector}, IframeEffective={iframe_for_step}"
                    )
    
                    try:
                        from execution_factory import get_executor_factory
                        from step_executor import enrich_execution_step, is_desktop_step, is_mobile_step
    
                        factory = get_executor_factory()
                        exec_step = enrich_execution_step(step)
                        if is_desktop_step(exec_step):
                            try:
                                desk_step = dict(exec_step)
                                desk_step["_case_name"] = (case.get("name") or "").strip()
                                desk_step["selector_value"] = selector_value
                                desk_step["input_value"] = input_value
                                desk_result = factory.execute_desktop_step(
                                    desk_step,
                                    selector_value=selector_value,
                                    input_value=input_value,
                                )
                            except Exception as desk_exc:
                                if _case_run_cancelled(user_id):
                                    raise Exception("用户已停止执行")
                                from desktop_visual_engine import VisualMatchFailed
    
                                step_duration = round(time.time() - step_start_time, 3)
                                step_screenshot = getattr(desk_exc, "failure_screenshot", "") or ""
                                step_results_list.append({
                                    'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                                    'action': action, 'selector_value': selector_value,
                                    'input_value': input_value, 'description': description,
                                    'status': 'error', 'error': str(desk_exc),
                                    'screenshot': step_screenshot, 'duration': step_duration,
                                    'automation_layer': 'desktop',
                                })
                                if isinstance(desk_exc, VisualMatchFailed):
                                    desk_exc.step_id = step.get('id')  # type: ignore[attr-defined]
                                raise
                            from step_executor import validate_desktop_step_result
    
                            validate_desktop_step_result(desk_result, action)
                            step_status = 'success'
                            step_error = ''
                            step_screenshot = (desk_result or {}).get('screenshot') or ''
                            if (desk_result or {}).get("resolved_via"):
                                uat_logger.info(
                                    "桌面步骤 #%s 定位方式: %s",
                                    step.get("id"),
                                    desk_result.get("resolved_via"),
                                )
                            step_duration = round(time.time() - step_start_time, 3)
                            step_results_list.append({
                                'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                                'action': action, 'selector_value': selector_value,
                                'input_value': input_value, 'description': description,
                                'status': step_status, 'error': step_error,
                                'screenshot': step_screenshot, 'duration': step_duration,
                                'automation_layer': 'desktop',
                            })
                            _case_job_update(
                                user_id,
                                completed_steps=len(step_results_list),
                                message=f"已完成 {len(step_results_list)}/{len(steps)} 步",
                            )
                            continue
                        if is_mobile_step(exec_step):
                            try:
                                mob_step = dict(exec_step)
                                mob_step["selector_value"] = selector_value
                                mob_step["input_value"] = input_value
                                mob_result = factory.execute_mobile_step(
                                    mob_step,
                                    selector_value=selector_value,
                                    input_value=input_value,
                                )
                            except Exception as mob_exc:
                                if _case_run_cancelled(user_id):
                                    raise Exception("用户已停止执行")
                                step_duration = round(time.time() - step_start_time, 3)
                                step_screenshot = (getattr(mob_exc, "failure_screenshot", None) or "")
                                step_results_list.append({
                                    'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                                    'action': action, 'selector_value': selector_value,
                                    'input_value': input_value, 'description': description,
                                    'status': 'error', 'error': str(mob_exc),
                                    'screenshot': step_screenshot, 'duration': step_duration,
                                    'automation_layer': 'android',
                                })
                                raise
                            from mobile_automation import validate_mobile_step_result
    
                            validate_mobile_step_result(mob_result, action)
                            step_status = 'success'
                            step_error = ''
                            step_screenshot = (mob_result or {}).get('screenshot') or ''
                            step_duration = round(time.time() - step_start_time, 3)
                            step_results_list.append({
                                'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                                'action': action, 'selector_value': selector_value,
                                'input_value': input_value, 'description': description,
                                'status': step_status, 'error': step_error,
                                'screenshot': step_screenshot, 'duration': step_duration,
                                'automation_layer': 'android',
                            })
                            _case_job_update(
                                user_id,
                                completed_steps=len(step_results_list),
                                message=f"已完成 {len(step_results_list)}/{len(steps)} 步",
                            )
                            continue
                    except ImportError:
                        pass
    
                    from desktop_automation import normalize_automation_layer as _norm_layer
    
                    if _norm_layer(step) == "desktop":
                        raise Exception(
                            f"桌面步骤未进入桌面执行器（action={action}，"
                            f"automation_layer={step.get('automation_layer')!r}）。"
                            f"请确认步骤自动化层为「桌面」后重试。"
                        )
                    if _norm_layer(step) == "android":
                        raise Exception(
                            f"Android 步骤未进入移动端执行器（action={action}，"
                            f"automation_layer={step.get('automation_layer')!r}）。"
                            f"请确认 ENABLE_MOBILE=1 且已安装 Appium 依赖。"
                        )
    
                    _ensure_browser_for_web_step()
                    
                    if action == 'navigate':
                        # 获取URL并进行有效性检查
                        raw_url = step.get('url') or step.get('input_value') or ''
                        fixed_url, url_err = _validate_and_fix_url(raw_url)
                        if url_err:
                            uat_logger.error(f"导航步骤URL无效: {url_err}")
                            raise Exception(url_err)
                        elif fixed_url:
                            uat_logger.log_automation_step("navigate", fixed_url, "导航到URL")
                            sync_navigate_to(fixed_url)
                        else:
                            uat_logger.warning("导航步骤URL为空，跳过")
                    elif action == 'click':
                                if not selector_value:
                                    raise Exception("点击步骤缺少选择器")
                                try:
                                    _repeat = _norm_click_repeat_count(step.get('click_repeat_count'))
                                    for _r in range(_repeat):
                                        with _case_run_lock:
                                            if bool(_case_run_jobs.get(user_id, {}).get('cancel_requested')):
                                                raise Exception("用户已停止执行")
                                        sync_click_element(
                                            selector_value,
                                            selector_type,
                                            iframe_selector=iframe_for_step,
                                            locator_candidates=locator_candidates,
                                        )
                                except Exception as click_error:
                                    uat_logger.error(f"执行点击操作时出错: {click_error}")
                                    raise
                    elif action == 'input':
                        if selector_value:
                            try:
                                try:
                                    safe_input_value = resolve_fill_step_text({
                                        'input_value': input_value,
                                        'description': step.get('description'),
                                    })
                                except Exception as fill_val_err:
                                    if input_value is None and step_description_implies_empty_input(
                                        step.get('description')
                                    ):
                                        safe_input_value = ""
                                    else:
                                        raise fill_val_err
                                uat_logger.info(f"🔍 准备执行输入操作: 步骤ID={step.get('id', 'unknown')}, 选择器类型={selector_type}, 选择器值={selector_value}, 输入值={safe_input_value!r}")
                                
                                # 🔥 添加详细的诊断信息
                                if len(selector_value) > 200:
                                    uat_logger.warning(f"⚠️ 检测到超长CSS选择器（{len(selector_value)}字符），建议优化选择器")
                                    uat_logger.warning(f"   当前选择器前100字符: {selector_value[:100]}...")
                                
                                # 🔥 检查选择器类型
                                if selector_type == "css" and "nth-child" in selector_value:
                                    uat_logger.warning(f"⚠️ 检测到使用nth-child定位，可能不够稳定，建议改用ID或类名")
                                
                                sync_fill_input(
                                        selector_value,
                                        safe_input_value,
                                        selector_type,
                                        iframe_selector=iframe_for_step,
                                        locator_candidates=locator_candidates,
                                    )
                                uat_logger.info(f"✅ 输入操作执行完成: 步骤ID={step.get('id', 'unknown')}")
                                
                            except Exception as input_error:
                                uat_logger.error(f"🔥 输入操作执行失败详情:")
                                uat_logger.error(f"   步骤ID: {step.get('id', 'unknown')}")
                                uat_logger.error(f"   选择器类型: {selector_type}")
                                uat_logger.error(f"   选择器值: {selector_value}")
                                uat_logger.error(f"   输入值: {input_value}")
                                uat_logger.error(f"   错误信息: {input_error}")
                                uat_logger.error(f"   iframe选择器: {iframe_for_step}")
                                
                                # 🔥 提供改进建议
                                if len(selector_value) > 200:
                                    uat_logger.error(f"   💡 建议: 选择器过长，尝试使用更简单的选择器，如ID、类名或文本定位")
                                if "nth-child" in selector_value:
                                    uat_logger.error(f"   💡 建议: 避免使用nth-child，改用更稳定的定位方式")
                                if "h-[" in selector_value or "w-[" in selector_value or "calc(" in selector_value:
                                    uat_logger.error(f"   💡 建议: 避免使用Tailwind动态类名，这些类名容易变化")
                                
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                        else:
                            uat_logger.error("输入操作缺少选择器，步骤不能执行")
                            raise Exception("输入操作缺少选择器")
                    elif action == 'batch_input':
                        pairs = parse_batch_input_lines(input_value or '')
                        if not pairs:
                            raise Exception("批量输入步骤缺少有效行（每行：选择器 + Tab 或逗号 + 文本，参见步骤说明）")
                        for bsel, bval in pairs:
                            with _case_run_lock:
                                if bool(_case_run_jobs.get(user_id, {}).get('cancel_requested')):
                                    raise Exception("用户已停止执行")
                            uat_logger.info(
                                f"批量输入: 选择器={bsel!r} 值长={len(bval or '')}"
                            )
                            sync_fill_input(
                                bsel,
                                bval,
                                selector_type,
                                iframe_selector=iframe_for_step,
                                locator_candidates=None,
                            )
                    elif action == 'hover':
                        if selector_value:
                            try:
                                sync_hover_element(selector_value, selector_type, iframe_selector=iframe_for_step)
                                # 悬停后等待页面响应
                                sync_wait_for_timeout(1000)
                            except Exception as hover_error:
                                uat_logger.error(f"执行悬停操作时出错: {hover_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                        else:
                            uat_logger.error("悬停操作缺少选择器")
                            raise Exception("悬停操作缺少选择器")
                    elif action == 'double_click':
                        if selector_value:
                            try:
                                sync_double_click_element(selector_value, selector_type, iframe_selector=iframe_for_step)
                                # 双击后等待页面响应
                                sync_wait_for_timeout(2000)
                            except Exception as double_click_error:
                                uat_logger.error(f"执行双击操作时出错: {double_click_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                        else:
                            uat_logger.error("双击操作缺少选择器")
                            raise Exception("双击操作缺少选择器")
                    elif action == 'right_click':
                        if selector_value:
                            try:
                                sync_right_click_element(selector_value, selector_type, iframe_selector=iframe_for_step)
                                # 右键点击后等待页面响应
                                sync_wait_for_timeout(1000)
                            except Exception as right_click_error:
                                uat_logger.error(f"执行右键点击操作时出错: {right_click_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                        else:
                            uat_logger.error("右键点击操作缺少选择器")
                            raise Exception("右键点击操作缺少选择器")
                    elif action == 'wait':
                        # 修复：wait操作应该等待时间，而不是等待选择器
                        if input_value:
                            try:
                                # 将输入值转换为毫秒（支持秒为单位）
                                wait_time = int(input_value) * 1000 if int(input_value) < 1000 else int(input_value)
                                sync_wait_for_timeout(wait_time)
                            except ValueError:
                                uat_logger.error(f"无效的等待时间值: {input_value}")
                                # 🔥 修复：无效的等待时间应该视为测试失败
                                raise Exception(f"无效的等待时间值: {input_value}")
                        else:
                            # 🔥 修复：如果没有输入值，应该视为失败（或者使用默认值但记录警告）
                            uat_logger.warning("等待操作缺少输入值，使用默认值1000毫秒")
                            sync_wait_for_timeout(1000)
                    elif action == 'select':
                        # 修复：添加下拉框选择操作
                        if selector_value and input_value:
                            try:
                                sync_select_option(selector_value, input_value, selector_type, iframe_selector=iframe_for_step)
                                # 选择后等待页面响应
                                sync_wait_for_timeout(1000)
                            except Exception as select_error:
                                uat_logger.error(f"执行下拉框选择操作时出错: {select_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                    elif action == 'date':
                        # 新增：日期选择器操作
                        if selector_value and input_value:
                            try:
                                sync_select_date(selector_value, input_value)
                                # 选择日期后等待页面响应
                                sync_wait_for_timeout(1000)
                            except Exception as date_error:
                                uat_logger.error(f"执行日期选择操作时出错: {date_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                    elif action == 'scroll':
                        try:
                            _run_db_step_scroll(
                                input_value or "",
                                iframe_selector=iframe_for_step,
                            )
                            # 滚动后等待页面响应
                            sync_wait_for_timeout(1500)
                        except Exception as scroll_error:
                            uat_logger.error(f"执行滚动操作时出错: {scroll_error}")
                            # 直接抛出错误，视为测试用例执行失败
                            raise
                    elif action == 'swipe':
                        if selector_value:
                            direction = 'up'
                            distance = 100
                            if input_value:
                                # 解析 input_value 中的方向和距离 (格式: direction:distance)
                                parts = input_value.split(':')
                                if len(parts) == 2:
                                    direction = parts[0]
                                    try:
                                        distance = int(parts[1])
                                    except ValueError:
                                        uat_logger.warning(f"无效的滑动距离值: {parts[1]}，使用默认值 100")
                                else:
                                    # 兼容旧格式 (只有方向)
                                    direction = input_value
                            try:
                                sync_swipe_element(selector_value, direction, distance, selector_type, iframe_selector=iframe_for_step)
                                # 滑动后等待页面响应
                                sync_wait_for_timeout(1500)
                            except Exception as swipe_error:
                                uat_logger.error(f"执行滑动操作时出错: {swipe_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                    elif action == 'verify':
                        # 验证操作处理
                        verify_type = input_value if input_value else 'auto'
                        try:
                            sync_verify_element(
                                selector=selector_value,
                                verify_type=verify_type,
                                selector_type=selector_type,
                                iframe_selector=iframe_for_step,
                                locator_candidates=locator_candidates,
                                captcha_max_attempts=step.get('captcha_max_attempts'),
                            )
                            # 验证后等待页面响应
                            sync_wait_for_timeout(1500)
                        except Exception as verify_error:
                            uat_logger.error(f"执行验证操作时出错: {verify_error}")
                            # 直接抛出错误，视为测试用例执行失败
                            raise
                    elif action == 'extract_text' or action == 'text_compare':
                        extracted_text, expected_text = _run_extract_text_automation_step(
                            action,
                            step,
                            selector_value,
                            input_value,
                            description,
                            selector_type,
                            iframe_for_step,
                            locator_candidates=locator_candidates,
                        )
                    elif action == 'extract_json':
                        if selector_value:
                            # 提取元素JSON数据
                            try:
                                json_data = sync_extract_element_json(selector_value, selector_type)
                                uat_logger.info(f"提取到JSON数据: {json_data}")
                                # 保存到extracted_text变量，以便在结果中显示
                                extracted_text = str(json_data)
                            except Exception as extract_error:
                                uat_logger.error(f"提取JSON数据失败: {extract_error}")
                                # 🔥 修复：提取JSON数据失败应该视为测试失败
                                raise Exception(f"提取JSON数据失败: {extract_error}")
                        else:
                            uat_logger.error("提取JSON数据时缺少选择器")
                            # 🔥 修复：缺少选择器应该视为测试失败
                            raise Exception("提取JSON数据时缺少选择器")
                        
                        # 提取后等待页面响应
                        sync_wait_for_timeout(1000)
                    elif action == 'assert':
                        extra_ex = _run_assert_automation_step(
                            step, selector_value, input_value, selector_type, iframe_for_step
                        )
                        if extra_ex is not None:
                            extracted_text = extra_ex
                    elif action == 'enter_iframe':
                        if selector_value:
                            # 进入 iframe：状态在 automation.current_iframe，下游步骤由 _effective_step_iframe_selector 继承
                            try:
                                sync_enter_iframe(selector_value, selector_type)
                                uat_logger.info(f"✅ 成功进入iframe: {selector_value}")
                            except Exception as enter_error:
                                uat_logger.error(f"执行进入iframe操作时出错: {enter_error}")
                                # 直接抛出错误，视为测试用例执行失败
                                raise
                        else:
                            uat_logger.warning("进入iframe操作缺少选择器")
                    elif action == 'exit_iframe':
                        # 跳出 iframe：清除 automation.current_iframe，后续步骤回到主文档（除非步骤自身勾选 iframe）
                        try:
                            sync_exit_iframe()
                            uat_logger.info("✅ 成功跳出iframe，返回主文档")
                        except Exception as exit_error:
                            uat_logger.error(f"执行跳出iframe操作时出错: {exit_error}")
                            # 直接抛出错误，视为测试用例执行失败
                            raise
                    else:
                        from desktop_automation import normalize_automation_layer as _norm_al_web
    
                        _layer_web = _norm_al_web(step)
                        _st_web = (selector_type or '').strip().lower()
                        if _layer_web == 'desktop' or _st_web == 'visual':
                            raise RuntimeError(
                                f"步骤 #{step.get('id')} 为桌面/visual 步骤，但未通过桌面执行器完成"
                                f"（action={action!r}，automation_layer={step.get('automation_layer')!r}）。"
                                "请在步骤编辑器中确认自动化层为「桌面」并保存后重试。"
                            )
    
                    # 🔥 修复：步骤执行到这里说明成功，更新状态为 success
                    step_status = 'success'
                    step_error = ''
                    
                    # ⭐⭐ 记录成功步骤结果
                    step_duration = round(time.time() - step_start_time, 3)
                    step_results_list.append({
                        'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                        'action': action, 'selector_value': selector_value,
                        'input_value': input_value, 'description': description,
                        'status': step_status, 'error': step_error,
                        'screenshot': step_screenshot, 'duration': step_duration
                    })
                    _case_job_update(
                        user_id,
                        completed_steps=len(step_results_list),
                        message=f"已完成 {len(step_results_list)}/{len(steps)} 步",
                    )
                
                if not steps:
                    raise RuntimeError("用例没有可执行的步骤")
    
                # 计算执行时间
                duration = round(time.time() - start_time, 2)
                
                uat_logger.info(f"测试用例 #{case_id} 运行成功，耗时: {duration}秒")
                
                # 保存运行历史记录，并写入步骤结果
                try:
                    run_id = db.create_run_history(case_id, 'success', duration, "", extracted_text, expected_text)
                    for sr in step_results_list:
                        db.create_step_result(run_id, sr['step_id'], sr['step_order'], sr['action'],
                            sr['selector_value'], sr['input_value'], sr['description'],
                            sr['status'], sr['error'], sr['screenshot'], sr['duration'])
                    try:
                        from ai_memory_store import ingest_successful_run, memory_ingest_run_success_enabled
    
                        if memory_ingest_run_success_enabled():
                            tid = db.get_user_tenant_id(user_id)
                            ingest_successful_run(
                                user_id,
                                tid,
                                case_id,
                                case.get('name') or '',
                                case.get('url') or '',
                                duration,
                                run_id,
                            )
                    except Exception as _mem_run_e:
                        uat_logger.debug("memory ingest successful run: %s", _mem_run_e)
                    sync_run_to_team_server(
                        case_id, 'success', duration, step_results=step_results_list
                    )
                except Exception as history_error:
                    uat_logger.warning(f"保存运行历史记录失败: {history_error}")
                
                if browser_started:
                    try:
                        sync_close_browser()
                    except Exception as close_error:
                        uat_logger.warning(f"关闭浏览器时出错: {close_error}")
                
                return jsonify({
                    'success': True,
                    'status': 'success',
                    'duration': duration,
                    'message': '测试用例运行成功',
                    'step_count': len(step_results_list)
                })
                
            except Exception as e:
                # 执行失败时的处理
                duration = round(time.time() - start_time, 2)
    
                # 用户已点「停止执行」/ 浏览器被手动关闭：不因关应用/断连等中间错误弹失败
                _inner_err = str(e)
                if _case_run_cancelled(user_id) or '用户已停止执行' in _inner_err or '浏览器已关闭' in _inner_err or '浏览器连接已断开' in _inner_err:
                    stop_msg = '浏览器已关闭' if '浏览器' in _inner_err else '用户已停止执行'
                    uat_logger.info(f"测试用例 #{case_id} 执行结束: {stop_msg}（忽略步骤异常: {_inner_err[:300]}）")
                    try:
                        run_id = db.create_run_history(
                            case_id, 'stopped', duration, stop_msg, extracted_text, expected_text
                        )
                        for sr in step_results_list:
                            db.create_step_result(
                                run_id,
                                sr['step_id'],
                                sr['step_order'],
                                sr['action'],
                                sr['selector_value'],
                                sr['input_value'],
                                sr['description'],
                                sr['status'],
                                sr['error'],
                                sr['screenshot'],
                                sr['duration'],
                            )
                    except Exception as history_error:
                        uat_logger.warning(f"保存停止运行历史失败: {history_error}")
                    try:
                        force_reset_execution_state()
                    except Exception:
                        pass
                    return jsonify({
                        'success': False,
                        'status': 'stopped',
                        'duration': duration,
                        'error': stop_msg,
                        'stopped': True,
                    })
    
                error_msg = str(e)
                _captcha_manual = False
                try:
                    from captcha_recovery import CaptchaManualRequiredError as _CMR
    
                    _cause_cap = e
                    while _cause_cap is not None:
                        if isinstance(_cause_cap, _CMR):
                            _captcha_manual = True
                            error_msg = str(_cause_cap)
                            break
                        _cause_cap = getattr(_cause_cap, "__cause__", None)
                except ImportError:
                    pass
                
                # 将会话断开类错误单独归类
                # 注意：不要用单词 page/target/context 等做子串匹配，否则 “timeout waiting for …” 等普通错误会被误判。
                _el = error_msg.lower()
                _disconnect_patterns = (
                    'has been closed',
                    'browser has been closed',
                    'browser was closed',
                    'browser closed',
                    'target page, context or browser has been closed',
                    'page was closed',
                    'context was closed',
                    'browser disconnected',
                    'connection closed',
                    'connection lost',
                    'websocket',
                    'econnreset',
                    'broken pipe',
                    'execution context was destroyed',
                    'session deleted',
                    'browser crashed',
                    ' crashed',
                    'disconnected',
                )
                if any(p in _el for p in _disconnect_patterns):
                    browser_closed_manually = True
                    error_msg = (
                        "浏览器或自动化连接已中断（无头模式下常见于进程崩溃、资源不足、超时或被系统结束；"
                        "不一定是手动关闭窗口）"
                    )
                    uat_logger.warning(
                        f"测试用例 #{case_id} 执行中断（会话/连接断开）: {error_msg} | 原始: {str(e)[:800]}"
                    )
                    # 🔥 浏览器被关闭时强制重置所有状态（包括执行锁和浏览器引用）
                    try:
                        force_reset_execution_state()
                    except Exception:
                        pass
                else:
                    uat_logger.error(f"测试用例 #{case_id} 运行失败: {error_msg}")
    
                # 失败时自动截图（桌面 visual 失败已生成元素 ROI 对比图，不再截浏览器整页）
                failure_screenshot = ''
                visual_match_failed = None
                _cause = e
                while _cause is not None:
                    try:
                        from desktop_visual_engine import VisualMatchFailed
                    except ImportError:
                        break
                    if isinstance(_cause, VisualMatchFailed):
                        visual_match_failed = _cause
                        break
                    _cause = getattr(_cause, '__cause__', None)
                if visual_match_failed and getattr(visual_match_failed, 'failure_screenshot', None):
                    failure_screenshot = visual_match_failed.failure_screenshot
                elif _captcha_manual:
                    _cap_shot = None
                    _c = e
                    while _c is not None:
                        if getattr(_c, "screenshot_path", None):
                            _cap_shot = _c.screenshot_path
                            break
                        _c = getattr(_c, "__cause__", None)
                    if _cap_shot:
                        failure_screenshot = _cap_shot
                        screenshots.append(_cap_shot)
                elif not browser_closed_manually:
                    try:
                        screenshot_dir = os.path.join(os.getcwd(), 'screenshots')
                        os.makedirs(screenshot_dir, exist_ok=True)
                        screenshot_path = os.path.join(screenshot_dir, f'fail_{case_id}_{int(time.time())}.png')
                        sync_take_screenshot(screenshot_path)
                        failure_screenshot = screenshot_path
                        screenshots.append(screenshot_path)
                        uat_logger.info(f"失败截图已保存: {screenshot_path}")
                    except Exception as ss_error:
                        uat_logger.warning(f"截图失败: {ss_error}")
    
                # ⭐⭐ 记录失败步骤：将当前步骤添加到列表（如果该步骤未被记录）
                # 失败的步骤在 raise 后跳过了步骤记录代码，需要在此处补充记录
                # 通过检查 step_results_list 的最后一条记录是否是该失败步骤
                already_recorded = (
                    step_results_list and
                    step_results_list[-1].get('step_id') == step.get('id')
                ) if 'step' in dir() else False
                if not already_recorded and 'step' in dir() and step:
                    failed_step_duration = round(time.time() - step_start_time, 3) if 'step_start_time' in dir() else 0
                    step_results_list.append({
                        'step_id': step.get('id'), 'step_order': step.get('step_order', 0),
                        'action': step.get('action', ''), 'selector_value': step.get('selector_value', ''),
                        'input_value': step.get('input_value', ''), 'description': step.get('description', ''),
                        'status': 'error', 'error': error_msg,
                        'screenshot': failure_screenshot, 'duration': failed_step_duration
                    })
                    uat_logger.error(f"⭐⭐ [步骤记录] 当前失败步骤 ID={step.get('id')} 已记录到列表")
                elif step_results_list and step_results_list[-1]['status'] == 'success':
                    # 备用：如果上述逆转属不到，把最后一条成功记录改为失败
                    step_results_list[-1]['status'] = 'error'
                    step_results_list[-1]['error'] = error_msg
                    step_results_list[-1]['screenshot'] = failure_screenshot
                
                # 保存运行历史记录（确保即使浏览器关闭也能保存）
                try:
                    run_id = db.create_run_history(case_id, 'error', duration, error_msg, extracted_text, expected_text)
                    # 更新 screenshots 字段
                    try:
                        conn = __import__('sqlite3').connect(db.db_path)
                        conn.execute("UPDATE run_history SET screenshots = ? WHERE id = ?",
                                     (json.dumps(screenshots), run_id))
                        conn.commit(); conn.close()
                    except Exception:
                        pass
                    for sr in step_results_list:
                        db.create_step_result(run_id, sr['step_id'], sr['step_order'], sr['action'],
                            sr['selector_value'], sr['input_value'], sr['description'],
                            sr['status'], sr['error'], sr['screenshot'], sr['duration'])
                    uat_logger.info(f"运行历史记录已保存，Run ID: {run_id}")
                except Exception as history_error:
                    uat_logger.error(f"保存运行历史记录失败: {history_error}")
                
                if browser_started and not browser_closed_manually and not _captcha_manual:
                    try:
                        sync_close_browser()
                    except Exception:
                        pass
                
                _failed_sid = None
                if visual_match_failed and 'step' in dir() and step:
                    _failed_sid = step.get('id')
                return jsonify({
                    'success': False,
                    'status': 'error',
                    'duration': duration,
                    'error': error_msg,
                    'browser_closed': browser_closed_manually,
                    'stopped': ('用户已停止执行' in error_msg),
                    'failure_screenshot': failure_screenshot or None,
                    'failed_step_id': _failed_sid,
                    'need_relearn': bool(
                        visual_match_failed and getattr(visual_match_failed, 'need_relearn', False)
                    ),
                    'best_score': float(
                        getattr(visual_match_failed, 'best_score', 0.0) if visual_match_failed else 0.0
                    ),
                })
                
        except Exception as e:
            # 最外层异常处理 - 确保历史记录被保存
            duration = round(time.time() - start_time, 2)
            _err_str = str(e)
            # 用户主动停止 / 浏览器被手动关闭 → 归类为"已停止"
            if _case_run_cancelled(user_id) or '用户已停止执行' in _err_str or '浏览器已关闭' in _err_str or '浏览器连接已断开' in _err_str:
                stop_msg = '浏览器已关闭' if '浏览器' in _err_str else '用户已停止执行'
                uat_logger.info(f"测试用例 #{case_id} 执行结束: {stop_msg}")
                try:
                    db.create_run_history(case_id, 'stopped', duration, stop_msg, extracted_text, expected_text)
                except Exception:
                    pass
                try:
                    force_reset_execution_state()
                except Exception:
                    pass
                return jsonify({
                    'success': False,
                    'status': 'stopped',
                    'duration': duration,
                    'error': stop_msg,
                    'stopped': True,
                })
            error_msg = str(e)
            uat_logger.error(f"运行测试用例时发生严重错误: {error_msg}")
            
            # 尝试保存运行历史记录
            try:
                run_id = db.create_run_history(case_id, 'error', duration, f"执行异常: {error_msg}", extracted_text, expected_text)
                uat_logger.info(f"异常情况下运行历史记录已保存，Run ID: {run_id}")
            except Exception as history_error:
                uat_logger.error(f"异常情况下保存运行历史记录失败: {history_error}")
            
            return jsonify({
                'success': False,
                'error': error_msg,
                'stopped': False,
            }), 500
        finally:
            try:
                from captcha_engine import set_captcha_status_callback
    
                set_captcha_status_callback(None)
            except ImportError:
                pass
            if playwright_lock_acquired:
                set_execution_in_progress(False)
                try:
                    _execution_lock.release()
                except RuntimeError:
                    pass
            if machine_lock_acquired:
                try:
                    from execution_lock import release as release_machine_lock
    
                    release_machine_lock()
                except ImportError:
                    pass
            with _case_run_lock:
                job = _case_run_jobs.get(user_id)
                if job:
                    job['active'] = False
                    job['finished_at'] = time.time()
                    job['duration'] = round(time.time() - job.get('started_at', time.time()), 2)
    
    
def _case_run_status_payload(user_id: int) -> dict:
    """当前用户 Web 用例步骤运行任务状态（与 /api/cases/current-run/status 响应体一致）。"""
    with _case_run_lock:
        job = _case_run_jobs.get(user_id)
    if not job:
        return {
            'success': True,
            'active': False,
            'total_steps': 0,
            'completed_steps': 0,
            'progress': 0,
            'message': '暂无运行任务',
        }

    total_steps = max(1, int(job.get('total_steps', 0) or 0))
    completed_steps = int(job.get('completed_steps', 0) or 0)
    progress = min(100, int((completed_steps / total_steps) * 100))
    return {
        'success': True,
        'active': bool(job.get('active')),
        'case_id': job.get('case_id'),
        'case_name': job.get('case_name', ''),
        'total_steps': int(job.get('total_steps', 0) or 0),
        'completed_steps': completed_steps,
        'current_step_order': int(job.get('current_step_order', 0) or 0),
        'current_action': job.get('current_action', ''),
        'progress': progress,
        'cancel_requested': bool(job.get('cancel_requested')),
        'message': job.get('message', ''),
        'duration': job.get('duration'),
    }


def _dataset_current_run_payload(user_id: int) -> dict:
    """当前用户数据驱动任务状态（与 /api/datasets/current-run/status 响应体一致）。"""
    run_id, job = _get_current_user_dataset_job(user_id)
    if not job:
        return {'success': True, 'active': False, 'message': '暂无运行任务'}

    finished = bool(job.get('finished'))
    total = int(job.get('total', 0) or 0)
    completed = int(job.get('completed', 0) or 0)
    return {
        'success': True,
        'active': not finished,
        'finished': finished,
        'run_id': run_id,
        'dataset_id': job.get('dataset_id'),
        'case_id': job.get('case_id'),
        'total': total,
        'completed': completed,
        'successful_rows': int(job.get('successful_rows', 0) or 0),
        'failed_rows': int(job.get('failed_rows', 0) or 0),
        'current_row_index': job.get('current_row_index'),
        'error': job.get('error'),
    }


@app.route('/api/ui/current-runs/status', methods=['GET'])
@login_required
@api_error_handler
def api_ui_current_runs_status():
    """合并用例与数据驱动的运行状态，供全站轮询一次请求拿到全部信息。"""
    uid = current_user.id
    return jsonify({
        'success': True,
        'case': _case_run_status_payload(uid),
        'dataset': _dataset_current_run_payload(uid),
    })


@app.route('/api/cases/current-run/status', methods=['GET'])
@login_required
@api_error_handler
def api_case_run_status():
    return jsonify(_case_run_status_payload(current_user.id))


@app.route('/api/cases/current-run/stop', methods=['POST'])
@login_required
@api_error_handler
def api_stop_case_run():
    user_id = current_user.id
    with _case_run_lock:
        job = _case_run_jobs.get(user_id)
        if not job or not job.get('active'):
            # 幂等：避免前端在状态切换瞬间看到“没有任务在运行”的误报弹窗
            return jsonify({'success': True, 'message': '当前没有运行任务'})
        job['cancel_requested'] = True
        job['active'] = False
        job['message'] = '已停止执行'
        total_steps = int(job.get('total_steps', 0) or 0)
        job['completed_steps'] = total_steps
        job['finished_at'] = time.time()
        job['duration'] = round(time.time() - job.get('started_at', time.time()), 2)

    threading.Thread(target=_force_stop_browser_async, daemon=True, name='stop-case-run').start()

    return jsonify({'success': True, 'message': '已发送停止请求'})

@app.route('/run-history', methods=['GET'])
@login_required
def run_history_page():
    """运行历史记录页面"""
    return render_template('run_history.html')


@app.route('/api/run-history', methods=['GET'])
@login_required
def get_run_history():
    """获取所有运行历史记录（支持分页、按测试用例ID过滤、按项目ID过滤和搜索）"""
    try:
        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        case_id = request.args.get('case_id', type=int)
        project_id = request.args.get('project_id', type=int)
        search_text = request.args.get('search_text', type=str)
        raw_status = request.args.get('status', type=str) or ''
        status_filter = raw_status if raw_status in ('passed', 'failed') else None
        
        db = Database()
        history = db.get_all_run_history(page, page_size, case_id, search_text, project_id, status_filter=status_filter)
        total = db.get_run_history_count(case_id, search_text, project_id, status_filter=status_filter)
        try:
            from ai_modules.execute.history_ops_summary import enrich_run_history_record

            history = [enrich_run_history_record(h) for h in (history or [])]
        except Exception:
            pass

        return jsonify({
            'success': True,
            'history': history,
            'total': total,
            'page': page,
            'page_size': page_size
        })
    except Exception as e:
        uat_logger.error(f"获取运行历史记录失败: {e}")
        return jsonify({
            'success': False,
            'error': f'获取运行历史记录失败: {str(e)}'
        }), 500

@app.route('/api/run-history/<int:record_id>', methods=['DELETE'])
def delete_run_history(record_id):
    """删除运行历史记录"""
    try:
        db = Database()
        success = db.delete_run_history(record_id)
        if success:
            return jsonify({
                'success': True,
                'message': '运行历史记录删除成功'
            })
        else:
            return jsonify({
                'success': False,
                'error': '运行历史记录不存在'
            }), 404
    except Exception as e:
        uat_logger.error(f"删除运行历史记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/run-history', methods=['DELETE'])
def delete_all_run_history():
    """删除所有运行历史记录"""
    try:
        db = Database()
        success = db.delete_all_run_history()
        if success:
            return jsonify({
                'success': True,
                'message': '所有运行历史记录删除成功'
            })
        else:
            return jsonify({
                'success': True,
                'message': '没有运行历史记录可删除'
            })
    except Exception as e:
        uat_logger.error(f"删除所有运行历史记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/run-history/<int:record_id>', methods=['GET'])
def get_run_history_detail(record_id):
    """获取运行历史记录详情（含门禁摘要 / 证据包与 CI 链接）。"""
    try:
        db = Database()
        record = db.get_run_history_detail(record_id)
        if record:
            try:
                from ai_modules.execute.history_ops_summary import enrich_run_history_record

                record = enrich_run_history_record(record)
            except Exception:
                pass
            return jsonify({
                'success': True,
                'record': record
            })
        else:
            return jsonify({
                'success': False,
                'error': '运行历史记录不存在'
            }), 404
    except Exception as e:
        uat_logger.error(f"获取运行历史记录详情失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/cases/<int:case_id>/run-history', methods=['GET'])
def get_case_run_history(case_id):
    """获取指定测试用例的运行历史记录"""
    try:
        db = Database()
        history = db.get_case_run_history(case_id)
        return jsonify({
            'success': True,
            'history': history
        })
    except Exception as e:
        uat_logger.error(f"获取测试用例运行历史记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== 测试报告API ====================

# 测试报告页面
@app.route('/test_report')
@login_required
def test_report():
    return render_template('test_report.html')


@app.route('/cicd')
@login_required
def cicd_page():
    """CI/CD 集成入口页（可发现触发与近期运行）。"""
    return render_template('cicd.html')


@app.route('/trace-hub')
@login_required
def trace_hub_page():
    """证据中心：门禁摘要、证据导出、Skill 沉淀入口。"""
    return render_template('trace_hub.html')


@app.route('/execution-farm')
@login_required
def execution_farm_page():
    """执行节点页：登记远程执行机 / Desktop Gateway。"""
    return render_template('execution_farm.html')

# API: 获取测试统计概览
@app.route('/api/report/overview', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_report_overview():
    """获取测试统计概览"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        
        report_generator = TestReportGenerator()
        overview = report_generator.get_statistics_overview(
            project_id, start_date, end_date, case_category
        )
        
        return jsonify({
            'success': True,
            'data': overview
        })
    except Exception as e:
        uat_logger.error(f"获取测试统计概览失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/report/ops-summary', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_report_ops_summary():
    """治理看板：HITL/Risk/证据/CI（含无 case 的跨端历史）。"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None and str(project_id).strip() != '':
            project_id = int(project_id)
        else:
            project_id = None
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        report_generator = TestReportGenerator()
        data = report_generator.get_ops_governance_summary(
            project_id, start_date, end_date, case_category
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        uat_logger.error(f"获取治理看板失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/report/customer-audit-pack', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@feature_required('customer_audit_export')
@api_error_handler
@log_api_request
def api_report_customer_audit_pack():
    """客户向审计交付包 ZIP：索引 + 治理摘要 + 关键失败/门禁 Trace（不美化）。"""
    from ai_modules.execute.customer_audit_pack import build_customer_audit_pack
    from flask import send_file

    data = request.get_json(silent=True) or {}
    q = request.args

    def _pick(key, default=None):
        if data.get(key) is not None and str(data.get(key)).strip() != '':
            return data.get(key)
        if q.get(key) is not None and str(q.get(key)).strip() != '':
            return q.get(key)
        return default

    project_id = _pick('project_id')
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'project_id 必须是整数'}), 400
    start_date = _pick('start_date')
    end_date = _pick('end_date')
    case_category = _pick('case_category')
    try:
        scan_limit = int(_pick('scan_limit', 500) or 500)
    except (TypeError, ValueError):
        scan_limit = 500
    try:
        embed_limit = int(_pick('embed_limit', 15) or 15)
    except (TypeError, ValueError):
        embed_limit = 15
    fmt = str(_pick('format', 'zip') or 'zip').lower()
    want_zip = fmt in ('zip', 'download', '')

    exported = build_customer_audit_pack(
        project_id=project_id,
        start_date=start_date,
        end_date=end_date,
        case_category=case_category,
        scan_limit=scan_limit,
        embed_limit=embed_limit,
        make_zip=want_zip,
    )
    if want_zip:
        zp = exported.get('zip_path')
        if not zp or not os.path.isfile(zp):
            return jsonify({
                'success': False,
                'error': exported.get('error') or '审计包 ZIP 生成失败',
                'data': exported,
            }), 500
        return send_file(
            zp,
            mimetype='application/zip',
            as_attachment=True,
            download_name=exported.get('download_name') or 'customer_audit_pack.zip',
        )
    return jsonify({'success': True, 'data': exported})

# API: 获取状态分布
@app.route('/api/report/status-distribution', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_status_distribution():
    """获取状态分布"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        
        report_generator = TestReportGenerator()
        status_dist = report_generator.get_status_distribution(
            project_id, start_date, end_date, case_category
        )
        
        return jsonify({
            'success': True,
            'data': status_dist
        })
    except Exception as e:
        uat_logger.error(f"获取状态分布失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API: 获取耗时分布
@app.route('/api/report/duration-distribution', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_duration_distribution():
    """获取耗时分布"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        
        report_generator = TestReportGenerator()
        duration_dist = report_generator.get_duration_distribution(
            project_id, start_date, end_date, case_category
        )
        
        return jsonify({
            'success': True,
            'data': duration_dist
        })
    except Exception as e:
        uat_logger.error(f"获取耗时分布失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API: 获取趋势数据
@app.route('/api/report/trend', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_trend_data():
    """获取趋势数据"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        days = request.args.get('days', 30)
        if days is not None:
            days = int(days)
        
        report_generator = TestReportGenerator()
        trend_data = report_generator.get_trend_data(project_id, days)
        
        return jsonify({
            'success': True,
            'data': trend_data
        })
    except Exception as e:
        uat_logger.error(f"获取趋势数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API: 获取用例统计
@app.route('/api/report/case-statistics', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_case_statistics():
    """获取用例统计"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        
        report_generator = TestReportGenerator()
        case_stats = report_generator.get_case_statistics(
            project_id, start_date, end_date, case_category
        )
        
        return jsonify({
            'success': True,
            'data': case_stats
        })
    except Exception as e:
        uat_logger.error(f"获取用例统计失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API: 获取项目统计
@app.route('/api/report/project-statistics', methods=['GET'])
@api_error_handler
@log_api_request
def api_get_project_statistics():
    """获取项目统计"""
    try:
        project_id = request.args.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        case_category = request.args.get('case_category')
        
        report_generator = TestReportGenerator()
        project_stats = report_generator.get_project_statistics(
            project_id, start_date, end_date, case_category
        )
        
        return jsonify({
            'success': True,
            'data': project_stats
        })
    except Exception as e:
        uat_logger.error(f"获取项目统计失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# API: 导出报告
@app.route('/api/report/export', methods=['POST'])
@api_error_handler
@log_api_request
def api_export_report():
    """导出测试报告"""
    try:
        data = request.get_json(silent=True) or {}
        format_type = data.get('format', 'html')
        project_id = data.get('project_id')
        if project_id is not None:
            project_id = int(project_id)
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        case_category = data.get('case_category')
        filename = data.get('filename')
        
        # 收集报告数据
        report_generator = TestReportGenerator()
        report_data = {
            'overview': report_generator.get_statistics_overview(
                project_id, start_date, end_date, case_category
            ),
            'status_distribution': report_generator.get_status_distribution(
                project_id, start_date, end_date, case_category
            ),
            'duration_distribution': report_generator.get_duration_distribution(
                project_id, start_date, end_date, case_category
            ),
            'case_statistics': report_generator.get_case_statistics(
                project_id, start_date, end_date, case_category
            ),
            'project_statistics': report_generator.get_project_statistics(
                project_id, start_date, end_date, case_category
            ),
        }
        
        # 导出报告
        exporter = ReportExporter()
        
        if format_type == 'html':
            filepath = exporter.export_to_html(report_data, filename)
        elif format_type == 'excel':
            filepath = exporter.export_to_excel(report_data, filename)
        elif format_type == 'pdf':
            filepath = exporter.export_to_pdf(report_data, filename)
        else:
            return jsonify({
                'success': False,
                'error': f'不支持的导出格式: {format_type}'
            }), 400
        
        return jsonify({
            'success': True,
            'filepath': filepath
        })
    except Exception as e:
        uat_logger.error(f"导出报告失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/report/export/download', methods=['GET', 'POST'])
@login_required
@api_error_handler
@log_api_request
def api_export_report_download():
    """导出报告并以附件下载，由浏览器/系统对话框选择保存位置。"""
    from flask import send_file

    try:
        data = request.get_json(silent=True) or {}
        if request.method == 'GET':
            data = {
                'format': request.args.get('format', 'html'),
                'project_id': request.args.get('project_id'),
                'start_date': request.args.get('start_date'),
                'end_date': request.args.get('end_date'),
                'case_category': request.args.get('case_category'),
                'filename': request.args.get('filename'),
            }
        format_type = (data.get('format') or 'html').strip().lower()
        project_id = data.get('project_id')
        if project_id is not None and str(project_id).strip() != '':
            project_id = int(project_id)
        else:
            project_id = None
        start_date = data.get('start_date') or None
        end_date = data.get('end_date') or None
        case_category = data.get('case_category') or None
        filename = data.get('filename')

        report_generator = TestReportGenerator()
        report_data = {
            'overview': report_generator.get_statistics_overview(
                project_id, start_date, end_date, case_category
            ),
            'status_distribution': report_generator.get_status_distribution(
                project_id, start_date, end_date, case_category
            ),
            'duration_distribution': report_generator.get_duration_distribution(
                project_id, start_date, end_date, case_category
            ),
            'case_statistics': report_generator.get_case_statistics(
                project_id, start_date, end_date, case_category
            ),
            'project_statistics': report_generator.get_project_statistics(
                project_id, start_date, end_date, case_category
            ),
        }
        exporter = ReportExporter()
        if format_type == 'html':
            filepath = exporter.export_to_html(report_data, filename)
            mime = 'text/html; charset=utf-8'
        elif format_type == 'excel':
            filepath = exporter.export_to_excel(report_data, filename)
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format_type == 'pdf':
            filepath = exporter.export_to_pdf(report_data, filename)
            mime = 'application/pdf'
        else:
            return jsonify({'success': False, 'error': f'不支持的导出格式: {format_type}'}), 400
        if not filepath or not os.path.isfile(filepath):
            return jsonify({'success': False, 'error': '导出文件未生成'}), 500
        return send_file(
            filepath,
            mimetype=mime,
            as_attachment=True,
            download_name=os.path.basename(filepath),
        )
    except Exception as e:
        uat_logger.error(f"导出报告下载失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 变量管理API ====================

@app.route('/api/variables', methods=['GET'])
@login_required
def api_get_variables():
    scope = request.args.get('scope')
    project_id = request.args.get('project_id', type=int)
    case_id = request.args.get('case_id', type=int)
    _db = Database()
    variables = _db.get_variables(scope, project_id, case_id)
    return jsonify({'success': True, 'variables': variables})

@app.route('/api/variables', methods=['POST'])
@login_required
def api_create_variable():
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    value = data.get('value', '')
    scope = data.get('scope', 'global')
    if not name:
        return jsonify({'success': False, 'error': '变量名不能为空'}), 400
    if scope not in ('global', 'project', 'case'):
        return jsonify({'success': False, 'error': '无效的作用域'}), 400
    _db = Database()
    var_id = _db.create_variable(name, value, scope, data.get('project_id'), data.get('case_id'), data.get('description', ''))
    return jsonify({'success': True, 'var_id': var_id})

@app.route('/api/variables/<int:var_id>', methods=['PUT'])
@login_required
def api_update_variable(var_id):
    data = request.get_json(silent=True) or {}
    _db = Database()
    success = _db.update_variable(var_id, data.get('name'), data.get('value'), data.get('description'))
    return jsonify({'success': success})

@app.route('/api/variables/<int:var_id>', methods=['DELETE'])
@login_required
def api_delete_variable(var_id):
    _db = Database()
    success = _db.delete_variable(var_id)
    return jsonify({'success': success})

# ==================== 步骤执行结果API ====================

@app.route('/api/run-history/<int:record_id>/steps', methods=['GET'])
def get_run_step_results(record_id):
    """获取某次运行的步骤级结果"""
    try:
        _db = Database()
        results = _db.get_step_results(record_id)
        return jsonify({'success': True, 'step_results': results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 定时调度API ====================

@app.route('/api/schedules', methods=['GET'])
@login_required
@feature_required('schedule')
def api_get_schedules():
    _db = Database()
    schedules = _db.get_all_schedules()
    return jsonify({'success': True, 'schedules': schedules})

@app.route('/api/schedules', methods=['POST'])
@login_required
@feature_required('schedule')
@role_required('admin', 'tester')
def api_create_schedule():
    try:
        data = request.get_json(silent=True) or {}
        name = data.get('name', '').strip()
        case_ids = data.get('case_ids', [])
        cron_expr = data.get('cron_expr', '').strip()
        is_active = data.get('is_active', 1)
        if not name or not case_ids or not cron_expr:
            return jsonify({'success': False, 'error': '名称、用例ID列表和cron表达式不能为空'}), 400
        _db = Database()
        project_id = data.get('project_id')
        raw_ec = data.get('execution_count', -1)
        try:
            execution_count = int(raw_ec)
        except (TypeError, ValueError):
            execution_count = -1
        schedule_id = _db.create_schedule(name, case_ids, cron_expr, project_id=project_id, is_active=is_active, execution_count=execution_count)
        _sync_schedule_job(schedule_id)
        return jsonify({'success': True, 'schedule_id': schedule_id})
    except Exception as e:
        uat_logger.error(f"创建定时任务失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedules/<int:schedule_id>', methods=['PUT'])
@login_required
@feature_required('schedule')
@role_required('admin', 'tester')
def api_update_schedule(schedule_id):
    try:
        data = request.get_json(silent=True) or {}
        _db = Database()
        raw_ec = data.get('execution_count')
        execution_count = None
        if raw_ec is not None:
            try:
                execution_count = int(raw_ec)
            except (TypeError, ValueError):
                execution_count = None
        success = _db.update_schedule(schedule_id,
            name=data.get('name'), cron_expr=data.get('cron_expr'),
            is_active=data.get('is_active'), case_ids=data.get('case_ids'),
            project_id=data.get('project_id'), execution_count=execution_count)
        if success:
            _sync_schedule_job(schedule_id)
        return jsonify({'success': success})
    except Exception as e:
        uat_logger.error(f"更新定时任务失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedules/<int:schedule_id>', methods=['DELETE'])
@login_required
@feature_required('schedule')
@role_required('admin', 'tester')
def api_delete_schedule(schedule_id):
    _db = Database()
    success = _db.delete_schedule(schedule_id)
    try:
        scheduler.remove_job(f'schedule_{schedule_id}')
    except Exception:
        pass
    return jsonify({'success': success})

@app.route('/api/schedules/<int:schedule_id>/run', methods=['POST'])
@login_required
@feature_required('schedule')
@role_required('admin', 'tester')
def api_run_schedule_now(schedule_id):
    """立即执行定时任务"""
    _db = Database()
    schedules = _db.get_all_schedules()
    schedule = None
    for s in schedules:
        if s['id'] == schedule_id:
            schedule = s
            break

    if not schedule:
        return jsonify({'success': False, 'error': '定时任务不存在'}), 404

    try:
        ec = int(schedule.get('execution_count', 0))
    except (TypeError, ValueError):
        ec = 0
    if ec == 0:
        return jsonify({'success': False, 'error': '剩余执行次数为 0（请设为 -1 无限次或大于 0）'}), 400

    try:
        # 异步执行
        import threading
        thread = threading.Thread(
            target=_run_scheduled_cases,
            args=(schedule_id, schedule['case_ids'])
        )
        thread.start()
        return jsonify({'success': True, 'message': '任务已开始执行'})
    except Exception as e:
        uat_logger.error(f"手动执行任务失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedules/<int:schedule_id>/history', methods=['GET'])
@login_required
@feature_required('schedule')
def api_get_schedule_history(schedule_id):
    """获取定时任务执行历史"""
    _db = Database()
    history = _db.get_schedule_history(schedule_id)
    return jsonify({'success': True, 'history': history})

# ==================== API Token 管理 ====================

@app.route('/api/tokens', methods=['GET'])
@login_required
@role_required('admin')
def api_get_tokens():
    _db = Database()
    tokens = _db.get_all_tokens()
    return jsonify({'success': True, 'tokens': tokens})

@app.route('/api/tokens', methods=['POST'])
@login_required
@role_required('admin')
def api_create_token():
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Token名称不能为空'}), 400
    token_val = secrets.token_urlsafe(32)
    _db = Database()
    token_id = _db.create_api_token(name, token_val, data.get('project_id'), data.get('expires_at'))
    return jsonify({'success': True, 'token_id': token_id, 'token': token_val,
                    'message': '请妥善保存此Token，它只显示一次！'})

@app.route('/api/tokens/<int:token_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def api_revoke_token(token_id):
    _db = Database()
    success = _db.revoke_token(token_id)
    return jsonify({'success': success})

# ==================== Webhook/CI 触发接口 ====================

def _ci_resolve_case_ids(data: dict, _db) -> Tuple[list, Optional[str]]:
    """从请求体解析 case_ids；返回 (ids, error)。默认排除 review_status=pending/rejected。"""
    include_pending = bool(
        data.get("include_pending_review")
        or data.get("include_pending")
        or data.get("allow_pending")
    )
    case_ids = data.get("case_ids") or data.get("caseIds") or []
    if isinstance(case_ids, str):
        case_ids = [x.strip() for x in case_ids.split(",") if x.strip()]
    if case_ids:
        out = []
        for x in case_ids:
            try:
                out.append(int(x))
            except (TypeError, ValueError):
                return [], f"无效 case_id: {x}"
        try:
            from ai_modules.code_intel.review import filter_ci_case_ids

            filtered = filter_ci_case_ids(_db, out, include_pending=include_pending)
            kept = filtered.get("kept") or []
            if not kept:
                skipped = filtered.get("skipped") or []
                return [], f"无可用用例（均未激活或不可用）: {skipped[:5]}"
            return kept, None
        except Exception:
            return out, None

    project_id = data.get("project_id")
    if project_id is not None:
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            return [], "project_id 无效"
        cases = _db.get_project_cases(project_id, case_type="ui") or []
        try:
            from ai_modules.code_intel.review import case_is_ci_eligible

            ids = [
                c["id"]
                for c in cases
                if isinstance(c, dict)
                and c.get("id") is not None
                and case_is_ci_eligible(c, include_pending=include_pending)
            ]
        except Exception:
            ids = [c["id"] for c in cases if isinstance(c, dict) and c.get("id") is not None]
        if not ids:
            return [], "项目没有可执行用例（或均为待审核）"
        return ids, None

    suite_id = data.get("suite_id") or data.get("suiteId")
    if suite_id is not None:
        return [], "suite_id 尚未接入，请改用 project_id 或 case_ids"

    return [], "请提供 case_ids 或 project_id"


def _ci_want_async(data: dict) -> bool:
    """async=true / sync=false / mode=async → 异步。默认同步（兼容 0c-1）。"""
    if data.get("async") is True:
        return True
    if data.get("sync") is False:
        return True
    mode = str(data.get("mode") or "").strip().lower()
    return mode in ("async", "background", "queue")


def _ci_run_batch_and_finalize(
    *,
    run_id: str,
    case_ids: list,
    project_id,
    trigger_source: str,
    build_id: str,
    git_sha: str,
    branch: str,
    suite_name: str,
    user_id,
    tenant_id,
    sync_mode: bool,
):
    """执行批量用例并写入 run 终态；可选投递 callback。"""
    from auth_batch_helpers import count_batch_gate_failures
    from ci_adapter import (
        build_run_record_from_batch,
        deliver_run_callback,
        finalize_run_from_batch,
        mark_run_running,
    )

    if not sync_mode:
        mark_run_running(run_id)

    _db = Database()
    _exec_ctx = ExecutionContext(
        user_id=user_id,
        tenant_id=tenant_id,
        trigger="ci",
        on_case_failure=_on_case_execution_failure,
        extra={
            "project_id": project_id,
            "build_id": build_id,
            "git_sha": git_sha,
            "trigger_source": trigger_source,
            "ci_run_id": run_id,
        },
    )
    try:
        results = sync_execute_multiple_test_cases(case_ids, _db, execution_context=_exec_ctx)
    except Exception as e:
        uat_logger.error(f"CI run 执行失败 run_id={run_id}: {e}")
        results = {
            "total_cases": len(case_ids),
            "successful_cases": 0,
            "failed_cases": len(case_ids),
            "case_results": [
                {
                    "case_id": cid,
                    "case_name": str(cid),
                    "status": "error",
                    "error": str(e),
                }
                for cid in case_ids
            ],
            "error": str(e),
        }
    finally:
        try:
            sync_close_browser()
        except Exception:
            pass

    if not isinstance(results, dict):
        results = {"case_results": [], "error": "执行返回无效"}

    try:
        gate_fails = count_batch_gate_failures(results.get("case_results") or [])
        if gate_fails and int(results.get("failed_cases") or 0) < gate_fails:
            results["failed_cases"] = gate_fails
            results["successful_cases"] = max(
                0, int(results.get("total_cases") or len(case_ids)) - gate_fails
            )
    except Exception:
        pass

    if sync_mode:
        record = build_run_record_from_batch(
            results,
            run_id=run_id,
            project_id=project_id,
            case_ids=case_ids,
            trigger_source=trigger_source,
            build_id=build_id,
            git_sha=git_sha,
            branch=branch,
            suite_name=suite_name,
        )
    else:
        record = finalize_run_from_batch(run_id, results, suite_name=suite_name)

    try:
        deliver_run_callback(run_id)
    except Exception as cb_err:
        uat_logger.warning(f"CI callback 投递异常 run_id={run_id}: {cb_err}")
    try:
        from ai_modules.enterprise.ci_unified_sync import on_testory_run_finished

        on_testory_run_finished(run_id)
    except Exception as sync_err:
        uat_logger.warning(f"CI sync 刷新异常 run_id={run_id}: {sync_err}")
    try:
        from ai_modules.code_intel.pipeline import on_ci_run_finished

        on_ci_run_finished(run_id, db_factory=lambda: Database())
    except Exception as heal_err:
        uat_logger.warning(f"code-change heal 钩子异常 run_id={run_id}: {heal_err}")
    from ci_adapter import get_run

    return get_run(run_id) or record


@app.route('/api/ci/runs', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_runs_list():
    """列出近期 CI 运行（供 CI/CD 页展示）。"""
    from ci_adapter import list_runs

    try:
        limit = int(request.args.get('limit') or 30)
    except (TypeError, ValueError):
        limit = 30
    runs = list_runs(limit=limit)
    return jsonify({'success': True, 'ok': True, 'runs': runs, 'count': len(runs)})


@app.route('/api/ci/sync', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
@api_error_handler
def api_ci_sync_list():
    """列出统一门禁同步会话。"""
    from ai_modules.enterprise.ci_unified_sync import list_syncs

    try:
        limit = int(request.args.get('limit') or 30)
    except (TypeError, ValueError):
        limit = 30
    rows = list_syncs(limit=limit)
    return jsonify({'ok': True, 'success': True, 'syncs': rows, 'count': len(rows)})


@app.route('/api/ci/sync', methods=['POST'])
@token_or_login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@feature_required('ci_integration')
@api_error_handler
def api_ci_sync_create():
    """创建统一门禁：绑定 Testory run 并/或触发 Jenkins 受理构建，两侧结果同步。"""
    from ai_modules.enterprise.ci_unified_sync import start_unified_sync

    data = request.get_json(silent=True) or {}
    params = data.get('parameters') if isinstance(data.get('parameters'), dict) else {}
    if not params and isinstance(data.get('params'), dict):
        params = data.get('params')
    policy = _ai_str(data.get('policy') or 'both_must_pass') or 'both_must_pass'
    out = start_unified_sync(
        policy=policy,
        testory_run_id=_ai_str(data.get('testory_run_id') or data.get('run_id')),
        jenkins_job=_ai_str(data.get('jenkins_job') or data.get('job_name') or data.get('job')),
        jenkins_parameters=params,
        label=_ai_str(data.get('label')),
        auto_poll=bool(data.get('auto_poll', True)),
    )
    code = 200 if out.get('ok') else 400
    return jsonify(out), code


@app.route('/api/ci/sync/<sync_id>', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
@api_error_handler
def api_ci_sync_get(sync_id):
    """查询统一门禁；默认刷新两侧状态。"""
    from ai_modules.enterprise.ci_unified_sync import get_sync, refresh_sync

    refresh = str(request.args.get('refresh') or '1').strip().lower() not in ('0', 'false', 'no')
    rec = refresh_sync(sync_id) if refresh else get_sync(sync_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'sync_not_found', 'sync_id': sync_id}), 404
    return jsonify({'ok': True, 'success': True, 'sync': rec})


@app.route('/api/ci/sync/<sync_id>/jenkins', methods=['POST'])
@token_or_login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@feature_required('ci_integration')
@api_error_handler
def api_ci_sync_jenkins_report(sync_id):
    """Jenkins 流水线主动回写构建结果（与轮询同步）。"""
    from ai_modules.enterprise.ci_unified_sync import apply_jenkins_result

    data = request.get_json(silent=True) or {}
    result = _ai_str(data.get('result') or data.get('status'))
    if not result:
        return jsonify({'ok': False, 'error': 'result_required'}), 400
    rec = apply_jenkins_result(
        sync_id,
        result=result,
        build_url=_ai_str(data.get('build_url') or data.get('url')),
        build_number=data.get('build_number') or data.get('number'),
        building=bool(data.get('building', False)),
    )
    if not rec:
        return jsonify({'ok': False, 'error': 'sync_not_found'}), 404
    return jsonify({'ok': True, 'success': True, 'sync': rec})


@app.route('/api/ci/jenkins/status', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
@api_error_handler
def api_ci_jenkins_status():
    """Jenkins 反向触发配置是否就绪（不暴露 token）。"""
    from ai_modules.enterprise.jenkins_trigger import jenkins_config_from_env, jenkins_configured

    cfg = jenkins_config_from_env()
    return jsonify({
        'ok': True,
        'configured': jenkins_configured(),
        'base_url': cfg.get('base_url') or '',
        'user': cfg.get('user') or '',
        'token_set': bool(cfg.get('token')),
        'disclaimer': (
            '配置 JENKINS_URL / JENKINS_USER / JENKINS_API_TOKEN 后，'
            '可从本平台触发 Jenkins Job；触发成功（受理）≠ Job 已通过。'
            '统一门禁请用 POST /api/ci/sync。'
        ),
    })


@app.route('/api/ci/jenkins/trigger', methods=['POST'])
@token_or_login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@feature_required('ci_integration')
@api_error_handler
def api_ci_jenkins_trigger():
    """从 Testory 触发 Jenkins Job；可选 create_sync 纳入统一门禁。"""
    from ai_modules.enterprise.jenkins_trigger import trigger_jenkins_job

    data = request.get_json(silent=True) or {}
    params = data.get('parameters') if isinstance(data.get('parameters'), dict) else {}
    if not params and isinstance(data.get('params'), dict):
        params = data.get('params')
    job_name = _ai_str(data.get('job_name') or data.get('job'))
    want_sync = bool(data.get('create_sync') or data.get('unified_gate') or data.get('sync'))
    if want_sync:
        from ai_modules.enterprise.ci_unified_sync import start_unified_sync

        out = start_unified_sync(
            policy=_ai_str(data.get('policy') or 'both_must_pass') or 'both_must_pass',
            testory_run_id=_ai_str(data.get('testory_run_id') or data.get('run_id')),
            jenkins_job=job_name,
            jenkins_parameters=params,
            label=_ai_str(data.get('label')),
            auto_poll=bool(data.get('auto_poll', True)),
        )
        code = 200 if out.get('ok') else 400
        return jsonify(out), code

    result = trigger_jenkins_job(
        job_name=job_name,
        parameters=params,
        base_url=_ai_str(data.get('base_url') or data.get('jenkins_url')),
        user=_ai_str(data.get('user')),
        token=_ai_str(data.get('token') or data.get('api_token')),
    )
    code = 200 if result.get('ok') else 400
    return jsonify(result), code


@app.route('/api/ci/runs', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_runs_create():
    """CI 触发执行。默认同步；async=true 时立即返回 queued（202），可轮询 + webhook。"""
    from ci_adapter import create_queued_run, new_run_id, public_run_view

    data = request.get_json(silent=True) or {}
    _db = Database()
    case_ids, err = _ci_resolve_case_ids(data, _db)
    if err:
        return jsonify({"ok": False, "success": False, "error": err, "status": "failed"}), 400

    trigger_source = str(data.get("trigger_source") or data.get("source") or "ci").strip() or "ci"
    build_id = str(data.get("build_id") or data.get("pipeline_id") or "").strip()
    git_sha = str(data.get("git_sha") or data.get("commit") or "").strip()
    branch = str(data.get("branch") or "").strip()
    project_id = data.get("project_id")
    suite_name = str(data.get("suite_name") or f"Testory-project-{project_id or 'adhoc'}").strip()
    callback_url = str(
        data.get("callback_url") or data.get("webhook_url") or data.get("callback") or ""
    ).strip()
    want_async = _ci_want_async(data)

    uid = current_user.id if current_user.is_authenticated else None
    tid = None
    if uid is not None:
        try:
            tid = _db.get_user_tenant_id(int(uid))
        except Exception:
            tid = None

    if want_async:
        queued = create_queued_run(
            project_id=project_id,
            case_ids=case_ids,
            trigger_source=trigger_source,
            build_id=build_id,
            git_sha=git_sha,
            branch=branch,
            suite_name=suite_name,
            callback_url=callback_url,
        )
        run_id = queued["run_id"]

        def _worker():
            try:
                _ci_run_batch_and_finalize(
                    run_id=run_id,
                    case_ids=case_ids,
                    project_id=project_id,
                    trigger_source=trigger_source,
                    build_id=build_id,
                    git_sha=git_sha,
                    branch=branch,
                    suite_name=suite_name,
                    user_id=uid,
                    tenant_id=tid,
                    sync_mode=False,
                )
            except Exception as e:
                uat_logger.error(f"CI async worker 失败 run_id={run_id}: {e}")
                try:
                    from ci_adapter import deliver_run_callback, update_run_fields

                    update_run_fields(
                        run_id,
                        status="failed",
                        gate_passed=False,
                        success=False,
                        batch_error=str(e)[:300],
                        finished_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                    )
                    deliver_run_callback(run_id)
                    try:
                        from ai_modules.enterprise.ci_unified_sync import on_testory_run_finished

                        on_testory_run_finished(run_id)
                    except Exception:
                        pass
                except Exception:
                    pass

        t = threading.Thread(target=_worker, name=f"ci-run-{run_id}", daemon=True)
        t.start()
        view = public_run_view(queued)
        return jsonify({
            "ok": True,
            "accepted": True,
            "async": True,
            "run_id": run_id,
            "job_id": run_id,
            "status": "queued",
            "success": False,
            "gate_passed": False,
            "poll_url": queued.get("poll_url"),
            "junit_url": queued.get("junit_url"),
            "report_url": queued.get("report_url"),
            "build_id": build_id or None,
            "callback_url": callback_url or None,
            "result": view,
            "message": "已入队，请轮询 poll_url 直至 status 为 success/failed",
        }), 202

    # 同步路径
    run_id = new_run_id()
    if callback_url:
        from ci_adapter import save_run

        save_run({
            "run_id": run_id,
            "status": "running",
            "callback_url": callback_url,
            "case_ids": case_ids,
            "project_id": project_id,
            "trigger_source": trigger_source,
            "build_id": build_id,
            "git_sha": git_sha,
            "branch": branch,
            "suite_name": suite_name,
            "poll_url": f"/api/ci/runs/{run_id}",
            "junit_url": f"/api/ci/runs/{run_id}/junit.xml",
            "report_url": f"/api/ci/runs/{run_id}",
            "async": False,
        })

    record = _ci_run_batch_and_finalize(
        run_id=run_id,
        case_ids=case_ids,
        project_id=project_id,
        trigger_source=trigger_source,
        build_id=build_id,
        git_sha=git_sha,
        branch=branch,
        suite_name=suite_name,
        user_id=uid,
        tenant_id=tid,
        sync_mode=True,
    )

    view = public_run_view(record or {})
    return jsonify({
        "ok": True,
        "accepted": True,
        "async": False,
        "run_id": (record or {}).get("run_id") or run_id,
        "job_id": (record or {}).get("run_id") or run_id,
        "status": (record or {}).get("status"),
        "success": bool((record or {}).get("gate_passed")),
        "gate_passed": bool((record or {}).get("gate_passed")),
        "passed": (record or {}).get("passed"),
        "failed": (record or {}).get("failed"),
        "total": (record or {}).get("total"),
        "poll_url": (record or {}).get("poll_url"),
        "junit_url": (record or {}).get("junit_url"),
        "report_url": (record or {}).get("report_url"),
        "build_id": build_id or None,
        "callback_status": (record or {}).get("callback_status"),
        "result": view,
    }), 200


@app.route('/api/ci/runs/<run_id>', methods=['GET'])
@token_or_login_required
def api_ci_runs_get(run_id):
    from ci_adapter import get_run, is_terminal_status, public_run_view

    record = get_run(run_id)
    if not record:
        return jsonify({"ok": False, "error": "CI run 不存在", "success": False}), 404
    view = public_run_view(record)
    terminal = is_terminal_status(record.get("status"))
    return jsonify({
        "ok": True,
        "run_id": record["run_id"],
        "status": record.get("status"),
        "terminal": terminal,
        "success": bool(record.get("gate_passed")) if terminal else False,
        "gate_passed": bool(record.get("gate_passed")) if terminal else False,
        "passed": record.get("passed"),
        "failed": record.get("failed"),
        "total": record.get("total"),
        "poll_url": record.get("poll_url"),
        "junit_url": record.get("junit_url"),
        "report_url": record.get("report_url"),
        "callback_status": record.get("callback_status"),
        "result": view,
    })


@app.route('/api/ci/runs/<run_id>/junit.xml', methods=['GET'])
@token_or_login_required
def api_ci_runs_junit(run_id):
    from ci_adapter import get_run

    record = get_run(run_id)
    if not record:
        return jsonify({"ok": False, "error": "CI run 不存在"}), 404
    xml_text = record.get("junit_xml") or ""
    return Response(
        xml_text,
        mimetype="application/xml",
        headers={
            "Content-Disposition": f'attachment; filename="testory-{run_id}-junit.xml"',
        },
    )


def _ci_code_change_run_trigger(
    *,
    case_ids,
    project_id=None,
    build_id="",
    git_sha="",
    branch="",
    trigger_source="code_change",
):
    """供 code-change 流水线异步触发 /api/ci/runs 等价逻辑，返回 run_id。"""
    from ci_adapter import create_queued_run

    if not case_ids:
        return None
    suite_name = f"Testory-code-change-{project_id or 'adhoc'}"
    queued = create_queued_run(
        project_id=project_id,
        case_ids=list(case_ids),
        trigger_source=trigger_source or "code_change",
        build_id=build_id or "",
        git_sha=git_sha or "",
        branch=branch or "",
        suite_name=suite_name,
        callback_url="",
    )
    run_id = queued["run_id"]
    uid = current_user.id if current_user.is_authenticated else None
    tid = None
    _db = Database()
    if uid is not None:
        try:
            tid = _db.get_user_tenant_id(int(uid))
        except Exception:
            tid = None

    def _worker():
        try:
            _ci_run_batch_and_finalize(
                run_id=run_id,
                case_ids=list(case_ids),
                project_id=project_id,
                trigger_source=trigger_source or "code_change",
                build_id=build_id or "",
                git_sha=git_sha or "",
                branch=branch or "",
                suite_name=suite_name,
                user_id=uid,
                tenant_id=tid,
                sync_mode=False,
            )
        except Exception as e:
            uat_logger.error(f"code-change CI worker 失败 run_id={run_id}: {e}")
            try:
                from ci_adapter import deliver_run_callback, update_run_fields

                update_run_fields(
                    run_id,
                    status="failed",
                    gate_passed=False,
                    success=False,
                    batch_error=str(e)[:300],
                    finished_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                )
                deliver_run_callback(run_id)
            except Exception:
                pass

    threading.Thread(target=_worker, name=f"cc-ci-{run_id}", daemon=True).start()
    return run_id


@app.route('/api/ci/code-change', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_list():
    """列出近期代码变更分析任务。"""
    from ai_modules.code_intel.task_store import list_tasks
    from ai_modules.code_intel.pipeline import public_task_view

    try:
        limit = int(request.args.get('limit') or 30)
    except (TypeError, ValueError):
        limit = 30
    rows = [public_task_view(r) for r in list_tasks(limit=limit)]
    return jsonify({'ok': True, 'success': True, 'tasks': rows, 'count': len(rows)})


@app.route('/api/ci/code-change', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_create():
    """异步代码变更分析：CI 传 changed_files/diff → ChangeImpactReport + 推荐用例。

    Body：project_id, git_sha, branch, repo, changed_files[], diff, mr_description,
    file_snippets{}, analyze_only(默认 true), trigger_run, generate_drafts, callback_url, use_llm。
    """
    from ai_modules.code_intel.pipeline import enqueue_code_change, public_task_view
    from ai_modules.code_intel.policy import check_rate_limit, resolve_use_llm

    data = request.get_json(silent=True) or {}
    ok_rl, rl_err = check_rate_limit(f"code-change:{request.remote_addr or 'na'}")
    if not ok_rl:
        return jsonify({'ok': False, 'success': False, 'error': rl_err}), 429

    files = data.get('changed_files') or data.get('files') or []
    if isinstance(files, str):
        files = [x.strip() for x in files.replace(';', '\n').splitlines() if x.strip()]

    uid = current_user.id if current_user.is_authenticated else None
    tid = None
    if uid is not None:
        try:
            tid = Database().get_user_tenant_id(int(uid))
        except Exception:
            tid = None

    payload = {
        'project_id': data.get('project_id'),
        'tenant_id': data.get('tenant_id') if data.get('tenant_id') is not None else tid,
        'repo': _ai_str(data.get('repo')),
        'branch': _ai_str(data.get('branch')),
        'git_sha': _ai_str(data.get('git_sha') or data.get('commit') or data.get('sha')),
        'mr_key': _ai_str(data.get('mr_key') or data.get('pr_url') or data.get('mr_url')),
        'mr_description': _ai_str(
            data.get('mr_description') or data.get('description') or data.get('message')
        ),
        'build_id': _ai_str(data.get('build_id') or data.get('pipeline_id')),
        'trigger_source': _ai_str(data.get('trigger_source') or data.get('source') or 'ci') or 'ci',
        'changed_files': files,
        'diff': data.get('diff') or data.get('patch') or '',
        'file_snippets': data.get('file_snippets') if isinstance(data.get('file_snippets'), dict) else {},
        'analyze_only': bool(data.get('analyze_only', True)),
        'generate_drafts': bool(data.get('generate_drafts') or data.get('generate_cases')),
        'trigger_run': bool(data.get('trigger_run') or data.get('run_recommended')),
        'callback_url': _ai_str(data.get('callback_url') or data.get('webhook_url')),
        'pr_url': _ai_str(data.get('pr_url') or data.get('mr_url')),
    }
    if payload['trigger_run']:
        payload['analyze_only'] = False
    if not payload['changed_files'] and not payload['diff'] and not payload['mr_description']:
        return jsonify({
            'ok': False, 'success': False,
            'error': '请提供 changed_files、diff 或 mr_description 之一',
        }), 400

    use_llm = resolve_use_llm(data.get('use_llm'))
    sync_mode = bool(data.get('sync')) and not bool(data.get('async', True))

    view = enqueue_code_change(
        payload,
        db_factory=lambda: Database(),
        run_trigger=_ci_code_change_run_trigger if payload['trigger_run'] else None,
        profile=None,
        use_llm=bool(use_llm),
        background=not sync_mode,
    )
    tid_task = view.get('task_id')
    status_code = 200 if sync_mode else 202
    return jsonify({
        'ok': True,
        'success': True,
        'accepted': True,
        'async': not sync_mode,
        'task_id': tid_task,
        'status': view.get('status'),
        'poll_url': view.get('poll_url') or (f'/api/ci/code-change/{tid_task}' if tid_task else None),
        'idempotent_hit': bool(view.get('idempotent_hit')),
        'result': public_task_view(view),
        'message': '分析已入队，请轮询 poll_url' if not sync_mode else '分析已完成',
    }), status_code


@app.route('/api/ci/code-change/metrics', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_metrics():
    from ai_modules.code_intel.metrics import collect_metrics
    try:
        limit = int(request.args.get('limit') or 200)
    except (TypeError, ValueError):
        limit = 200
    return jsonify({'ok': True, 'success': True, 'metrics': collect_metrics(limit=limit)})


@app.route('/api/ci/code-change/cleanup', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_cleanup():
    from ai_modules.code_intel.task_store import cleanup_expired_tasks
    data = request.get_json(silent=True) or {}
    ttl = data.get('ttl_days')
    try:
        ttl_i = int(ttl) if ttl is not None else None
    except (TypeError, ValueError):
        ttl_i = None
    result = cleanup_expired_tasks(ttl_days=ttl_i)
    return jsonify({'ok': True, 'success': True, **result})


@app.route('/api/ci/code-change/<task_id>', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_get(task_id):
    from ai_modules.code_intel.task_store import get_task
    from ai_modules.code_intel.pipeline import public_task_view, attach_heal_proposals_for_run

    rec = get_task(task_id)
    if not rec:
        return jsonify({'ok': False, 'success': False, 'error': '任务不存在'}), 404

    refresh_heal = str(request.args.get('refresh_heal') or '').strip() in ('1', 'true', 'yes')
    if refresh_heal:
        try:
            attach_heal_proposals_for_run(task_id, db_factory=lambda: Database())
            rec = get_task(task_id) or rec
        except Exception as e:
            uat_logger.warning(f"refresh_heal failed: {e}")

    view = public_task_view(rec)
    terminal = str(rec.get('status') or '') in ('done', 'failed')
    impact = rec.get('impact') if isinstance(rec.get('impact'), dict) else {}
    return jsonify({
        'ok': True,
        'success': True,
        'task_id': rec.get('task_id'),
        'status': rec.get('status'),
        'terminal': terminal,
        'impact': impact,
        'impact_summary': impact.get('summary'),
        'recommended_case_ids': rec.get('recommended_case_ids') or [],
        'at_risk_case_ids': rec.get('at_risk_case_ids') or [],
        'draft_case_ids': rec.get('draft_case_ids') or [],
        'draft_preview': rec.get('draft_preview') or [],
        'ci_run_id': rec.get('ci_run_id'),
        'heal_proposals': rec.get('heal_proposals') or [],
        'warnings': rec.get('warnings') or [],
        'poll_url': rec.get('poll_url'),
        'result': view,
    })


@app.route('/api/ci/code-change/<task_id>/trigger-run', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_trigger_run(task_id):
    """对已完成分析的任务，用 recommended_case_ids 触发 CI run。"""
    from ai_modules.code_intel.task_store import get_task, update_task

    rec = get_task(task_id)
    if not rec:
        return jsonify({'ok': False, 'error': '任务不存在'}), 404
    case_ids = list(rec.get('recommended_case_ids') or [])
    data = request.get_json(silent=True) or {}
    if data.get('case_ids'):
        try:
            case_ids = [int(x) for x in data['case_ids']]
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'case_ids 无效'}), 400
    if not case_ids:
        return jsonify({'ok': False, 'error': '无推荐用例可执行'}), 400

    try:
        from ai_modules.code_intel.review import filter_ci_case_ids
        filtered = filter_ci_case_ids(Database(), case_ids, include_pending=False)
        case_ids = filtered.get('kept') or []
        if not case_ids:
            return jsonify({
                'ok': False,
                'error': '推荐用例均未激活（pending/rejected），请先 POST /api/ci/cases/<id>/review',
                'skipped': filtered.get('skipped'),
            }), 400
    except Exception:
        pass

    run_id = _ci_code_change_run_trigger(
        case_ids=case_ids,
        project_id=rec.get('project_id'),
        build_id=str(data.get('build_id') or rec.get('build_id') or ''),
        git_sha=str(rec.get('git_sha') or ''),
        branch=str(rec.get('branch') or ''),
        trigger_source='code_change',
    )
    update_task(task_id, ci_run_id=run_id, trigger_run=True, analyze_only=False)
    return jsonify({
        'ok': True,
        'success': True,
        'task_id': task_id,
        'ci_run_id': run_id,
        'case_ids': case_ids,
        'poll_url': f'/api/ci/runs/{run_id}',
        'message': '已触发推荐回归；终态后可 GET ...?refresh_heal=1 获取自愈提案',
    }), 202


@app.route('/api/ci/code-change/<task_id>/generate-drafts', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_generate_drafts(task_id):
    """基于已有分析结果生成并落库「待审核」草稿用例。"""
    from ai_modules.code_intel.task_store import get_task, update_task
    from ai_modules.code_intel.generate_from_code import (
        generate_cases_from_code,
        save_code_drafts_pending,
    )
    from ai_modules.code_intel.match_cases import load_project_cases_for_match

    rec = get_task(task_id)
    if not rec:
        return jsonify({'ok': False, 'error': '任务不存在'}), 404
    impact = rec.get('impact') if isinstance(rec.get('impact'), dict) else None
    signals = rec.get('signals') if isinstance(rec.get('signals'), dict) else None
    if not impact or not signals:
        return jsonify({'ok': False, 'error': '任务尚未完成影响分析'}), 400
    if impact.get('is_rollback'):
        return jsonify({'ok': False, 'error': '回滚变更不生成新用例'}), 400
    project_id = rec.get('project_id')
    if project_id is None:
        return jsonify({'ok': False, 'error': '需要 project_id'}), 400

    _db = Database()
    blobs = []
    try:
        for c in load_project_cases_for_match(_db, int(project_id)):
            blobs.append(f"{c.get('name','')} {c.get('description','')}")
    except Exception:
        pass

    use_llm = True
    data = request.get_json(silent=True) or {}
    if data.get('use_llm') is False:
        use_llm = False

    drafts, warns = generate_cases_from_code(
        signals=signals,
        impact=impact,
        diff=str(rec.get('diff') or ''),
        git_sha=str(rec.get('git_sha') or ''),
        use_llm=use_llm,
        existing_case_blobs=blobs,
        file_snippets=rec.get('file_snippets') if isinstance(rec.get('file_snippets'), dict) else {},
    )
    if not drafts:
        return jsonify({
            'ok': True, 'success': True, 'created_case_ids': [], 'warnings': warns, 'count': 0,
        })

    uid = current_user.id if current_user.is_authenticated else 0
    saved = save_code_drafts_pending(
        _db,
        project_id=int(project_id),
        drafts=drafts,
        user_id=int(uid or 0),
        git_sha=str(rec.get('git_sha') or ''),
    )
    ids = list(saved.get('created_case_ids') or [])
    prev = list(rec.get('draft_case_ids') or [])
    update_task(
        task_id,
        draft_case_ids=prev + ids,
        generate_drafts=True,
        warnings=list(rec.get('warnings') or []) + warns + list(saved.get('warnings') or []),
    )
    return jsonify({
        'ok': True,
        'success': True,
        'created_case_ids': ids,
        'count': len(ids),
        'warnings': warns + list(saved.get('warnings') or []),
        'message': '草稿已写入，描述含 [review_status:pending]，默认不进 CI 绿灯',
    })


@app.route('/api/ci/code-change/<task_id>/heal-proposals/<proposal_id>/ack', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_code_change_heal_ack(task_id, proposal_id):
    """确认自愈提案需人工处理（不自动改步骤/不假绿）。"""
    from ai_modules.code_intel.task_store import get_task, update_task
    from ai_modules.code_intel.heal_bridge import apply_heal_proposal_noop

    rec = get_task(task_id)
    if not rec:
        return jsonify({'ok': False, 'error': '任务不存在'}), 404
    proposals = list(rec.get('heal_proposals') or [])
    found = None
    for p in proposals:
        if isinstance(p, dict) and str(p.get('proposal_id')) == str(proposal_id):
            found = p
            break
    if not found:
        return jsonify({'ok': False, 'error': '提案不存在'}), 404
    out = apply_heal_proposal_noop(found)
    found['status'] = 'acked_manual'
    found['acked_at'] = time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())
    update_task(task_id, heal_proposals=proposals)
    return jsonify({'ok': True, 'success': True, **out, 'proposal': found})


@app.route('/api/ci/webhooks/github', methods=['POST'])
def api_ci_webhook_github():
    """GitHub push/PR webhook → 入队 code-change（需 TESTORY_GITHUB_WEBHOOK_SECRET）。"""
    from ai_modules.code_intel.webhooks import parse_webhook
    from ai_modules.code_intel.pipeline import enqueue_code_change
    from ai_modules.code_intel.policy import check_ip_allowed, check_rate_limit

    ok_ip, ip_err = check_ip_allowed(request.remote_addr or "")
    if not ok_ip:
        return jsonify({'ok': False, 'error': ip_err}), 403
    ok_rl, rl_err = check_rate_limit(f"wh-github:{request.remote_addr or 'na'}")
    if not ok_rl:
        return jsonify({'ok': False, 'error': rl_err}), 429

    body = request.get_data(cache=False, as_text=False) or b''
    headers = {k: v for k, v in request.headers.items()}
    norm, err, code = parse_webhook(provider='github', headers=headers, body=body)
    if err and code != 202:
        return jsonify({'ok': False, 'error': err}), code
    if not norm:
        return jsonify({'ok': True, 'ignored': True, 'message': err or 'ignored'}), 202

    data = request.get_json(silent=True) or {}
    project_id = request.args.get('project_id') or data.get('project_id')
    try:
        project_id = int(project_id) if project_id is not None else None
    except (TypeError, ValueError):
        project_id = None

    payload = dict(norm)
    payload['project_id'] = project_id
    payload['mr_key'] = str(norm.get('pr_url') or '')
    payload['analyze_only'] = True
    payload['generate_drafts'] = False
    payload['trigger_run'] = False

    view = enqueue_code_change(
        payload,
        db_factory=lambda: Database(),
        run_trigger=None,
        use_llm=True,
        background=True,
    )
    return jsonify({
        'ok': True,
        'accepted': True,
        'task_id': view.get('task_id'),
        'poll_url': view.get('poll_url'),
        'status': view.get('status'),
        'project_id': view.get('project_id') or project_id,
    }), 202


@app.route('/api/ci/webhooks/gitlab', methods=['POST'])
def api_ci_webhook_gitlab():
    """GitLab push/MR webhook → 入队 code-change（需 TESTORY_GITLAB_WEBHOOK_SECRET）。"""
    from ai_modules.code_intel.webhooks import parse_webhook
    from ai_modules.code_intel.pipeline import enqueue_code_change
    from ai_modules.code_intel.policy import check_ip_allowed, check_rate_limit

    ok_ip, ip_err = check_ip_allowed(request.remote_addr or "")
    if not ok_ip:
        return jsonify({'ok': False, 'error': ip_err}), 403
    ok_rl, rl_err = check_rate_limit(f"wh-gitlab:{request.remote_addr or 'na'}")
    if not ok_rl:
        return jsonify({'ok': False, 'error': rl_err}), 429

    body = request.get_data(cache=False, as_text=False) or b''
    headers = {k: v for k, v in request.headers.items()}
    norm, err, code = parse_webhook(provider='gitlab', headers=headers, body=body)
    if err and code != 202:
        return jsonify({'ok': False, 'error': err}), code
    if not norm:
        return jsonify({'ok': True, 'ignored': True, 'message': err or 'ignored'}), 202

    project_id = request.args.get('project_id')
    try:
        project_id = int(project_id) if project_id is not None else None
    except (TypeError, ValueError):
        project_id = None

    payload = dict(norm)
    payload['project_id'] = project_id
    payload['mr_key'] = str(norm.get('pr_url') or '')
    payload['analyze_only'] = True

    view = enqueue_code_change(
        payload,
        db_factory=lambda: Database(),
        run_trigger=None,
        use_llm=True,
        background=True,
    )
    return jsonify({
        'ok': True,
        'accepted': True,
        'task_id': view.get('task_id'),
        'poll_url': view.get('poll_url'),
        'status': view.get('status'),
        'project_id': view.get('project_id') or project_id,
    }), 202


@app.route('/api/ci/repo-map', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_repo_map_list():
    from ai_modules.code_intel.repo_map import list_mappings
    rows = list_mappings()
    return jsonify({'ok': True, 'success': True, 'mappings': rows, 'count': len(rows)})


@app.route('/api/ci/repo-map', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_repo_map_upsert():
    from ai_modules.code_intel.repo_map import upsert_mapping
    data = request.get_json(silent=True) or {}
    try:
        entry = upsert_mapping(
            repo=_ai_str(data.get('repo')),
            project_id=int(data.get('project_id')),
            tenant_id=data.get('tenant_id'),
            label=_ai_str(data.get('label')),
            default_branch=_ai_str(data.get('default_branch') or data.get('branch')),
        )
    except (TypeError, ValueError) as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    return jsonify({'ok': True, 'success': True, 'mapping': entry})


@app.route('/api/ci/repo-map', methods=['DELETE'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_repo_map_delete():
    from ai_modules.code_intel.repo_map import delete_mapping
    data = request.get_json(silent=True) or {}
    repo = _ai_str(data.get('repo') or request.args.get('repo'))
    if not repo:
        return jsonify({'ok': False, 'error': 'repo 必填'}), 400
    ok = delete_mapping(repo)
    return jsonify({'ok': ok, 'success': ok, 'deleted': ok})


@app.route('/api/ci/cases/pending', methods=['GET'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_cases_pending():
    from ai_modules.code_intel.review import list_pending_cases
    project_id = request.args.get('project_id')
    if project_id is None:
        return jsonify({'ok': False, 'error': 'project_id 必填'}), 400
    try:
        pid = int(project_id)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'project_id 无效'}), 400
    rows = list_pending_cases(Database(), pid)
    return jsonify({'ok': True, 'success': True, 'cases': rows, 'count': len(rows)})


@app.route('/api/ci/cases/<int:case_id>/review', methods=['POST'])
@token_or_login_required
@feature_required('ci_integration')
def api_ci_case_review(case_id):
    """激活/拒绝待审核用例。body: {status: active|rejected|pending}"""
    data = request.get_json(silent=True) or {}
    status = _ai_str(data.get('status') or data.get('review_status')).lower()
    if status not in ('pending', 'active', 'rejected'):
        return jsonify({'ok': False, 'error': 'status 须为 pending|active|rejected'}), 400
    _db = Database()
    ok = _db.set_case_review_status(
        int(case_id),
        status,
        git_sha=_ai_str(data.get('git_sha') or data.get('source_commit')),
    )
    if not ok:
        return jsonify({'ok': False, 'error': '更新失败（用例不存在或库未迁移）'}), 400
    case = _db.get_test_case_v2(int(case_id))
    return jsonify({
        'ok': True,
        'success': True,
        'case_id': case_id,
        'review_status': status,
        'case': case,
        'message': '已激活，可进入 CI 门禁' if status == 'active' else f'已标记为 {status}',
    })


@app.route('/api/trigger/<int:project_id>', methods=['POST'])
@token_or_login_required
def api_trigger_project(project_id):
    """CI/CD 触发指定项目的所有用例执行（兼容旧路径；推荐 /api/ci/runs）。"""
    from auth_batch_helpers import count_batch_gate_failures
    from ci_adapter import build_run_record_from_batch, public_run_view

    try:
        _db = Database()
        cases = _db.get_project_cases(project_id, case_type="ui")
        if not cases:
            return jsonify({'success': False, 'ok': False, 'error': '项目没有测试用例', 'status': 'failed'}), 400
        case_ids = [c['id'] for c in cases]
        uat_logger.info(f"CI/CD 触发项目 #{project_id} 执行，共 {len(case_ids)} 个用例")
        uid = current_user.id if current_user.is_authenticated else None
        tid = None
        if uid is not None:
            try:
                tid = _db.get_user_tenant_id(int(uid))
            except Exception:
                tid = None
        data = request.get_json(silent=True) or {}
        _exec_ctx = ExecutionContext(
            user_id=uid,
            tenant_id=tid,
            trigger="ci",
            on_case_failure=_on_case_execution_failure,
            extra={"project_id": project_id, "build_id": data.get("build_id")},
        )
        results = sync_execute_multiple_test_cases(case_ids, _db, execution_context=_exec_ctx)
        try:
            sync_close_browser()
        except Exception:
            pass
        if not isinstance(results, dict):
            results = {"case_results": [], "error": "执行返回无效"}
        record = build_run_record_from_batch(
            results,
            project_id=project_id,
            case_ids=case_ids,
            trigger_source=str(data.get("trigger_source") or "ci"),
            build_id=str(data.get("build_id") or ""),
            git_sha=str(data.get("git_sha") or ""),
            suite_name=f"Testory-project-{project_id}",
        )
        gate_ok = bool(record.get("gate_passed"))
        return jsonify({
            'ok': True,
            'success': gate_ok,
            'gate_passed': gate_ok,
            'status': record.get('status'),
            'run_id': record['run_id'],
            'execution_id': record['run_id'],  # 兼容旧 GitLab 脚本字段名
            'junit_url': record.get('junit_url'),
            'poll_url': record.get('poll_url'),
            'results': results,
            'result': public_run_view(record),
            'failed_cases': count_batch_gate_failures(results.get('case_results') or []),
        })
    except Exception as e:
        uat_logger.error(f"CI 触发执行失败: {e}")
        return jsonify({'success': False, 'ok': False, 'error': str(e), 'status': 'failed'}), 500

@app.route('/api/trigger/cases', methods=['POST'])
@token_or_login_required
def api_trigger_cases():
    """CI/CD 触发指定用例列表执行（兼容旧路径；推荐 /api/ci/runs）。"""
    from auth_batch_helpers import count_batch_gate_failures
    from ci_adapter import build_run_record_from_batch, public_run_view

    try:
        data = request.get_json(silent=True) or {}
        case_ids = data.get('case_ids', [])
        if not case_ids:
            return jsonify({'success': False, 'ok': False, 'error': '缺少 case_ids 参数', 'status': 'failed'}), 400
        _db = Database()
        uat_logger.info(f"CI/CD 触发用例列表执行: {case_ids}")
        uid = current_user.id if current_user.is_authenticated else None
        tid = None
        if uid is not None:
            try:
                tid = _db.get_user_tenant_id(int(uid))
            except Exception:
                tid = None
        _exec_ctx = ExecutionContext(
            user_id=uid,
            tenant_id=tid,
            trigger="ci",
            on_case_failure=_on_case_execution_failure,
        )
        results = sync_execute_multiple_test_cases(case_ids, _db, execution_context=_exec_ctx)
        try:
            sync_close_browser()
        except Exception:
            pass
        if not isinstance(results, dict):
            results = {"case_results": [], "error": "执行返回无效"}
        record = build_run_record_from_batch(
            results,
            case_ids=case_ids,
            trigger_source=str(data.get("trigger_source") or "ci"),
            build_id=str(data.get("build_id") or ""),
            git_sha=str(data.get("git_sha") or ""),
        )
        gate_ok = bool(record.get("gate_passed"))
        return jsonify({
            'ok': True,
            'success': gate_ok,
            'gate_passed': gate_ok,
            'status': record.get('status'),
            'run_id': record['run_id'],
            'execution_id': record['run_id'],
            'junit_url': record.get('junit_url'),
            'poll_url': record.get('poll_url'),
            'results': results,
            'result': public_run_view(record),
            'failed_cases': count_batch_gate_failures(results.get('case_results') or []),
        })
    except Exception as e:
        uat_logger.error(f"CI 触发执行失败: {e}")
        return jsonify({'success': False, 'ok': False, 'error': str(e), 'status': 'failed'}), 500

# ==================== 数据驱动测试接口 ====================

def _dataset_job_update(run_id: str, **kwargs):
    with _dataset_run_lock:
        if run_id in _dataset_run_jobs:
            _dataset_run_jobs[run_id].update(kwargs)


def _dataset_run_worker(run_id: str, user_id: int, dataset_id: int, case_id: int):
    """后台线程：按数据行执行用例，每行更新进度。"""
    import time as _time

    machine_lock_acquired = False

    def resolve_with_row(text, row_dict):
        if not text:
            return text
        import re

        def replace(m):
            key = m.group(1).strip()
            if key.startswith('row.'):
                field = key[4:]
                return str(row_dict.get(field, m.group(0)))
            return m.group(0)

        return re.sub(r'\{\{(.+?)\}\}', replace, text)

    try:
        _db = Database()
        dataset = _db.get_dataset(dataset_id)
        if not dataset:
            _dataset_job_update(run_id, finished=True, success=False, error='数据集不存在')
            return
        rows = _db.get_data_rows(dataset_id)
        if not rows:
            _dataset_job_update(run_id, finished=True, success=False, error='数据集没有数据行')
            return
        case = _db.get_test_case_v2(case_id)
        if not case:
            _dataset_job_update(run_id, finished=True, success=False, error='测试用例不存在')
            return
        if _app_case_type(case) == 'api':
            _dataset_job_update(
                run_id,
                finished=True,
                success=False,
                error='数据驱动执行仅支持 Web 用例；接口用例请使用接口测试运行入口。',
            )
            return
        steps = _db.get_case_steps(case_id)
        if not steps:
            _dataset_job_update(run_id, finished=True, success=False, error='测试用例没有步骤')
            return

        machine_lock_acquired = False
        try:
            from execution_lock import acquire as acquire_machine_lock

            machine_lock_acquired = acquire_machine_lock(
                owner=f"dataset:{dataset_id}:case:{case_id}", timeout_sec=120
            )
            if not machine_lock_acquired:
                _dataset_job_update(
                    run_id,
                    finished=True,
                    success=False,
                    error='本机已有自动化任务在执行，请稍后再试。',
                )
                return
        except ImportError:
            pass

        with _dataset_run_lock:
            if run_id in _dataset_run_jobs:
                _dataset_run_jobs[run_id]['total'] = len(rows)

        uat_logger.info(f"数据驱动执行：数据集#{dataset_id}，用例#{case_id}，共{len(rows)}行数据")

        run_results = []
        total_success = 0
        total_fail = 0
        cancelled = False

        for row_info in rows:
            with _dataset_run_lock:
                if _dataset_run_jobs.get(run_id, {}).get('cancel_requested'):
                    cancelled = True
                    break
            row_data = {
                k: _normalize_dataset_cell_value(v)
                for k, v in (row_info.get('data') or {}).items()
            }
            row_index = row_info['row_index']
            _dataset_job_update(run_id, current_row_index=row_index)

            uat_logger.info(f"[数据驱动] 准备执行第{row_index}行数据，重启浏览器...")
            try:
                sync_close_browser()
                uat_logger.info(f"[数据驱动] 浏览器已关闭")
            except Exception as e:
                uat_logger.debug(f"[数据驱动] 关闭浏览器时出错（可忽略）: {e}")

            _time.sleep(0.5)
            uat_logger.info(f"[数据驱动] 启动新浏览器实例...")

            start_time = _time.time()
            status = 'success'
            error_msg = ''
            step_results_list = []

            try:
                sync_start_browser()
                initial_nav_url, nav_source = _resolve_case_navigation_url_for_data_row(
                    _db,
                    case,
                    case_id,
                    steps,
                    row_data,
                    case.get('project_id'),
                    resolve_with_row,
                )
                if initial_nav_url:
                    uat_logger.log_automation_step(
                        'navigate',
                        initial_nav_url,
                        f'数据驱动 行{row_index} 测试开始时导航({nav_source})',
                    )
                    sync_navigate_to(initial_nav_url)
                else:
                    uat_logger.warning(
                        f'[数据驱动] 行{row_index} 未解析到用例初始 URL（case.url 或步骤中首个有效 navigate），'
                        '若首步为 enter_iframe 等需已有页面，可能失败；请配置用例地址或在首步增加导航。'
                    )

                for step_idx, step in enumerate(steps):
                    with _dataset_run_lock:
                        if _dataset_run_jobs.get(run_id, {}).get('cancel_requested'):
                            raise Exception('用户已停止执行')
                    selector_value = resolve_with_row(
                        _db.resolve_variables(step.get('selector_value', ''), project_id=case.get('project_id'), case_id=case_id),
                        row_data
                    )
                    input_value = resolve_with_row(
                        _db.resolve_variables(step.get('input_value', ''), project_id=case.get('project_id'), case_id=case_id),
                        row_data
                    )
                    url_value = resolve_with_row(
                        _db.resolve_variables(step.get('url', '') or '', project_id=case.get('project_id'), case_id=case_id),
                        row_data,
                    )
                    description = resolve_with_row(
                        _db.resolve_variables(step.get('description', '') or '', project_id=case.get('project_id'), case_id=case_id),
                        row_data,
                    )

                    step_start = _time.time()
                    step_status = 'success'
                    step_error = ''
                    try:
                        action = step.get('action', '')
                        selector_type = step.get('selector_type', 'css')
                        iframe_sel = _effective_step_iframe_selector(
                            automation,
                            _db,
                            step,
                            case.get('project_id'),
                            case_id,
                            row_resolve_fn=lambda t: resolve_with_row(t, row_data),
                        )
                        if action == 'navigate':
                            raw_url = (url_value or input_value or case.get('url') or '').strip()
                            fixed_url, url_err = _validate_and_fix_url(raw_url)
                            if url_err:
                                raise Exception(url_err)
                            if fixed_url:
                                sync_navigate_to(fixed_url)
                        elif action == 'click':
                            if selector_value:
                                _repeat = _norm_click_repeat_count(step.get('click_repeat_count'))
                                for _r in range(_repeat):
                                    with _dataset_run_lock:
                                        if _dataset_run_jobs.get(run_id, {}).get('cancel_requested'):
                                            raise Exception('用户已停止执行')
                                    sync_click_element(
                                        selector_value,
                                        selector_type,
                                        iframe_selector=iframe_sel,
                                        locator_candidates=step.get('locator_candidates') or None,
                                    )
                        elif action == 'input':
                            if selector_value:
                                safe_input_value = resolve_fill_step_text({
                                    'input_value': input_value,
                                    'description': step.get('description'),
                                })
                                sync_fill_input(
                                    selector_value,
                                    safe_input_value,
                                    selector_type,
                                    iframe_selector=iframe_sel,
                                    locator_candidates=step.get('locator_candidates') or None,
                                )
                            else:
                                raise Exception("输入操作缺少选择器")
                        elif action == 'batch_input':
                            b_pairs = parse_batch_input_lines(input_value or '')
                            if not b_pairs:
                                raise Exception("批量输入步骤缺少有效行")
                            for bsel, bval in b_pairs:
                                with _dataset_run_lock:
                                    if _dataset_run_jobs.get(run_id, {}).get('cancel_requested'):
                                        raise Exception('用户已停止执行')
                                sync_fill_input(
                                    bsel,
                                    bval,
                                    selector_type,
                                    iframe_selector=iframe_sel,
                                    locator_candidates=None,
                                )
                        elif action == 'hover':
                            if selector_value:
                                sync_hover_element(selector_value, selector_type, iframe_selector=iframe_sel)
                                sync_wait_for_timeout(1000)
                            else:
                                raise Exception("悬停操作缺少选择器")
                        elif action == 'double_click':
                            if selector_value:
                                sync_double_click_element(selector_value, selector_type, iframe_selector=iframe_sel)
                                sync_wait_for_timeout(2000)
                            else:
                                raise Exception("双击操作缺少选择器")
                        elif action == 'right_click':
                            if selector_value:
                                sync_right_click_element(selector_value, selector_type, iframe_selector=iframe_sel)
                                sync_wait_for_timeout(1000)
                            else:
                                raise Exception("右键点击操作缺少选择器")
                        elif action == 'wait':
                            if input_value:
                                try:
                                    wait_time = (
                                        int(input_value) * 1000
                                        if int(input_value) < 1000
                                        else int(input_value)
                                    )
                                    sync_wait_for_timeout(wait_time)
                                except ValueError:
                                    raise Exception(f"无效的等待时间值: {input_value}")
                            else:
                                sync_wait_for_timeout(1000)
                        elif action == 'select':
                            if selector_value and input_value:
                                sync_select_option(
                                    selector_value,
                                    input_value,
                                    selector_type,
                                    iframe_selector=iframe_sel,
                                )
                                sync_wait_for_timeout(1000)
                        elif action == 'date':
                            if selector_value and input_value:
                                sync_select_date(selector_value, input_value)
                                sync_wait_for_timeout(1000)
                        elif action == 'scroll':
                            _run_db_step_scroll(input_value or "", iframe_selector=iframe_sel)
                            sync_wait_for_timeout(1500)
                        elif action == 'swipe':
                            if selector_value:
                                direction = 'up'
                                distance = 100
                                if input_value:
                                    parts = input_value.split(':')
                                    if len(parts) == 2:
                                        direction = parts[0]
                                        try:
                                            distance = int(parts[1])
                                        except ValueError:
                                            uat_logger.warning(
                                                f"无效的滑动距离值: {parts[1]}，使用默认值 100"
                                            )
                                    else:
                                        direction = input_value
                                sync_swipe_element(
                                    selector_value,
                                    direction,
                                    distance,
                                    selector_type,
                                    iframe_selector=iframe_sel,
                                )
                                sync_wait_for_timeout(1500)
                            else:
                                raise Exception("滑动操作缺少选择器")
                        elif action == 'verify':
                            verify_type = input_value if input_value else 'auto'
                            sync_verify_element(
                                selector=selector_value,
                                verify_type=verify_type,
                                selector_type=selector_type,
                                iframe_selector=iframe_sel,
                                locator_candidates=step.get('locator_candidates') or None,
                            )
                            sync_wait_for_timeout(1500)
                        elif action == 'extract_text' or action == 'text_compare':
                            _run_extract_text_automation_step(
                                action,
                                step,
                                selector_value,
                                input_value,
                                description,
                                selector_type,
                                iframe_sel,
                                locator_candidates=step.get('locator_candidates') or None,
                            )
                        elif action == 'extract_json':
                            if selector_value:
                                sync_extract_element_json(selector_value, selector_type)
                                sync_wait_for_timeout(1000)
                            else:
                                raise Exception("提取JSON数据时缺少选择器")
                        elif action == 'assert':
                            _run_assert_automation_step(
                                step, selector_value, input_value, selector_type, iframe_sel
                            )
                        elif action == 'enter_iframe':
                            if selector_value:
                                sync_enter_iframe(selector_value, selector_type)
                            else:
                                raise Exception('进入 iframe 步骤缺少选择器')
                        elif action == 'exit_iframe':
                            sync_exit_iframe()
                        elif action:
                            raise Exception(
                                f"数据驱动执行不支持的操作类型「{action}」。"
                                "支持的类型：navigate, click, input, batch_input, hover, double_click, right_click, wait, select, date, scroll, swipe, verify, extract_text, text_compare, extract_json, assert, enter_iframe, exit_iframe。"
                            )
                    except Exception as e:
                        step_status = 'failed'
                        step_error = str(e)
                        status = 'failed'
                        error_msg = f"行{row_index} 步骤{step_idx+1}({step.get('action','')}) 失败: {e}"

                    step_results_list.append({
                        'step_order': step_idx + 1,
                        'action': step.get('action', ''),
                        'selector_value': selector_value,
                        'input_value': input_value,
                        'status': step_status,
                        'error': step_error,
                        'duration': round(_time.time() - step_start, 3)
                    })
                    if status == 'failed':
                        break
            except Exception as e:
                status = 'failed'
                error_msg = str(e)

            duration = round(_time.time() - start_time, 3)
            history_id = _db.create_run_history(
                case_id=case_id,
                status=status,
                duration=duration,
                error=f"[数据驱动 行{row_index}] {error_msg}" if error_msg else '',
                extracted_text=f"数据驱动 行{row_index}: {str(row_data)}"
            )

            if status == 'success':
                total_success += 1
            else:
                total_fail += 1

            run_results.append({
                'row_index': row_index,
                'row_data': row_data,
                'status': status,
                'error': error_msg,
                'duration': duration,
                'history_id': history_id,
                'steps': step_results_list
            })

            _dataset_job_update(
                run_id,
                completed=len(run_results),
                successful_rows=total_success,
                failed_rows=total_fail,
                current_row_index=row_index,
            )

        try:
            sync_close_browser()
        except Exception:
            pass

        if cancelled:
            _dataset_job_update(
                run_id,
                finished=True,
                success=False,
                error='用户已停止执行',
                completed=len(run_results),
                results=run_results,
            )
            return

        uat_logger.info(f"数据驱动执行完成：成功{total_success}行，失败{total_fail}行")
        _dataset_job_update(
            run_id,
            finished=True,
            success=True,
            successful_rows=total_success,
            failed_rows=total_fail,
            completed=len(rows),
            results=run_results,
            error=None,
        )
    except Exception as e:
        uat_logger.log_exception("dataset_run_worker", e)
        try:
            sync_close_browser()
        except Exception:
            pass
        _dataset_job_update(run_id, finished=True, success=False, error=str(e))
    finally:
        if machine_lock_acquired:
            try:
                from execution_lock import release as release_machine_lock

                release_machine_lock()
            except ImportError:
                pass


@app.route('/data-driven')
@login_required
@feature_required('data_driven')
def data_driven_page():
    """数据驱动测试管理页面"""
    return render_template('data_driven.html')

@app.route('/api/datasets', methods=['GET'])
@login_required
def api_list_datasets():
    """获取数据集列表"""
    _db = Database()
    case_id = request.args.get('case_id', type=int)
    project_id = request.args.get('project_id', type=int)
    datasets = _db.get_all_datasets(case_id=case_id, project_id=project_id)
    return jsonify({'success': True, 'datasets': datasets})

@app.route('/api/datasets', methods=['POST'])
@login_required
def api_create_dataset():
    """创建数据集（手动创建，不含数据行）"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '数据集名称不能为空'}), 400
    _db = Database()
    dataset_id = _db.create_dataset(
        name=name,
        case_id=data.get('case_id'),
        project_id=data.get('project_id'),
        description=data.get('description', '')
    )
    return jsonify({'success': True, 'dataset_id': dataset_id})

@app.route('/api/datasets/<int:dataset_id>', methods=['GET'])
@login_required
def api_get_dataset(dataset_id):
    """获取数据集详情及数据行"""
    _db = Database()
    dataset = _db.get_dataset(dataset_id)
    if not dataset:
        return jsonify({'success': False, 'error': '数据集不存在'}), 404
    rows = _db.get_data_rows(dataset_id)
    dataset['rows'] = rows
    return jsonify({'success': True, 'dataset': dataset})

@app.route('/api/datasets/<int:dataset_id>', methods=['DELETE'])
@login_required
def api_delete_dataset(dataset_id):
    """删除数据集"""
    _db = Database()
    success = _db.delete_dataset(dataset_id)
    return jsonify({'success': success})


def _normalize_dataset_cell_value(v):
    """
    将上传文件中的单元格转为写入数据集的字符串。
    - Excel 日期单元格在 openpyxl data_only 下常为 datetime，str() 会变成 \"YYYY-MM-DD 00:00:00\"，
      与页面控件回读不一致；午夜时间统一压成纯日期 YYYY-MM-DD。
    - 所有结果 strip，去掉 CSV/粘贴带来的首尾空格。
    - 字符串若已是 \"日期 + 00:00:00\"（及 .0 小数秒）也压成纯日期。
    """
    import re
    from datetime import date, datetime, time as dtime

    if v is None:
        return ''
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, datetime):
        vv = v.replace(microsecond=0) if v.microsecond else v
        if vv.time() == dtime(0, 0, 0):
            return vv.strftime('%Y-%m-%d')
        return vv.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    if re.fullmatch(r'\d{4}-\d{2}-\d{2}\s+00:00:00(?:\.0+)?', s):
        return s[:10]
    return s


@app.route('/api/datasets/upload', methods=['POST'])
@login_required
def api_upload_dataset():
    """上传 CSV 或 Excel 文件创建数据集"""
    import csv
    import io
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '缺少文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    filename = file.filename.lower()
    name = request.form.get('name', '').strip() or file.filename.rsplit('.', 1)[0]
    case_id = request.form.get('case_id', type=int)
    project_id = request.form.get('project_id', type=int)
    description = request.form.get('description', '')

    rows_data = []
    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                rows_data.append({
                    (k.strip() if k else k): _normalize_dataset_cell_value(val)
                    for k, val in row.items()
                })
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [
                            (str(c).strip() if c is not None else f'col{i}')
                            for i, c in enumerate(row)
                        ]
                    else:
                        row_dict = {
                            headers[i]: _normalize_dataset_cell_value(v)
                            for i, v in enumerate(row)
                            if i < len(headers)
                        }
                        rows_data.append(row_dict)
                wb.close()
            except ImportError:
                return jsonify({'success': False, 'error': '解析 Excel 需要安装 openpyxl: pip install openpyxl==3.1.2'}), 400
        else:
            return jsonify({'success': False, 'error': '只支持 .csv / .xlsx / .xls 格式'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败: {str(e)}'}), 400

    if not rows_data:
        return jsonify({'success': False, 'error': '文件中没有数据行'}), 400

    _db = Database()
    dataset_id = _db.create_dataset(name=name, case_id=case_id, project_id=project_id, description=description)
    count = _db.add_data_rows(dataset_id, rows_data)
    columns = list(rows_data[0].keys()) if rows_data else []
    return jsonify({'success': True, 'dataset_id': dataset_id, 'row_count': count, 'columns': columns})

@app.route('/api/datasets/<int:dataset_id>/run', methods=['POST'])
@login_required
def api_run_dataset(dataset_id):
    """启动数据驱动后台任务，立即返回 run_id；前端轮询 /api/datasets/run-status/<run_id> 更新进度"""
    data = request.get_json(silent=True) or {}
    case_id = data.get('case_id')
    if not case_id:
        return jsonify({'success': False, 'error': '缺少 case_id 参数'}), 400

    try:
        case_id_int = int(case_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'case_id 无效'}), 400

    _db = Database()
    dataset = _db.get_dataset(dataset_id)
    if not dataset:
        return jsonify({'success': False, 'error': '数据集不存在'}), 404

    rows = _db.get_data_rows(dataset_id)
    if not rows:
        return jsonify({'success': False, 'error': '数据集没有数据行'}), 400

    case = _db.get_test_case_v2(case_id_int)
    if not case:
        return jsonify({'success': False, 'error': '测试用例不存在'}), 404

    steps = _db.get_case_steps(case_id_int)
    if not steps:
        return jsonify({'success': False, 'error': '测试用例没有步骤'}), 400

    run_id = secrets.token_urlsafe(24)
    uid = current_user.id
    with _dataset_run_lock:
        _dataset_run_jobs[run_id] = {
            'user_id': uid,
            'dataset_id': dataset_id,
            'case_id': case_id_int,
            'finished': False,
            'success': None,
            'error': None,
            'total': len(rows),
            'completed': 0,
            'successful_rows': 0,
            'failed_rows': 0,
            'results': [],
            'current_row_index': None,
            'cancel_requested': False,
        }

    thread = threading.Thread(
        target=_dataset_run_worker,
        args=(run_id, uid, dataset_id, case_id_int),
        daemon=True,
        name=f'dataset-run-{run_id[:10]}',
    )
    thread.start()

    return jsonify({
        'success': True,
        'run_id': run_id,
        'total': len(rows),
    })


@app.route('/api/datasets/run-status/<run_id>', methods=['GET'])
@login_required
def api_dataset_run_status(run_id):
    """查询数据驱动任务进度；完成后首次请求返回明细并移除任务缓存"""
    with _dataset_run_lock:
        job = _dataset_run_jobs.get(run_id)
    if not job:
        return jsonify({'success': False, 'error': '任务不存在或已结束'}), 404
    if job.get('user_id') != current_user.id:
        return jsonify({'success': False, 'error': '无权访问该任务'}), 403

    resp = {
        'success': True,
        'finished': job['finished'],
        'total': job['total'],
        'completed': job['completed'],
        'successful_rows': job['successful_rows'],
        'failed_rows': job['failed_rows'],
        'current_row_index': job.get('current_row_index'),
    }

    if not job['finished']:
        return jsonify(resp)

    resp['run_success'] = bool(job.get('success'))
    if job.get('error'):
        resp['error'] = job['error']
    if job.get('success'):
        resp['results'] = job.get('results') or []
        resp['successful_rows'] = job['successful_rows']
        resp['failed_rows'] = job['failed_rows']
    with _dataset_run_lock:
        _dataset_run_jobs.pop(run_id, None)

    return jsonify(resp)


@app.route('/api/datasets/run-stop/<run_id>', methods=['POST'])
@login_required
def api_dataset_run_stop(run_id):
    with _dataset_run_lock:
        job = _dataset_run_jobs.get(run_id)
        if not job:
            return jsonify({'success': False, 'error': '任务不存在或已结束'}), 404
        if job.get('user_id') != current_user.id:
            return jsonify({'success': False, 'error': '无权访问该任务'}), 403
        if job.get('finished'):
            return jsonify({'success': False, 'error': '任务已结束'}), 400
        job['cancel_requested'] = True
        job['finished'] = True
        job['success'] = False
        job['error'] = '用户已停止执行'
        job['completed'] = int(job.get('total', 0) or 0)
        job['current_row_index'] = None
    threading.Thread(target=_force_stop_browser_async, daemon=True, name='stop-dataset-run').start()
    return jsonify({'success': True, 'message': '已发送停止请求'})


def _get_current_user_dataset_job(user_id: int):
    """获取当前用户最近的一个数据驱动任务（优先 active）。"""
    with _dataset_run_lock:
        items = list(_dataset_run_jobs.items())
    for run_id, job in reversed(items):
        if job.get('user_id') != user_id:
            continue
        return run_id, job
    return None, None


@app.route('/api/datasets/current-run/status', methods=['GET'])
@login_required
def api_dataset_current_run_status():
    """查询当前用户的数据驱动任务状态（用于跨页面恢复显示/停止）。"""
    return jsonify(_dataset_current_run_payload(current_user.id))


@app.route('/api/datasets/current-run/stop', methods=['POST'])
@login_required
def api_dataset_current_run_stop():
    """停止当前用户最近的数据驱动任务（幂等）。"""
    run_id, job = _get_current_user_dataset_job(current_user.id)
    if not job:
        return jsonify({'success': True, 'message': '当前没有运行任务'})
    if job.get('finished'):
        return jsonify({'success': True, 'message': '任务已结束'})
    return api_dataset_run_stop(run_id)


# ==================== 调度器初始化 ====================

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger

    scheduler = BackgroundScheduler(timezone='Asia/Shanghai')

    def _run_scheduled_cases(schedule_id: int, case_ids: list, retry_count: int = 0):
        """调度器回调：执行定时用例（支持重跑）"""
        import datetime
        import time
        from notifications import notify

        _db = Database()
        # 仅首次触发消耗执行次数，重试不重复扣减
        if retry_count == 0:
            try:
                consume_result = _db.consume_schedule_execution(schedule_id)
            except Exception as consume_err:
                uat_logger.error(f"⏰ 定时任务 #{schedule_id} 扣减执行次数失败: {consume_err}")
                return

            if not consume_result.get('allowed'):
                reason = consume_result.get('reason')
                uat_logger.warning(f"⏰ 定时任务 #{schedule_id} 跳过执行: {reason}")
                try:
                    # 若任务已失效/不存在，尝试移除本地job，避免后续重复触发
                    scheduler.remove_job(f'schedule_{schedule_id}')
                except Exception:
                    pass
                return

            if consume_result.get('unlimited'):
                uat_logger.info(f"⏰ 定时任务 #{schedule_id} 执行次数模式: 无限次")
            else:
                uat_logger.info(f"⏰ 定时任务 #{schedule_id} 执行次数已扣减，剩余: {consume_result.get('remaining')}")
                if consume_result.get('exhausted'):
                    uat_logger.info(f"⏰ 定时任务 #{schedule_id} 达到执行次数上限，已自动禁用")
                    try:
                        scheduler.remove_job(f'schedule_{schedule_id}')
                    except Exception:
                        pass

        schedule = None
        for s in _db.get_all_schedules():
            if s['id'] == schedule_id:
                schedule = s
                break

        max_retries = schedule['retry_count'] if schedule else 3
        retry_interval = schedule['retry_interval'] if schedule else 5
        schedule_name = schedule['name'] if schedule else f'任务#{schedule_id}'

        # 创建执行历史记录
        history_id = _db.create_schedule_history(
            schedule_id, case_ids, 'running',
            retry_count, max_retries
        )

        start_time = time.time()
        uat_logger.info(f"⏰ 定时任务 #{schedule_id} 开始执行（第{retry_count + 1}次尝试），用例: {case_ids}")

        if not case_ids:
            _db.update_schedule_history(history_id, 'failed', '调度未配置任何用例')
            uat_logger.error(f"⏰ 定时任务 #{schedule_id} 无用例，记失败")
            notify('schedule_failed', {
                'schedule_name': schedule_name,
                'retry_count': retry_count,
                'success_count': 0,
                'failed_count': 0,
                'total_count': 0,
                'error': '调度未配置任何用例',
                'executed_at': beijing_now_iso()
            })
            return

        machine_lock_acquired = False
        try:
            from execution_lock import acquire as acquire_machine_lock

            machine_lock_acquired = acquire_machine_lock(
                owner=f"schedule:{schedule_id}", timeout_sec=120
            )
            if not machine_lock_acquired:
                _db.update_schedule_history(
                    history_id,
                    'failed',
                    '本机已有自动化任务在执行，跳过本次调度',
                )
                uat_logger.warning(
                    "⏰ 定时任务 #%s 因本机执行锁占用而跳过", schedule_id
                )
                return
        except ImportError:
            pass

        try:
            _spid = schedule.get("project_id") if schedule else None
            _exec_ctx = ExecutionContext(
                user_id=None,
                tenant_id=None,
                trigger="schedule",
                on_case_failure=_on_case_execution_failure,
                extra={"schedule_id": schedule_id, "project_id": _spid},
            )
            results = sync_execute_multiple_test_cases(case_ids, _db, execution_context=_exec_ctx)
            successful = results.get('successful_cases', 0)
            failed = results.get('failed_cases', 0)
            duration = time.time() - start_time
            case_results = results.get('case_results', []) or []

            # 门禁：仅全部 success 才记调度成功（warning/stopped 等不得当绿）
            try:
                from auth_batch_helpers import count_batch_gate_failures, is_execution_gate_success
                gate_failures = count_batch_gate_failures(case_results)
            except ImportError:
                gate_failures = failed
                is_execution_gate_success = lambda s: (s or "") == "success"  # noqa: E731

            # 为每个用例发送单独的通知（case_success 或 case_failed）
            for case_result in case_results:
                case_name = case_result.get('case_name', '未知用例')
                case_status = case_result.get('status', 'unknown')
                case_duration = case_result.get('execution_time', 0)
                case_error = case_result.get('error', '')
                
                if is_execution_gate_success(case_status):
                    notify('case_success', {
                        'case_name': case_name,
                        'duration': case_duration,
                        'executed_at': beijing_now_iso()
                    })
                else:
                    notify('case_failed', {
                        'case_name': case_name,
                        'error': case_error or f'执行未通过（status={case_status}）',
                        'executed_at': beijing_now_iso()
                    })

            if gate_failures == 0 and len(case_results) == len(case_ids) and failed == 0:
                # 全部成功
                _db.update_schedule_history(history_id, 'success')
                uat_logger.info(f"⏰ 定时任务 #{schedule_id} 完成，全部成功")
                # 发送成功通知
                notify('schedule_success', {
                    'schedule_name': schedule_name,
                    'success_count': successful,
                    'total_count': successful + failed,
                    'duration': duration,
                    'executed_at': beijing_now_iso()
                })
            elif retry_count < max_retries:
                # 有失败且还可以重跑
                _fail_n = max(int(failed or 0), int(gate_failures or 0))
                uat_logger.warning(f"⏰ 定时任务 #{schedule_id} 部分失败，{_fail_n}个用例未通过门禁，将在{retry_interval}分钟后重试（{retry_count + 1}/{max_retries}）")
                _db.update_schedule_history(history_id, 'retrying', f'{_fail_n}个用例未通过门禁，准备重试')

                # 等待后重跑
                time.sleep(retry_interval * 60)
                _run_scheduled_cases(schedule_id, case_ids, retry_count + 1)
            else:
                # 达到最大重跑次数，最终失败
                _fail_n = max(int(failed or 0), int(gate_failures or 0))
                _db.update_schedule_history(history_id, 'failed', f'达到最大重试次数，{_fail_n}个用例未通过门禁')
                uat_logger.error(f"⏰ 定时任务 #{schedule_id} 达到最大重试次数，执行失败")
                # 发送失败通知（成功和失败都需要发送对应通知）
                notify('schedule_failed', {
                    'schedule_name': schedule_name,
                    'retry_count': retry_count,
                    'success_count': successful,
                    'failed_count': _fail_n,
                    'total_count': max(successful + failed, len(case_results)),
                    'error': f'{_fail_n}个用例未通过门禁，达到最大重试次数',
                    'executed_at': beijing_now_iso()
                })

        except Exception as e:
            error_msg = str(e)
            uat_logger.error(f"⏰ 定时任务 #{schedule_id} 执行失败: {error_msg}")

            if retry_count < max_retries:
                uat_logger.warning(f"⏰ 定时任务 #{schedule_id} 将在{retry_interval}分钟后重试（{retry_count + 1}/{max_retries}）")
                _db.update_schedule_history(history_id, 'retrying', f'执行异常: {error_msg}')
                time.sleep(retry_interval * 60)
                _run_scheduled_cases(schedule_id, case_ids, retry_count + 1)
            else:
                _db.update_schedule_history(history_id, 'failed', f'达到最大重试次数: {error_msg}')
                # 发送失败通知
                notify('schedule_failed', {
                    'schedule_name': schedule_name,
                    'retry_count': retry_count,
                    'error': error_msg,
                    'executed_at': beijing_now_iso()
                })
        finally:
            if machine_lock_acquired:
                try:
                    from execution_lock import release as release_machine_lock

                    release_machine_lock()
                except ImportError:
                    pass
            try:
                sync_close_browser()
            except Exception:
                pass
            last_run = utc_sql_str()
            _db.update_schedule(schedule_id, last_run=last_run)

    def _register_schedule_job(schedule_id: int, case_ids: list, cron_expr: str):
        """注册/更新定时任务"""
        job_id = f'schedule_{schedule_id}'
        try:
            scheduler.remove_job(job_id)
        except Exception:
            pass
        parts = cron_expr.strip().split()
        # 将 Quartz 风格的 ? 替换为 *
        parts = [p if p != '?' else '*' for p in parts]
        try:
            if len(parts) == 6:
                # 6个部分：秒 分 时 日 月 周
                second, minute, hour, day, month, day_of_week = parts
                trigger = CronTrigger(second=second, minute=minute, hour=hour,
                                      day=day, month=month, day_of_week=day_of_week)
            elif len(parts) == 5:
                # 5个部分：分 时 日 月 周
                minute, hour, day, month, day_of_week = parts
                trigger = CronTrigger(minute=minute, hour=hour, day=day,
                                      month=month, day_of_week=day_of_week)
            else:
                uat_logger.warning(f"定时任务 #{schedule_id} cron表达式格式不正确: {cron_expr}")
                return
            scheduler.add_job(
                _run_scheduled_cases,
                trigger,
                args=[schedule_id, case_ids],
                id=job_id,
                replace_existing=True,
                max_instances=1,
                coalesce=True,
                misfire_grace_time=30
            )
            uat_logger.info(f"定时任务 #{schedule_id} 已注册，cron: {cron_expr}")
        except Exception as e:
            uat_logger.error(f"注册定时任务 #{schedule_id} 失败: {e}")

    def _sync_schedule_job(schedule_id: int):
        """根据数据库状态注册或移除 APScheduler 任务（启用、次数、Cron 变更时调用）。"""
        if scheduler is None or not getattr(scheduler, "running", False):
            return
        _db = Database()
        sched = None
        for s in _db.get_all_schedules():
            if s["id"] == schedule_id:
                sched = s
                break
        job_id = f"schedule_{schedule_id}"
        try:
            scheduler.remove_job(job_id)
        except Exception:
            pass
        if not sched:
            return
        try:
            ec = int(sched.get("execution_count", 0))
        except (TypeError, ValueError):
            ec = 0
        if sched.get("is_active") and ec != 0:
            _register_schedule_job(schedule_id, sched["case_ids"], sched["cron_expr"])

    # 启动调度器并加载已有任务（仅在实际提供 HTTP 的进程启动，避免 Flask debug 重载父子双进程各启一调度器）
    _werkzeug_main = os.environ.get("WERKZEUG_RUN_MAIN")
    _start_apscheduler = _werkzeug_main == "true" or (_werkzeug_main is None and not app.debug)
    if _start_apscheduler:
        scheduler.start()
        _init_db = Database()
        for sched in _init_db.get_active_schedules():
            try:
                try:
                    ec = int(sched.get("execution_count", 0))
                except (TypeError, ValueError):
                    ec = 0
                if ec == 0:
                    continue
                _register_schedule_job(sched["id"], sched["case_ids"], sched["cron_expr"])
            except Exception as e:
                uat_logger.warning(f"加载定时任务 #{sched['id']} 失败: {e}")
        uat_logger.info("APScheduler 调度器已启动")
    else:
        uat_logger.info("APScheduler：当前进程为 Flask 重载监视进程，跳过启动，避免定时任务重复触发")

except ImportError:
    uat_logger.warning("APScheduler 未安装，定时执行功能不可用。运行: pip install APScheduler==3.10.4")
    scheduler = None
    def _register_schedule_job(*args, **kwargs):
        pass

    def _sync_schedule_job(*args, **kwargs):
        pass


# ==================== License 管理 API ====================

@app.route('/license')
@login_required
def license_page():
    """License 管理页面"""
    return render_template('license.html')


@app.route('/user-management')
@login_required
@role_required('admin')
def user_management_page():
    """用户管理页面（仅管理员）"""
    return render_template('user_management.html')


@app.route('/api/license/info', methods=['GET'])
@login_required
@api_error_handler
def api_get_user_license_info():
    """获取当前用户的 License 信息和使用情况"""
    try:
        # 使用缓存的 license 信息，提高性能
        # 只有在激活新 license 时才需要清除缓存
        license_info = license_manager.get_current_license()
        limits = license_manager.get_limits()
        
        # 获取今日使用统计
        _db = Database()
        today_stats = _db.get_user_usage_stats(current_user.id)
        
        return jsonify({
            'success': True,
            'info': {
                'license_type': license_info.license_type,
                'issued_to': license_info.issued_to,
                'issued_at': license_info.issued_at,
                'expires_at': license_info.expires_at,
                'features': license_info.features,
                'effective_features': limits.get('effective_features'),
                'open_core_features_unlocked': limits.get('open_core_features_unlocked'),
            },
            'limits': limits,
            'usage': {
                'today_executions': today_stats.get('execution_count', 0) if today_stats else 0,
                'today_created_cases': today_stats.get('created_cases', 0) if today_stats else 0
            }
        })
    except Exception as e:
        uat_logger.log_exception('api_get_user_license_info', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/license/activate', methods=['POST'])
@login_required
def api_activate_license():
    """用户激活 License"""
    data = request.get_json(silent=True) or {}
    license_key = data.get('license_key', '').strip()
    
    if not license_key:
        return jsonify({'success': False, 'error': '请输入 License 密钥'}), 400

    from deployment_config import is_client_mode, is_server_mode
    from instance_identity import get_instance_id, get_machine_id
    from platform_sync import report_license_activation

    binding_type = ""
    binding_id = ""
    if is_server_mode():
        binding_type = "instance"
        binding_id = get_instance_id()
    elif is_client_mode():
        binding_type = "machine"
        binding_id = get_machine_id()
    else:
        binding_type = "machine"
        binding_id = get_machine_id()

    result = license_manager.activate_license_key(license_key, binding_type, binding_id)
    if not result['valid']:
        return jsonify({'success': False, 'error': result['message']}), 400

    info = result.get('info')
    activation_synced = False
    activation_sync_hint = ''
    if info and info.license_id and binding_id:
        from deployment_config import get_platform_admin_url

        activation_synced = report_license_activation(
            info.license_id, binding_type, binding_id
        )
        if not activation_synced:
            admin_url = get_platform_admin_url()
            if not admin_url:
                activation_sync_hint = (
                    '本地已激活，但未配置 PLATFORM_ADMIN_URL，创始人控制面看不到激活记录。'
                    '请在用户数据目录 .env 中设置，例如 PLATFORM_ADMIN_URL=http://127.0.0.1:5100'
                )
            else:
                activation_sync_hint = (
                    f'本地已激活，但未能连接创始人控制面（{admin_url}）。'
                    '请确认 platform_admin 已启动且网络可达；重启 Testory 后会自动补报一次。'
                )

    limits = license_manager.get_limits()
    return jsonify({
        'success': True,
        'message': 'License 激活成功',
        'license_type': info.license_type if info else '',
        'expires_at': info.expires_at if info else '',
        'activation_synced': activation_synced,
        'activation_sync_hint': activation_sync_hint,
        'license': {
            'type': info.license_type if info else '',
            'features': limits.get('features') or [],
            'limits': {
                'max_projects': limits.get('max_projects'),
                'max_cases_per_project': limits.get('max_cases_per_project'),
                'max_executions_per_day': limits.get('max_executions_per_day'),
            },
        },
    })


@app.route('/api/license', methods=['GET'])
@login_required
@role_required('admin')
def api_get_license_info():
    """获取当前 License 信息"""
    license_info = license_manager.get_current_license()
    limits = license_manager.get_limits()
    return jsonify({
        'success': True,
        'license': {
            'type': license_info.license_type,
            'issued_to': license_info.issued_to,
            'issued_at': license_info.issued_at,
            'expires_at': license_info.expires_at,
            'features': license_info.features,
            'limits': limits
        }
    })


@app.route('/api/license', methods=['POST'])
@login_required
@role_required('admin')
def api_upload_license():
    """上传并激活 License"""
    data = request.get_json(silent=True) or {}
    license_str = data.get('license', '').strip()

    if not license_str:
        return jsonify({'success': False, 'error': 'License 不能为空'}), 400

    # 验证 License
    result = license_manager.validate_license(license_str)
    if not result['valid']:
        return jsonify({'success': False, 'error': result['message']}), 400

    from deployment_config import is_client_mode, is_server_mode
    from instance_identity import get_instance_id, get_machine_id

    binding_type = ""
    binding_id = ""
    if is_server_mode():
        binding_type = "instance"
        binding_id = get_instance_id()
    elif is_client_mode():
        binding_type = "machine"
        binding_id = get_machine_id()

    result = license_manager.activate_license_key(license_str, binding_type, binding_id)
    if not result['valid']:
        return jsonify({'success': False, 'error': result['message']}), 400

    info = result.get('info')
    if info and info.license_id and binding_id:
        from platform_sync import report_license_activation

        report_license_activation(info.license_id, binding_type, binding_id)

    return jsonify({
        'success': True,
        'message': 'License 激活成功',
        'license': {
            'type': info.license_type if info else '',
            'expires_at': info.expires_at if info else '',
            'license_id': info.license_id if info else '',
        }
    })


# ==================== 审计日志 API ====================

def _collect_audit_export_rows(target_type=None, username=None, max_rows=50000):
    """分页拉取审计日志，用于导出（最多 max_rows 条）。"""
    _db = Database()
    rows_out = []
    page = 1
    page_size = 2000
    while len(rows_out) < max_rows:
        logs = _db.get_audit_logs(
            target_type=target_type,
            username=username,
            page=page,
            page_size=page_size,
        )
        if not logs:
            break
        for log in logs:
            details = log.get('details') or ''
            if details is not None and not isinstance(details, str):
                details = str(details)
            if isinstance(details, str) and len(details) > 8000:
                details = details[:8000] + '...'
            created = log.get('created_at') or ''
            rows_out.append([
                created,
                log.get('username') or '',
                log.get('action') or '',
                log.get('target_type') or '',
                log.get('target_id') if log.get('target_id') is not None else '',
                log.get('ip_address') or '',
                details,
            ])
            if len(rows_out) >= max_rows:
                return rows_out
        if len(logs) < page_size:
            break
        page += 1
    return rows_out


def _audit_logs_export_response(fmt: str):
    """生成审计日志导出响应，fmt 为 csv 或 xlsx。"""
    import io
    import csv
    from datetime import datetime

    fmt = (fmt or '').lower()
    target_type = (request.args.get('target_type') or '').strip() or None
    username = (request.args.get('username') or '').strip() or None

    header = ['时间', '用户', '操作', '目标类型', '目标ID', 'IP', '详情']
    data_rows = _collect_audit_export_rows(target_type=target_type, username=username)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(header)
        writer.writerows(data_rows)
        filename = f'audit_logs_{ts}.csv'
        payload = output.getvalue().encode('utf-8-sig')
        return Response(
            payload,
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Cache-Control': 'no-store',
            },
        )

    if fmt in ('xlsx', 'excel'):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({'success': False, 'error': '请先安装 openpyxl: pip install openpyxl'}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = 'audit_logs'
        ws.append(header)
        for row in data_rows:
            ws.append(row)
        bold = Font(bold=True)
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = bold
        ws.freeze_panes = 'A2'
        widths = (22, 14, 28, 14, 12, 18, 72)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f'audit_logs_{ts}.xlsx'
        return Response(
            bio.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Cache-Control': 'no-store',
            },
        )

    return jsonify({'success': False, 'error': '不支持的导出格式'}), 400


def _resolve_runtime_log_path(source: str):
    from pathlib import Path

    raw_dir = os.environ.get('UAT_DATA_DIR', '').strip()
    log_dir = (Path(raw_dir) / 'logs') if raw_dir else Path('logs')
    source = (source or 'platform').strip().lower()
    if source == 'backend':
        return log_dir / 'backend_startup.log'
    today = datetime.datetime.now().strftime('%Y%m%d')
    return log_dir / f'uat_platform_{today}.log'


@app.route('/runtime-logs')
@login_required
def runtime_logs_page():
    """桌面版运行日志 tail 页面（Tauri / 本地客户端）。"""
    return render_template('runtime_logs.html')


@app.route('/stream/logs')
@login_required
def stream_runtime_logs():
    """SSE tail 平台或 backend 启动日志。"""
    import json
    import time
    from flask import Response, stream_with_context

    source = (request.args.get('source') or 'platform').strip().lower()
    log_path = _resolve_runtime_log_path(source)
    tail_bytes = max(4096, min(int(request.args.get('tail_bytes', 65536)), 512 * 1024))

    def _emit(event: str, payload: dict):
        return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

    def _read_tail_lines(path, max_bytes: int):
        if not path.is_file():
            return []
        try:
            size = path.stat().st_size
            with path.open('rb') as fp:
                if size > max_bytes:
                    fp.seek(size - max_bytes)
                    fp.readline()
                chunk = fp.read().decode('utf-8', errors='replace')
            return [ln for ln in chunk.splitlines() if ln != '']
        except OSError:
            return []

    @stream_with_context
    def generate():
        yield _emit('meta', {'path': str(log_path), 'source': source})
        idx = 0
        for line in _read_tail_lines(log_path, tail_bytes):
            yield _emit('line', {'idx': idx, 'text': line})
            idx += 1

        if not log_path.is_file():
            yield _emit('line', {'idx': idx, 'text': '[日志文件尚未创建，等待写入…]'})
            idx += 1

        pos = 0
        while True:
            if not log_path.is_file():
                time.sleep(1.0)
                continue
            try:
                with log_path.open('r', encoding='utf-8', errors='replace') as fp:
                    fp.seek(pos)
                    while True:
                        line = fp.readline()
                        if not line:
                            pos = fp.tell()
                            break
                        text = line.rstrip('\n\r')
                        if text:
                            yield _emit('line', {'idx': idx, 'text': text})
                            idx += 1
            except OSError:
                pass
            time.sleep(0.45)

    headers = {
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'X-Accel-Buffering': 'no',
    }
    return Response(generate(), mimetype='text/event-stream', headers=headers)


@app.route('/api/audit-logs', methods=['GET'])
@login_required
@role_required('admin')
@feature_required('audit_log')
def api_get_audit_logs():
    """获取审计日志（仅管理员）。指定 format=csv 或 format=xlsx 时导出文件（与列表筛选一致）。"""
    fmt = (request.args.get('format') or '').strip().lower()
    if fmt in ('csv', 'xlsx', 'excel'):
        if fmt == 'excel':
            fmt = 'xlsx'
        return _audit_logs_export_response(fmt)

    page = max(1, request.args.get('page', 1, type=int) or 1)
    page_size = request.args.get('page_size', 50, type=int) or 50
    page_size = min(200, max(1, page_size))
    target_type = (request.args.get('target_type') or '').strip() or None
    username = (request.args.get('username') or '').strip() or None

    _db = Database()
    logs = _db.get_audit_logs(
        target_type=target_type,
        username=username,
        page=page,
        page_size=page_size,
    )
    total = _db.get_audit_logs_count(target_type=target_type, username=username)

    return jsonify({
        'success': True,
        'logs': logs,
        'total': total,
        'page': page,
        'page_size': page_size
    })


@app.route('/api/audit-logs/export', methods=['GET'])
@login_required
@role_required('admin')
@feature_required('audit_log')
def api_export_audit_logs():
    """兼容旧版前端路径：等同于 format=csv。"""
    return _audit_logs_export_response('csv')


# ==================== 项目成员管理 API ====================

@app.route('/api/projects/<int:project_id>/members', methods=['GET'])
@login_required
@project_access_required(min_role='viewer')
def api_get_project_members(project_id):
    """获取项目成员列表"""
    _db = Database()
    members = _db.get_project_members(project_id)
    return jsonify({'success': True, 'members': members})


@app.route('/api/projects/<int:project_id>/members', methods=['POST'])
@login_required
@project_access_required(min_role='owner')
@audit_log('ADD_PROJECT_MEMBER', 'project_member')
def api_add_project_member(project_id):
    """添加项目成员"""
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    role = data.get('role', 'editor')

    if not username:
        return jsonify({'success': False, 'error': '用户名不能为空'}), 400

    if role not in ('viewer', 'editor', 'owner'):
        return jsonify({'success': False, 'error': '无效的角色'}), 400

    _db = Database()
    user = _db.get_user_by_username(username)
    if not user:
        return jsonify({'success': False, 'error': '用户不存在'}), 404

    success = _db.add_project_member(project_id, user['id'], role)
    if success:
        return jsonify({'success': True, 'message': '成员添加成功'})
    else:
        return jsonify({'success': False, 'error': '成员已存在'}), 409


@app.route('/api/projects/<int:project_id>/members/<int:user_id>', methods=['PUT'])
@login_required
@project_access_required(min_role='owner')
@audit_log('UPDATE_PROJECT_MEMBER', 'project_member')
def api_update_project_member(project_id, user_id):
    """更新项目成员角色"""
    data = request.get_json(silent=True) or {}
    role = data.get('role')

    if role not in ('viewer', 'editor', 'owner'):
        return jsonify({'success': False, 'error': '无效的角色'}), 400

    _db = Database()
    success = _db.update_project_member_role(project_id, user_id, role)
    return jsonify({'success': success})


@app.route('/api/projects/<int:project_id>/members/<int:user_id>', methods=['DELETE'])
@login_required
@project_access_required(min_role='owner')
@audit_log('REMOVE_PROJECT_MEMBER', 'project_member')
def api_remove_project_member(project_id, user_id):
    """移除项目成员"""
    if user_id == current_user.id:
        return jsonify({'success': False, 'error': '不能移除自己'}), 400

    _db = Database()
    success = _db.remove_project_member(project_id, user_id)
    return jsonify({'success': success})


# ==================== 健康检查 API ====================

@app.route('/api/health', methods=['GET'])
def api_health():
    """存活探针：仅表示进程可响应（不访问数据库，便于与就绪探针区分）。"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.datetime.now().isoformat(),
        'version': '2.0.0',
    })


@app.route('/api/startup/status', methods=['GET'])
def api_startup_status():
    """桌面壳启动页轮询：分阶段进度文案。"""
    try:
        from desktop_startup import mark_app_ready, startup_status_payload

        payload = startup_status_payload()
        try:
            if db.ping():
                payload["database"] = "ok"
                if payload.get("phase") == "booting":
                    payload["message"] = "正在启动服务…"
            else:
                payload["database"] = "pending"
                payload["message"] = "正在初始化数据库…"
        except Exception:
            payload["database"] = "pending"
            payload["message"] = "正在初始化数据库…"
        if payload.get("ready"):
            mark_app_ready()
        return jsonify({"success": True, **payload})
    except Exception as exc:
        return jsonify(
            {
                "success": True,
                "phase": "ready",
                "message": "准备就绪",
                "ready": True,
                "error": str(exc),
            }
        )


@app.route('/api/health/ready', methods=['GET'])
def api_health_ready():
    """就绪探针：校验 SQLite 可连接（编排/负载均衡建议用此 URL）。"""
    payload = {
        'timestamp': datetime.datetime.now().isoformat(),
    }
    if sys.platform == 'win32':
        try:
            from desktop_runtime import desktop_runtime_available, desktop_runtime_unavailable_reason

            payload['desktop_runtime'] = {
                'available': bool(desktop_runtime_available()),
                'reason': desktop_runtime_unavailable_reason() or None,
                'python_executable': sys.executable,
            }
        except Exception as exc:
            payload['desktop_runtime'] = {
                'available': False,
                'reason': str(exc),
                'python_executable': sys.executable,
            }
    try:
        from embedded_browser_client import embedded_gateway_config, embedded_gateway_enabled

        base, secret, pub_ws = embedded_gateway_config()
        payload['embedded_browser'] = {
            'configured': embedded_gateway_enabled(),
            'gateway_url': base or None,
            'public_ws_base': pub_ws or None,
        }
        if embedded_gateway_enabled():
            import socket
            from urllib.parse import urlparse

            parsed = urlparse(base)
            host = parsed.hostname or '127.0.0.1'
            port = parsed.port or 8765
            try:
                with socket.create_connection((host, port), timeout=0.5):
                    payload['embedded_browser']['listening'] = True
            except OSError:
                payload['embedded_browser']['listening'] = False
    except Exception as exc:
        payload['embedded_browser'] = {'configured': False, 'error': str(exc)}

    if not db.ping():
        payload['status'] = 'unready'
        payload['database'] = 'unavailable'
        return jsonify(payload), 503
    payload['status'] = 'ready'
    payload['database'] = 'ok'
    return jsonify(payload)


# ==================== 通知配置 API ====================

@app.route('/api/notifications/configs', methods=['GET'])
@login_required
@role_required('admin')
def api_get_notification_configs():
    """获取通知配置列表"""
    db = Database()
    configs = db.get_all_notification_configs()
    return jsonify({'success': True, 'configs': configs})


@app.route('/api/notifications/configs', methods=['POST'])
@login_required
@role_required('admin')
@audit_log('CREATE_NOTIFICATION_CONFIG', 'notification')
def api_create_notification_config():
    """创建通知配置"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    config_type = data.get('type', '').strip()
    config_data = data.get('config', {})
    events = data.get('events', [])
    is_active = 1 if data.get('is_active', True) else 0

    if not name or not config_type:
        return jsonify({'success': False, 'error': '名称和类型不能为空'}), 400

    db = Database()
    config_id = db.create_notification_config(name, config_type, config_data, events, is_active)

    # 重新加载通知管理器配置
    from notification_manager import notification_manager
    notification_manager.reload_configs()

    return jsonify({'success': True, 'message': '配置创建成功', 'id': config_id})


@app.route('/api/notifications/configs/<int:config_id>', methods=['PUT'])
@login_required
@role_required('admin')
@audit_log('UPDATE_NOTIFICATION_CONFIG', 'notification')
def api_update_notification_config(config_id):
    """更新通知配置"""
    data = request.get_json(silent=True) or {}

    updates = {}
    if 'name' in data:
        updates['name'] = data['name'].strip()
    if 'type' in data:
        updates['type'] = data['type'].strip()
    if 'config' in data:
        updates['config'] = data['config']
    if 'events' in data:
        updates['events'] = data['events']
    if 'is_active' in data:
        updates['is_active'] = 1 if data['is_active'] else 0

    if not updates:
        return jsonify({'success': False, 'error': '没有要更新的字段'}), 400

    db = Database()
    success = db.update_notification_config(config_id, **updates)

    if success:
        # 重新加载通知管理器配置
        from notification_manager import notification_manager
        notification_manager.reload_configs()
        return jsonify({'success': True, 'message': '配置更新成功'})
    else:
        return jsonify({'success': False, 'error': '配置不存在'}), 404


@app.route('/api/notifications/configs/<int:config_id>', methods=['DELETE'])
@login_required
@role_required('admin')
@audit_log('DELETE_NOTIFICATION_CONFIG', 'notification')
def api_delete_notification_config(config_id):
    """删除通知配置"""
    db = Database()
    success = db.delete_notification_config(config_id)

    if success:
        # 重新加载通知管理器配置
        from notification_manager import notification_manager
        notification_manager.reload_configs()
        return jsonify({'success': True, 'message': '配置删除成功'})
    else:
        return jsonify({'success': False, 'error': '配置不存在'}), 404


@app.route('/api/notifications/test', methods=['POST'])
@login_required
@role_required('admin')
def api_test_notification():
    """测试通知"""
    from notification_manager import notification_manager, NotificationConfig

    data = request.get_json(silent=True) or {}
    config_data = data.get('config', {})
    config_type = config_data.get('type', 'webhook')

    # 根据类型构建配置
    if config_type == 'email':
        config = NotificationConfig(
            name='test',
            type='email',
            webhook_url='',  # email 不需要 webhook_url
            secret=None,
            enabled=True
        )
        # 邮件配置存储在 template 字段
        config.template = json.dumps({
            'smtp_server': config_data.get('smtp_server', ''),
            'smtp_port': config_data.get('smtp_port', 587),
            'username': config_data.get('username', ''),
            'password': config_data.get('password', ''),
            'to_emails': config_data.get('to_emails', [])
        })
    else:
        config = NotificationConfig(
            name='test',
            type=config_type,
            webhook_url=config_data.get('webhook_url', ''),
            secret=config_data.get('secret') or None,
            enabled=True
        )

    success = notification_manager.send_notification(
        config,
        title="测试通知",
        content="这是一条测试通知消息",
        data={'test': True}
    )

    return jsonify({
        'success': success,
        'message': '发送成功' if success else '发送失败'
    })


# ==================== 浏览器支持 API ====================

@app.route('/api/browsers', methods=['GET'])
@login_required
def api_get_browsers():
    """获取支持的浏览器列表"""
    from browser_manager import BrowserManager
    browsers = BrowserManager.get_available_browsers()
    devices = BrowserManager.get_device_presets()
    return jsonify({
        'success': True,
        'browsers': browsers,
        'device_presets': devices
    })


# ==================== 断言类型 API ====================

@app.route('/api/assertion-types', methods=['GET'])
@login_required
def api_get_assertion_types():
    """获取支持的断言类型"""
    types = [
        {'id': 'text_equals', 'name': '文本相等', 'category': 'text'},
        {'id': 'text_contains', 'name': '文本包含', 'category': 'text'},
        {'id': 'text_regex', 'name': '文本正则匹配', 'category': 'text'},
        {'id': 'element_exists', 'name': '元素存在', 'category': 'element'},
        {'id': 'element_visible', 'name': '元素可见', 'category': 'element'},
        {'id': 'element_attr', 'name': '元素属性', 'category': 'element'},
        {'id': 'element_css', 'name': '元素CSS样式', 'category': 'element'},
        {'id': 'element_count', 'name': '元素数量', 'category': 'element'},
        {'id': 'url_equals', 'name': 'URL相等', 'category': 'url'},
        {'id': 'url_contains', 'name': 'URL包含', 'category': 'url'},
        {'id': 'api_status', 'name': 'API状态码', 'category': 'api'},
        {'id': 'api_json', 'name': 'API JSON响应', 'category': 'api'},
        {'id': 'javascript', 'name': 'JavaScript执行', 'category': 'script'}
    ]
    return jsonify({'success': True, 'types': types})


# ==================== SSO 单点登录 API ====================

@app.route('/sso-settings')
@login_required
@role_required('admin')
@feature_required('sso')
def sso_settings_page():
    """返回 SSO 设置页面"""
    return render_template('sso_settings.html')


@app.route('/api/sso/configs', methods=['GET'])
@login_required
@role_required('admin')
@feature_required('sso')
def api_get_sso_configs():
    """获取 SSO 配置列表"""
    try:
        from sso_manager import sso_manager
        configs = sso_manager.get_sso_configs()
        return jsonify({'success': True, 'configs': configs})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/configs', methods=['POST'])
@login_required
@role_required('admin')
@feature_required('sso')
@audit_log('CREATE_SSO_CONFIG', 'sso_config')
def api_create_sso_config():
    """创建 SSO 配置"""
    try:
        from sso_manager import sso_manager
        data = request.get_json(silent=True) or {}
        
        if not data.get('provider_type'):
            return jsonify({'success': False, 'error': '请选择 SSO 类型'}), 400
        if not data.get('name'):
            return jsonify({'success': False, 'error': '请输入配置名称'}), 400
        
        config_id = sso_manager.create_sso_config(data)
        return jsonify({'success': True, 'config_id': config_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/configs/<int:config_id>', methods=['GET'])
@login_required
@role_required('admin')
@feature_required('sso')
def api_get_sso_config_detail(config_id):
    """获取单个 SSO 配置详情（用于编辑）"""
    try:
        from sso_manager import sso_manager
        config = sso_manager.get_sso_config(config_id)
        if not config:
            return jsonify({'success': False, 'error': 'SSO 配置不存在'}), 404
        return jsonify({'success': True, 'config': config.__dict__})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/configs/<int:config_id>', methods=['PUT'])
@login_required
@role_required('admin')
@feature_required('sso')
@audit_log('UPDATE_SSO_CONFIG', 'sso_config')
def api_update_sso_config(config_id):
    """更新 SSO 配置"""
    try:
        from sso_manager import sso_manager
        data = request.get_json(silent=True) or {}
        success = sso_manager.update_sso_config(config_id, data)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/configs/<int:config_id>', methods=['DELETE'])
@login_required
@role_required('admin')
@feature_required('sso')
@audit_log('DELETE_SSO_CONFIG', 'sso_config')
def api_delete_sso_config(config_id):
    """删除 SSO 配置"""
    try:
        from sso_manager import sso_manager
        success = sso_manager.delete_sso_config(config_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/login/<int:config_id>', methods=['GET'])
def api_sso_login_url(config_id):
    """获取 SSO 登录 URL"""
    try:
        from sso_manager import sso_manager, SSOProviderType
        config = sso_manager.get_sso_config(config_id)
        
        if not config:
            return jsonify({'success': False, 'error': 'SSO 配置不存在'}), 404
        
        redirect_uri = request.host_url.rstrip('/') + f'/api/sso/callback/{config_id}'
        
        if config.provider_type == SSOProviderType.WECOM.value:
            login_url = sso_manager.wecom_get_login_url(config, redirect_uri)
        elif config.provider_type == SSOProviderType.OAUTH2.value:
            login_url = sso_manager.oauth2_get_login_url(config)
            if not login_url:
                return jsonify({'success': False, 'error': 'OAuth2 授权地址或回调地址未配置完整'}), 400
        else:
            return jsonify({'success': False, 'error': '该 SSO 类型不支持跳转登录'}), 400
        
        return jsonify({'success': True, 'login_url': login_url})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sso/callback/<int:config_id>', methods=['GET'])
def api_sso_callback(config_id):
    """处理 SSO 登录回调"""
    try:
        from urllib.parse import quote

        from sso_manager import sso_manager, SSOProviderType

        def _err_redirect(msg: str):
            return redirect('/login?error=' + quote(msg, safe=''))

        config = sso_manager.get_sso_config(config_id)
        if not config:
            return _err_redirect('SSO配置不存在')

        code = request.args.get('code')
        state = request.args.get('state')

        # 企业微信 / OAuth2：校验 state，防止 CSRF 与伪造回调
        if config.provider_type in (SSOProviderType.WECOM.value, SSOProviderType.OAUTH2.value):
            if not state:
                return _err_redirect('缺少安全参数state请重新登录')
            verified_id = sso_manager.verify_state(state)
            if verified_id != config_id:
                uat_logger.warning(
                    'SSO state 校验失败: path_config_id=%s verified_id=%s', config_id, verified_id
                )
                return _err_redirect('登录状态已失效请重新登录')

        if not code:
            return _err_redirect('授权失败未返回code')

        success, user_info, message = sso_manager.authenticate(config_id, code=code)

        if not success:
            uat_logger.error(f"SSO 登录失败: {message}")
            try:
                from auth_audit import ACTION_SSO_LOGIN_FAILURE, record_auth_audit

                record_auth_audit(
                    action=ACTION_SSO_LOGIN_FAILURE,
                    username=str((user_info or {}).get('username') or ''),
                    user_id=0,
                    ip_address=request.remote_addr,
                    details={
                        'method': 'sso',
                        'provider': getattr(config, 'provider_type', None),
                        'config_id': config_id,
                        'reason': (message or '')[:300],
                    },
                )
            except Exception:
                pass
            return _err_redirect(message or 'SSO登录失败')

        # 获取或创建用户
        user_id = sso_manager.get_or_create_user(
            config.provider_type,
            user_info,
            tenant_id=config.tenant_id
        )
        
        # 登录用户
        _db = Database()
        user_data = _db.get_user_by_id(user_id)
        if user_data:
            user = UserModel(user_data)
            login_user(user, remember=True)
            _db.update_user_last_login(user_id)
            uat_logger.info(f"SSO 用户 {user_info.get('username')} 登录成功")
            try:
                from auth_audit import ACTION_SSO_LOGIN_SUCCESS, record_auth_audit

                record_auth_audit(
                    action=ACTION_SSO_LOGIN_SUCCESS,
                    username=user_data.get('username') or user_info.get('username') or '',
                    user_id=user_id,
                    ip_address=request.remote_addr,
                    details={
                        'method': 'sso',
                        'provider': getattr(config, 'provider_type', None),
                        'config_id': config_id,
                    },
                    db=_db,
                )
            except Exception:
                pass
            return redirect('/')
        
        return redirect('/login?error=user_not_found')
    except Exception as e:
        uat_logger.error(f"SSO 回调处理失败: {e}")
        return redirect('/login?error=sso_error')


@app.route('/api/sso/ldap-login', methods=['POST'])
def api_sso_ldap_login():
    """处理 LDAP 登录"""
    try:
        from sso_manager import sso_manager
        data = request.get_json(silent=True) or {}
        config_id = data.get('config_id')
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not config_id:
            return jsonify({'success': False, 'error': '请选择 LDAP 配置'}), 400
        if not username or not password:
            return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400
        
        success, user_info, message = sso_manager.authenticate(
            config_id, username=username, password=password
        )
        
        if not success:
            try:
                from auth_audit import ACTION_LDAP_LOGIN_FAILURE, record_auth_audit

                record_auth_audit(
                    action=ACTION_LDAP_LOGIN_FAILURE,
                    username=username,
                    user_id=0,
                    ip_address=request.remote_addr,
                    details={
                        'method': 'ldap',
                        'config_id': config_id,
                        'reason': (message or '')[:300],
                    },
                )
            except Exception:
                pass
            return jsonify({'success': False, 'error': message}), 401
        
        # 获取或创建用户
        config = sso_manager.get_sso_config(config_id)
        user_id = sso_manager.get_or_create_user(
            config.provider_type,
            user_info,
            tenant_id=config.tenant_id
        )
        
        # 登录用户
        _db = Database()
        user_data = _db.get_user_by_id(user_id)
        if user_data:
            user = UserModel(user_data)
            login_user(user, remember=True)
            _db.update_user_last_login(user_id)
            uat_logger.info(f"LDAP 用户 {username} 登录成功")
            try:
                from auth_audit import ACTION_LDAP_LOGIN_SUCCESS, record_auth_audit

                record_auth_audit(
                    action=ACTION_LDAP_LOGIN_SUCCESS,
                    username=user_data.get('username') or username,
                    user_id=user_id,
                    ip_address=request.remote_addr,
                    details={'method': 'ldap', 'config_id': config_id},
                    db=_db,
                )
            except Exception:
                pass
            return jsonify({'success': True, 'user': {'id': user_id, 'username': user_data['username']}})
        
        return jsonify({'success': False, 'error': '用户创建失败'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 支付集成 API ====================

@app.route('/pricing')
def pricing_page():
    """返回定价页面"""
    blocked = guard_billing_route()
    if blocked:
        return blocked
    return render_template('pricing.html')


@app.route('/api/payment/plans', methods=['GET'])
def api_get_payment_plans():
    """获取套餐列表"""
    try:
        from payment_manager import payment_manager
        plans = payment_manager.get_plan_list()
        return jsonify({'success': True, 'plans': plans})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/orders', methods=['POST'])
@login_required
@audit_log('CREATE_ORDER', 'order')
def api_create_order():
    """创建订单"""
    try:
        from payment_manager import payment_manager
        data = request.get_json(silent=True) or {}
        
        plan_type = data.get('plan_type')
        period = data.get('period', 'monthly')
        quantity = data.get('quantity', 1)
        
        if not plan_type:
            return jsonify({'success': False, 'error': '请选择套餐类型'}), 400
        
        order = payment_manager.create_order(
            user_id=current_user.id,
            plan_type=plan_type,
            period=period,
            quantity=quantity
        )
        
        return jsonify({'success': True, 'order': order})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/orders', methods=['GET'])
@login_required
def api_get_user_orders():
    """获取用户订单列表"""
    try:
        from payment_manager import payment_manager
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        
        orders, total = payment_manager.get_user_orders(current_user.id, page, page_size)
        
        return jsonify({
            'success': True,
            'orders': orders,
            'total': total,
            'page': page,
            'page_size': page_size
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/orders/<order_no>', methods=['GET'])
@login_required
def api_get_order_detail(order_no):
    """获取订单详情"""
    try:
        from payment_manager import payment_manager
        order = payment_manager.get_order(order_no=order_no)
        
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404
        
        if order['user_id'] != current_user.id and current_user.role != 'admin':
            return jsonify({'success': False, 'error': '无权限查看此订单'}), 403
        
        # 已支付订单返回自动生成的 License Key（按订单唯一）
        license_key = None
        if order.get('status') == 'paid':
            license_key = payment_manager.get_or_create_order_license(order_no)
        return jsonify({'success': True, 'order': order, 'license_key': license_key})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/orders/<order_no>/cancel', methods=['POST'])
@login_required
def api_cancel_order(order_no):
    """取消订单"""
    try:
        from payment_manager import payment_manager
        ok, err = payment_manager.cancel_order(order_no, current_user.id)
        if not ok:
            return jsonify({'success': False, 'error': err}), 400
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/pay/<order_no>', methods=['POST'])
@login_required
def api_create_payment(order_no):
    """创建支付"""
    try:
        from payment_manager import payment_manager, PaymentMethod
        data = request.get_json(silent=True) or {}
        payment_method = data.get('payment_method', 'alipay')
        
        order = payment_manager.get_order(order_no=order_no)
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404
        
        if order['user_id'] != current_user.id:
            return jsonify({'success': False, 'error': '无权限支付此订单'}), 403
        
        if payment_method == PaymentMethod.ALIPAY.value:
            result = payment_manager.create_alipay_payment(order_no)
        elif payment_method == PaymentMethod.WECHAT.value:
            result = payment_manager.create_wechat_payment(order_no)
        else:
            return jsonify({'success': False, 'error': '不支持的支付方式'}), 400
        
        return jsonify(result)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/callback/alipay', methods=['POST'])
def api_alipay_callback():
    """支付宝支付回调"""
    try:
        from payment_manager import payment_manager
        params = request.form.to_dict()
        
        success, order_no = payment_manager.verify_alipay_callback(params)
        
        if success:
            uat_logger.info(f"支付宝支付成功: 订单 {order_no}")
            return 'success'
        else:
            uat_logger.warning(f"支付宝支付回调失败: 订单 {order_no}")
            return 'fail'
    except Exception as e:
        uat_logger.error(f"支付宝回调处理异常: {e}")
        return 'fail'


@app.route('/api/payment/callback/wechat', methods=['POST'])
def api_wechat_callback():
    """微信支付回调"""
    try:
        from payment_manager import payment_manager
        xml_data = request.data.decode('utf-8')
        
        success, order_no = payment_manager.verify_wechat_callback(xml_data)
        
        if success:
            uat_logger.info(f"微信支付成功: 订单 {order_no}")
            return '<xml><return_code><![CDATA[SUCCESS]]></return_code></xml>'
        else:
            uat_logger.warning(f"微信支付回调失败")
            return '<xml><return_code><![CDATA[FAIL]]></return_code></xml>'
    except Exception as e:
        uat_logger.error(f"微信回调处理异常: {e}")
        return '<xml><return_code><![CDATA[FAIL]]></return_code></xml>'


@app.route('/api/payment/subscription', methods=['GET'])
@login_required
def api_get_subscription():
    """获取当前用户订阅信息"""
    try:
        from payment_manager import payment_manager
        subscription = payment_manager.get_user_subscription(current_user.id)
        return jsonify({'success': True, 'subscription': subscription})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment/mock/complete/<order_no>', methods=['POST'])
@login_required
def api_mock_complete_payment(order_no):
    """模拟支付成功（开发/演示使用）"""
    try:
        from payment_manager import payment_manager, PaymentStatus, PaymentMethod
        order = payment_manager.get_order(order_no=order_no)
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404
        if order['user_id'] != current_user.id and current_user.role != 'admin':
            return jsonify({'success': False, 'error': '无权限操作此订单'}), 403
        if order['status'] == PaymentStatus.PAID.value:
            return jsonify({'success': True, 'message': '订单已支付'})

        txid = f"MOCK-{int(time.time())}-{current_user.id}"
        ok = payment_manager.update_order_status(
            order_no=order_no,
            status=PaymentStatus.PAID.value,
            payment_method=PaymentMethod.MANUAL.value,
            transaction_id=txid
        )
        return jsonify({'success': bool(ok), 'transaction_id': txid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/payment/<order_no>')
@login_required
def payment_page(order_no):
    blocked = guard_billing_route()
    if blocked:
        return blocked
    """支付页面"""
    return render_template('payment.html', order_no=order_no)


@app.route('/payment/orders')
@login_required
def payment_orders_page():
    """订单列表页面"""
    blocked = guard_billing_route()
    if blocked:
        return blocked
    return render_template('payment_orders.html')


# ==================== 审计日志页面 ====================

@app.route('/audit-logs')
@login_required
@role_required('admin')
@feature_required('audit_log')
def audit_logs_page():
    """审计日志页面"""
    return render_template('audit_logs.html')


# ==================== 缺陷管理 API ====================

@app.route('/defects')
@login_required
@feature_required('defect_management')
def defects_page():
    """缺陷管理页面"""
    return render_template('defects.html')

@app.route('/defects/<int:defect_id>')
@login_required
def defect_detail_page(defect_id):
    """缺陷详情页面"""
    return render_template('defect_detail.html', defect_id=defect_id)

@app.route('/api/defects', methods=['GET'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_get_defects():
    """获取缺陷列表"""
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status')
    assignee_id = request.args.get('assignee_id', type=int)
    severity = request.args.get('severity')
    case_id = request.args.get('case_id', type=int)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    
    _db = Database()
    result = _db.get_defects(
        project_id=project_id, status=status, assignee_id=assignee_id,
        severity=severity, case_id=case_id, page=page, page_size=page_size
    )
    return jsonify({'success': True, **result})

@app.route('/api/defects', methods=['POST'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_DEFECT', 'defect')
def api_create_defect():
    """创建缺陷"""
    data = request.get_json(silent=True) or {}
    
    project_id = data.get('project_id')
    title = data.get('title', '').strip()
    
    if not project_id:
        return jsonify({'success': False, 'error': '项目ID不能为空'}), 400
    if not title:
        return jsonify({'success': False, 'error': '缺陷标题不能为空'}), 400
    
    _db = Database()
    defect_id = _db.create_defect(
        project_id=project_id,
        title=title,
        reporter_id=current_user.id,
        description=data.get('description', ''),
        severity=data.get('severity', 'medium'),
        priority=data.get('priority', 'medium'),
        assignee_id=data.get('assignee_id'),
        case_id=data.get('case_id'),
        run_history_id=data.get('run_history_id'),
        step_result_id=data.get('step_result_id'),
        error_message=data.get('error_message', ''),
        screenshots=data.get('screenshots', ''),
        environment=data.get('environment', ''),
        browser_info=data.get('browser_info', ''),
        reproduce_steps=data.get('reproduce_steps', ''),
        expected_result=data.get('expected_result', ''),
        actual_result=data.get('actual_result', ''),
        status=data.get('status', 'open')
    )
    
    return jsonify({'success': True, 'defect_id': defect_id})

@app.route('/api/defects/batch_from_cases', methods=['POST'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
def api_batch_create_defects_from_cases():
    """从选中的测试用例批量创建缺陷"""
    data = request.get_json(silent=True) or {}
    
    case_ids = data.get('case_ids', [])
    project_id = data.get('project_id')
    
    if not case_ids or not isinstance(case_ids, list):
        return jsonify({'success': False, 'error': '请选择至少一个测试用例'}), 400
    if not project_id:
        return jsonify({'success': False, 'error': '项目ID不能为空'}), 400
    
    title_template = data.get('title_template', '').strip()
    description = data.get('description', '').strip()
    severity = data.get('severity', 'medium')
    priority = data.get('priority', 'medium')
    assignee_id = data.get('assignee_id')
    environment = data.get('environment', '').strip()
    expected_result = data.get('expected_result', '').strip()
    actual_result = data.get('actual_result', '').strip()
    
    valid_severities = ['critical', 'high', 'medium', 'low']
    if severity not in valid_severities:
        severity = 'medium'
    valid_priorities = ['urgent', 'high', 'medium', 'low']
    if priority not in valid_priorities:
        priority = 'medium'
    
    _db = Database()
    created_ids = _db.batch_create_defects_from_cases(
        case_ids=case_ids,
        project_id=project_id,
        reporter_id=current_user.id,
        title_template=title_template,
        description=description,
        severity=severity,
        priority=priority,
        assignee_id=assignee_id,
        environment=environment,
        expected_result=expected_result,
        actual_result=actual_result
    )
    
    return jsonify({
        'success': True,
        'created_count': len(created_ids),
        'defect_ids': created_ids,
        'message': f'成功创建 {len(created_ids)} 个缺陷记录'
    })

@app.route('/api/defects/<int:defect_id>', methods=['GET'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_get_defect(defect_id):
    """获取缺陷详情"""
    _db = Database()
    defect = _db.get_defect(defect_id)
    if not defect:
        return jsonify({'success': False, 'error': '缺陷不存在'}), 404
    return jsonify({'success': True, 'defect': defect})

@app.route('/api/defects/<int:defect_id>', methods=['PUT'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('UPDATE_DEFECT', 'defect')
def api_update_defect(defect_id):
    """更新缺陷"""
    data = request.get_json(silent=True) or {}
    
    _db = Database()
    success = _db.update_defect(defect_id, current_user.id, **data)
    
    if success:
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '更新失败'}), 400

@app.route('/api/defects/<int:defect_id>', methods=['DELETE'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'project_manager')
@api_error_handler
@log_api_request
@audit_log('DELETE_DEFECT', 'defect')
def api_delete_defect(defect_id):
    """删除缺陷"""
    _db = Database()
    _db.delete_defect(defect_id)
    return jsonify({'success': True})

@app.route('/api/defects/<int:defect_id>/status', methods=['PUT'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('UPDATE_DEFECT_STATUS', 'defect')
def api_update_defect_status(defect_id):
    """更新缺陷状态（状态流转）"""
    data = request.get_json(silent=True) or {}
    new_status = data.get('status')
    
    if not new_status:
        return jsonify({'success': False, 'error': '状态不能为空'}), 400
    
    valid_statuses = ['open', 'in_progress', 'resolved', 'closed', 'reopened']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': f'无效状态，允许的状态: {valid_statuses}'}), 400
    
    _db = Database()
    
    # 先获取当前缺陷状态
    current_defect = _db.get_defect(defect_id)
    if not current_defect:
        return jsonify({'success': False, 'error': '缺陷不存在'}), 404
    
    success = _db.update_defect_status(defect_id, current_user.id, new_status)
    
    if success:
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '状态更新失败，可能是状态未发生变化'}), 400

@app.route('/api/defects/<int:defect_id>/comments', methods=['GET'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_get_defect_comments(defect_id):
    """获取缺陷评论"""
    _db = Database()
    comments = _db.get_defect_comments(defect_id)
    return jsonify({'success': True, 'comments': comments})

@app.route('/api/defects/<int:defect_id>/comments', methods=['POST'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_add_defect_comment(defect_id):
    """添加缺陷评论"""
    data = request.get_json(silent=True) or {}
    content = data.get('content', '').strip()
    
    if not content:
        return jsonify({'success': False, 'error': '评论内容不能为空'}), 400
    
    _db = Database()
    comment_id = _db.add_defect_comment(defect_id, current_user.id, content)
    return jsonify({'success': True, 'comment_id': comment_id})

@app.route('/api/defects/<int:defect_id>/history', methods=['GET'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_get_defect_history(defect_id):
    """获取缺陷状态变更历史"""
    _db = Database()
    history = _db.get_defect_history(defect_id)
    return jsonify({'success': True, 'history': history})

@app.route('/api/defects/from-failure', methods=['POST'])
@login_required
@feature_required('defect_management')
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('CREATE_DEFECT_FROM_FAILURE', 'defect')
def api_create_defect_from_failure():
    """从失败用例一键创建缺陷"""
    data = request.get_json(silent=True) or {}
    run_history_id = data.get('run_history_id')
    
    if not run_history_id:
        return jsonify({'success': False, 'error': '执行历史ID不能为空'}), 400
    
    _db = Database()
    defect_id = _db.create_defect_from_failure(
        run_history_id=run_history_id,
        reporter_id=current_user.id,
        title=data.get('title'),
        assignee_id=data.get('assignee_id')
    )
    
    if defect_id:
        return jsonify({'success': True, 'defect_id': defect_id})
    return jsonify({'success': False, 'error': '创建缺陷失败，可能执行历史不存在'}), 400

@app.route('/api/defects/statistics', methods=['GET'])
@login_required
@feature_required('defect_management')
@api_error_handler
@log_api_request
def api_get_defect_statistics():
    """获取缺陷统计数据"""
    project_id = request.args.get('project_id', type=int)
    _db = Database()
    stats = _db.get_defect_statistics(project_id)
    return jsonify({'success': True, 'statistics': stats})


# ==================== 用例导入导出 API ====================

@app.route('/api/cases/export/excel', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_export_cases_excel():
    """导出用例到Excel"""
    try:
        from case_importer import case_importer
        
        data = request.get_json(silent=True) or {}
        project_id = data.get('project_id')
        case_ids = data.get('case_ids', [])
        
        if not project_id and not case_ids:
            return jsonify({'success': False, 'error': '请指定项目ID或用例ID列表'}), 400
        
        filepath = case_importer.export_cases_to_excel(
            project_id=project_id,
            case_ids=case_ids if case_ids else None
        )
        
        return jsonify({
            'success': True,
            'filepath': filepath,
            'download_url': f'/api/download/{os.path.basename(filepath)}'
        })
    except ImportError as e:
        return jsonify({'success': False, 'error': '请先安装 openpyxl: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cases/export/json', methods=['POST'])
@login_required
@api_error_handler
@log_api_request
def api_export_cases_json():
    """导出用例到JSON"""
    try:
        from case_importer import case_importer
        
        data = request.get_json(silent=True) or {}
        project_id = data.get('project_id')
        case_ids = data.get('case_ids', [])
        
        if not project_id and not case_ids:
            return jsonify({'success': False, 'error': '请指定项目ID或用例ID列表'}), 400
        
        filepath = case_importer.export_cases_to_json(
            project_id=project_id,
            case_ids=case_ids if case_ids else None
        )
        
        return jsonify({
            'success': True,
            'filepath': filepath,
            'download_url': f'/api/download/{os.path.basename(filepath)}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cases/<int:case_id>/export/yaml-preview', methods=['GET'])
@login_required
@api_error_handler
@log_api_request
def api_export_case_yaml_preview(case_id: int):
    """Phase 5：用例 YAML 预览（人工审阅，非执行格式）。"""
    from case_yaml_export import build_case_yaml_preview

    _db = Database()
    case = _db.get_test_case_v2(case_id)
    if not case:
        return jsonify({'success': False, 'error': '用例不存在'}), 404
    if case.get('project_id') and not _db.check_project_access(current_user.id, case['project_id'], 'viewer'):
        return jsonify({'success': False, 'error': '无权限'}), 403
    yaml_text = build_case_yaml_preview(_db, case_id)
    if not yaml_text:
        return jsonify({'success': False, 'error': '无法生成预览'}), 500
    return jsonify({'success': True, 'yaml': yaml_text, 'case_id': case_id})


@app.route('/api/projects/<int:project_id>/export/yaml-preview', methods=['GET'])
@login_required
@project_access_required(min_role='viewer')
@api_error_handler
@log_api_request
def api_export_project_yaml_preview(project_id: int):
    """Phase 5：按项目/单元批量 YAML 预览。"""
    from case_yaml_export import build_project_yaml_preview

    unit_id = request.args.get('unit_id')
    yaml_text = build_project_yaml_preview(Database(), project_id, unit_id=unit_id)
    if yaml_text is None:
        return jsonify({'success': False, 'error': '无法生成预览'}), 404
    return jsonify({'success': True, 'yaml': yaml_text, 'project_id': project_id})


@app.route('/api/cases/import/excel', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('IMPORT_CASES', 'case')
def api_import_cases_excel():
    """从Excel导入用例"""
    try:
        from case_importer import case_importer
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请上传文件'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '请上传Excel文件(.xlsx或.xls)'}), 400
        
        project_id = request.form.get('project_id', type=int)
        
        # 保存上传的文件（Windows上需要先关闭临时文件再写入）
        import tempfile
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        file.save(tmp_path)
        
        try:
            results = case_importer.import_cases_from_excel(tmp_path, project_id)
            return jsonify({
                'success': True,
                'cases_created': results['cases_created'],
                'steps_created': results['steps_created'],
                'errors': results['errors']
            })
        finally:
            # 尝试删除临时文件，失败不影响返回结果
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
            
    except ImportError as e:
        return jsonify({'success': False, 'error': '请先安装 openpyxl: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cases/import/json', methods=['POST'])
@login_required
@role_required('admin', 'tester', 'project_manager', 'test_lead')
@api_error_handler
@log_api_request
@audit_log('IMPORT_CASES', 'case')
def api_import_cases_json():
    """从JSON导入用例"""
    try:
        from case_importer import case_importer
        
        data = request.get_json(silent=True) or {}
        project_id = data.get('project_id')
        json_data = data.get('data')
        
        if not project_id:
            return jsonify({'success': False, 'error': '请指定项目ID'}), 400
        if not json_data:
            return jsonify({'success': False, 'error': '请提供JSON数据'}), 400
        
        results = case_importer.import_cases_from_json(
            json_data=json.dumps(json_data) if isinstance(json_data, dict) else json_data,
            project_id=project_id
        )
        
        return jsonify({
            'success': True,
            'cases_created': results['cases_created'],
            'steps_created': results['steps_created'],
            'errors': results['errors']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cases/template/excel', methods=['GET'])
@login_required
def api_get_excel_template():
    """获取Excel导入模板"""
    try:
        from case_importer import case_importer
        filepath = case_importer.generate_excel_template()
        return jsonify({
            'success': True,
            'download_url': f'/api/download/{os.path.basename(filepath)}'
        })
    except ImportError:
        return jsonify({'success': False, 'error': '请先安装 openpyxl'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cases/template/json', methods=['GET'])
@login_required
def api_get_json_template():
    """获取JSON导入模板"""
    try:
        from case_importer import case_importer
        template = case_importer.generate_json_template()
        return jsonify({
            'success': True,
            'template': json.loads(template)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download/<filename>')
@login_required
def api_download_file(filename):
    """下载导出文件（限制在 exports 目录内，防止路径穿越）"""
    from flask import abort, send_from_directory
    from werkzeug.utils import secure_filename

    safe_name = secure_filename(os.path.basename(filename))
    if not safe_name:
        abort(404)
    export_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'exports'))
    candidate = os.path.abspath(os.path.join(export_dir, safe_name))
    try:
        if os.path.commonpath([export_dir, candidate]) != export_dir:
            abort(404)
    except ValueError:
        abort(404)
    if not os.path.isfile(candidate):
        abort(404)
    return send_from_directory(export_dir, safe_name, as_attachment=True)


def _run_case_worker_sync(case_id: int, user_id: int) -> dict:
    """Worker / 内部 API 同步执行用例（跳过 server 入队）。"""
    with app.test_request_context(
        f"/api/cases/{case_id}/run",
        method="POST",
        json={},
    ):
        from flask import g
        from flask_login import login_user

        g.force_local_run = True
        user_data = Database().get_user_by_id(user_id)
        if user_data:
            login_user(UserModel(user_data))
        try:
            resp = api_run_case(case_id)
            if isinstance(resp, tuple):
                rv, _code = resp
                if hasattr(rv, "get_json"):
                    return rv.get_json() or {}
                return {"success": False, "error": "empty response"}
            if hasattr(resp, "get_json"):
                return resp.get_json() or {}
            return {"success": False, "error": "unexpected response"}
        finally:
            g.force_local_run = False


# ---------- 组件管理 API ----------
import threading

_comp_install_lock = threading.Lock()
_comp_install_status = {}


@app.route("/api/components", methods=["GET"])
def api_components_list():
    """列出所有可选组件及其安装状态。"""
    try:
        from components_manager import list_components
        components = list_components()
        return jsonify({"success": True, "components": components})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/components/<component_id>/install", methods=["POST"])
def api_component_install(component_id: str):
    """安装指定组件（异步，后台执行）。"""
    try:
        from components_manager import is_installed, COMPONENT_DEFS, install

        if component_id not in COMPONENT_DEFS:
            return jsonify({"success": False, "error": f"未知组件: {component_id}"}), 400

        if is_installed(component_id):
            return jsonify({"success": True, "installed": True, "message": "组件已安装"})

        with _comp_install_lock:
            if _comp_install_status.get(component_id, {}).get("status") == "installing":
                return jsonify({"success": True, "status": "installing", "message": "正在安装中..."})

            _comp_install_status[component_id] = {
                "status": "installing",
                "percent": 0,
                "message": "准备安装...",
            }

        def _do_install():
            def _progress(status, percent, message):
                _comp_install_status[component_id] = {
                    "status": status,
                    "percent": percent,
                    "message": message,
                }

            try:
                ok = install(component_id, progress=_progress)
                if ok:
                    _comp_install_status[component_id] = {
                        "status": "done",
                        "percent": 100,
                        "message": "安装成功",
                    }
                else:
                    last_msg = _comp_install_status.get(component_id, {}).get("message", "安装失败")
                    _comp_install_status[component_id] = {
                        "status": "error",
                        "percent": 0,
                        "message": last_msg,
                    }
            except Exception as e:
                _comp_install_status[component_id] = {
                    "status": "error",
                    "percent": 0,
                    "message": f"安装异常: {e}",
                }

        t = threading.Thread(target=_do_install, daemon=True)
        t.start()

        return jsonify({"success": True, "status": "installing", "message": "开始安装"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/components/<component_id>/status", methods=["GET"])
def api_component_status(component_id: str):
    """查询组件安装进度状态。"""
    try:
        from components_manager import is_installed, COMPONENT_DEFS

        if component_id not in COMPONENT_DEFS:
            return jsonify({"success": False, "error": f"未知组件: {component_id}"}), 400

        status = _comp_install_status.get(component_id, {})
        installed = is_installed(component_id)
        return jsonify({
            "success": True,
            "installed": installed,
            "status": status.get("status", "idle" if not installed else "done"),
            "percent": status.get("percent", 100 if installed else 0),
            "message": status.get("message", ""),
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


from deployment_hooks import start_background_workers, wire_internal_runner  # noqa: E402

wire_internal_runner(app, _run_case_worker_sync)
start_background_workers()

try:
    from desktop_startup import mark_app_ready, schedule_deferred_gateway_boot

    if __import__("desktop_startup", fromlist=["desktop_lazy_gateway_boot"]).desktop_lazy_gateway_boot():
        schedule_deferred_gateway_boot()
    else:
        mark_app_ready()
except Exception:
    pass


if __name__ == '__main__':
    _host = os.environ.get('FLASK_RUN_HOST', '0.0.0.0')
    _debug = os.environ.get('FLASK_DEBUG', 'false').lower() in ('1', 'true', 'yes')
    _is_tauri = os.environ.get('TESTORY_TAURI_MODE', '0').strip() == '1'

    if _is_tauri:
        # 修复：Tauri 模式下 Flask 监听 0.0.0.0 + 固定端口（默认 5000），
        # 这样手机端可通过 PC 局域网 IP 访问 PC 端进行移动端同步。
        # 之前强制绑定 127.0.0.1 + 随机端口会导致手机无法连接。
        _bind_host = os.environ.get('FLASK_RUN_HOST', '0.0.0.0')
        if _bind_host in ('::', ''):
            _bind_host = '0.0.0.0'
        _host = _bind_host
        # 固定端口（与 src-tauri/src/flask_process.rs 的 FLASK_RUN_PORT=5000 保持一致）
        _port = int(os.environ.get('FLASK_RUN_PORT', '5000'))
        # 仍然写入端口文件，供 Rust 端 health-check 与 future-proof 使用
        _port_file = os.environ.get('TESTORY_FLASK_PORT_FILE', '').strip()
        if _port_file:
            from pathlib import Path

            _pf = Path(_port_file)
            _pf.parent.mkdir(parents=True, exist_ok=True)
            _pf.write_text(str(_port), encoding='utf-8')
    else:
        _port = int(os.environ.get('FLASK_RUN_PORT', '5000'))

    app.run(debug=_debug, host=_host, port=_port, threaded=True)




